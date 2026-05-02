"""dialogs/voucher_dialogs.py — 凭证录入对话框和辅助核算"""
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, QTimer, QSize, QPoint
from PySide6.QtGui import QColor, QFont, QPalette, QIcon, QPixmap, QPainter, QPen, QCursor

from db import get_db, seed_client_accounts, log_action, VOUCHER_TEMPLATES
from utils import (lbl, sep, card, fmt_amt, cn_amount,
                   NoScrollSpinBox, NoScrollDoubleSpinBox,
                   infer_account_type_direction)

# openpyxl imported lazily inside each export function


class HoverTipPopup(QFrame):
    def __init__(self, parent=None):
        super().__init__(
            parent,
            Qt.Tool | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.NoDropShadowWindowHint
        )
        self.setAttribute(Qt.WA_ShowWithoutActivating)
        self.setAttribute(Qt.WA_TransparentForMouseEvents)
        self.setStyleSheet(
            "QFrame{background:#ffffff;border:1px solid #3d6fdb;border-radius:6px;}"
        )
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 6, 10, 6)
        self.label = QLabel("")
        self.label.setStyleSheet(
            "QLabel{color:#2d5dc8;font-size:12px;font-weight:bold;border:none;background:transparent;}"
        )
        layout.addWidget(self.label)
        self.hide()


class HoverTipButton(QToolButton):
    _tip_popup = None

    @classmethod
    def _popup(cls):
        if cls._tip_popup is None:
            cls._tip_popup = HoverTipPopup()
        return cls._tip_popup

    def __init__(self):
        super().__init__()
        self.setMouseTracking(True)

    def _show_tip(self):
        tip = getattr(self, "_hover_tip_text", "")
        if not tip:
            return
        QToolTip.hideText()
        popup = self._popup()
        popup.label.setText(tip)
        popup.adjustSize()
        popup.move(QCursor.pos() + QPoint(8, 10))
        popup.show()
        popup.raise_()

    def enterEvent(self, event):
        self._show_tip()
        super().enterEvent(event)

    def mouseMoveEvent(self, event):
        popup = self._popup()
        if popup.isVisible():
            popup.move(QCursor.pos() + QPoint(8, 10))
        super().mouseMoveEvent(event)

    def leaveEvent(self, event):
        QToolTip.hideText()
        self._popup().hide()
        super().leaveEvent(event)

class VoucherDialog(QDialog):
    """新增/编辑凭证 — 智一会计风格"""
    def __init__(self, parent=None, client_id=None, period=None, voucher_id=None):
        super().__init__(parent)
        self.client_id = client_id; self.period = period; self.voucher_id = voucher_id
        self.setWindowTitle("编辑凭证" if voucher_id else "新增凭证")
        self.setMinimumSize(860, 580)
        self._accounts = self._fetch_accounts()
        self._templates = VOUCHER_TEMPLATES
        from PySide6.QtCore import QStringListModel
        self._account_completion_model = QStringListModel(
            [f"{a['code']}  {a['full_name']}" for a in self._accounts], self)
        self._row_completers = []
        self._build()
        if voucher_id: self._load_voucher()
        else: self._add_row(); self._add_row()

    def _fetch_accounts(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name,full_name,direction FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        r = [dict(x) for x in c.fetchall()]
        # Load aux config: account_code -> list of (dim_id, dim_name, items)
        c.execute("""SELECT ac.account_code, ad.id as dim_id, ad.name as dim_name
            FROM account_aux_config ac
            JOIN aux_dimensions ad ON ad.id=ac.dimension_id
            WHERE ac.client_id=? ORDER BY ac.account_code, ad.sort_order""",
                  (self.client_id,))
        self._aux_config = {}   # account_code -> [{dim_id, dim_name}]
        for row in c.fetchall():
            self._aux_config.setdefault(row["account_code"], []).append(
                {"dim_id": row["dim_id"], "dim_name": row["dim_name"]})
        # Load all aux items per dimension
        c.execute("SELECT id, dimension_id, name, code FROM aux_items WHERE client_id=? AND is_active=1 ORDER BY dimension_id,id",
                  (self.client_id,))
        self._aux_items = {}    # dim_id -> [{id, name, code}]
        for row in c.fetchall():
            self._aux_items.setdefault(row["dimension_id"], []).append(
                {"id": row["id"], "name": row["name"], "code": row["code"] or ""})
        # 建立非末级科目集合（有子科目的父科目，不允许直接录入凭证）
        all_codes = {a['code'] for a in r}
        self._parent_codes = set()
        for code in all_codes:
            parts = code.split('.')
            for depth in range(1, len(parts)):
                self._parent_codes.add('.'.join(parts[:depth]))
        conn.close(); return r

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(16,14,16,14); L.setSpacing(10)

        # Header bar
        hdr = QHBoxLayout()
        # Voucher number + date + preparer
        self.lbl_no = lbl("新 建", bold=True, color="#3d6fdb", size=14)
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True); self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.preparer_lbl = lbl("未来", color="#888")
        self.attach_spin = NoScrollSpinBox(); self.attach_spin.setRange(0,999)
        self.attach_spin.setSuffix(" 张"); self.attach_spin.setFixedWidth(70)
        hdr.addWidget(self.lbl_no); hdr.addSpacing(16)
        hdr.addWidget(lbl("日期：")); hdr.addWidget(self.date_edit)
        hdr.addWidget(lbl("  制单：")); hdr.addWidget(self.preparer_lbl)
        hdr.addWidget(lbl("  附单据")); hdr.addWidget(self.attach_spin)
        hdr.addStretch()
        # Template button
        tpl_btn = QPushButton("凭证模板 ▼"); tpl_btn.setObjectName("btn_outline")
        tpl_btn.clicked.connect(self._show_template_menu)
        b_save_new = QPushButton("保存并新增"); b_save_new.setObjectName("btn_primary")
        b_save_new.clicked.connect(lambda: self._save(and_new=True))
        b_save = QPushButton("保 存"); b_save.setObjectName("btn_primary")
        b_save.clicked.connect(lambda: self._save(and_new=False))
        hdr.addWidget(tpl_btn); hdr.addWidget(b_save_new); hdr.addWidget(b_save)
        L.addLayout(hdr)
        L.addWidget(sep())

        # Entry table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["摘  要","科  目","核算对象","借方金额","贷方金额"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(2,160); self.table.setColumnWidth(3,130); self.table.setColumnWidth(4,130)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setMinimumHeight(280)
        L.addWidget(self.table)

        # Add/del row buttons
        row_btns = QHBoxLayout()
        b_add = QPushButton("＋ 增行"); b_add.setObjectName("btn_outline")
        b_add.clicked.connect(lambda: self._add_row())
        b_del = QPushButton("－ 删行"); b_del.setObjectName("btn_gray")
        b_del.clicked.connect(lambda: self._del_row())
        row_btns.addWidget(b_add); row_btns.addWidget(b_del); row_btns.addStretch()
        L.addLayout(row_btns)
        L.addWidget(sep())

        # Total row
        tot = QHBoxLayout()
        tot.addWidget(lbl("合计金额：")); self.lbl_cn = lbl("零元整", bold=True, color="#1e2130")
        tot.addWidget(self.lbl_cn); tot.addStretch()
        self.lbl_debit = lbl("借方合计：0.00", color="#3d6fdb", bold=True)
        self.lbl_credit = lbl("贷方合计：0.00", color="#e05252", bold=True)
        self.lbl_balance = lbl("✓ 借贷平衡", color="#52c41a", bold=True)
        tot.addWidget(self.lbl_debit); tot.addSpacing(20)
        tot.addWidget(self.lbl_credit); tot.addSpacing(20)
        tot.addWidget(self.lbl_balance)
        L.addLayout(tot)

    def _add_row(self, summary="", acct_code="", acct_name="", debit=0, credit=0, aux_data=None):
        """aux_data: list of (dim_id, item_id, item_name) saved previously"""
        i = self.table.rowCount(); self.table.insertRow(i)
        self.table.setRowHeight(i, 44)

        summary_edit = QLineEdit(summary)
        summary_edit.setPlaceholderText("摘要...")

        # ── 科目搜索框（自定义 completer，支持编号/名称任意位置模糊匹配） ──
        acct_edit = QLineEdit()
        acct_edit.setPlaceholderText("输入编号或名称搜索...")
        acct_edit._code = acct_code  # store selected code

        # Build display list and lookup maps
        display_list = [f"{a['code']}  {a['full_name']}" for a in self._accounts]
        code_by_display = {f"{a['code']}  {a['full_name']}": a['code'] for a in self._accounts}
        display_by_code = {a['code']: f"{a['code']}  {a['full_name']}" for a in self._accounts}

        from PySide6.QtWidgets import QCompleter
        completer = QCompleter(self._account_completion_model, acct_edit)
        completer.setFilterMode(Qt.MatchContains)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.popup().setMinimumWidth(380)
        acct_edit.setCompleter(completer)
        acct_edit._completer = completer
        self._row_completers.append(completer)

        # Restore pre-filled value
        if acct_code and acct_code in display_by_code:
            acct_edit.setText(display_by_code[acct_code])

        def on_acct_selected(text):
            code = code_by_display.get(text, "")
            acct_edit._code = code
            rebuild_aux(code)

        def on_acct_edited(text):
            # If user cleared the field, clear code
            if not text:
                acct_edit._code = ""
                rebuild_aux("")
            # If exact match, update code
            if text in code_by_display:
                acct_edit._code = code_by_display[text]
                rebuild_aux(acct_edit._code)
            if text:
                completer.setCompletionPrefix(text)
                completer.complete()

        completer.activated.connect(on_acct_selected)
        acct_edit.textChanged.connect(on_acct_edited)

        # ── 辅助核算 ──
        aux_container = QWidget()
        aux_layout = QHBoxLayout(aux_container)
        aux_layout.setContentsMargins(2,2,2,2); aux_layout.setSpacing(4)
        aux_container._combos = []

        def rebuild_aux(code, saved=None):
            for w in aux_container._combos:
                w[1].setParent(None)
            aux_container._combos.clear()
            dims = self._aux_config.get(code, [])
            for dim in dims:
                cb = QComboBox(); cb.setMinimumWidth(130)
                cb.addItem("— 选核算对象 —", None)
                for it in self._aux_items.get(dim["dim_id"], []):
                    display = f"{it['code']} {it['name']}" if it['code'] else it['name']
                    cb.addItem(display, it["id"])
                cb.setToolTip(dim["dim_name"])
                if saved:
                    for sv_dim, sv_id, sv_name in saved:
                        if sv_dim == dim["dim_id"]:
                            for ki in range(cb.count()):
                                if cb.itemData(ki) == sv_id:
                                    cb.setCurrentIndex(ki); break
                aux_layout.addWidget(cb)
                aux_container._combos.append((dim["dim_id"], cb))
            if not dims:
                aux_layout.addWidget(QLabel("—"))

        rebuild_aux(acct_code, aux_data)

        # ── 借方金额（支持 = 键自动补平） ──
        d_spin = NoScrollDoubleSpinBox(); d_spin.setRange(0,999999999); d_spin.setDecimals(2)
        d_spin.setSpecialValueText(""); d_spin.setValue(debit)
        d_spin.valueChanged.connect(self._update_totals)

        orig_d_key = d_spin.keyPressEvent
        def d_key(event, _d=d_spin):
            if event.text() == "=":
                self._auto_balance_debit(_d)
            else:
                orig_d_key(event)
        d_spin.keyPressEvent = d_key

        # ── 贷方金额（支持 = 键自动补平） ──
        cr_spin = NoScrollDoubleSpinBox(); cr_spin.setRange(0,999999999); cr_spin.setDecimals(2)
        cr_spin.setSpecialValueText(""); cr_spin.setValue(credit)
        cr_spin.valueChanged.connect(self._update_totals)

        orig_cr_key = cr_spin.keyPressEvent
        def cr_key(event, _cr=cr_spin):
            if event.text() == "=":
                self._auto_balance_credit(_cr)
            else:
                orig_cr_key(event)
        cr_spin.keyPressEvent = cr_key

        self.table.setCellWidget(i,0,summary_edit)
        self.table.setCellWidget(i,1,acct_edit)
        self.table.setCellWidget(i,2,aux_container)
        self.table.setCellWidget(i,3,d_spin)
        self.table.setCellWidget(i,4,cr_spin)
        self._update_totals()

    def _auto_balance_debit(self, target_spin):
        """在借方按 = 键：将 target_spin 设为令借贷平衡所需的金额（以贷方合计为准）。"""
        tc = td_other = 0
        for i in range(self.table.rowCount()):
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            if cw: tc += cw.value()
            if dw and dw is not target_spin: td_other += dw.value()
        needed = max(0, round(tc - td_other, 2))
        target_spin.setValue(needed)
        self._update_totals()

    def _auto_balance_credit(self, target_spin):
        """在贷方按 = 键：将 target_spin 设为令借贷平衡所需的金额（以借方合计为准）。"""
        td = tc_other = 0
        for i in range(self.table.rowCount()):
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            if dw: td += dw.value()
            if cw and cw is not target_spin: tc_other += cw.value()
        needed = max(0, round(td - tc_other, 2))
        target_spin.setValue(needed)
        self._update_totals()

    def _del_row(self):
        row = self.table.currentRow()
        if row < 0:
            row = self.table.rowCount() - 1
        if row >= 0 and self.table.rowCount() > 1:
            self.table.removeRow(row)
            self._update_totals()

    def _update_totals(self):
        td = tc = 0
        for i in range(self.table.rowCount()):
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            if dw: td += dw.value()
            if cw: tc += cw.value()
        self.lbl_debit.setText(f"借方合计：{td:,.2f}")
        self.lbl_credit.setText(f"贷方合计：{tc:,.2f}")
        balanced = abs(td-tc) < 0.005
        self.lbl_balance.setText("✓ 借贷平衡" if balanced else f"✗ 差额 {abs(td-tc):,.2f}")
        self.lbl_balance.setStyleSheet(f"color:{'#52c41a' if balanced else '#ff4d4f'};font-weight:bold;")
        self.lbl_cn.setText(cn_amount(td))

    def _load_templates_from_db(self):
        """Load user-saved templates from database, merged with built-in templates."""
        conn = get_db(); c = conn.cursor()
        try:
            c.execute("SELECT name, entries FROM voucher_templates WHERE client_id=? OR client_id IS NULL ORDER BY id",
                      (self.client_id,))
            db_tmpls = c.fetchall()
        except Exception:
            db_tmpls = []
        conn.close()
        import json
        user_tmpls = []
        for row in db_tmpls:
            try:
                entries = json.loads(row["entries"])
                user_tmpls.append((row["name"], entries))
            except Exception:
                pass
        # Merge: built-in first, then user-saved
        self._templates = list(VOUCHER_TEMPLATES) + user_tmpls

    def _show_template_menu(self):
        self._load_templates_from_db()
        menu = QMenu(self)
        for name, _ in self._templates:
            menu.addAction(name)
        menu.addSeparator()
        b_save_tpl = menu.addAction("📌 存为模板...")
        b_del_tpl  = menu.addAction("🗑 删除模板...")
        act = menu.exec(self.sender().mapToGlobal(self.sender().rect().bottomLeft()))
        if not act: return

        if act is b_save_tpl:
            self._save_as_template()
            return
        if act is b_del_tpl:
            self._delete_template()
            return

        # Apply selected template
        for name, entries in self._templates:
            if act.text() == name:
                self.table.setRowCount(0)
                for entry in entries:
                    # Support both old tuple format and new dict format
                    if isinstance(entry, dict):
                        self._add_row(entry.get("summary",""), entry.get("code",""),
                                      entry.get("name",""), entry.get("debit",0), entry.get("credit",0))
                    else:
                        s, code, _, d, cr = entry
                        self._add_row(s, code, "", d, cr)
                break

    def _save_as_template(self):
        """Save current voucher entries as a reusable template."""
        import json
        # Collect current entries
        entries = []
        for i in range(self.table.rowCount()):
            sw = self.table.cellWidget(i,0); aw = self.table.cellWidget(i,1)
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            code = getattr(aw, '_code', "") or "" if aw else ""
            summary = sw.text().strip() if sw else ""
            d = dw.value() if dw else 0
            cr = cw.value() if cw else 0
            if not code and d == 0 and cr == 0: continue
            # Find account name
            aname = ""
            for a in self._accounts:
                if a['code'] == code: aname = a['full_name']; break
            entries.append({"summary": summary, "code": code, "name": aname,
                            "debit": d, "credit": cr})
        if not entries:
            QMessageBox.warning(self, "提示", "请先填写凭证分录再保存为模板"); return

        name, ok = QInputDialog.getText(self, "保存为模板", "模板名称：",
                                         text=entries[0].get("summary","") or "新模板")
        if not ok or not name.strip(): return
        name = name.strip()

        # Check duplicate name
        all_names = [t[0] for t in self._templates]
        if name in all_names:
            if QMessageBox.question(self, "覆盖确认", f"模板【{name}】已存在，是否覆盖？",
                    QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return
            conn = get_db()
            conn.execute("DELETE FROM voucher_templates WHERE name=? AND (client_id=? OR client_id IS NULL)",
                         (name, self.client_id))
            conn.commit(); conn.close()

        conn = get_db()
        conn.execute("INSERT INTO voucher_templates(client_id, name, entries) VALUES(?,?,?)",
                     (self.client_id, name, json.dumps(entries, ensure_ascii=False)))
        conn.commit(); conn.close()
        QMessageBox.information(self, "成功", f"已保存模板：{name}")

    def _delete_template(self):
        """Delete a user-saved template."""
        conn = get_db(); c = conn.cursor()
        try:
            c.execute("SELECT id, name FROM voucher_templates WHERE client_id=? OR client_id IS NULL ORDER BY id",
                      (self.client_id,))
            user_tmpls = c.fetchall()
        except Exception:
            user_tmpls = []
        conn.close()
        if not user_tmpls:
            QMessageBox.information(self, "提示", "没有可删除的自定义模板（内置模板不能删除）")
            return
        names = [r["name"] for r in user_tmpls]
        name, ok = QInputDialog.getItem(self, "删除模板", "选择要删除的模板：", names, editable=False)
        if not ok or not name: return
        conn = get_db()
        conn.execute("DELETE FROM voucher_templates WHERE name=? AND (client_id=? OR client_id IS NULL)",
                     (name, self.client_id))
        conn.commit(); conn.close()
        QMessageBox.information(self, "成功", f"已删除模板：{name}")

    def _load_voucher(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM vouchers WHERE id=?", (self.voucher_id,))
        v = c.fetchone()
        self.lbl_no.setText(v["voucher_no"])
        self.date_edit.setDate(QDate.fromString(v["date"],"yyyy-MM-dd"))
        self.attach_spin.setValue(v["attachment_count"] or 0)
        c.execute("SELECT * FROM voucher_entries WHERE voucher_id=? ORDER BY line_no", (self.voucher_id,))
        entries = c.fetchall()
        # Load aux data per entry
        entry_ids = [e["id"] for e in entries]
        aux_by_entry = {}
        if entry_ids:
            placeholders = ",".join("?" * len(entry_ids))
            c.execute(f"SELECT entry_id,dimension_id,aux_item_id,aux_item_name FROM voucher_entry_aux WHERE entry_id IN ({placeholders})",
                      entry_ids)
            for row in c.fetchall():
                aux_by_entry.setdefault(row["entry_id"], []).append(
                    (row["dimension_id"], row["aux_item_id"], row["aux_item_name"]))
        conn.close()
        for e in entries:
            aux_data = aux_by_entry.get(e["id"], None)
            self._add_row(e["summary"] or "", e["account_code"] or "", e["account_name"] or "",
                          e["debit"] or 0, e["credit"] or 0, aux_data=aux_data)

    def _save(self, and_new=False):
        # Validate
        entries = []
        for i in range(self.table.rowCount()):
            sw  = self.table.cellWidget(i,0); aw  = self.table.cellWidget(i,1)
            auxw= self.table.cellWidget(i,2)
            dw  = self.table.cellWidget(i,3); cw  = self.table.cellWidget(i,4)
            if not aw: continue
            code = getattr(aw, '_code', "") or ""
            d = dw.value() if dw else 0; cr = cw.value() if cw else 0
            if not code and d == 0 and cr == 0: continue
            if not code: QMessageBox.warning(self,"提示",f"第{i+1}行科目不能为空"); return
            aname = ""
            for a in self._accounts:
                if a['code'] == code: aname = a['full_name']; break
            # Collect aux selections
            aux_sel = []   # (dim_id, item_id, item_name)
            if auxw:
                for dim_id, cb in auxw._combos:
                    item_id = cb.currentData()
                    item_name = cb.currentText() if item_id else ""
                    if item_id:
                        aux_sel.append((dim_id, item_id, item_name))
            entries.append((i, sw.text().strip() if sw else "", code, aname, d, cr, aux_sel))

        if not entries: QMessageBox.warning(self,"提示","请至少填写一行分录"); return

        # ── 校验一：同行借贷方不能同时有值 ──
        for i, sw, code, aname, d, cr, aux_sel in entries:
            if d > 0.005 and cr > 0.005:
                QMessageBox.warning(self, "录入错误",
                    f"第{i+1}行【{code} {aname}】借方和贷方不能同时有金额。\n"
                    "请将其拆分为两行，或清空其中一方。")
                return

        # ── 校验二：只允许录入末级科目 ──
        parent_codes = getattr(self, '_parent_codes', set())
        for i, sw, code, aname, d, cr, aux_sel in entries:
            if code in parent_codes:
                # Find any child code to show as hint
                children = [a['code'] for a in self._accounts
                            if a['code'].startswith(code + '.')]
                hint = f"\n例如应使用子科目：{children[0]}" if children else ""
                QMessageBox.warning(self, "科目层级错误",
                    f"第{i+1}行【{code} {aname}】下设有子科目，不能直接使用父科目录凭证。"
                    f"请选择最末级科目。{hint}")
                return

        # ── 校验三：绑定辅助核算的科目必须选择核算对象 ──
        for i, sw, code, aname, d, cr, aux_sel in entries:
            required_dims = self._aux_config.get(code, [])
            if not required_dims:
                continue
            selected_dim_ids = {dim_id for dim_id, _, _ in aux_sel}
            missing = [dim['dim_name'] for dim in required_dims
                       if dim['dim_id'] not in selected_dim_ids]
            if missing:
                QMessageBox.warning(self, "辅助核算未填写",
                    f"第{i+1}行【{code} {aname}】已绑定辅助核算，"
                    f"以下核算对象必须填写：\n\n" + "\n".join(f"· {m}" for m in missing))
                return

        td = sum(e[4] for e in entries); tc = sum(e[5] for e in entries)
        if abs(td-tc) > 0.005:
            QMessageBox.warning(self,"借贷不平",f"借方合计 {td:.2f} ≠ 贷方合计 {tc:.2f}\n请检查金额"); return

        conn = get_db(); c = conn.cursor()
        dt = self.date_edit.date().toString("yyyy-MM-dd")
        reverted_to_pending = False

        if self.voucher_id:
            c.execute("SELECT voucher_no, status FROM vouchers WHERE id=?", (self.voucher_id,))
            voucher = c.fetchone()
            if not voucher:
                conn.close()
                QMessageBox.warning(self, "提示", "凭证不存在或已被删除，请刷新后重试")
                return
            vno = voucher["voucher_no"]
            new_status = voucher["status"]
            if voucher["status"] == "已审核":
                new_status = "待审核"
                reverted_to_pending = True
            c.execute("UPDATE vouchers SET date=?,attachment_count=?,status=? WHERE id=?",
                      (dt, self.attach_spin.value(), new_status, self.voucher_id))
            c.execute("DELETE FROM voucher_entries WHERE voucher_id=?", (self.voucher_id,))
            vid = self.voucher_id
        else:
            c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=?",
                      (self.client_id, self.period))
            n = c.fetchone()[0] + 1
            vno = f"记-{n:03d}"
            c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,attachment_count) VALUES(?,?,?,?,?)",
                      (self.client_id, self.period, vno, dt, self.attach_spin.value()))
            vid = c.lastrowid

        for ln, summary, code, aname, d, cr, aux_sel in entries:
            c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                      (vid, ln, summary, code, aname, d, cr))
            entry_id = c.lastrowid
            for dim_id, item_id, item_name in aux_sel:
                c.execute("INSERT INTO voucher_entry_aux(entry_id,dimension_id,aux_item_id,aux_item_name) VALUES(?,?,?,?)",
                          (entry_id, dim_id, item_id, item_name))
        action = "编辑凭证" if self.voucher_id else "新增凭证"
        detail = f"凭证号:{vno} 借方合计:{td:.2f}"
        if reverted_to_pending:
            detail += " 修改后自动回退为待审核"
        log_action(conn, self.client_id,
                   action, "voucher", vid, detail)
        conn.commit(); conn.close()
        self.saved_and_new = and_new
        if reverted_to_pending:
            QMessageBox.information(self, "已回退待审核", "该凭证原状态为“已审核”，修改后已自动回退为“待审核”，请重新审核。")
        self.accept()

# ── Pages ──────────────────────────────────────────────────────────────────

# ── 辅助核算管理页（嵌入 VoucherPage 的 Tab） ─────────────────────────────

class AuxItemDialog(QDialog):
    """新增/编辑 核算对象"""
    def __init__(self, parent, client_id, dimension_id, item=None):
        super().__init__(parent)
        self.client_id = client_id
        self.dimension_id = dimension_id
        self.item = item
        self.setWindowTitle("编辑核算对象" if item else "新增核算对象")
        self.setMinimumWidth(360)
        self._build()
        if item: self._load()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(22,20,22,20); L.setSpacing(12)
        L.addWidget(lbl("核算对象信息", bold=True, size=14))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.f_name    = QLineEdit(); self.f_name.setPlaceholderText("名称（必填）")
        self.f_code    = QLineEdit(); self.f_code.setPlaceholderText("编码（可选）")
        self.f_contact = QLineEdit(); self.f_contact.setPlaceholderText("联系人")
        self.f_phone   = QLineEdit(); self.f_phone.setPlaceholderText("电话")
        F.addRow("名称 *",  self.f_name)
        F.addRow("编码",    self.f_code)
        F.addRow("联系人",  self.f_contact)
        F.addRow("电话",    self.f_phone)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        bc = QPushButton("取消"); bc.setObjectName("btn_gray")
        bs = QPushButton("保存"); bs.setObjectName("btn_primary")
        bc.clicked.connect(self.reject); bs.clicked.connect(self._save)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        self.f_name.setText(self.item["name"] or "")
        self.f_code.setText(self.item["code"] or "")
        self.f_contact.setText(self.item["contact"] or "")
        self.f_phone.setText(self.item["phone"] or "")

    def _save(self):
        name = self.f_name.text().strip()
        if not name: QMessageBox.warning(self, "提示", "名称不能为空"); return
        conn = get_db(); c = conn.cursor()
        d = (name, self.f_code.text().strip(), self.f_contact.text().strip(),
             self.f_phone.text().strip())
        if self.item:
            c.execute("UPDATE aux_items SET name=?,code=?,contact=?,phone=? WHERE id=?",
                      d + (self.item["id"],))
        else:
            c.execute("INSERT INTO aux_items(client_id,dimension_id,name,code,contact,phone)"
                      " VALUES(?,?,?,?,?,?)",
                      (self.client_id, self.dimension_id) + d)
        conn.commit(); conn.close(); self.accept()


class AuxPage(QWidget):
    """辅助核算管理：维度 + 对象 + 科目绑定"""
    def __init__(self):
        super().__init__()
        self.client_id = None
        self._cur_dim_id = None
        L = QHBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # ── 左栏：维度列表 ──
        left = QWidget(); left.setFixedWidth(200)
        left.setStyleSheet("background:#f7f9fc; border-right:1px solid #e8ecf2;")
        ll = QVBoxLayout(left); ll.setContentsMargins(0,0,0,0); ll.setSpacing(0)
        hdr_l = QWidget(); hdr_l.setStyleSheet("background:#fff; border-bottom:1px solid #eee;")
        hl = QHBoxLayout(hdr_l); hl.setContentsMargins(12,10,8,10)
        hl.addWidget(lbl("核算维度", bold=True)); hl.addStretch()
        b_adddim = self._make_icon_btn(
            self._theme_icon("list-add", QStyle.SP_FileDialogNewFolder),
            "新增维度", primary=True, size=28)
        b_adddim.clicked.connect(self._add_dim)
        hl.addWidget(b_adddim); ll.addWidget(hdr_l)

        self.dim_list = QListWidget()
        self.dim_list.setStyleSheet(
            "QListWidget{border:none;background:#f7f9fc;}"
            "QListWidget::item{padding:10px 14px;border-bottom:1px solid #eef0f4;}"
            "QListWidget::item:selected{background:#e6f0ff;color:#3d6fdb;font-weight:bold;}")
        self.dim_list.currentRowChanged.connect(self._on_dim_changed)
        ll.addWidget(self.dim_list)

        # 维度右键菜单
        self.dim_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.dim_list.customContextMenuRequested.connect(self._dim_context_menu)
        L.addWidget(left)

        # ── 右栏：对象列表 + 科目绑定 ──
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(0,0,0,0); rl.setSpacing(0)

        # Right-side tab: 对象管理 | 往来对账
        self.right_tabs = QTabWidget()
        self.right_tabs.setStyleSheet(
            "QTabBar::tab{padding:8px 20px;color:#888;border:none;background:transparent;"
            "border-bottom:2px solid transparent;}"
            "QTabBar::tab:selected{color:#3d6fdb;border-bottom:2px solid #3d6fdb;}"
            "QTabWidget::pane{border:none;}")

        # ── Tab A: 核算对象管理 ──
        tab_mgr = QWidget(); tl = QVBoxLayout(tab_mgr); tl.setContentsMargins(20,14,20,14); tl.setSpacing(10)

        hdr_r = QHBoxLayout()
        self.dim_title = lbl("请选择左侧维度", bold=True, size=15)
        hdr_r.addWidget(self.dim_title); hdr_r.addStretch()
        self.b_additem = self._make_icon_btn(self._glyph_icon("add", "#ffffff"), "新增对象", primary=True)
        self.b_additem.clicked.connect(self._add_item)
        b_exp = self._make_icon_btn(self._glyph_icon("export"), "导出对象列表")
        b_exp.clicked.connect(self._export_items)
        hdr_r.addWidget(b_exp); hdr_r.addWidget(self.b_additem)
        tl.addLayout(hdr_r)

        f1 = card(); v1 = QVBoxLayout(f1); v1.setContentsMargins(0,0,0,0)
        self.item_tbl = QTableWidget(); self.item_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.item_tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.item_tbl.setShowGrid(False); self.item_tbl.verticalHeader().setVisible(False)
        self.item_tbl.setColumnCount(5)
        self.item_tbl.setHorizontalHeaderLabels(["编码","名称","联系人","电话","操作"])
        hh = self.item_tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        self.item_tbl.setColumnWidth(0,80); self.item_tbl.setColumnWidth(2,90)
        self.item_tbl.setColumnWidth(3,110); self.item_tbl.setColumnWidth(4,160)
        v1.addWidget(self.item_tbl); tl.addWidget(f1)

        # 科目绑定区
        tl.addWidget(lbl("绑定科目（凭证录入时，选中这些科目将显示本维度的对象选择）",
                         color="#666", size=12))
        f2 = card(); v2 = QVBoxLayout(f2); v2.setContentsMargins(12,10,12,10); v2.setSpacing(8)
        bind_hdr = QHBoxLayout()
        bind_hdr.addWidget(lbl("已绑定科目", bold=True)); bind_hdr.addStretch()
        b_bind = self._make_icon_btn(self._glyph_icon("bind"), "绑定科目")
        b_bind.clicked.connect(self._bind_account)
        bind_hdr.addWidget(b_bind); v2.addLayout(bind_hdr)
        self.bind_list = QListWidget()
        self.bind_list.setStyleSheet(
            "QListWidget{border:1px solid #eee;border-radius:5px;background:#fff;}"
            "QListWidget::item{padding:6px 10px;}"
            "QListWidget::item:selected{background:#e6f0ff;}")
        self.bind_list.setMaximumHeight(120)
        self.bind_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.bind_list.customContextMenuRequested.connect(self._unbind_context_menu)
        v2.addWidget(self.bind_list)
        tl.addWidget(f2)
        self.right_tabs.addTab(tab_mgr, "核算对象管理")

        # ── Tab B: 往来对账报表 ──
        tab_rpt = QWidget(); tr = QVBoxLayout(tab_rpt); tr.setContentsMargins(20,14,20,14); tr.setSpacing(10)

        rpt_hdr = QHBoxLayout()
        rpt_hdr.addWidget(lbl("往来对账表", bold=True, size=15)); rpt_hdr.addStretch()
        b_rpt_exp = self._make_icon_btn(self._glyph_icon("export"), "导出往来对账")
        b_rpt_exp.clicked.connect(self._export_aux_report)
        rpt_hdr.addWidget(b_rpt_exp); tr.addLayout(rpt_hdr)

        fr = QHBoxLayout(); fr.setSpacing(8)
        fr.addWidget(lbl("维度:"))
        self.rpt_dim_combo = QComboBox(); self.rpt_dim_combo.setMinimumWidth(120)
        fr.addWidget(self.rpt_dim_combo)
        fr.addWidget(lbl("期间:"))
        self.rpt_period_edit = QLineEdit(); self.rpt_period_edit.setFixedWidth(100)
        self.rpt_period_edit.setPlaceholderText("如 2026-03")
        fr.addWidget(self.rpt_period_edit)
        fr.addWidget(lbl("科目:"))
        self.rpt_acct_combo = QComboBox(); self.rpt_acct_combo.setMinimumWidth(200)
        self.rpt_acct_combo.addItem("全部科目", "")
        fr.addWidget(self.rpt_acct_combo)
        b_q = self._make_icon_btn(self._glyph_icon("query", "#ffffff"), "查询", primary=True)
        b_q.clicked.connect(self._load_aux_report)
        fr.addWidget(b_q); fr.addStretch()
        tr.addLayout(fr)

        f3 = card(); v3 = QVBoxLayout(f3); v3.setContentsMargins(0,0,0,0)
        self.aux_rpt_tbl = QTableWidget(); self.aux_rpt_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.aux_rpt_tbl.setShowGrid(True); self.aux_rpt_tbl.verticalHeader().setVisible(False)
        self.aux_rpt_tbl.setColumnCount(6)
        self.aux_rpt_tbl.setHorizontalHeaderLabels(["核算对象","科目","期初余额","本期借方","本期贷方","期末余额"])
        hh3 = self.aux_rpt_tbl.horizontalHeader()
        hh3.setSectionResizeMode(QHeaderView.Interactive)
        hh3.setSectionResizeMode(0, QHeaderView.Stretch)
        hh3.setSectionResizeMode(1, QHeaderView.Stretch)
        for ci in range(2,6): self.aux_rpt_tbl.setColumnWidth(ci, 110)
        v3.addWidget(self.aux_rpt_tbl); tr.addWidget(f3)
        self.right_tabs.addTab(tab_rpt, "往来对账")

        rl.addWidget(self.right_tabs)
        L.addWidget(right)

    def _theme_icon(self, primary, fallback):
        icon = QIcon.fromTheme(primary)
        if icon.isNull():
            icon = self.style().standardIcon(fallback)
        return icon

    def _glyph_icon(self, kind, color="#3d6fdb"):
        size = 18
        pm = QPixmap(size, size)
        pm.fill(Qt.transparent)
        p = QPainter(pm)
        p.setRenderHint(QPainter.Antialiasing)
        pen = QPen(QColor(color))
        pen.setWidth(2)
        pen.setCapStyle(Qt.RoundCap)
        pen.setJoinStyle(Qt.RoundJoin)
        p.setPen(pen)

        if kind == "add":
            p.drawLine(9, 3, 9, 15)
            p.drawLine(3, 9, 15, 9)
        elif kind == "edit":
            p.drawLine(4, 13, 13, 4)
            p.drawLine(11, 4, 13, 6)
            p.drawLine(4, 11, 6, 13)
            p.drawLine(3, 15, 7, 15)
        elif kind == "delete":
            p.drawLine(4, 4, 14, 14)
            p.drawLine(14, 4, 4, 14)
        elif kind == "export":
            p.drawLine(9, 3, 9, 11)
            p.drawLine(6, 8, 9, 11)
            p.drawLine(12, 8, 9, 11)
            p.drawLine(4, 15, 14, 15)
        elif kind == "bind":
            p.drawArc(2, 6, 7, 6, 270 * 16, 180 * 16)
            p.drawArc(9, 6, 7, 6, 90 * 16, 180 * 16)
            p.drawLine(7, 9, 11, 9)
        elif kind == "query":
            p.drawEllipse(3, 3, 8, 8)
            p.drawLine(10, 10, 14, 14)

        p.end()
        return QIcon(pm)

    def _make_icon_btn(self, icon, tooltip, primary=False, danger=False, size=30):
        btn = HoverTipButton()
        btn.setFixedSize(size, size)
        btn.setIcon(icon)
        btn.setIconSize(QSize(18, 18))
        btn.setToolButtonStyle(Qt.ToolButtonIconOnly)
        if primary:
            btn.setStyleSheet(
                "QToolButton{background:#3d6fdb;color:#fff;border:none;border-radius:6px;}"
                "QToolButton:hover{background:#2d5dc8;}"
            )
        elif danger:
            btn.setStyleSheet(
                "QToolButton{background:#ff4d4f;color:#fff;border:none;border-radius:6px;}"
                "QToolButton:hover{background:#e63b3d;}"
            )
        else:
            btn.setStyleSheet(
                "QToolButton{background:transparent;color:#3d6fdb;border:1px solid #3d6fdb;border-radius:6px;}"
                "QToolButton:hover{background:#eef3ff;}"
            )
        btn._hover_tip_text = tooltip
        btn.setToolTip("")
        return btn

    def set_client(self, client_id, period=""):
        self.client_id = client_id
        self._period = period
        self._cur_dim_id = None
        self._dims = []
        self._items = []
        self._bindings = []
        self._load_dims()
        self._refresh_rpt_combos()

    def ensure_dimension(self, dim_name):
        """Create the dimension when missing and return its id."""
        if not self.client_id:
            raise ValueError("请先选择客户")
        name = (dim_name or "").strip()
        if not name:
            raise ValueError("维度名称不能为空")

        conn = get_db(); c = conn.cursor()
        c.execute("SELECT id FROM aux_dimensions WHERE client_id=? AND name=?",
                  (self.client_id, name))
        row = c.fetchone()
        if row:
            dim_id = row["id"]
        else:
            c.execute("INSERT INTO aux_dimensions(client_id,name) VALUES(?,?)",
                      (self.client_id, name))
            conn.commit()
            dim_id = c.lastrowid
        conn.close()
        self._load_dims()
        self._refresh_rpt_combos()
        return dim_id

    def bind_account_dimension(self, account_code, dimension_id):
        """Bind an account to a dimension if it is not already bound."""
        if not self.client_id:
            raise ValueError("请先选择客户")
        conn = get_db()
        conn.execute("""INSERT OR IGNORE INTO account_aux_config(client_id,account_code,dimension_id)
                        VALUES(?,?,?)""",
                     (self.client_id, account_code, dimension_id))
        conn.commit()
        conn.close()
        if self._cur_dim_id == dimension_id:
            self._load_bindings()

    def focus_dimension(self, dimension_id):
        """Select a dimension row by id."""
        for idx, dim in enumerate(getattr(self, "_dims", [])):
            if dim["id"] == dimension_id:
                self.dim_list.setCurrentRow(idx)
                self.right_tabs.setCurrentIndex(0)
                return

    def _refresh_rpt_combos(self):
        if not self.client_id: return
        # Period
        from datetime import datetime
        now = datetime.now()
        if self._period:
            self.rpt_period_edit.setText(self._period)
        else:
            self.rpt_period_edit.setText(f"{now.year}-{now.month:02d}")
        # Dimensions
        self.rpt_dim_combo.clear()
        for d in self._dims:
            self.rpt_dim_combo.addItem(d["name"], d["id"])
        # Accounts
        self.rpt_acct_combo.clear(); self.rpt_acct_combo.addItem("全部科目", "")
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code", (self.client_id,))
        for a in c.fetchall():
            self.rpt_acct_combo.addItem(f"{a['code']} {a['name']}", a['code'])
        conn.close()

    def _load_dims(self):
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM aux_dimensions WHERE client_id=? ORDER BY sort_order,id",
                  (self.client_id,))
        dims = c.fetchall(); conn.close()
        self.dim_list.clear()
        self._dims = [dict(d) for d in dims]
        for d in self._dims:
            self.dim_list.addItem(d["name"])
        if self._dims:
            self.dim_list.setCurrentRow(0)

    def _on_dim_changed(self, row):
        if row < 0 or row >= len(self._dims): return
        self._cur_dim_id = self._dims[row]["id"]
        self.dim_title.setText(f"【{self._dims[row]['name']}】核算对象")
        self._load_items()
        self._load_bindings()

    def _add_dim(self):
        if not self.client_id: return
        name, ok = QInputDialog.getText(self, "新增维度", "维度名称（如：客户、员工、项目）：")
        if not ok or not name.strip(): return
        conn = get_db()
        try:
            conn.execute("INSERT INTO aux_dimensions(client_id,name) VALUES(?,?)",
                         (self.client_id, name.strip()))
            conn.commit()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"维度已存在或保存失败：{e}")
        finally:
            conn.close()
        self._load_dims()

    def _dim_context_menu(self, pos):
        row = self.dim_list.currentRow()
        if row < 0: return
        menu = QMenu(self)
        menu.addAction("重命名").triggered.connect(lambda: self._rename_dim(row))
        menu.addAction("删除维度").triggered.connect(lambda: self._del_dim(row))
        menu.exec(self.dim_list.mapToGlobal(pos))

    def _rename_dim(self, row):
        old = self._dims[row]["name"]
        name, ok = QInputDialog.getText(self, "重命名维度", "新名称：", text=old)
        if not ok or not name.strip(): return
        conn = get_db()
        conn.execute("UPDATE aux_dimensions SET name=? WHERE id=?",
                     (name.strip(), self._dims[row]["id"]))
        conn.commit(); conn.close(); self._load_dims()

    def _del_dim(self, row):
        if QMessageBox.question(self, "确认",
                f"删除维度【{self._dims[row]['name']}】及其所有对象和绑定？",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db(); did = self._dims[row]["id"]
        conn.execute("DELETE FROM voucher_entry_aux WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM account_aux_config WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM aux_items WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM aux_dimensions WHERE id=?", (did,))
        conn.commit(); conn.close(); self._load_dims()

    def _load_items(self):
        if not self._cur_dim_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM aux_items WHERE dimension_id=? AND client_id=? ORDER BY id",
                  (self._cur_dim_id, self.client_id))
        rows = c.fetchall(); conn.close()
        self._items = [dict(r) for r in rows]
        self.item_tbl.setRowCount(len(rows))
        for i, r in enumerate(self._items):
            self.item_tbl.setRowHeight(i, 40)
            for j, v in enumerate([r["code"] or "", r["name"], r["contact"] or "", r["phone"] or ""]):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter if j != 1 else Qt.AlignLeft | Qt.AlignVCenter)
                self.item_tbl.setItem(i, j, it)
            bw = QWidget(); bl = QHBoxLayout(bw); bl.setContentsMargins(4,3,4,3); bl.setSpacing(4)
            b_ed = self._make_icon_btn(
                self._glyph_icon("edit"),
                "编辑对象", size=30)
            b_ed.clicked.connect(lambda _, rr=r: self._edit_item(rr))
            b_dl = self._make_icon_btn(
                self._glyph_icon("delete", "#ffffff"),
                "删除对象", danger=True, size=30)
            b_dl.clicked.connect(lambda _, rid=r["id"]: self._del_item(rid))
            bl.addWidget(b_ed); bl.addWidget(b_dl); bl.addStretch()
            self.item_tbl.setCellWidget(i, 4, bw)

    def _add_item(self):
        if not self._cur_dim_id:
            QMessageBox.information(self, "提示", "请先选择左侧维度"); return
        d = AuxItemDialog(self, self.client_id, self._cur_dim_id)
        if d.exec(): self._load_items()

    def _edit_item(self, r):
        d = AuxItemDialog(self, self.client_id, self._cur_dim_id, item=r)
        if d.exec(): self._load_items()

    def _del_item(self, item_id):
        if QMessageBox.question(self, "确认", "删除该核算对象？",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db()
        conn.execute("DELETE FROM aux_items WHERE id=?", (item_id,))
        conn.commit(); conn.close(); self._load_items()

    def _load_aux_report(self):
        if not self.client_id: return
        dim_id = self.rpt_dim_combo.currentData()
        period = self.rpt_period_edit.text().strip()
        acct_filter = self.rpt_acct_combo.currentData() or ""
        if not dim_id or not period:
            QMessageBox.information(self, "提示", "请选择维度并填写期间"); return
        conn = get_db(); c = conn.cursor()
        # Get account direction info for opening balance computation
        c.execute("SELECT code,direction,opening_debit,opening_credit FROM accounts WHERE client_id=?",
                  (self.client_id,))
        acct_info = {r["code"]: r for r in c.fetchall()}
        sql = """
            SELECT ai.name AS item_name, ai.id AS item_id,
                   e.account_code, e.account_name,
                   SUM(e.debit) AS td, SUM(e.credit) AS tc
            FROM voucher_entry_aux ea
            JOIN aux_items ai ON ai.id=ea.aux_item_id
            JOIN voucher_entries e ON e.id=ea.entry_id
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE ea.dimension_id=? AND v.client_id=? AND v.period=?
        """
        params = [dim_id, self.client_id, period]
        if acct_filter:
            sql += " AND e.account_code=?"; params.append(acct_filter)
        sql += " GROUP BY ai.id, e.account_code ORDER BY ai.name, e.account_code"
        c.execute(sql, params)
        rows = c.fetchall(); conn.close()

        self.aux_rpt_tbl.setRowCount(len(rows) + 1)
        td_tot = tc_tot = 0
        for i, r in enumerate(rows):
            self.aux_rpt_tbl.setRowHeight(i, 36)
            td = r["td"] or 0; tc = r["tc"] or 0
            ai = acct_info.get(r["account_code"])
            od = (ai["opening_debit"] or 0) - (ai["opening_credit"] or 0) if ai else 0
            if ai and ai["direction"] == "贷": od = -od
            ending = od + td - tc
            td_tot += td; tc_tot += tc
            vals = [r["item_name"], f"{r['account_code']} {r['account_name']}",
                    fmt_amt(od), fmt_amt(td), fmt_amt(tc), fmt_amt(ending)]
            for j, v in enumerate(vals):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter if j<=1 else Qt.AlignRight|Qt.AlignVCenter)
                if j == 3 and td: it.setForeground(QColor("#3d6fdb"))
                if j == 4 and tc: it.setForeground(QColor("#e05252"))
                if j == 5:
                    it.setForeground(QColor("#3d6fdb") if ending > 0 else
                                     QColor("#e05252") if ending < 0 else QColor("#888"))
                self.aux_rpt_tbl.setItem(i, j, it)
        # Totals
        n = len(rows); self.aux_rpt_tbl.setRowHeight(n, 38)
        for j, v in enumerate(["合计", "", "", fmt_amt(td_tot), fmt_amt(tc_tot), fmt_amt(td_tot-tc_tot)]):
            it = QTableWidgetItem(v)
            it.setFont(QFont("", weight=QFont.Bold)); it.setBackground(QColor("#f5f7fa"))
            it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter if j >= 2 else Qt.AlignLeft|Qt.AlignVCenter)
            self.aux_rpt_tbl.setItem(n, j, it)

    def _export_aux_report(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        period = self.rpt_period_edit.text().strip() or "report"
        path, _ = QFileDialog.getSaveFileName(self, "保存", f"往来对账_{period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "往来对账"
        hdrs = ["核算对象","科目","期初余额","本期借方","本期贷方","期末余额"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(1, ci, h); cell.font = XFont(bold=True, color="FFFFFF")
            cell.fill = fill; cell.alignment = Alignment(horizontal="center")
        for ri in range(self.aux_rpt_tbl.rowCount()):
            ws.append([self.aux_rpt_tbl.item(ri, ci).text() if self.aux_rpt_tbl.item(ri, ci) else ""
                       for ci in range(6)])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _export_items(self):
        if not self._cur_dim_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        path, _ = QFileDialog.getSaveFileName(self, "保存", "核算对象.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["编码", "名称", "联系人", "电话"])
        for r in self._items:
            ws.append([r["code"], r["name"], r["contact"], r["phone"]])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _load_bindings(self):
        if not self._cur_dim_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT a.code, a.name FROM account_aux_config ac
            JOIN accounts a ON a.client_id=ac.client_id AND a.code=ac.account_code
            WHERE ac.client_id=? AND ac.dimension_id=? ORDER BY a.code""",
                  (self.client_id, self._cur_dim_id))
        rows = c.fetchall(); conn.close()
        self.bind_list.clear()
        self._bindings = [dict(r) for r in rows]
        for r in self._bindings:
            self.bind_list.addItem(f"{r['code']}  {r['name']}")

    def _bind_account(self):
        if not self._cur_dim_id:
            QMessageBox.information(self, "提示", "请先选择维度"); return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code, name FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        accts = c.fetchall(); conn.close()
        # Already bound codes
        bound = {r["code"] for r in self._bindings}
        items = [f"{a['code']}  {a['name']}" for a in accts if a['code'] not in bound]
        if not items:
            QMessageBox.information(self, "提示", "所有科目均已绑定该维度"); return
        item, ok = QInputDialog.getItem(self, "绑定科目", "选择要绑定的科目：", items, editable=True)
        if not ok or not item: return
        code = item.split()[0]
        conn = get_db()
        try:
            conn.execute("INSERT INTO account_aux_config(client_id,account_code,dimension_id) VALUES(?,?,?)",
                         (self.client_id, code, self._cur_dim_id))
            conn.commit()
        except Exception:
            pass
        finally:
            conn.close()
        self._load_bindings()

    def _unbind_context_menu(self, pos):
        row = self.bind_list.currentRow()
        if row < 0: return
        menu = QMenu(self); menu.addAction("解除绑定")
        act = menu.exec(self.bind_list.mapToGlobal(pos))
        if act:
            code = self._bindings[row]["code"]
            conn = get_db()
            conn.execute("DELETE FROM account_aux_config WHERE client_id=? AND account_code=? AND dimension_id=?",
                         (self.client_id, code, self._cur_dim_id))
            conn.commit(); conn.close(); self._load_bindings()
