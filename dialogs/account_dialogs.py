"""dialogs/account_dialogs.py — 科目编辑和Excel导入"""
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, QTimer
from PySide6.QtGui import QColor, QFont, QPalette

from db import get_db, seed_client_accounts, log_action, VOUCHER_TEMPLATES
from utils import (lbl, sep, card, fmt_amt, cn_amount,
                   NoScrollSpinBox, NoScrollDoubleSpinBox,
                   infer_account_type_direction)

# openpyxl imported lazily inside each export function

class AccountEditDialog(QDialog):
    def __init__(self, parent, client_id, account=None, parent_acct=None):
        super().__init__(parent)
        self.client_id = client_id; self.account = account; self.parent_acct = parent_acct
        self.setWindowTitle("编辑科目" if account else "新增科目")
        self.setMinimumWidth(420); self._build()
        if account: self._load()
        elif parent_acct: self._prefill_from_parent()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(24,20,24,20); L.setSpacing(12)
        L.addWidget(lbl("科目信息", bold=True, size=15))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.code = QLineEdit(); self.code.setPlaceholderText("如 1002.01")
        self.name = QLineEdit(); self.name.setPlaceholderText("科目名称（必填）")
        self.full_name = QLineEdit(); self.full_name.setPlaceholderText("完整名称，如 银行存款-工商银行")
        self.type_cb = QComboBox()
        self.type_cb.addItems(["资产","负债","所有者权益","成本","收入","费用"])
        self.dir_cb = QComboBox(); self.dir_cb.addItems(["借","贷"])
        self.parent_cb = QComboBox(); self.parent_cb.addItem("（无上级）","")
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code",(self.client_id,))
        for r in c.fetchall():
            self.parent_cb.addItem(f"{r['code']}  {r['name']}", r['code'])
        conn.close()
        self.od = NoScrollDoubleSpinBox(); self.od.setRange(0,9999999999); self.od.setDecimals(2); self.od.setPrefix("¥ ")
        self.oc = NoScrollDoubleSpinBox(); self.oc.setRange(0,9999999999); self.oc.setDecimals(2); self.oc.setPrefix("¥ ")
        F.addRow("科目编号 *", self.code); F.addRow("科目名称 *", self.name)
        F.addRow("完整名称", self.full_name); F.addRow("科目类型", self.type_cb)
        F.addRow("余额方向", self.dir_cb); F.addRow("上级科目", self.parent_cb)
        F.addRow("期初借方", self.od); F.addRow("期初贷方", self.oc)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        bc = QPushButton("取消"); bc.setObjectName("btn_gray")
        bs = QPushButton("保存"); bs.setObjectName("btn_primary")
        bc.clicked.connect(self.reject); bs.clicked.connect(self._save)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        r = self.account
        self.code.setText(r["code"] or ""); self.name.setText(r["name"] or "")
        self.full_name.setText(r["full_name"] or "")
        idx = self.type_cb.findText(r["type"] or ""); self.type_cb.setCurrentIndex(max(0,idx))
        idx2 = self.dir_cb.findText(r["direction"] or "借"); self.dir_cb.setCurrentIndex(max(0,idx2))
        self.od.setValue(r["opening_debit"] or 0); self.oc.setValue(r["opening_credit"] or 0)
        if r["parent_code"]:
            for i in range(self.parent_cb.count()):
                if self.parent_cb.itemData(i) == r["parent_code"]:
                    self.parent_cb.setCurrentIndex(i); break

    def _prefill_from_parent(self):
        p = self.parent_acct
        self.code.setText(p["code"] + ".01")
        idx = self.type_cb.findText(p["type"] or ""); self.type_cb.setCurrentIndex(max(0,idx))
        idx2 = self.dir_cb.findText(p["direction"] or "借"); self.dir_cb.setCurrentIndex(max(0,idx2))
        for i in range(self.parent_cb.count()):
            if self.parent_cb.itemData(i) == p["code"]:
                self.parent_cb.setCurrentIndex(i); break

    def _save(self):
        code = self.code.text().strip(); name = self.name.text().strip()
        if not code or not name: QMessageBox.warning(self,"提示","编号和名称不能为空"); return
        full = self.full_name.text().strip() or name
        parent = self.parent_cb.currentData() or None
        level = (parent.count(".")+2) if parent else 1
        conn = get_db(); c = conn.cursor()
        try:
            if self.account:
                c.execute("""UPDATE accounts SET code=?,name=?,full_name=?,type=?,direction=?,
                    parent_code=?,level=?,opening_debit=?,opening_credit=? WHERE id=?""",
                    (code,name,full,self.type_cb.currentText(),self.dir_cb.currentText(),
                     parent,level,self.od.value(),self.oc.value(),self.account["id"]))
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,direction,
                    parent_code,level,opening_debit,opening_credit) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self.client_id,code,name,full,self.type_cb.currentText(),
                     self.dir_cb.currentText(),parent,level,self.od.value(),self.oc.value()))
            conn.commit(); conn.close(); self.accept()
        except Exception as e:
            conn.close(); QMessageBox.warning(self,"错误",f"保存失败：{e}")


class ImportExcelDialog(QDialog):
    """三合一导入：凭证 / 科目余额表期初 / 银行日记账"""

    def __init__(self, parent, client_id):
        super().__init__(parent)
        self.client_id = client_id
        self.setWindowTitle("从Excel导入历史数据")
        self.setMinimumSize(780, 560)
        self._build()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # Tab bar
        tab_bar = QWidget(); tab_bar.setStyleSheet("background:#f5f7fa;border-bottom:1px solid #e4e8f0;")
        tl = QHBoxLayout(tab_bar); tl.setContentsMargins(16,0,0,0); tl.setSpacing(0)
        self._itabs = []
        for name in ["记账凭证", "科目期初余额", "银行日记账"]:
            b = QPushButton(name)
            b.setStyleSheet("""QPushButton{background:transparent;color:#888;border:none;
                padding:11px 18px;border-bottom:2px solid transparent;}
                QPushButton:hover{color:#3d6fdb;}
                QPushButton[active=true]{color:#3d6fdb;border-bottom:2px solid #3d6fdb;font-weight:bold;}""")
            b.clicked.connect(lambda _,n=name: self._switch_itab(n))
            tl.addWidget(b); self._itabs.append(b)
        tl.addStretch()
        L.addWidget(tab_bar)

        self.istack = QStackedWidget(); L.addWidget(self.istack)
        self._build_voucher_import()
        self._build_balance_import()
        self._build_bank_import()
        self._switch_itab("记账凭证")

    def _switch_itab(self, name):
        mapping = {"记账凭证":0,"科目期初余额":1,"银行日记账":2}
        for b in self._itabs:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        self.istack.setCurrentIndex(mapping[name])

    # ── helpers ──
    def _make_tab(self, info_text):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,16,20,16); L.setSpacing(10)
        info = QLabel(info_text)
        info.setStyleSheet("background:#f6f8ff;border-radius:6px;padding:10px 14px;font-size:12px;color:#444;")
        info.setWordWrap(True); L.addWidget(info)
        log = QTextEdit(); log.setReadOnly(True)
        log.setStyleSheet("font-size:12px;font-family:monospace;background:#fafafa;")
        log.setPlaceholderText("导入日志将显示在这里…")
        L.addWidget(log)
        close_btn = QPushButton("关闭"); close_btn.setObjectName("btn_gray")
        close_btn.clicked.connect(self.accept)
        row = QHBoxLayout(); row.addStretch(); row.addWidget(close_btn); L.addLayout(row)
        return w, L, log

    def _pick_xls(self):
        path, _ = QFileDialog.getOpenFileName(self,"选择文件","","Excel(*.xls *.xlsx)")
        return path

    # ── Tab 1: 记账凭证 ──
    def _build_voucher_import(self):
        info = ("支持格式：从用友/金蝶等软件导出的记账凭证XLS文件。\n"
                "识别规则：每张凭证以[日期:...凭证字号:记-xxx]开头，下方各行为分录，[合计]行结束。\n"
                "也支持通用模板格式（A=期间 B=凭证号 C=日期 D=摘要 E=科目编号 F=科目名 G=借方 H=贷方）。")
        w, L, self.v_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_tmpl = QPushButton("↓ 下载通用模板"); b_tmpl.setObjectName("btn_outline")
        b_tmpl.clicked.connect(self._dl_voucher_template)
        b_imp = QPushButton("导入凭证文件"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_vouchers)
        btn_row.addWidget(b_tmpl); btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _dl_voucher_template(self):
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        path, _ = QFileDialog.getSaveFileName(self,"保存模板","凭证导入模板.xlsx","Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="凭证数据"
        hdrs = ["期间(YYYY-MM)","凭证号","日期(YYYY-MM-DD)","摘要","科目编号","科目名称","借方","贷方"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell=ws.cell(1,ci,h); cell.font=XFont(bold=True,color="FFFFFF")
            cell.fill=fill; cell.alignment=Alignment(horizontal="center")
        samples = [
            ("2025-01","记-001","2025-01-15","发放工资","2211.001","应付职工薪酬-工资",5000,0),
            ("2025-01","记-001","2025-01-15","发放工资","1002.001","银行存款-光大银行",0,5000),
            ("2025-01","记-002","2025-01-20","收到货款","1002.001","银行存款-光大银行",10000,0),
            ("2025-01","记-002","2025-01-20","收到货款","6001.001","主营业务收入-咨询",0,10000),
        ]
        for r in samples: ws.append(list(r))
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=20
        wb.save(path); self.v_log.append(f"✓ 模板已保存: {path}")

    def _import_vouchers(self):
        path = self._pick_xls()
        if not path: return
        self.v_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.v_log.append(f"✗ 读取失败: {e}"); return

        # Detect format: 用友-style vs generic template
        # 用友 style: row with "日期:…凭证字号:" header rows
        is_yonyou = any("凭证字号" in str(df.iloc[r,1]) for r in range(min(10,len(df))))

        conn = get_db(); c = conn.cursor()
        ok = skip = err = 0

        if is_yonyou:
            # Parse 用友/金蝶 style
            # Extract period from title row (row 0, col 1: "2026年03期 凭证")
            title = str(df.iloc[0,1])
            import re
            pm = re.search(r"(\d{4})年(\d{2})期", title)
            period = f"{pm.group(1)}-{pm.group(2)}" if pm else "2026-01"
            self.v_log.append(f"检测到期间: {period}，开始解析…")

            cur_vno = None; cur_date = ""; entries = []

            def flush(vno, date, ents):
                nonlocal ok, skip, err
                if not vno or not ents: return
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self.client_id, period, vno))
                if c.fetchone():
                    self.v_log.append(f"  跳过 {vno}（已存在）"); skip += 1; return
                td = sum(e[3] for e in ents); tc = sum(e[4] for e in ents)
                if abs(td-tc) > 0.01:
                    self.v_log.append(f"  ✗ {vno} 借贷不平 差{td-tc:.2f}，跳过"); err += 1; return
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status) VALUES(?,?,?,?,?)",
                          (self.client_id,period,vno,date,"已审核"))
                vid = c.lastrowid
                for ln,ent in enumerate(ents,1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln)+ent)
                ok += 1
                self.v_log.append(f"  ✓ {vno}  {len(ents)}行  借={td:.2f}")

            for ri in range(len(df)):
                row = df.iloc[ri]
                cell1 = str(row.iloc[1]).strip()
                if "凭证字号" in cell1:
                    flush(cur_vno, cur_date, entries)
                    entries = []
                    dm = re.search(r"日期:(\S+)", cell1)
                    nm = re.search(r"凭证字号:(\S+)", cell1)
                    cur_date = dm.group(1) if dm else period+"-28"
                    cur_vno  = nm.group(1).split()[0] if nm else None
                elif cell1 in ("合计：","合计:","") or not cell1:
                    continue
                else:
                    # entry row: col1=summary, col2=account, col3=debit, col4=credit
                    acct_full = str(row.iloc[2]).strip()
                    if not acct_full: continue
                    parts = acct_full.split(" ", 1)
                    code = parts[0]; aname = parts[1] if len(parts)>1 else acct_full
                    # Preserve auxiliary-account separators from source files.
                    code_norm = code
                    try: d = float(row.iloc[3]) if row.iloc[3] else 0
                    except: d = 0
                    try: cr = float(row.iloc[4]) if row.iloc[4] else 0
                    except: cr = 0
                    if d == 0 and cr == 0: continue
                    entries.append((cell1, code_norm, aname, d, cr))
            flush(cur_vno, cur_date, entries)

        else:
            # Generic template format (cols: period,vno,date,summary,code,name,debit,credit)
            from collections import OrderedDict
            vouchers = OrderedDict()
            n_cols = df.shape[1]
            for ri in range(1, len(df)):
                row = df.iloc[ri]
                def gcol(i, default=""):
                    try: v = str(row.iloc[i]).strip() if i < n_cols else default; return v if v != "nan" else default
                    except: return default
                def gcol_f(i):
                    try: v = row.iloc[i] if i < n_cols else 0; return float(v) if v and str(v) != "nan" else 0
                    except: return 0
                period  = gcol(0); vno = gcol(1); date = gcol(2)
                summary = gcol(3); code = gcol(4); aname = gcol(5)
                d = gcol_f(6); cr = gcol_f(7)
                if not period or not vno or not code: continue
                key = (period, vno)
                if key not in vouchers:
                    vouchers[key] = {"period":period,"vno":vno,"date":date,"entries":[]}
                vouchers[key]["entries"].append((summary,code,aname,d,cr))
            for (period,vno),v in vouchers.items():
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self.client_id,period,vno))
                if c.fetchone(): skip+=1; continue
                ents = v["entries"]
                td=sum(e[3] for e in ents); tc=sum(e[4] for e in ents)
                if abs(td-tc)>0.01: err+=1; continue
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status) VALUES(?,?,?,?,?)",
                          (self.client_id,period,vno,v["date"],"已审核"))
                vid=c.lastrowid
                for ln,ent in enumerate(ents,1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln)+ent)
                ok+=1

        if ok:
            log_action(conn, self.client_id, "批量导入凭证", "import", "",
                       f"导入{ok}张凭证，跳过{skip}张，失败{err}张")
        conn.commit(); conn.close()
        self.v_log.append(f"\n✅ 完成：导入 {ok} 张，跳过 {skip} 张，失败 {err} 张")

    # ── Tab 2: 科目余额表期初 ──
    def _build_balance_import(self):
        info = ("支持格式：用友/金蝶导出的「科目余额表」XLS。\n"
                "识别列：科目编号（第2列）、科目名称（第3列）、期初借方（第4列）、期初贷方（第5列）。\n"
                "导入后自动创建不存在的科目，并设置期初余额。")
        w, L, self.b_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_imp = QPushButton("导入科目余额表"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_balance)
        btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _import_balance(self):
        path = self._pick_xls()
        if not path: return
        self.b_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.b_log.append(f"✗ 读取失败: {e}"); return

        import re
        # Find data rows: col1=科目编号(numeric-ish), col2=科目名称, col3=期初借方, col4=期初贷方
        # Header rows are rows 0-5 typically; data starts where col1 looks like an account code
        conn = get_db(); c = conn.cursor()
        created = updated = skipped = 0

        for ri in range(len(df)):
            row = df.iloc[ri]
            code = str(row.iloc[1]).strip()
            name = str(row.iloc[2]).strip()
            if not code or not name: continue
            # Account codes: start with digit, may contain dots or underscores (auxiliary dims)
            if not re.match(r"^\d[\d._]*$", code): continue
            # Skip obvious non-account rows (e.g. "2026年03期" matched by looser regex)
            if len(code) < 4 or not code[:4].isdigit(): continue
            try:
                od = float(str(row.iloc[3]).replace(",","")) if row.iloc[3] else 0
            except: od = 0
            try:
                oc = float(str(row.iloc[4]).replace(",","")) if row.iloc[4] else 0
            except: oc = 0

            # Determine account type from code + name
            acct_type, direction = infer_account_type_direction(code, name)

            # Compute level and parent from code (treat _ same as . for hierarchy)
            normalized = code.replace("_", ".")
            level = normalized.count(".") + 1
            parent = code.rsplit(".", 1)[0] if "." in code else (
                     code.rsplit("_", 1)[0] if "_" in code else None)

            c.execute("SELECT id FROM accounts WHERE client_id=? AND code=?",
                      (self.client_id, code))
            existing = c.fetchone()
            if existing:
                c.execute("UPDATE accounts SET opening_debit=?,opening_credit=?,name=?,full_name=?,type=?,direction=? WHERE id=?",
                          (od, oc, name, name, acct_type, direction, existing[0]))
                updated += 1
                self.b_log.append(f"  ↻ {code} {name}  期初借={od:.2f} 贷={oc:.2f}")
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,direction,
                    parent_code,level,opening_debit,opening_credit) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self.client_id,code,name,name,acct_type,direction,parent,level,od,oc))
                created += 1
                self.b_log.append(f"  ✓ 新建 {code} {name}  期初借={od:.2f} 贷={oc:.2f}")

        conn.commit(); conn.close()
        self.b_log.append(f"\n✅ 完成：新建科目 {created} 个，更新期初 {updated} 个，跳过 {skipped} 个")

    # ── Tab 3: 银行日记账 ──
    def _build_bank_import(self):
        info = ("支持格式：用友/金蝶导出的「银行存款日记账 / 明细账」XLS。\n"
                "识别列：科目（第2列）、日期（第3列）、凭证号（第4列）、摘要（第5列）、借方（第6列）、贷方（第7列）、余额（第9列）。\n"
                "导入后可在「记账(凭证)→明细账」中查看，并可与实际银行流水对照。")
        w, L, self.k_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_imp = QPushButton("导入银行日记账"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_bank)
        btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _import_bank(self):
        path = self._pick_xls()
        if not path: return
        self.k_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.k_log.append(f"✗ 读取失败: {e}"); return

        import re
        conn = get_db(); c = conn.cursor()
        ok = skip = 0

        # 自动识别日期列位置
        date_col = None; data_start = 0
        for ri in range(min(15, len(df))):
            for ci in range(min(df.shape[1], 6)):
                v = str(df.iloc[ri, ci]).strip()
                if re.match(r"\d{4}[-/]\d{2}[-/]\d{2}", v):
                    date_col = ci; data_start = ri; break
            if date_col is not None: break

        if date_col is None:
            self.k_log.append("✗ 未能识别日期列，请确认文件格式"); conn.close(); return

        # 根据 date_col 判断布局
        if date_col == 2:   # 用友明细账
            acct_col=1; vno_col=3; summ_col=4; d_col=5; cr_col=6; bal_col=8
        elif date_col == 0: # 通用银行流水
            acct_col=None; vno_col=None; summ_col=1; d_col=2; cr_col=3; bal_col=4
        else:
            acct_col=None; vno_col=None; summ_col=date_col+1
            d_col=date_col+2; cr_col=date_col+3; bal_col=date_col+4

        def gcol(row, ci, default=""):
            try:
                v = str(row.iloc[ci]).strip() if ci is not None and ci < len(row) else default
                return v if v not in ("nan","") else default
            except: return default

        def flt(v):
            try: return float(str(v).replace(",","").strip())
            except: return 0.0

        for ri in range(data_start, len(df)):
            row = df.iloc[ri]
            raw_date = gcol(row, date_col).replace("/","-")
            if not re.match(r"\d{4}-\d{2}-\d{2}", raw_date): continue
            summary = gcol(row, summ_col)
            if summary in ("期初余额","本月合计","本年累计","合计",""): continue
            d  = flt(gcol(row, d_col))
            cr = flt(gcol(row, cr_col))
            bal_v = gcol(row, bal_col) if bal_col and bal_col < df.shape[1] else ""
            bal = flt(bal_v) if bal_v else None
            if d == 0 and cr == 0: continue
            vno = gcol(row, vno_col) if vno_col else ""
            acct_raw = gcol(row, acct_col) if acct_col else "1002"
            parts = acct_raw.split(" ", 1)
            acct_code = parts[0] if re.match(r"^\d[\d.]*$", parts[0]) else "1002"
            acct_name = parts[1] if len(parts) > 1 else "银行存款"
            # 去重
            c.execute("""SELECT id FROM bank_statements
                WHERE client_id=? AND date=? AND description=? AND debit=? AND credit=?""",
                (self.client_id, raw_date, summary, d, cr))
            if c.fetchone(): skip += 1; continue
            c.execute("""INSERT INTO bank_statements
                (client_id,account_code,account_name,date,voucher_no,
                 description,debit,credit,balance,source)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (self.client_id, acct_code, acct_name, raw_date,
                 vno, summary, d, cr, bal, "import"))
            ok += 1
            self.k_log.append(f"  ✓ {raw_date}  {summary[:20]}  借={d:.2f}  贷={cr:.2f}")

        conn.commit(); conn.close()
        self.k_log.append(f"\n✅ 完成：导入 {ok} 条，跳过重复 {skip} 条")


