"""dialogs.py — 所有对话框和辅助核算页面
包含: ClientDialog, AccountInitDialog, VoucherDialog,
       AuxItemDialog, AuxPage, ImportAccountSetDialog,
       AccountEditDialog, ImportExcelDialog
"""
import sys, os
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, QTimer
from PySide6.QtGui import QColor, QFont, QPalette

from db import get_db, seed_client_accounts, log_action, VOUCHER_TEMPLATES
from utils import (SS, lbl, sep, card, fmt_amt, cn_amount,
                   NoScrollSpinBox, NoScrollDoubleSpinBox,
                   infer_account_type_direction)

# openpyxl imported lazily inside each export function

class ClientDialog(QDialog):
    def __init__(self, parent=None, client=None):
        super().__init__(parent)
        self.client = client
        self.setWindowTitle("新建客户" if not client else "编辑客户")
        self.setMinimumWidth(400)
        self._build()
        if client: self._load()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(24,24,24,24); L.setSpacing(14)
        L.addWidget(lbl("客户信息", bold=True, size=15))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.name = QLineEdit(); self.name.setPlaceholderText("公司全称（必填）")
        self.code = QLineEdit(); self.code.setPlaceholderText("如 ZY")
        self.typ = QComboBox(); self.typ.addItems(["小规模纳税人","一般纳税人","其他"])
        self.taxid = QLineEdit(); self.taxid.setPlaceholderText("统一社会信用代码")
        self.contact = QLineEdit(); self.phone = QLineEdit()
        F.addRow("公司名称 *", self.name); F.addRow("助记码", self.code)
        F.addRow("客户类型", self.typ);   F.addRow("税号", self.taxid)
        F.addRow("联系人", self.contact); F.addRow("电话", self.phone)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        b_cancel = QPushButton("取消"); b_cancel.setObjectName("btn_gray")
        b_save = QPushButton("保 存"); b_save.setObjectName("btn_primary")
        b_cancel.clicked.connect(self.reject); b_save.clicked.connect(self._save)
        row.addWidget(b_cancel); row.addWidget(b_save); L.addLayout(row)

    def _load(self):
        c = self.client
        self.name.setText(c["name"] or ""); self.code.setText(c["short_code"] or "")
        self.taxid.setText(c["tax_id"] or ""); self.contact.setText(c["contact"] or "")
        self.phone.setText(c["phone"] or "")
        idx = self.typ.findText(c["client_type"] or ""); self.typ.setCurrentIndex(max(0,idx))

    def _save(self):
        n = self.name.text().strip()
        if not n: QMessageBox.warning(self,"提示","公司名称不能为空"); return
        conn = get_db(); c = conn.cursor()
        d = (n, self.code.text().strip(), self.typ.currentText(),
             self.taxid.text().strip(), self.contact.text().strip(), self.phone.text().strip())
        if self.client:
            c.execute("UPDATE clients SET name=?,short_code=?,client_type=?,tax_id=?,contact=?,phone=? WHERE id=?",
                      d+(self.client["id"],))
        else:
            c.execute("INSERT INTO clients(name,short_code,client_type,tax_id,contact,phone) VALUES(?,?,?,?,?,?)",d)
            cid = c.lastrowid
            seed_client_accounts(cid, conn)   # reuse same connection — no lock
        conn.commit(); conn.close(); self.accept()


class AccountInitDialog(QDialog):
    """科目期初设置"""
    def __init__(self, parent, client_id, period):
        super().__init__(parent)
        self.client_id = client_id; self.period = period
        self.setWindowTitle("科目期初余额"); self.setMinimumSize(680, 500)
        self._build()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(16,16,16,16); L.setSpacing(10)
        L.addWidget(lbl("科目期初余额设置", bold=True, size=14))
        L.addWidget(lbl(f"期间：{self.period}  （仅显示末级科目）", color="#888"))
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["科目编号","科目名称","期初借方","期初贷方"])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        hh.setMinimumSectionSize(80)
        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(2, 160)
        self.table.setColumnWidth(3, 160)
        self.table.verticalHeader().setVisible(False)
        self._load()
        L.addWidget(self.table)
        row = QHBoxLayout(); row.addStretch()
        bs = QPushButton("保存期初"); bs.setObjectName("btn_primary")
        bc = QPushButton("关闭"); bc.setObjectName("btn_gray")
        bs.clicked.connect(self._save); bc.clicked.connect(self.accept)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT id,code,name,opening_debit,opening_credit FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        rows = c.fetchall(); conn.close()
        self.table.setRowCount(len(rows))
        self._ids = []
        for i,r in enumerate(rows):
            self.table.setRowHeight(i, 40)          # 行高足够容纳输入框
            self._ids.append(r["id"])
            code_it = QTableWidgetItem(r["code"])
            code_it.setForeground(QColor("#3d6fdb"))
            name_it = QTableWidgetItem(r["name"])
            self.table.setItem(i,0,code_it)
            self.table.setItem(i,1,name_it)
            # Spinbox with explicit minimum size so numbers are readable
            def make_spin(val):
                sp = NoScrollDoubleSpinBox()
                sp.setRange(0, 9999999999)
                sp.setDecimals(2)
                sp.setValue(val or 0)
                sp.setMinimumHeight(32)
                sp.setMinimumWidth(140)
                sp.setAlignment(Qt.AlignRight)
                sp.setStyleSheet("QDoubleSpinBox{padding:4px 8px;font-size:13px;}")
                return sp
            d_spin  = make_spin(r["opening_debit"])
            cr_spin = make_spin(r["opening_credit"])
            self.table.setCellWidget(i,2,d_spin)
            self.table.setCellWidget(i,3,cr_spin)

    def _save(self):
        conn = get_db(); c = conn.cursor()
        for i,aid in enumerate(self._ids):
            d = self.table.cellWidget(i,2).value()
            cr = self.table.cellWidget(i,3).value()
            c.execute("UPDATE accounts SET opening_debit=?,opening_credit=? WHERE id=?", (d,cr,aid))
        conn.commit(); conn.close()
        QMessageBox.information(self,"成功","期初余额已保存")


class VoucherDialog(QDialog):
    """新增/编辑凭证 — 智一会计风格"""
    def __init__(self, parent=None, client_id=None, period=None, voucher_id=None):
        super().__init__(parent)
        self.client_id = client_id; self.period = period; self.voucher_id = voucher_id
        self.setWindowTitle("编辑凭证" if voucher_id else "新增凭证")
        self.setMinimumSize(860, 580)
        self._accounts = self._fetch_accounts()
        self._templates = VOUCHER_TEMPLATES
        from PySide6.QtCore import QStringListModel
        self._account_completion_model = QStringListModel(
            [f"{a['code']}  {a['full_name']}" for a in self._accounts], self)
        self._row_completers = []
        self._build()
        if voucher_id: self._load_voucher()
        else: self._add_row(); self._add_row()

    def _fetch_accounts(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name,full_name,direction FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        r = [dict(x) for x in c.fetchall()]
        # Load aux config: account_code -> list of (dim_id, dim_name, items)
        c.execute("""SELECT ac.account_code, ad.id as dim_id, ad.name as dim_name
            FROM account_aux_config ac
            JOIN aux_dimensions ad ON ad.id=ac.dimension_id
            WHERE ac.client_id=? ORDER BY ac.account_code, ad.sort_order""",
                  (self.client_id,))
        self._aux_config = {}   # account_code -> [{dim_id, dim_name}]
        for row in c.fetchall():
            self._aux_config.setdefault(row["account_code"], []).append(
                {"dim_id": row["dim_id"], "dim_name": row["dim_name"]})
        # Load all aux items per dimension
        c.execute("SELECT id, dimension_id, name, code FROM aux_items WHERE client_id=? AND is_active=1 ORDER BY dimension_id,id",
                  (self.client_id,))
        self._aux_items = {}    # dim_id -> [{id, name, code}]
        for row in c.fetchall():
            self._aux_items.setdefault(row["dimension_id"], []).append(
                {"id": row["id"], "name": row["name"], "code": row["code"] or ""})
        conn.close(); return r

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(16,14,16,14); L.setSpacing(10)

        # Header bar
        hdr = QHBoxLayout()
        # Voucher number + date + preparer
        self.lbl_no = lbl("新 建", bold=True, color="#3d6fdb", size=14)
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True); self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.preparer_lbl = lbl("未来", color="#888")
        self.attach_spin = NoScrollSpinBox(); self.attach_spin.setRange(0,999)
        self.attach_spin.setSuffix(" 张"); self.attach_spin.setFixedWidth(70)
        hdr.addWidget(self.lbl_no); hdr.addSpacing(16)
        hdr.addWidget(lbl("日期：")); hdr.addWidget(self.date_edit)
        hdr.addWidget(lbl("  制单：")); hdr.addWidget(self.preparer_lbl)
        hdr.addWidget(lbl("  附单据")); hdr.addWidget(self.attach_spin)
        hdr.addStretch()
        # Template button
        tpl_btn = QPushButton("凭证模板 ▼"); tpl_btn.setObjectName("btn_outline")
        tpl_btn.clicked.connect(self._show_template_menu)
        b_save_new = QPushButton("保存并新增"); b_save_new.setObjectName("btn_primary")
        b_save_new.clicked.connect(lambda: self._save(and_new=True))
        b_save = QPushButton("保 存"); b_save.setObjectName("btn_primary")
        b_save.clicked.connect(lambda: self._save(and_new=False))
        hdr.addWidget(tpl_btn); hdr.addWidget(b_save_new); hdr.addWidget(b_save)
        L.addLayout(hdr)
        L.addWidget(sep())

        # Entry table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["摘  要","科  目","核算对象","借方金额","贷方金额"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(2,160); self.table.setColumnWidth(3,130); self.table.setColumnWidth(4,130)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setMinimumHeight(280)
        L.addWidget(self.table)

        # Add/del row buttons
        row_btns = QHBoxLayout()
        b_add = QPushButton("＋ 增行"); b_add.setObjectName("btn_outline")
        b_add.clicked.connect(lambda: self._add_row())
        b_del = QPushButton("－ 删行"); b_del.setObjectName("btn_gray")
        b_del.clicked.connect(lambda: self._del_row())
        row_btns.addWidget(b_add); row_btns.addWidget(b_del); row_btns.addStretch()
        L.addLayout(row_btns)
        L.addWidget(sep())

        # Total row
        tot = QHBoxLayout()
        tot.addWidget(lbl("合计金额：")); self.lbl_cn = lbl("零元整", bold=True, color="#1e2130")
        tot.addWidget(self.lbl_cn); tot.addStretch()
        self.lbl_debit = lbl("借方合计：0.00", color="#3d6fdb", bold=True)
        self.lbl_credit = lbl("贷方合计：0.00", color="#e05252", bold=True)
        self.lbl_balance = lbl("✓ 借贷平衡", color="#52c41a", bold=True)
        tot.addWidget(self.lbl_debit); tot.addSpacing(20)
        tot.addWidget(self.lbl_credit); tot.addSpacing(20)
        tot.addWidget(self.lbl_balance)
        L.addLayout(tot)

    def _add_row(self, summary="", acct_code="", acct_name="", debit=0, credit=0, aux_data=None):
        """aux_data: list of (dim_id, item_id, item_name) saved previously"""
        i = self.table.rowCount(); self.table.insertRow(i)
        self.table.setRowHeight(i, 44)

        summary_edit = QLineEdit(summary)
        summary_edit.setPlaceholderText("摘要...")

        # ── 科目搜索框（自定义 completer，支持编号/名称任意位置模糊匹配） ──
        acct_edit = QLineEdit()
        acct_edit.setPlaceholderText("输入编号或名称搜索...")
        acct_edit._code = acct_code  # store selected code

        # Build display list and lookup maps
        display_list = [f"{a['code']}  {a['full_name']}" for a in self._accounts]
        code_by_display = {f"{a['code']}  {a['full_name']}": a['code'] for a in self._accounts}
        display_by_code = {a['code']: f"{a['code']}  {a['full_name']}" for a in self._accounts}

        from PySide6.QtWidgets import QCompleter
        completer = QCompleter(self._account_completion_model, acct_edit)
        completer.setFilterMode(Qt.MatchContains)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.popup().setMinimumWidth(380)
        acct_edit.setCompleter(completer)
        acct_edit._completer = completer
        self._row_completers.append(completer)

        # Restore pre-filled value
        if acct_code and acct_code in display_by_code:
            acct_edit.setText(display_by_code[acct_code])

        def on_acct_selected(text):
            code = code_by_display.get(text, "")
            acct_edit._code = code
            rebuild_aux(code)

        def on_acct_edited(text):
            # If user cleared the field, clear code
            if not text:
                acct_edit._code = ""
                rebuild_aux("")
            # If exact match, update code
            if text in code_by_display:
                acct_edit._code = code_by_display[text]
                rebuild_aux(acct_edit._code)
            if text:
                completer.setCompletionPrefix(text)
                completer.complete()

        completer.activated.connect(on_acct_selected)
        acct_edit.textChanged.connect(on_acct_edited)

        # ── 辅助核算 ──
        aux_container = QWidget()
        aux_layout = QHBoxLayout(aux_container)
        aux_layout.setContentsMargins(2,2,2,2); aux_layout.setSpacing(4)
        aux_container._combos = []

        def rebuild_aux(code, saved=None):
            for w in aux_container._combos:
                w[1].setParent(None)
            aux_container._combos.clear()
            dims = self._aux_config.get(code, [])
            for dim in dims:
                cb = QComboBox(); cb.setMinimumWidth(130)
                cb.addItem("— 选核算对象 —", None)
                for it in self._aux_items.get(dim["dim_id"], []):
                    display = f"{it['code']} {it['name']}" if it['code'] else it['name']
                    cb.addItem(display, it["id"])
                cb.setToolTip(dim["dim_name"])
                if saved:
                    for sv_dim, sv_id, sv_name in saved:
                        if sv_dim == dim["dim_id"]:
                            for ki in range(cb.count()):
                                if cb.itemData(ki) == sv_id:
                                    cb.setCurrentIndex(ki); break
                aux_layout.addWidget(cb)
                aux_container._combos.append((dim["dim_id"], cb))
            if not dims:
                aux_layout.addWidget(QLabel("—"))

        rebuild_aux(acct_code, aux_data)

        # ── 借方金额 ──
        d_spin = NoScrollDoubleSpinBox(); d_spin.setRange(0,999999999); d_spin.setDecimals(2)
        d_spin.setSpecialValueText(""); d_spin.setValue(debit)
        d_spin.valueChanged.connect(self._update_totals)

        # ── 贷方金额（支持 = 键自动补平） ──
        cr_spin = NoScrollDoubleSpinBox(); cr_spin.setRange(0,999999999); cr_spin.setDecimals(2)
        cr_spin.setSpecialValueText(""); cr_spin.setValue(credit)
        cr_spin.valueChanged.connect(self._update_totals)

        # Patch keyPressEvent on cr_spin to handle "=" auto-balance
        _row_ref = [i]  # mutable ref so lambda can find the row
        orig_key = cr_spin.keyPressEvent
        def cr_key(event, _cr=cr_spin):
            if event.text() == "=":
                self._auto_balance_credit(_cr)
            else:
                orig_key(event)
        cr_spin.keyPressEvent = cr_key

        self.table.setCellWidget(i,0,summary_edit)
        self.table.setCellWidget(i,1,acct_edit)
        self.table.setCellWidget(i,2,aux_container)
        self.table.setCellWidget(i,3,d_spin)
        self.table.setCellWidget(i,4,cr_spin)
        self._update_totals()

    def _auto_balance_credit(self, target_spin):
        """按 = 键时，将 target_spin 设为令借贷平衡所需的金额。"""
        td = tc_other = 0
        for i in range(self.table.rowCount()):
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            if dw: td += dw.value()
            if cw and cw is not target_spin: tc_other += cw.value()
        needed = max(0, round(td - tc_other, 2))
        target_spin.setValue(needed)
        self._update_totals()

    def _del_row(self):
        row = self.table.currentRow()
        if row < 0:
            row = self.table.rowCount() - 1
        if row >= 0 and self.table.rowCount() > 1:
            self.table.removeRow(row)
            self._update_totals()

    def _update_totals(self):
        td = tc = 0
        for i in range(self.table.rowCount()):
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            if dw: td += dw.value()
            if cw: tc += cw.value()
        self.lbl_debit.setText(f"借方合计：{td:,.2f}")
        self.lbl_credit.setText(f"贷方合计：{tc:,.2f}")
        balanced = abs(td-tc) < 0.005
        self.lbl_balance.setText("✓ 借贷平衡" if balanced else f"✗ 差额 {abs(td-tc):,.2f}")
        self.lbl_balance.setStyleSheet(f"color:{'#52c41a' if balanced else '#ff4d4f'};font-weight:bold;")
        self.lbl_cn.setText(cn_amount(td))

    def _load_templates_from_db(self):
        """Load user-saved templates from database, merged with built-in templates."""
        conn = get_db(); c = conn.cursor()
        try:
            c.execute("SELECT name, entries FROM voucher_templates WHERE client_id=? OR client_id IS NULL ORDER BY id",
                      (self.client_id,))
            db_tmpls = c.fetchall()
        except Exception:
            db_tmpls = []
        conn.close()
        import json
        user_tmpls = []
        for row in db_tmpls:
            try:
                entries = json.loads(row["entries"])
                user_tmpls.append((row["name"], entries))
            except Exception:
                pass
        # Merge: built-in first, then user-saved
        self._templates = list(VOUCHER_TEMPLATES) + user_tmpls

    def _show_template_menu(self):
        self._load_templates_from_db()
        menu = QMenu(self)
        for name, _ in self._templates:
            menu.addAction(name)
        menu.addSeparator()
        b_save_tpl = menu.addAction("📌 存为模板...")
        b_del_tpl  = menu.addAction("🗑 删除模板...")
        act = menu.exec(self.sender().mapToGlobal(self.sender().rect().bottomLeft()))
        if not act: return

        if act is b_save_tpl:
            self._save_as_template()
            return
        if act is b_del_tpl:
            self._delete_template()
            return

        # Apply selected template
        for name, entries in self._templates:
            if act.text() == name:
                self.table.setRowCount(0)
                for entry in entries:
                    # Support both old tuple format and new dict format
                    if isinstance(entry, dict):
                        self._add_row(entry.get("summary",""), entry.get("code",""),
                                      entry.get("name",""), entry.get("debit",0), entry.get("credit",0))
                    else:
                        s, code, _, d, cr = entry
                        self._add_row(s, code, "", d, cr)
                break

    def _save_as_template(self):
        """Save current voucher entries as a reusable template."""
        import json
        # Collect current entries
        entries = []
        for i in range(self.table.rowCount()):
            sw = self.table.cellWidget(i,0); aw = self.table.cellWidget(i,1)
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            code = getattr(aw, '_code', "") or "" if aw else ""
            summary = sw.text().strip() if sw else ""
            d = dw.value() if dw else 0
            cr = cw.value() if cw else 0
            if not code and d == 0 and cr == 0: continue
            # Find account name
            aname = ""
            for a in self._accounts:
                if a['code'] == code: aname = a['full_name']; break
            entries.append({"summary": summary, "code": code, "name": aname,
                            "debit": d, "credit": cr})
        if not entries:
            QMessageBox.warning(self, "提示", "请先填写凭证分录再保存为模板"); return

        name, ok = QInputDialog.getText(self, "保存为模板", "模板名称：",
                                         text=entries[0].get("summary","") or "新模板")
        if not ok or not name.strip(): return
        name = name.strip()

        # Check duplicate name
        all_names = [t[0] for t in self._templates]
        if name in all_names:
            if QMessageBox.question(self, "覆盖确认", f"模板【{name}】已存在，是否覆盖？",
                    QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return
            conn = get_db()
            conn.execute("DELETE FROM voucher_templates WHERE name=? AND (client_id=? OR client_id IS NULL)",
                         (name, self.client_id))
            conn.commit(); conn.close()

        conn = get_db()
        conn.execute("INSERT INTO voucher_templates(client_id, name, entries) VALUES(?,?,?)",
                     (self.client_id, name, json.dumps(entries, ensure_ascii=False)))
        conn.commit(); conn.close()
        QMessageBox.information(self, "成功", f"已保存模板：{name}")

    def _delete_template(self):
        """Delete a user-saved template."""
        conn = get_db(); c = conn.cursor()
        try:
            c.execute("SELECT id, name FROM voucher_templates WHERE client_id=? OR client_id IS NULL ORDER BY id",
                      (self.client_id,))
            user_tmpls = c.fetchall()
        except Exception:
            user_tmpls = []
        conn.close()
        if not user_tmpls:
            QMessageBox.information(self, "提示", "没有可删除的自定义模板（内置模板不能删除）")
            return
        names = [r["name"] for r in user_tmpls]
        name, ok = QInputDialog.getItem(self, "删除模板", "选择要删除的模板：", names, editable=False)
        if not ok or not name: return
        conn = get_db()
        conn.execute("DELETE FROM voucher_templates WHERE name=? AND (client_id=? OR client_id IS NULL)",
                     (name, self.client_id))
        conn.commit(); conn.close()
        QMessageBox.information(self, "成功", f"已删除模板：{name}")

    def _load_voucher(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM vouchers WHERE id=?", (self.voucher_id,))
        v = c.fetchone()
        self.lbl_no.setText(v["voucher_no"])
        self.date_edit.setDate(QDate.fromString(v["date"],"yyyy-MM-dd"))
        self.attach_spin.setValue(v["attachment_count"] or 0)
        c.execute("SELECT * FROM voucher_entries WHERE voucher_id=? ORDER BY line_no", (self.voucher_id,))
        entries = c.fetchall()
        # Load aux data per entry
        entry_ids = [e["id"] for e in entries]
        aux_by_entry = {}
        if entry_ids:
            placeholders = ",".join("?" * len(entry_ids))
            c.execute(f"SELECT entry_id,dimension_id,aux_item_id,aux_item_name FROM voucher_entry_aux WHERE entry_id IN ({placeholders})",
                      entry_ids)
            for row in c.fetchall():
                aux_by_entry.setdefault(row["entry_id"], []).append(
                    (row["dimension_id"], row["aux_item_id"], row["aux_item_name"]))
        conn.close()
        for e in entries:
            aux_data = aux_by_entry.get(e["id"], None)
            self._add_row(e["summary"] or "", e["account_code"] or "", e["account_name"] or "",
                          e["debit"] or 0, e["credit"] or 0, aux_data=aux_data)

    def _save(self, and_new=False):
        # Validate
        entries = []
        for i in range(self.table.rowCount()):
            sw  = self.table.cellWidget(i,0); aw  = self.table.cellWidget(i,1)
            auxw= self.table.cellWidget(i,2)
            dw  = self.table.cellWidget(i,3); cw  = self.table.cellWidget(i,4)
            if not aw: continue
            code = getattr(aw, '_code', "") or ""
            d = dw.value() if dw else 0; cr = cw.value() if cw else 0
            if not code and d == 0 and cr == 0: continue
            if not code: QMessageBox.warning(self,"提示",f"第{i+1}行科目不能为空"); return
            aname = ""
            for a in self._accounts:
                if a['code'] == code: aname = a['full_name']; break
            # Collect aux selections
            aux_sel = []   # (dim_id, item_id, item_name)
            if auxw:
                for dim_id, cb in auxw._combos:
                    item_id = cb.currentData()
                    item_name = cb.currentText() if item_id else ""
                    if item_id:
                        aux_sel.append((dim_id, item_id, item_name))
            entries.append((i, sw.text().strip() if sw else "", code, aname, d, cr, aux_sel))

        if not entries: QMessageBox.warning(self,"提示","请至少填写一行分录"); return
        td = sum(e[4] for e in entries); tc = sum(e[5] for e in entries)
        if abs(td-tc) > 0.005:
            QMessageBox.warning(self,"借贷不平",f"借方合计 {td:.2f} ≠ 贷方合计 {tc:.2f}\n请检查金额"); return

        conn = get_db(); c = conn.cursor()
        dt = self.date_edit.date().toString("yyyy-MM-dd")
        reverted_to_pending = False

        if self.voucher_id:
            c.execute("SELECT voucher_no, status FROM vouchers WHERE id=?", (self.voucher_id,))
            voucher = c.fetchone()
            if not voucher:
                conn.close()
                QMessageBox.warning(self, "提示", "凭证不存在或已被删除，请刷新后重试")
                return
            vno = voucher["voucher_no"]
            new_status = voucher["status"]
            if voucher["status"] == "已审核":
                new_status = "待审核"
                reverted_to_pending = True
            c.execute("UPDATE vouchers SET date=?,attachment_count=?,status=? WHERE id=?",
                      (dt, self.attach_spin.value(), new_status, self.voucher_id))
            c.execute("DELETE FROM voucher_entries WHERE voucher_id=?", (self.voucher_id,))
            vid = self.voucher_id
        else:
            c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=?",
                      (self.client_id, self.period))
            n = c.fetchone()[0] + 1
            vno = f"记-{n:03d}"
            c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,attachment_count) VALUES(?,?,?,?,?)",
                      (self.client_id, self.period, vno, dt, self.attach_spin.value()))
            vid = c.lastrowid

        for ln, summary, code, aname, d, cr, aux_sel in entries:
            c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                      (vid, ln, summary, code, aname, d, cr))
            entry_id = c.lastrowid
            for dim_id, item_id, item_name in aux_sel:
                c.execute("INSERT INTO voucher_entry_aux(entry_id,dimension_id,aux_item_id,aux_item_name) VALUES(?,?,?,?)",
                          (entry_id, dim_id, item_id, item_name))
        action = "编辑凭证" if self.voucher_id else "新增凭证"
        detail = f"凭证号:{vno} 借方合计:{td:.2f}"
        if reverted_to_pending:
            detail += " 修改后自动回退为待审核"
        log_action(conn, self.client_id,
                   action, "voucher", vid, detail)
        conn.commit(); conn.close()
        self.saved_and_new = and_new
        if reverted_to_pending:
            QMessageBox.information(self, "已回退待审核", "该凭证原状态为“已审核”，修改后已自动回退为“待审核”，请重新审核。")
        self.accept()

# ── Pages ──────────────────────────────────────────────────────────────────

# ── 辅助核算管理页（嵌入 VoucherPage 的 Tab） ─────────────────────────────

class AuxItemDialog(QDialog):
    """新增/编辑 核算对象"""
    def __init__(self, parent, client_id, dimension_id, item=None):
        super().__init__(parent)
        self.client_id = client_id
        self.dimension_id = dimension_id
        self.item = item
        self.setWindowTitle("编辑核算对象" if item else "新增核算对象")
        self.setMinimumWidth(360)
        self._build()
        if item: self._load()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(22,20,22,20); L.setSpacing(12)
        L.addWidget(lbl("核算对象信息", bold=True, size=14))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.f_name    = QLineEdit(); self.f_name.setPlaceholderText("名称（必填）")
        self.f_code    = QLineEdit(); self.f_code.setPlaceholderText("编码（可选）")
        self.f_contact = QLineEdit(); self.f_contact.setPlaceholderText("联系人")
        self.f_phone   = QLineEdit(); self.f_phone.setPlaceholderText("电话")
        F.addRow("名称 *",  self.f_name)
        F.addRow("编码",    self.f_code)
        F.addRow("联系人",  self.f_contact)
        F.addRow("电话",    self.f_phone)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        bc = QPushButton("取消"); bc.setObjectName("btn_gray")
        bs = QPushButton("保存"); bs.setObjectName("btn_primary")
        bc.clicked.connect(self.reject); bs.clicked.connect(self._save)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        self.f_name.setText(self.item["name"] or "")
        self.f_code.setText(self.item["code"] or "")
        self.f_contact.setText(self.item["contact"] or "")
        self.f_phone.setText(self.item["phone"] or "")

    def _save(self):
        name = self.f_name.text().strip()
        if not name: QMessageBox.warning(self, "提示", "名称不能为空"); return
        conn = get_db(); c = conn.cursor()
        d = (name, self.f_code.text().strip(), self.f_contact.text().strip(),
             self.f_phone.text().strip())
        if self.item:
            c.execute("UPDATE aux_items SET name=?,code=?,contact=?,phone=? WHERE id=?",
                      d + (self.item["id"],))
        else:
            c.execute("INSERT INTO aux_items(client_id,dimension_id,name,code,contact,phone)"
                      " VALUES(?,?,?,?,?,?)",
                      (self.client_id, self.dimension_id) + d)
        conn.commit(); conn.close(); self.accept()


class AuxPage(QWidget):
    """辅助核算管理：维度 + 对象 + 科目绑定"""
    def __init__(self):
        super().__init__()
        self.client_id = None
        self._cur_dim_id = None
        L = QHBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # ── 左栏：维度列表 ──
        left = QWidget(); left.setFixedWidth(200)
        left.setStyleSheet("background:#f7f9fc; border-right:1px solid #e8ecf2;")
        ll = QVBoxLayout(left); ll.setContentsMargins(0,0,0,0); ll.setSpacing(0)
        hdr_l = QWidget(); hdr_l.setStyleSheet("background:#fff; border-bottom:1px solid #eee;")
        hl = QHBoxLayout(hdr_l); hl.setContentsMargins(12,10,8,10)
        hl.addWidget(lbl("核算维度", bold=True)); hl.addStretch()
        b_adddim = QPushButton("＋"); b_adddim.setObjectName("btn_primary")
        b_adddim.setFixedSize(28,28); b_adddim.setToolTip("新增维度")
        b_adddim.clicked.connect(self._add_dim)
        hl.addWidget(b_adddim); ll.addWidget(hdr_l)

        self.dim_list = QListWidget()
        self.dim_list.setStyleSheet(
            "QListWidget{border:none;background:#f7f9fc;}"
            "QListWidget::item{padding:10px 14px;border-bottom:1px solid #eef0f4;}"
            "QListWidget::item:selected{background:#e6f0ff;color:#3d6fdb;font-weight:bold;}")
        self.dim_list.currentRowChanged.connect(self._on_dim_changed)
        ll.addWidget(self.dim_list)

        # 维度右键菜单
        self.dim_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.dim_list.customContextMenuRequested.connect(self._dim_context_menu)
        L.addWidget(left)

        # ── 右栏：对象列表 + 科目绑定 ──
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(0,0,0,0); rl.setSpacing(0)

        # Right-side tab: 对象管理 | 往来对账
        self.right_tabs = QTabWidget()
        self.right_tabs.setStyleSheet(
            "QTabBar::tab{padding:8px 20px;color:#888;border:none;background:transparent;"
            "border-bottom:2px solid transparent;}"
            "QTabBar::tab:selected{color:#3d6fdb;border-bottom:2px solid #3d6fdb;}"
            "QTabWidget::pane{border:none;}")

        # ── Tab A: 核算对象管理 ──
        tab_mgr = QWidget(); tl = QVBoxLayout(tab_mgr); tl.setContentsMargins(20,14,20,14); tl.setSpacing(10)

        hdr_r = QHBoxLayout()
        self.dim_title = lbl("请选择左侧维度", bold=True, size=15)
        hdr_r.addWidget(self.dim_title); hdr_r.addStretch()
        self.b_additem = QPushButton("＋ 新增对象"); self.b_additem.setObjectName("btn_primary")
        self.b_additem.clicked.connect(self._add_item)
        b_exp = QPushButton("导出Excel"); b_exp.setObjectName("btn_outline")
        b_exp.clicked.connect(self._export_items)
        hdr_r.addWidget(b_exp); hdr_r.addWidget(self.b_additem)
        tl.addLayout(hdr_r)

        f1 = card(); v1 = QVBoxLayout(f1); v1.setContentsMargins(0,0,0,0)
        self.item_tbl = QTableWidget(); self.item_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.item_tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.item_tbl.setShowGrid(False); self.item_tbl.verticalHeader().setVisible(False)
        self.item_tbl.setColumnCount(5)
        self.item_tbl.setHorizontalHeaderLabels(["编码","名称","联系人","电话","操作"])
        hh = self.item_tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        self.item_tbl.setColumnWidth(0,80); self.item_tbl.setColumnWidth(2,90)
        self.item_tbl.setColumnWidth(3,110); self.item_tbl.setColumnWidth(4,160)
        v1.addWidget(self.item_tbl); tl.addWidget(f1)

        # 科目绑定区
        tl.addWidget(lbl("绑定科目（凭证录入时，选中这些科目将显示本维度的对象选择）",
                         color="#666", size=12))
        f2 = card(); v2 = QVBoxLayout(f2); v2.setContentsMargins(12,10,12,10); v2.setSpacing(8)
        bind_hdr = QHBoxLayout()
        bind_hdr.addWidget(lbl("已绑定科目", bold=True)); bind_hdr.addStretch()
        b_bind = QPushButton("＋ 绑定科目"); b_bind.setObjectName("btn_outline")
        b_bind.clicked.connect(self._bind_account)
        bind_hdr.addWidget(b_bind); v2.addLayout(bind_hdr)
        self.bind_list = QListWidget()
        self.bind_list.setStyleSheet(
            "QListWidget{border:1px solid #eee;border-radius:5px;background:#fff;}"
            "QListWidget::item{padding:6px 10px;}"
            "QListWidget::item:selected{background:#e6f0ff;}")
        self.bind_list.setMaximumHeight(120)
        self.bind_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.bind_list.customContextMenuRequested.connect(self._unbind_context_menu)
        v2.addWidget(self.bind_list)
        tl.addWidget(f2)
        self.right_tabs.addTab(tab_mgr, "核算对象管理")

        # ── Tab B: 往来对账报表 ──
        tab_rpt = QWidget(); tr = QVBoxLayout(tab_rpt); tr.setContentsMargins(20,14,20,14); tr.setSpacing(10)

        rpt_hdr = QHBoxLayout()
        rpt_hdr.addWidget(lbl("往来对账表", bold=True, size=15)); rpt_hdr.addStretch()
        b_rpt_exp = QPushButton("导出Excel"); b_rpt_exp.setObjectName("btn_outline")
        b_rpt_exp.clicked.connect(self._export_aux_report)
        rpt_hdr.addWidget(b_rpt_exp); tr.addLayout(rpt_hdr)

        fr = QHBoxLayout(); fr.setSpacing(8)
        fr.addWidget(lbl("维度:"))
        self.rpt_dim_combo = QComboBox(); self.rpt_dim_combo.setMinimumWidth(120)
        fr.addWidget(self.rpt_dim_combo)
        fr.addWidget(lbl("期间:"))
        self.rpt_period_edit = QLineEdit(); self.rpt_period_edit.setFixedWidth(100)
        self.rpt_period_edit.setPlaceholderText("如 2026-03")
        fr.addWidget(self.rpt_period_edit)
        fr.addWidget(lbl("科目:"))
        self.rpt_acct_combo = QComboBox(); self.rpt_acct_combo.setMinimumWidth(200)
        self.rpt_acct_combo.addItem("全部科目", "")
        fr.addWidget(self.rpt_acct_combo)
        b_q = QPushButton("查询"); b_q.setObjectName("btn_primary")
        b_q.clicked.connect(self._load_aux_report)
        fr.addWidget(b_q); fr.addStretch()
        tr.addLayout(fr)

        f3 = card(); v3 = QVBoxLayout(f3); v3.setContentsMargins(0,0,0,0)
        self.aux_rpt_tbl = QTableWidget(); self.aux_rpt_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.aux_rpt_tbl.setShowGrid(True); self.aux_rpt_tbl.verticalHeader().setVisible(False)
        self.aux_rpt_tbl.setColumnCount(6)
        self.aux_rpt_tbl.setHorizontalHeaderLabels(["核算对象","科目","期初余额","本期借方","本期贷方","期末余额"])
        hh3 = self.aux_rpt_tbl.horizontalHeader()
        hh3.setSectionResizeMode(QHeaderView.Interactive)
        hh3.setSectionResizeMode(0, QHeaderView.Stretch)
        hh3.setSectionResizeMode(1, QHeaderView.Stretch)
        for ci in range(2,6): self.aux_rpt_tbl.setColumnWidth(ci, 110)
        v3.addWidget(self.aux_rpt_tbl); tr.addWidget(f3)
        self.right_tabs.addTab(tab_rpt, "往来对账")

        rl.addWidget(self.right_tabs)
        L.addWidget(right)

    def set_client(self, client_id, period=""):
        self.client_id = client_id
        self._period = period
        self._cur_dim_id = None
        self._dims = []
        self._items = []
        self._bindings = []
        self._load_dims()
        self._refresh_rpt_combos()

    def _refresh_rpt_combos(self):
        if not self.client_id: return
        # Period
        from datetime import datetime
        now = datetime.now()
        if self._period:
            self.rpt_period_edit.setText(self._period)
        else:
            self.rpt_period_edit.setText(f"{now.year}-{now.month:02d}")
        # Dimensions
        self.rpt_dim_combo.clear()
        for d in self._dims:
            self.rpt_dim_combo.addItem(d["name"], d["id"])
        # Accounts
        self.rpt_acct_combo.clear(); self.rpt_acct_combo.addItem("全部科目", "")
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code", (self.client_id,))
        for a in c.fetchall():
            self.rpt_acct_combo.addItem(f"{a['code']} {a['name']}", a['code'])
        conn.close()

    def _load_dims(self):
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM aux_dimensions WHERE client_id=? ORDER BY sort_order,id",
                  (self.client_id,))
        dims = c.fetchall(); conn.close()
        self.dim_list.clear()
        self._dims = [dict(d) for d in dims]
        for d in self._dims:
            self.dim_list.addItem(d["name"])
        if self._dims:
            self.dim_list.setCurrentRow(0)

    def _on_dim_changed(self, row):
        if row < 0 or row >= len(self._dims): return
        self._cur_dim_id = self._dims[row]["id"]
        self.dim_title.setText(f"【{self._dims[row]['name']}】核算对象")
        self._load_items()
        self._load_bindings()

    def _add_dim(self):
        if not self.client_id: return
        name, ok = QInputDialog.getText(self, "新增维度", "维度名称（如：客户、员工、项目）：")
        if not ok or not name.strip(): return
        conn = get_db()
        try:
            conn.execute("INSERT INTO aux_dimensions(client_id,name) VALUES(?,?)",
                         (self.client_id, name.strip()))
            conn.commit()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"维度已存在或保存失败：{e}")
        finally:
            conn.close()
        self._load_dims()

    def _dim_context_menu(self, pos):
        row = self.dim_list.currentRow()
        if row < 0: return
        menu = QMenu(self)
        menu.addAction("重命名").triggered.connect(lambda: self._rename_dim(row))
        menu.addAction("删除维度").triggered.connect(lambda: self._del_dim(row))
        menu.exec(self.dim_list.mapToGlobal(pos))

    def _rename_dim(self, row):
        old = self._dims[row]["name"]
        name, ok = QInputDialog.getText(self, "重命名维度", "新名称：", text=old)
        if not ok or not name.strip(): return
        conn = get_db()
        conn.execute("UPDATE aux_dimensions SET name=? WHERE id=?",
                     (name.strip(), self._dims[row]["id"]))
        conn.commit(); conn.close(); self._load_dims()

    def _del_dim(self, row):
        if QMessageBox.question(self, "确认",
                f"删除维度【{self._dims[row]['name']}】及其所有对象和绑定？",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db(); did = self._dims[row]["id"]
        conn.execute("DELETE FROM voucher_entry_aux WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM account_aux_config WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM aux_items WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM aux_dimensions WHERE id=?", (did,))
        conn.commit(); conn.close(); self._load_dims()

    def _load_items(self):
        if not self._cur_dim_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM aux_items WHERE dimension_id=? AND client_id=? ORDER BY id",
                  (self._cur_dim_id, self.client_id))
        rows = c.fetchall(); conn.close()
        self._items = [dict(r) for r in rows]
        self.item_tbl.setRowCount(len(rows))
        for i, r in enumerate(self._items):
            self.item_tbl.setRowHeight(i, 40)
            for j, v in enumerate([r["code"] or "", r["name"], r["contact"] or "", r["phone"] or ""]):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter if j != 1 else Qt.AlignLeft | Qt.AlignVCenter)
                self.item_tbl.setItem(i, j, it)
            bw = QWidget(); bl = QHBoxLayout(bw); bl.setContentsMargins(4,3,4,3); bl.setSpacing(4)
            b_ed = QPushButton("✏ 编辑"); b_ed.setObjectName("btn_outline"); b_ed.setFixedSize(64,26)
            b_ed.clicked.connect(lambda _, rr=r: self._edit_item(rr))
            b_dl = QPushButton("🗑"); b_dl.setObjectName("btn_red"); b_dl.setFixedSize(30,26)
            b_dl.clicked.connect(lambda _, rid=r["id"]: self._del_item(rid))
            bl.addWidget(b_ed); bl.addWidget(b_dl); bl.addStretch()
            self.item_tbl.setCellWidget(i, 4, bw)

    def _add_item(self):
        if not self._cur_dim_id:
            QMessageBox.information(self, "提示", "请先选择左侧维度"); return
        d = AuxItemDialog(self, self.client_id, self._cur_dim_id)
        if d.exec(): self._load_items()

    def _edit_item(self, r):
        d = AuxItemDialog(self, self.client_id, self._cur_dim_id, item=r)
        if d.exec(): self._load_items()

    def _del_item(self, item_id):
        if QMessageBox.question(self, "确认", "删除该核算对象？",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db()
        conn.execute("DELETE FROM aux_items WHERE id=?", (item_id,))
        conn.commit(); conn.close(); self._load_items()

    def _load_aux_report(self):
        if not self.client_id: return
        dim_id = self.rpt_dim_combo.currentData()
        period = self.rpt_period_edit.text().strip()
        acct_filter = self.rpt_acct_combo.currentData() or ""
        if not dim_id or not period:
            QMessageBox.information(self, "提示", "请选择维度并填写期间"); return
        conn = get_db(); c = conn.cursor()
        # Get account direction info for opening balance computation
        c.execute("SELECT code,direction,opening_debit,opening_credit FROM accounts WHERE client_id=?",
                  (self.client_id,))
        acct_info = {r["code"]: r for r in c.fetchall()}
        sql = """
            SELECT ai.name AS item_name, ai.id AS item_id,
                   e.account_code, e.account_name,
                   SUM(e.debit) AS td, SUM(e.credit) AS tc
            FROM voucher_entry_aux ea
            JOIN aux_items ai ON ai.id=ea.aux_item_id
            JOIN voucher_entries e ON e.id=ea.entry_id
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE ea.dimension_id=? AND v.client_id=? AND v.period=?
        """
        params = [dim_id, self.client_id, period]
        if acct_filter:
            sql += " AND e.account_code=?"; params.append(acct_filter)
        sql += " GROUP BY ai.id, e.account_code ORDER BY ai.name, e.account_code"
        c.execute(sql, params)
        rows = c.fetchall(); conn.close()

        self.aux_rpt_tbl.setRowCount(len(rows) + 1)
        td_tot = tc_tot = 0
        for i, r in enumerate(rows):
            self.aux_rpt_tbl.setRowHeight(i, 36)
            td = r["td"] or 0; tc = r["tc"] or 0
            ai = acct_info.get(r["account_code"])
            od = (ai["opening_debit"] or 0) - (ai["opening_credit"] or 0) if ai else 0
            if ai and ai["direction"] == "贷": od = -od
            ending = od + td - tc
            td_tot += td; tc_tot += tc
            vals = [r["item_name"], f"{r['account_code']} {r['account_name']}",
                    fmt_amt(od), fmt_amt(td), fmt_amt(tc), fmt_amt(ending)]
            for j, v in enumerate(vals):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter if j<=1 else Qt.AlignRight|Qt.AlignVCenter)
                if j == 3 and td: it.setForeground(QColor("#3d6fdb"))
                if j == 4 and tc: it.setForeground(QColor("#e05252"))
                if j == 5:
                    it.setForeground(QColor("#3d6fdb") if ending > 0 else
                                     QColor("#e05252") if ending < 0 else QColor("#888"))
                self.aux_rpt_tbl.setItem(i, j, it)
        # Totals
        n = len(rows); self.aux_rpt_tbl.setRowHeight(n, 38)
        for j, v in enumerate(["合计", "", "", fmt_amt(td_tot), fmt_amt(tc_tot), fmt_amt(td_tot-tc_tot)]):
            it = QTableWidgetItem(v)
            it.setFont(QFont("", weight=QFont.Bold)); it.setBackground(QColor("#f5f7fa"))
            it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter if j >= 2 else Qt.AlignLeft|Qt.AlignVCenter)
            self.aux_rpt_tbl.setItem(n, j, it)

    def _export_aux_report(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        period = self.rpt_period_edit.text().strip() or "report"
        path, _ = QFileDialog.getSaveFileName(self, "保存", f"往来对账_{period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "往来对账"
        hdrs = ["核算对象","科目","期初余额","本期借方","本期贷方","期末余额"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(1, ci, h); cell.font = XFont(bold=True, color="FFFFFF")
            cell.fill = fill; cell.alignment = Alignment(horizontal="center")
        for ri in range(self.aux_rpt_tbl.rowCount()):
            ws.append([self.aux_rpt_tbl.item(ri, ci).text() if self.aux_rpt_tbl.item(ri, ci) else ""
                       for ci in range(6)])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _export_items(self):
        if not self._cur_dim_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        path, _ = QFileDialog.getSaveFileName(self, "保存", "核算对象.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["编码", "名称", "联系人", "电话"])
        for r in self._items:
            ws.append([r["code"], r["name"], r["contact"], r["phone"]])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _load_bindings(self):
        if not self._cur_dim_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT a.code, a.name FROM account_aux_config ac
            JOIN accounts a ON a.client_id=ac.client_id AND a.code=ac.account_code
            WHERE ac.client_id=? AND ac.dimension_id=? ORDER BY a.code""",
                  (self.client_id, self._cur_dim_id))
        rows = c.fetchall(); conn.close()
        self.bind_list.clear()
        self._bindings = [dict(r) for r in rows]
        for r in self._bindings:
            self.bind_list.addItem(f"{r['code']}  {r['name']}")

    def _bind_account(self):
        if not self._cur_dim_id:
            QMessageBox.information(self, "提示", "请先选择维度"); return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code, name FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        accts = c.fetchall(); conn.close()
        # Already bound codes
        bound = {r["code"] for r in self._bindings}
        items = [f"{a['code']}  {a['name']}" for a in accts if a['code'] not in bound]
        if not items:
            QMessageBox.information(self, "提示", "所有科目均已绑定该维度"); return
        item, ok = QInputDialog.getItem(self, "绑定科目", "选择要绑定的科目：", items, editable=True)
        if not ok or not item: return
        code = item.split()[0]
        conn = get_db()
        try:
            conn.execute("INSERT INTO account_aux_config(client_id,account_code,dimension_id) VALUES(?,?,?)",
                         (self.client_id, code, self._cur_dim_id))
            conn.commit()
        except Exception:
            pass
        finally:
            conn.close()
        self._load_bindings()

    def _unbind_context_menu(self, pos):
        row = self.bind_list.currentRow()
        if row < 0: return
        menu = QMenu(self); menu.addAction("解除绑定")
        act = menu.exec(self.bind_list.mapToGlobal(pos))
        if act:
            code = self._bindings[row]["code"]
            conn = get_db()
            conn.execute("DELETE FROM account_aux_config WHERE client_id=? AND account_code=? AND dimension_id=?",
                         (self.client_id, code, self._cur_dim_id))
            conn.commit(); conn.close(); self._load_bindings()


class ImportAccountSetDialog(QDialog):
    """账套导入向导 — 四步骤，支持科目余额表/凭证/银行日记账"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("导入账套")
        self.setMinimumSize(860, 640)
        self.resize(960, 720)
        self._files = {}        # key -> path
        self._client_id = None
        self._build()

    def showEvent(self, event):
        super().showEvent(event)
        self._fit_to_screen()

    def _fit_to_screen(self):
        # macOS 下模态对话框有时会按默认几何打开，导致底部按钮栏超出可视区。
        screen = self.windowHandle().screen() if self.windowHandle() else QApplication.primaryScreen()
        if not screen:
            return
        avail = screen.availableGeometry()
        max_width = max(520, avail.width() - 80)
        max_height = max(420, avail.height() - 80)
        width = min(max(self.width(), 860), max_width)
        height = min(max(self.height(), 640), max_height)
        self.resize(width, height)
        x = avail.x() + (avail.width() - width) // 2
        y = avail.y() + (avail.height() - height) // 2
        self.move(max(avail.x(), x), max(avail.y(), y))

    # ─────────────────────────── UI 骨架 ───────────────────────────
    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # 顶部标题
        title_bar = QWidget()
        title_bar.setStyleSheet("background:#1c2340;")
        tl = QHBoxLayout(title_bar); tl.setContentsMargins(24,16,24,16)
        tl.addWidget(lbl("📥  账套导入向导", bold=True, color="#fff", size=16))
        tl.addSpacing(12)
        tl.addWidget(lbl("从用友 / 金蝶导出文件一键建立账套", color="#8b93ae", size=12))
        tl.addStretch(); L.addWidget(title_bar)

        # 左侧步骤导航 + 右侧内容
        body = QWidget(); bl = QHBoxLayout(body)
        bl.setContentsMargins(0,0,0,0); bl.setSpacing(0)

        nav = QWidget(); nav.setFixedWidth(176)
        nav.setStyleSheet("background:#f7f9fc;border-right:1px solid #e4e8f0;")
        nl = QVBoxLayout(nav); nl.setContentsMargins(0,14,0,14); nl.setSpacing(2)
        self._step_btns = []
        for num, text in [("1","客户信息"),("2","选择文件"),("3","预览确认"),("4","导入完成")]:
            sb = QPushButton(f"  {num}   {text}")
            sb.setStyleSheet("""QPushButton{background:transparent;color:#888;border:none;
                text-align:left;padding:11px 14px;border-left:3px solid transparent;}
                QPushButton[active=true]{background:#e6f0ff;color:#3d6fdb;
                border-left:3px solid #3d6fdb;font-weight:bold;}""")
            sb.setProperty("active","false"); nl.addWidget(sb)
            self._step_btns.append(sb)
        nl.addStretch(); bl.addWidget(nav)

        self.right_scroll = QScrollArea()
        self.right_scroll.setWidgetResizable(True)
        self.right_scroll.setFrameShape(QFrame.NoFrame)
        self.right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.right = QStackedWidget()
        self.right_scroll.setWidget(self.right)
        bl.addWidget(self.right_scroll)
        # stretch=1 让 body 撑满剩余空间，foot 保持固定高度不被挤走
        L.addWidget(body, 1)

        # 底部按钮栏 — 固定高度，确保始终可见
        foot = QWidget()
        foot.setObjectName("import_foot")
        foot.setMinimumHeight(76)
        foot.setMaximumHeight(76)
        fl = QHBoxLayout(foot)
        fl.setContentsMargins(20, 14, 20, 14)
        fl.setSpacing(10)
        fl.addStretch()
        self.btn_back = QPushButton("← 上一步"); self.btn_back.setObjectName("btn_gray")
        self.btn_back.setMinimumHeight(36)
        self.btn_back.setMinimumWidth(120)
        self.btn_back.clicked.connect(self._prev_step)
        self.btn_next = QPushButton("下一步 →"); self.btn_next.setObjectName("btn_primary")
        self.btn_next.setMinimumHeight(36)
        self.btn_next.setMinimumWidth(120)
        self.btn_next.clicked.connect(self._next_step)
        self.btn_close = QPushButton("关 闭"); self.btn_close.setObjectName("btn_gray")
        self.btn_close.setMinimumHeight(36)
        self.btn_close.setMinimumWidth(96)
        self.btn_close.clicked.connect(self.accept); self.btn_close.setVisible(False)
        fl.addWidget(self.btn_back); fl.addWidget(self.btn_next); fl.addWidget(self.btn_close)
        L.addWidget(foot, 0)   # stretch=0，高度由 setFixedHeight 决定

        self._build_step1(); self._build_step2(); self._build_step3(); self._build_step4()
        self._goto_step(0)

    # ─────────────────── Step 1 · 客户信息 ───────────────────
    def _build_step1(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(14)
        L.addWidget(lbl("填写客户（账套）基本信息", bold=True, size=14))
        hint = QLabel("  新建账套，或选择已有客户追加导入。同名账套自动追加，不会重复创建。")
        hint.setStyleSheet("background:#f6f8ff;color:#444;border-radius:6px;"
                           "padding:8px 12px;font-size:12px;")
        hint.setWordWrap(True); L.addWidget(hint)

        mode_row = QHBoxLayout()
        self.mode_new   = QPushButton("✦ 新建账套")
        self.mode_exist = QPushButton("◈ 追加到已有账套")
        for b in [self.mode_new, self.mode_exist]:
            b.setCheckable(True)
            b.setStyleSheet("""QPushButton{background:#f5f7fa;color:#555;
                border:1px solid #d9d9d9;border-radius:6px;padding:8px 18px;}
                QPushButton:checked{background:#3d6fdb;color:#fff;border-color:#3d6fdb;}""")
        self.mode_new.setChecked(True)
        self.mode_new.clicked.connect(lambda: self._toggle_mode(True))
        self.mode_exist.clicked.connect(lambda: self._toggle_mode(False))
        mode_row.addWidget(self.mode_new); mode_row.addWidget(self.mode_exist)
        mode_row.addStretch(); L.addLayout(mode_row)

        # 新建表单
        self.form_new = QWidget(); fn = QFormLayout(self.form_new)
        fn.setSpacing(10); fn.setLabelAlignment(Qt.AlignRight)
        self.f_name  = QLineEdit(); self.f_name.setPlaceholderText("公司全称（必填）")
        self.f_code  = QLineEdit(); self.f_code.setPlaceholderText("助记码，如 ZY")
        self.f_type  = QComboBox()
        self.f_type.addItems(["小规模纳税人","一般纳税人","其他"])
        self.f_taxid = QLineEdit(); self.f_taxid.setPlaceholderText("统一社会信用代码")
        fn.addRow("公司名称 *", self.f_name); fn.addRow("助记码", self.f_code)
        fn.addRow("客户类型", self.f_type);   fn.addRow("税号",    self.f_taxid)
        L.addWidget(self.form_new)

        # 已有账套
        self.form_exist = QWidget(); fe = QFormLayout(self.form_exist)
        fe.setSpacing(10); fe.setLabelAlignment(Qt.AlignRight)
        self.f_exist_combo = QComboBox(); self.f_exist_combo.setMinimumWidth(280)
        fe.addRow("选择已有账套", self.f_exist_combo)
        warn = QLabel("  ⚠ 追加模式：同期间同凭证号自动跳过，不会重复写入。")
        warn.setStyleSheet("color:#ad6800;font-size:12px;")
        fe.addRow("", warn)
        self.form_exist.setVisible(False); L.addWidget(self.form_exist)
        L.addStretch()
        self.right.addWidget(w)
        self._reload_exist_clients()

    def _reload_exist_clients(self):
        self.f_exist_combo.clear()
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT id,name FROM clients ORDER BY name")
        for r in c.fetchall():
            self.f_exist_combo.addItem(r["name"], r["id"])
        conn.close()

    def _toggle_mode(self, is_new):
        self.mode_new.setChecked(is_new); self.mode_exist.setChecked(not is_new)
        self.form_new.setVisible(is_new); self.form_exist.setVisible(not is_new)

    # ─────────────────── Step 2 · 选择文件 ───────────────────
    def _build_step2(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(10)
        L.addWidget(lbl("选择要导入的账务文件", bold=True, size=14))
        hint = QLabel("  支持用友T3/T6/U8、金蝶KIS/K3导出的 XLS/XLSX 文件。"
                       "所有类型均可选，选得越多导入越完整。")
        hint.setStyleSheet("background:#f6f8ff;color:#444;border-radius:6px;"
                           "padding:8px 12px;font-size:12px;")
        hint.setWordWrap(True); L.addWidget(hint)

        file_types = [
            ("balance", "📊 科目余额表",
             "含科目编号、名称、期初余额 — 自动建科目并写期初数（必选）"),
            ("voucher", "📝 记账凭证",
             "全部凭证分录，自动识别用友 / 通用模板格式"),
            ("bank",    "🏦 银行存款日记账",
             "银行流水逐笔明细，写入 bank_statements 表供对账"),
            ("ledger",  "📒 总账 / 明细账",
             "按科目汇总的发生额与余额，用于核对"),
            ("income",  "💹 利润表",
             "收入费用汇总，仅用于人工核对，不写入账套"),
            ("bs",      "🏛 资产负债表",
             "资产负债期末数，仅用于人工核对，不写入账套"),
        ]
        self._file_path_lbls = {}
        for key, label, desc in file_types:
            row_w = QFrame()
            row_w.setStyleSheet("QFrame{background:#fff;border:1px solid #e8ecf2;"
                                "border-radius:8px;}")
            rl = QHBoxLayout(row_w); rl.setContentsMargins(16,10,16,10); rl.setSpacing(10)
            vv = QVBoxLayout()
            vv.addWidget(lbl(label, bold=True))
            vv.addWidget(lbl(desc, color="#888", size=12))
            rl.addLayout(vv); rl.addStretch()
            path_lbl = QLabel("未选择")
            path_lbl.setStyleSheet("color:#bbb;font-size:11px;min-width:160px;max-width:220px;")
            b_sel = QPushButton("选择"); b_sel.setObjectName("btn_outline"); b_sel.setFixedWidth(72)
            b_clr = QPushButton("✕");   b_clr.setObjectName("btn_gray");    b_clr.setFixedSize(26,26)
            b_sel.clicked.connect(lambda _,k=key,pl=path_lbl: self._pick_file(k, pl))
            b_clr.clicked.connect(lambda _,k=key,pl=path_lbl: self._clear_file(k, pl))
            rl.addWidget(path_lbl); rl.addWidget(b_sel); rl.addWidget(b_clr)
            self._file_path_lbls[key] = path_lbl
            L.addWidget(row_w)
        L.addStretch()
        self.right.addWidget(w)

    def _pick_file(self, key, lbl_w):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择文件", "", "Excel 文件 (*.xls *.xlsx)")
        if not path: return
        self._files[key] = path
        import os; lbl_w.setText(os.path.basename(path))
        lbl_w.setStyleSheet("color:#3d6fdb;font-size:11px;")

    def _clear_file(self, key, lbl_w):
        self._files.pop(key, None)
        lbl_w.setText("未选择"); lbl_w.setStyleSheet("color:#bbb;font-size:11px;")

    # ─────────────────── Step 3 · 预览确认 ───────────────────
    def _build_step3(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(10)
        L.addWidget(lbl("预览与确认", bold=True, size=14))
        self.preview_info = QLabel("")
        self.preview_info.setStyleSheet("background:#f6f8ff;color:#333;border-radius:6px;"
                                        "padding:10px 14px;font-size:12px;")
        self.preview_info.setWordWrap(True); L.addWidget(self.preview_info)
        L.addWidget(lbl("数据预览（首个文件前 20 行）:", bold=True, size=12))
        self.preview_tbl = QTableWidget()
        self.preview_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.preview_tbl.setShowGrid(True)
        self.preview_tbl.verticalHeader().setVisible(False)
        L.addWidget(self.preview_tbl)
        self.right.addWidget(w)

    def _refresh_preview(self):
        lines = []
        if self.mode_new.isChecked():
            lines.append(f"新建账套：{self.f_name.text().strip() or '（未填）'}")
        else:
            lines.append(f"追加到：{self.f_exist_combo.currentText()}")
        labels = {"balance":"科目余额表","voucher":"记账凭证","bank":"银行日记账",
                  "ledger":"总账","income":"利润表","bs":"资产负债表"}
        for k, v in labels.items():
            import os
            mark = f"✓  {os.path.basename(self._files[k])}" if k in self._files else "○  跳过"
            lines.append(f"{v}：{mark}")
        self.preview_info.setText("\n".join(lines))

        first = next((k for k in ["balance","voucher","bank","ledger"] if k in self._files), None)
        if not first:
            self.preview_tbl.setRowCount(0); self.preview_tbl.setColumnCount(0); return
        try:
            df = self._read_df(self._files[first])
            df = df.iloc[:20]
            cols = min(df.shape[1], 10)
            self.preview_tbl.setColumnCount(cols); self.preview_tbl.setRowCount(len(df))
            for ri, row in df.iterrows():
                self.preview_tbl.setRowHeight(ri, 26)
                for ci in range(cols):
                    self.preview_tbl.setItem(ri, ci, QTableWidgetItem(str(row.iloc[ci])))
            self.preview_tbl.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeToContents)
        except Exception as e:
            self.preview_tbl.setRowCount(1); self.preview_tbl.setColumnCount(1)
            self.preview_tbl.setItem(0, 0, QTableWidgetItem(f"预览失败：{e}"))

    # ─────────────────── Step 4 · 导入完成 ───────────────────
    def _build_step4(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(10)
        self.result_icon  = lbl("⏳", size=40, color="#3d6fdb")
        self.result_icon.setAlignment(Qt.AlignCenter)
        self.result_title = lbl("正在导入…", bold=True, size=15)
        self.result_title.setAlignment(Qt.AlignCenter)
        L.addStretch()
        L.addWidget(self.result_icon); L.addWidget(self.result_title); L.addSpacing(8)
        self.log_box = QTextEdit(); self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(
            "font-size:12px;font-family:monospace;background:#fafafa;border-radius:6px;")
        L.addWidget(self.log_box); L.addStretch()
        self.right.addWidget(w)

    # ─────────────────── 步骤控制 ───────────────────
    def _goto_step(self, idx):
        self._cur_step = idx
        self.right.setCurrentIndex(idx)
        for i, b in enumerate(self._step_btns):
            b.setProperty("active", "true" if i == idx else "false")
            b.style().unpolish(b); b.style().polish(b)
        self.btn_back.setVisible(0 < idx < 3)
        self.btn_next.setVisible(idx < 3)
        self.btn_close.setVisible(idx == 3)
        if idx == 2: self._refresh_preview()
        if idx == 3: QTimer.singleShot(100, self._do_import)

    def _prev_step(self):
        if self._cur_step > 0: self._goto_step(self._cur_step - 1)

    def _next_step(self):
        if self._cur_step == 0:
            if self.mode_new.isChecked() and not self.f_name.text().strip():
                QMessageBox.warning(self, "提示", "请填写公司名称"); return
        if self._cur_step == 1 and not self._files:
            QMessageBox.information(self, "提示", "请至少选择一个文件"); return
        self._goto_step(self._cur_step + 1)

    # ─────────────────── 工具方法 ───────────────────
    def _read_df(self, path):
        import pandas as pd
        engine = "xlrd" if path.endswith(".xls") else "openpyxl"
        try:
            return pd.read_excel(path, header=None, dtype=str,
                                 engine=engine).fillna("")
        except Exception:
            import xlrd
            wb = xlrd.open_workbook(path, ignore_workbook_corruption=True)
            ws = wb.sheet_by_index(0)
            data = [[str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                    for r in range(ws.nrows)]
            import pandas as pd
            return pd.DataFrame(data).fillna("")

    def _flt(self, v):
        try: return float(str(v).replace(",", "").strip())
        except: return 0.0

    def _log(self, text):
        self.log_box.append(text)
        QApplication.processEvents()

    # ─────────────────── 核心导入 ───────────────────
    def _do_import(self):
        log = self._log
        log("🚀 开始导入账套…")
        conn = get_db(); c = conn.cursor()
        ok_total = err_total = 0
        try:
            # ── 创建 / 获取客户 ──
            if self.mode_new.isChecked():
                name  = self.f_name.text().strip()
                c.execute("SELECT id FROM clients WHERE name=?", (name,))
                row = c.fetchone()
                if row:
                    self._client_id = row[0]
                    log(f"ℹ 账套已存在，追加数据到：{name}")
                else:
                    c.execute(
                        "INSERT INTO clients(name,short_code,client_type,tax_id)"
                        " VALUES(?,?,?,?)",
                        (name, self.f_code.text().strip(),
                         self.f_type.currentText(),
                         self.f_taxid.text().strip()))
                    self._client_id = c.lastrowid
                    seed_client_accounts(self._client_id, conn)
                    conn.commit()
                    log(f"✓ 新建账套：{name}（ID={self._client_id}）")
            else:
                self._client_id = self.f_exist_combo.currentData()
                log(f"ℹ 追加到已有账套 ID={self._client_id}")

            # ── 1. 科目余额表 ──
            if "balance" in self._files:
                log("\n─── 导入科目余额表 ───")
                ok, err = self._imp_balance(conn, c)
                ok_total += ok; err_total += err

            # ── 2. 记账凭证 ──
            if "voucher" in self._files:
                log("\n─── 导入记账凭证 ───")
                ok, err = self._imp_voucher(conn, c)
                ok_total += ok; err_total += err

            # ── 3. 银行日记账 ──
            if "bank" in self._files:
                log("\n─── 导入银行日记账 ───")
                ok, err = self._imp_bank(conn, c)
                ok_total += ok; err_total += err

            # ── 4. 总账（补充科目期初） ──
            if "ledger" in self._files:
                log("\n─── 导入总账（补充期初） ───")
                ok, err = self._imp_ledger(conn, c)
                ok_total += ok; err_total += err

            # ── 5. 报表类 — 仅提示 ──
            for k, v in [("income","利润表"),("bs","资产负债表")]:
                if k in self._files:
                    log(f"\n○ {v} 已选择 — 仅供人工核对，不写入账套")

            log_action(conn, self._client_id, "账套导入", "import",
                       str(self._client_id),
                       f"文件:{list(self._files.keys())} 成功:{ok_total} 失败:{err_total}")
            conn.commit()
            log(f"\n✅ 导入完成！成功 {ok_total} 项，跳过/失败 {err_total} 项。")
            self.result_icon.setText("✅")
            self.result_title.setText("导入成功！")
        except Exception as e:
            conn.rollback()
            import traceback
            log(f"\n❌ 导入异常：{e}\n{traceback.format_exc()}")
            self.result_icon.setText("❌")
            self.result_title.setText("导入失败，请查看日志")
        finally:
            conn.close()

    # ── 科目余额表 ──
    def _imp_balance(self, conn, c):
        import re
        df = self._read_df(self._files["balance"])
        ok = err = 0
        for ri in range(len(df)):
            row = df.iloc[ri]
            code = str(row.iloc[1]).strip()
            name = str(row.iloc[2]).strip() if df.shape[1] > 2 else ""
            if not code or not name: continue
            if not re.match(r"^\d[\d._]*$", code): continue
            if len(code) < 4 or not code[:4].isdigit(): continue
            od = self._flt(row.iloc[3]) if df.shape[1] > 3 else 0
            oc = self._flt(row.iloc[4]) if df.shape[1] > 4 else 0
            # 推断类型 & 方向
            atype, direction = infer_account_type_direction(code, name)
            normalized = code.replace("_",".")
            level  = normalized.count(".") + 1
            parent = (code.rsplit(".",1)[0] if "." in code
                      else code.rsplit("_",1)[0] if "_" in code else None)
            c.execute("SELECT id FROM accounts WHERE client_id=? AND code=?",
                      (self._client_id, code))
            ex = c.fetchone()
            if ex:
                c.execute("UPDATE accounts SET name=?,full_name=?,type=?,direction=?,"
                          "opening_debit=?,opening_credit=? WHERE id=?",
                          (name, name, atype, direction, od, oc, ex[0]))
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,
                    direction,parent_code,level,opening_debit,opening_credit)
                    VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self._client_id,code,name,name,atype,direction,parent,level,od,oc))
            ok += 1
            if ok <= 5 or ok % 30 == 0:
                self._log(f"  {code}  {name}  借={od:.2f}  贷={oc:.2f}")
        self._log(f"  → 处理科目 {ok} 个")
        return ok, err

    # ── 记账凭证 ──
    def _imp_voucher(self, conn, c):
        import re
        df = self._read_df(self._files["voucher"])
        ok = skip = err = 0

        # 检测格式
        is_yonyou = any(
            "凭证字号" in str(df.iloc[r, 1] if df.shape[1] > 1 else "")
            for r in range(min(10, len(df))))

        title = str(df.iloc[0, 1]) if df.shape[1] > 1 else ""
        pm = re.search(r"(\d{4})年(\d{2})期", title)
        default_period = f"{pm.group(1)}-{pm.group(2)}" if pm else "2026-01"

        if is_yonyou:
            cur_vno = None; cur_date = ""; entries = []

            def flush(vno, date, ents):
                nonlocal ok, skip, err
                if not vno or not ents: return
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self._client_id, default_period, vno))
                if c.fetchone(): skip += 1; return
                td = sum(e[3] for e in ents); tc = sum(e[4] for e in ents)
                if abs(td - tc) > 0.01: err += 1; return
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status)"
                          " VALUES(?,?,?,?,?)",
                          (self._client_id, default_period, vno, date, "已审核"))
                vid = c.lastrowid
                for ln, ent in enumerate(ents, 1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,"
                              "account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid, ln) + ent)
                ok += 1
                if ok <= 5 or ok % 20 == 0:
                    self._log(f"  ✓ {vno}  {len(ents)} 行  借={td:.2f}")

            for ri in range(len(df)):
                row = df.iloc[ri]
                cell1 = str(row.iloc[1] if df.shape[1] > 1 else "").strip()
                if "凭证字号" in cell1:
                    flush(cur_vno, cur_date, entries); entries = []
                    dm = re.search(r"日期:(\S+)", cell1)
                    nm = re.search(r"凭证字号:(\S+)", cell1)
                    cur_date = dm.group(1) if dm else default_period + "-28"
                    cur_vno  = nm.group(1).split()[0] if nm else None
                elif not cell1 or cell1.startswith("合计"): continue
                else:
                    af = str(row.iloc[2] if df.shape[1] > 2 else "").strip()
                    if not af: continue
                    parts = af.split(" ", 1)
                    # Preserve auxiliary-account separators from source files.
                    code = parts[0]
                    aname = parts[1] if len(parts) > 1 else af
                    d  = self._flt(row.iloc[3]) if df.shape[1] > 3 else 0
                    cr = self._flt(row.iloc[4]) if df.shape[1] > 4 else 0
                    if d == 0 and cr == 0: continue
                    entries.append((cell1, code, aname, d, cr))
            flush(cur_vno, cur_date, entries)
        else:
            # 通用模板格式
            from collections import OrderedDict
            vouchers = OrderedDict()
            n = df.shape[1]
            for ri in range(1, len(df)):
                row = df.iloc[ri]
                def gcol(i, default=""):
                    try:
                        v = str(row.iloc[i]).strip() if i < n else default
                        return v if v not in ("nan","") else default
                    except: return default
                def gcolf(i):
                    try:
                        v = row.iloc[i] if i < n else 0
                        return self._flt(v) if v and str(v) not in ("nan","") else 0
                    except: return 0
                period = gcol(0); vno = gcol(1); date = gcol(2)
                summ   = gcol(3); code = gcol(4); aname = gcol(5)
                d = gcolf(6); cr = gcolf(7)
                if not period or not vno or not code: continue
                key = (period, vno)
                if key not in vouchers:
                    vouchers[key] = {"period":period,"vno":vno,"date":date,"entries":[]}
                vouchers[key]["entries"].append((summ, code, aname, d, cr))
            for (period, vno), v in vouchers.items():
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self._client_id, period, vno))
                if c.fetchone(): skip += 1; continue
                ents = v["entries"]
                td = sum(e[3] for e in ents); tc = sum(e[4] for e in ents)
                if abs(td - tc) > 0.01: err += 1; continue
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status)"
                          " VALUES(?,?,?,?,?)",
                          (self._client_id, period, vno, v["date"], "已审核"))
                vid = c.lastrowid
                for ln, ent in enumerate(ents, 1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,"
                              "account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid, ln) + ent)
                ok += 1

        self._log(f"  → 凭证导入 {ok} 张，跳过 {skip}，失败 {err}")
        return ok, skip + err

    # ── 银行日记账 → bank_statements ──
    def _imp_bank(self, conn, c):
        import re
        df = self._read_df(self._files["bank"])
        ok = skip = err = 0

        # 自动检测数据列布局
        # 用友明细账格式：col1=科目, col2=日期, col3=凭证号, col4=摘要,
        #                 col5=借方, col6=贷方, col8=余额
        # 通用银行流水格式：col0=日期, col1=摘要, col2=借/收入, col3=贷/支出, col4=余额

        # 先扫描找到第一个有 YYYY-MM-DD 格式日期的列和行
        date_col = None; data_start = 0
        for ri in range(min(15, len(df))):
            for ci in range(min(df.shape[1], 6)):
                v = str(df.iloc[ri, ci]).strip()
                if re.match(r"\d{4}[-/]\d{2}[-/]\d{2}", v):
                    date_col = ci; data_start = ri; break
            if date_col is not None: break

        if date_col is None:
            self._log("  ✗ 未能识别日期列，请确认文件格式"); return 0, 1

        # 根据 date_col 位置判断布局
        if date_col == 2:
            # 用友明细账：日期在 col2
            acct_col=1; vno_col=3; summ_col=4; d_col=5; cr_col=6; bal_col=8
        elif date_col == 0:
            # 通用银行流水：日期在 col0
            acct_col=None; vno_col=None; summ_col=1; d_col=2; cr_col=3; bal_col=4
        else:
            acct_col=None; vno_col=None; summ_col=date_col+1
            d_col=date_col+2; cr_col=date_col+3; bal_col=date_col+4

        def gcol(row, ci, default=""):
            try:
                v = str(row.iloc[ci]).strip() if ci is not None and ci < len(row) else default
                return v if v not in ("nan","") else default
            except: return default

        for ri in range(data_start, len(df)):
            row = df.iloc[ri]
            raw_date = gcol(row, date_col)
            # 标准化日期
            raw_date = raw_date.replace("/","-")
            if not re.match(r"\d{4}-\d{2}-\d{2}", raw_date): continue
            summary = gcol(row, summ_col)
            if summary in ("期初余额","本月合计","本年累计","合计",""): continue
            d  = self._flt(gcol(row, d_col))
            cr = self._flt(gcol(row, cr_col))
            bal = self._flt(gcol(row, bal_col)) if bal_col and bal_col < df.shape[1] else None
            if d == 0 and cr == 0: continue
            vno  = gcol(row, vno_col) if vno_col else ""
            # 科目信息
            acct_raw  = gcol(row, acct_col) if acct_col else "1002"
            parts = acct_raw.split(" ", 1)
            acct_code = parts[0] if re.match(r"^\d[\d.]*$", parts[0]) else "1002"
            acct_name = parts[1] if len(parts) > 1 else "银行存款"
            # 去重检查（日期+摘要+借+贷）
            c.execute("""SELECT id FROM bank_statements
                WHERE client_id=? AND date=? AND description=?
                AND debit=? AND credit=?""",
                (self._client_id, raw_date, summary, d, cr))
            if c.fetchone(): skip += 1; continue
            c.execute("""INSERT INTO bank_statements
                (client_id,account_code,account_name,date,voucher_no,
                 description,debit,credit,balance,source)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (self._client_id, acct_code, acct_name, raw_date,
                 vno, summary, d, cr, bal, "import"))
            ok += 1
            if ok <= 5 or ok % 50 == 0:
                self._log(f"  ✓ {raw_date}  {summary[:18]}  "
                          f"借={d:.2f}  贷={cr:.2f}")

        self._log(f"  → 银行流水导入 {ok} 条，跳过重复 {skip} 条")
        return ok, skip

    # ── 总账（补充期初） ──
    def _imp_ledger(self, conn, c):
        import re
        df = self._read_df(self._files["ledger"])
        ok = 0
        for ri in range(len(df)):
            row = df.iloc[ri]
            code = str(row.iloc[1]).strip() if df.shape[1] > 1 else ""
            name = str(row.iloc[2]).strip() if df.shape[1] > 2 else ""
            if not code or not name: continue
            if not re.match(r"^\d[\d._]*$", code): continue
            if len(code) < 4 or not code[:4].isdigit(): continue
            od = self._flt(row.iloc[3]) if df.shape[1] > 3 else 0
            oc = self._flt(row.iloc[4]) if df.shape[1] > 4 else 0
            if od == 0 and oc == 0: continue
            c.execute("SELECT id,opening_debit,opening_credit FROM accounts"
                      " WHERE client_id=? AND code=?", (self._client_id, code))
            ex = c.fetchone()
            if ex and ex["opening_debit"] == 0 and ex["opening_credit"] == 0:
                c.execute("UPDATE accounts SET opening_debit=?,opening_credit=? WHERE id=?",
                          (od, oc, ex["id"]))
                ok += 1
        self._log(f"  → 补充科目期初 {ok} 个")
        return ok, 0




class AccountEditDialog(QDialog):
    def __init__(self, parent, client_id, account=None, parent_acct=None):
        super().__init__(parent)
        self.client_id = client_id; self.account = account; self.parent_acct = parent_acct
        self.setWindowTitle("编辑科目" if account else "新增科目")
        self.setMinimumWidth(420); self._build()
        if account: self._load()
        elif parent_acct: self._prefill_from_parent()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(24,20,24,20); L.setSpacing(12)
        L.addWidget(lbl("科目信息", bold=True, size=15))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.code = QLineEdit(); self.code.setPlaceholderText("如 1002.01")
        self.name = QLineEdit(); self.name.setPlaceholderText("科目名称（必填）")
        self.full_name = QLineEdit(); self.full_name.setPlaceholderText("完整名称，如 银行存款-工商银行")
        self.type_cb = QComboBox()
        self.type_cb.addItems(["资产","负债","所有者权益","成本","收入","费用"])
        self.dir_cb = QComboBox(); self.dir_cb.addItems(["借","贷"])
        self.parent_cb = QComboBox(); self.parent_cb.addItem("（无上级）","")
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code",(self.client_id,))
        for r in c.fetchall():
            self.parent_cb.addItem(f"{r['code']}  {r['name']}", r['code'])
        conn.close()
        self.od = NoScrollDoubleSpinBox(); self.od.setRange(0,9999999999); self.od.setDecimals(2); self.od.setPrefix("¥ ")
        self.oc = NoScrollDoubleSpinBox(); self.oc.setRange(0,9999999999); self.oc.setDecimals(2); self.oc.setPrefix("¥ ")
        F.addRow("科目编号 *", self.code); F.addRow("科目名称 *", self.name)
        F.addRow("完整名称", self.full_name); F.addRow("科目类型", self.type_cb)
        F.addRow("余额方向", self.dir_cb); F.addRow("上级科目", self.parent_cb)
        F.addRow("期初借方", self.od); F.addRow("期初贷方", self.oc)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        bc = QPushButton("取消"); bc.setObjectName("btn_gray")
        bs = QPushButton("保存"); bs.setObjectName("btn_primary")
        bc.clicked.connect(self.reject); bs.clicked.connect(self._save)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        r = self.account
        self.code.setText(r["code"] or ""); self.name.setText(r["name"] or "")
        self.full_name.setText(r["full_name"] or "")
        idx = self.type_cb.findText(r["type"] or ""); self.type_cb.setCurrentIndex(max(0,idx))
        idx2 = self.dir_cb.findText(r["direction"] or "借"); self.dir_cb.setCurrentIndex(max(0,idx2))
        self.od.setValue(r["opening_debit"] or 0); self.oc.setValue(r["opening_credit"] or 0)
        if r["parent_code"]:
            for i in range(self.parent_cb.count()):
                if self.parent_cb.itemData(i) == r["parent_code"]:
                    self.parent_cb.setCurrentIndex(i); break

    def _prefill_from_parent(self):
        p = self.parent_acct
        self.code.setText(p["code"] + ".01")
        idx = self.type_cb.findText(p["type"] or ""); self.type_cb.setCurrentIndex(max(0,idx))
        idx2 = self.dir_cb.findText(p["direction"] or "借"); self.dir_cb.setCurrentIndex(max(0,idx2))
        for i in range(self.parent_cb.count()):
            if self.parent_cb.itemData(i) == p["code"]:
                self.parent_cb.setCurrentIndex(i); break

    def _save(self):
        code = self.code.text().strip(); name = self.name.text().strip()
        if not code or not name: QMessageBox.warning(self,"提示","编号和名称不能为空"); return
        full = self.full_name.text().strip() or name
        parent = self.parent_cb.currentData() or None
        level = (parent.count(".")+2) if parent else 1
        conn = get_db(); c = conn.cursor()
        try:
            if self.account:
                c.execute("""UPDATE accounts SET code=?,name=?,full_name=?,type=?,direction=?,
                    parent_code=?,level=?,opening_debit=?,opening_credit=? WHERE id=?""",
                    (code,name,full,self.type_cb.currentText(),self.dir_cb.currentText(),
                     parent,level,self.od.value(),self.oc.value(),self.account["id"]))
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,direction,
                    parent_code,level,opening_debit,opening_credit) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self.client_id,code,name,full,self.type_cb.currentText(),
                     self.dir_cb.currentText(),parent,level,self.od.value(),self.oc.value()))
            conn.commit(); conn.close(); self.accept()
        except Exception as e:
            conn.close(); QMessageBox.warning(self,"错误",f"保存失败：{e}")


class ImportExcelDialog(QDialog):
    """三合一导入：凭证 / 科目余额表期初 / 银行日记账"""

    def __init__(self, parent, client_id):
        super().__init__(parent)
        self.client_id = client_id
        self.setWindowTitle("从Excel导入历史数据")
        self.setMinimumSize(780, 560)
        self._build()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # Tab bar
        tab_bar = QWidget(); tab_bar.setStyleSheet("background:#f5f7fa;border-bottom:1px solid #e4e8f0;")
        tl = QHBoxLayout(tab_bar); tl.setContentsMargins(16,0,0,0); tl.setSpacing(0)
        self._itabs = []
        for name in ["记账凭证", "科目期初余额", "银行日记账"]:
            b = QPushButton(name)
            b.setStyleSheet("""QPushButton{background:transparent;color:#888;border:none;
                padding:11px 18px;border-bottom:2px solid transparent;}
                QPushButton:hover{color:#3d6fdb;}
                QPushButton[active=true]{color:#3d6fdb;border-bottom:2px solid #3d6fdb;font-weight:bold;}""")
            b.clicked.connect(lambda _,n=name: self._switch_itab(n))
            tl.addWidget(b); self._itabs.append(b)
        tl.addStretch()
        L.addWidget(tab_bar)

        self.istack = QStackedWidget(); L.addWidget(self.istack)
        self._build_voucher_import()
        self._build_balance_import()
        self._build_bank_import()
        self._switch_itab("记账凭证")

    def _switch_itab(self, name):
        mapping = {"记账凭证":0,"科目期初余额":1,"银行日记账":2}
        for b in self._itabs:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        self.istack.setCurrentIndex(mapping[name])

    # ── helpers ──
    def _make_tab(self, info_text):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,16,20,16); L.setSpacing(10)
        info = QLabel(info_text)
        info.setStyleSheet("background:#f6f8ff;border-radius:6px;padding:10px 14px;font-size:12px;color:#444;")
        info.setWordWrap(True); L.addWidget(info)
        log = QTextEdit(); log.setReadOnly(True)
        log.setStyleSheet("font-size:12px;font-family:monospace;background:#fafafa;")
        log.setPlaceholderText("导入日志将显示在这里…")
        L.addWidget(log)
        close_btn = QPushButton("关闭"); close_btn.setObjectName("btn_gray")
        close_btn.clicked.connect(self.accept)
        row = QHBoxLayout(); row.addStretch(); row.addWidget(close_btn); L.addLayout(row)
        return w, L, log

    def _pick_xls(self):
        path, _ = QFileDialog.getOpenFileName(self,"选择文件","","Excel(*.xls *.xlsx)")
        return path

    # ── Tab 1: 记账凭证 ──
    def _build_voucher_import(self):
        info = ("支持格式：从用友/金蝶等软件导出的记账凭证XLS文件。\n"
                "识别规则：每张凭证以[日期:...凭证字号:记-xxx]开头，下方各行为分录，[合计]行结束。\n"
                "也支持通用模板格式（A=期间 B=凭证号 C=日期 D=摘要 E=科目编号 F=科目名 G=借方 H=贷方）。")
        w, L, self.v_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_tmpl = QPushButton("↓ 下载通用模板"); b_tmpl.setObjectName("btn_outline")
        b_tmpl.clicked.connect(self._dl_voucher_template)
        b_imp = QPushButton("📂 导入凭证文件"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_vouchers)
        btn_row.addWidget(b_tmpl); btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _dl_voucher_template(self):
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        path, _ = QFileDialog.getSaveFileName(self,"保存模板","凭证导入模板.xlsx","Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="凭证数据"
        hdrs = ["期间(YYYY-MM)","凭证号","日期(YYYY-MM-DD)","摘要","科目编号","科目名称","借方","贷方"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell=ws.cell(1,ci,h); cell.font=XFont(bold=True,color="FFFFFF")
            cell.fill=fill; cell.alignment=Alignment(horizontal="center")
        samples = [
            ("2025-01","记-001","2025-01-15","发放工资","2211.001","应付职工薪酬-工资",5000,0),
            ("2025-01","记-001","2025-01-15","发放工资","1002.001","银行存款-光大银行",0,5000),
            ("2025-01","记-002","2025-01-20","收到货款","1002.001","银行存款-光大银行",10000,0),
            ("2025-01","记-002","2025-01-20","收到货款","6001.001","主营业务收入-咨询",0,10000),
        ]
        for r in samples: ws.append(list(r))
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=20
        wb.save(path); self.v_log.append(f"✓ 模板已保存: {path}")

    def _import_vouchers(self):
        path = self._pick_xls()
        if not path: return
        self.v_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.v_log.append(f"✗ 读取失败: {e}"); return

        # Detect format: 用友-style vs generic template
        # 用友 style: row with "日期:…凭证字号:" header rows
        is_yonyou = any("凭证字号" in str(df.iloc[r,1]) for r in range(min(10,len(df))))

        conn = get_db(); c = conn.cursor()
        ok = skip = err = 0

        if is_yonyou:
            # Parse 用友/金蝶 style
            # Extract period from title row (row 0, col 1: "2026年03期 凭证")
            title = str(df.iloc[0,1])
            import re
            pm = re.search(r"(\d{4})年(\d{2})期", title)
            period = f"{pm.group(1)}-{pm.group(2)}" if pm else "2026-01"
            self.v_log.append(f"检测到期间: {period}，开始解析…")

            cur_vno = None; cur_date = ""; entries = []

            def flush(vno, date, ents):
                nonlocal ok, skip, err
                if not vno or not ents: return
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self.client_id, period, vno))
                if c.fetchone():
                    self.v_log.append(f"  跳过 {vno}（已存在）"); skip += 1; return
                td = sum(e[3] for e in ents); tc = sum(e[4] for e in ents)
                if abs(td-tc) > 0.01:
                    self.v_log.append(f"  ✗ {vno} 借贷不平 差{td-tc:.2f}，跳过"); err += 1; return
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status) VALUES(?,?,?,?,?)",
                          (self.client_id,period,vno,date,"已审核"))
                vid = c.lastrowid
                for ln,ent in enumerate(ents,1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln)+ent)
                ok += 1
                self.v_log.append(f"  ✓ {vno}  {len(ents)}行  借={td:.2f}")

            for ri in range(len(df)):
                row = df.iloc[ri]
                cell1 = str(row.iloc[1]).strip()
                if "凭证字号" in cell1:
                    flush(cur_vno, cur_date, entries)
                    entries = []
                    dm = re.search(r"日期:(\S+)", cell1)
                    nm = re.search(r"凭证字号:(\S+)", cell1)
                    cur_date = dm.group(1) if dm else period+"-28"
                    cur_vno  = nm.group(1).split()[0] if nm else None
                elif cell1 in ("合计：","合计:","") or not cell1:
                    continue
                else:
                    # entry row: col1=summary, col2=account, col3=debit, col4=credit
                    acct_full = str(row.iloc[2]).strip()
                    if not acct_full: continue
                    parts = acct_full.split(" ", 1)
                    code = parts[0]; aname = parts[1] if len(parts)>1 else acct_full
                    # Preserve auxiliary-account separators from source files.
                    code_norm = code
                    try: d = float(row.iloc[3]) if row.iloc[3] else 0
                    except: d = 0
                    try: cr = float(row.iloc[4]) if row.iloc[4] else 0
                    except: cr = 0
                    if d == 0 and cr == 0: continue
                    entries.append((cell1, code_norm, aname, d, cr))
            flush(cur_vno, cur_date, entries)

        else:
            # Generic template format (cols: period,vno,date,summary,code,name,debit,credit)
            from collections import OrderedDict
            vouchers = OrderedDict()
            n_cols = df.shape[1]
            for ri in range(1, len(df)):
                row = df.iloc[ri]
                def gcol(i, default=""):
                    try: v = str(row.iloc[i]).strip() if i < n_cols else default; return v if v != "nan" else default
                    except: return default
                def gcol_f(i):
                    try: v = row.iloc[i] if i < n_cols else 0; return float(v) if v and str(v) != "nan" else 0
                    except: return 0
                period  = gcol(0); vno = gcol(1); date = gcol(2)
                summary = gcol(3); code = gcol(4); aname = gcol(5)
                d = gcol_f(6); cr = gcol_f(7)
                if not period or not vno or not code: continue
                key = (period, vno)
                if key not in vouchers:
                    vouchers[key] = {"period":period,"vno":vno,"date":date,"entries":[]}
                vouchers[key]["entries"].append((summary,code,aname,d,cr))
            for (period,vno),v in vouchers.items():
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self.client_id,period,vno))
                if c.fetchone(): skip+=1; continue
                ents = v["entries"]
                td=sum(e[3] for e in ents); tc=sum(e[4] for e in ents)
                if abs(td-tc)>0.01: err+=1; continue
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status) VALUES(?,?,?,?,?)",
                          (self.client_id,period,vno,v["date"],"已审核"))
                vid=c.lastrowid
                for ln,ent in enumerate(ents,1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln)+ent)
                ok+=1

        if ok:
            log_action(conn, self.client_id, "批量导入凭证", "import", "",
                       f"导入{ok}张凭证，跳过{skip}张，失败{err}张")
        conn.commit(); conn.close()
        self.v_log.append(f"\n✅ 完成：导入 {ok} 张，跳过 {skip} 张，失败 {err} 张")

    # ── Tab 2: 科目余额表期初 ──
    def _build_balance_import(self):
        info = ("支持格式：用友/金蝶导出的「科目余额表」XLS。\n"
                "识别列：科目编号（第2列）、科目名称（第3列）、期初借方（第4列）、期初贷方（第5列）。\n"
                "导入后自动创建不存在的科目，并设置期初余额。")
        w, L, self.b_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_imp = QPushButton("📂 导入科目余额表"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_balance)
        btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _import_balance(self):
        path = self._pick_xls()
        if not path: return
        self.b_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.b_log.append(f"✗ 读取失败: {e}"); return

        import re
        # Find data rows: col1=科目编号(numeric-ish), col2=科目名称, col3=期初借方, col4=期初贷方
        # Header rows are rows 0-5 typically; data starts where col1 looks like an account code
        conn = get_db(); c = conn.cursor()
        created = updated = skipped = 0

        for ri in range(len(df)):
            row = df.iloc[ri]
            code = str(row.iloc[1]).strip()
            name = str(row.iloc[2]).strip()
            if not code or not name: continue
            # Account codes: start with digit, may contain dots or underscores (auxiliary dims)
            if not re.match(r"^\d[\d._]*$", code): continue
            # Skip obvious non-account rows (e.g. "2026年03期" matched by looser regex)
            if len(code) < 4 or not code[:4].isdigit(): continue
            try:
                od = float(str(row.iloc[3]).replace(",","")) if row.iloc[3] else 0
            except: od = 0
            try:
                oc = float(str(row.iloc[4]).replace(",","")) if row.iloc[4] else 0
            except: oc = 0

            # Determine account type from code + name
            acct_type, direction = infer_account_type_direction(code, name)

            # Compute level and parent from code (treat _ same as . for hierarchy)
            normalized = code.replace("_", ".")
            level = normalized.count(".") + 1
            parent = code.rsplit(".", 1)[0] if "." in code else (
                     code.rsplit("_", 1)[0] if "_" in code else None)

            c.execute("SELECT id FROM accounts WHERE client_id=? AND code=?",
                      (self.client_id, code))
            existing = c.fetchone()
            if existing:
                c.execute("UPDATE accounts SET opening_debit=?,opening_credit=?,name=?,full_name=?,type=?,direction=? WHERE id=?",
                          (od, oc, name, name, acct_type, direction, existing[0]))
                updated += 1
                self.b_log.append(f"  ↻ {code} {name}  期初借={od:.2f} 贷={oc:.2f}")
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,direction,
                    parent_code,level,opening_debit,opening_credit) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self.client_id,code,name,name,acct_type,direction,parent,level,od,oc))
                created += 1
                self.b_log.append(f"  ✓ 新建 {code} {name}  期初借={od:.2f} 贷={oc:.2f}")

        conn.commit(); conn.close()
        self.b_log.append(f"\n✅ 完成：新建科目 {created} 个，更新期初 {updated} 个，跳过 {skipped} 个")

    # ── Tab 3: 银行日记账 ──
    def _build_bank_import(self):
        info = ("支持格式：用友/金蝶导出的「银行存款日记账 / 明细账」XLS。\n"
                "识别列：科目（第2列）、日期（第3列）、凭证号（第4列）、摘要（第5列）、借方（第6列）、贷方（第7列）、余额（第9列）。\n"
                "导入后可在「记账(凭证)→明细账」中查看，并可与实际银行流水对照。")
        w, L, self.k_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_imp = QPushButton("📂 导入银行日记账"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_bank)
        btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _import_bank(self):
        path = self._pick_xls()
        if not path: return
        self.k_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.k_log.append(f"✗ 读取失败: {e}"); return

        import re
        conn = get_db(); c = conn.cursor()
        ok = skip = 0

        # 自动识别日期列位置
        date_col = None; data_start = 0
        for ri in range(min(15, len(df))):
            for ci in range(min(df.shape[1], 6)):
                v = str(df.iloc[ri, ci]).strip()
                if re.match(r"\d{4}[-/]\d{2}[-/]\d{2}", v):
                    date_col = ci; data_start = ri; break
            if date_col is not None: break

        if date_col is None:
            self.k_log.append("✗ 未能识别日期列，请确认文件格式"); conn.close(); return

        # 根据 date_col 判断布局
        if date_col == 2:   # 用友明细账
            acct_col=1; vno_col=3; summ_col=4; d_col=5; cr_col=6; bal_col=8
        elif date_col == 0: # 通用银行流水
            acct_col=None; vno_col=None; summ_col=1; d_col=2; cr_col=3; bal_col=4
        else:
            acct_col=None; vno_col=None; summ_col=date_col+1
            d_col=date_col+2; cr_col=date_col+3; bal_col=date_col+4

        def gcol(row, ci, default=""):
            try:
                v = str(row.iloc[ci]).strip() if ci is not None and ci < len(row) else default
                return v if v not in ("nan","") else default
            except: return default

        def flt(v):
            try: return float(str(v).replace(",","").strip())
            except: return 0.0

        for ri in range(data_start, len(df)):
            row = df.iloc[ri]
            raw_date = gcol(row, date_col).replace("/","-")
            if not re.match(r"\d{4}-\d{2}-\d{2}", raw_date): continue
            summary = gcol(row, summ_col)
            if summary in ("期初余额","本月合计","本年累计","合计",""): continue
            d  = flt(gcol(row, d_col))
            cr = flt(gcol(row, cr_col))
            bal_v = gcol(row, bal_col) if bal_col and bal_col < df.shape[1] else ""
            bal = flt(bal_v) if bal_v else None
            if d == 0 and cr == 0: continue
            vno = gcol(row, vno_col) if vno_col else ""
            acct_raw = gcol(row, acct_col) if acct_col else "1002"
            parts = acct_raw.split(" ", 1)
            acct_code = parts[0] if re.match(r"^\d[\d.]*$", parts[0]) else "1002"
            acct_name = parts[1] if len(parts) > 1 else "银行存款"
            # 去重
            c.execute("""SELECT id FROM bank_statements
                WHERE client_id=? AND date=? AND description=? AND debit=? AND credit=?""",
                (self.client_id, raw_date, summary, d, cr))
            if c.fetchone(): skip += 1; continue
            c.execute("""INSERT INTO bank_statements
                (client_id,account_code,account_name,date,voucher_no,
                 description,debit,credit,balance,source)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (self.client_id, acct_code, acct_name, raw_date,
                 vno, summary, d, cr, bal, "import"))
            ok += 1
            self.k_log.append(f"  ✓ {raw_date}  {summary[:20]}  借={d:.2f}  贷={cr:.2f}")

        conn.commit(); conn.close()
        self.k_log.append(f"\n✅ 完成：导入 {ok} 条，跳过重复 {skip} 条")


