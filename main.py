import sys, os

# ── 启动前清理旧的字节码缓存，防止 .pyc 不一致导致静默崩溃 ──
_here = os.path.dirname(os.path.abspath(__file__))
_pycache = os.path.join(_here, '__pycache__')
if os.path.isdir(_pycache):
    import shutil, glob
    for _f in glob.glob(os.path.join(_pycache, 'main.*.pyc')):
        try: os.remove(_f)
        except: pass

sys.path.insert(0, _here)
from db import init_db, get_db, seed_client_accounts, STANDARD_ACCOUNTS, VOUCHER_TEMPLATES, log_action
from datetime import datetime, date
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, Signal, QTimer
from PySide6.QtGui import QColor, QFont, QBrush, QPalette
# openpyxl imported lazily inside each export function (saves ~600ms startup)

# ── Stylesheet ──────────────────────────────────────────────────────────────
SS = """
* { font-family:'Microsoft YaHei','PingFang SC',sans-serif; font-size:13px; color:#1e2130; }
QMainWindow,QWidget#root { background:#f0f2f5; }
/* Sidebar */
QWidget#sidebar { background:#1c2340; }
QPushButton#nav { background:transparent; color:#8b93ae; border:none;
    text-align:left; padding:11px 20px; border-radius:0; }
QPushButton#nav:hover { background:#252d4a; color:#fff; }
QPushButton#nav[active=true] { background:#2d3760; color:#fff;
    border-left:3px solid #4e7df4; padding-left:17px; }
QLabel#logo { color:#fff; font-size:17px; font-weight:bold;
    padding:22px 20px 6px 20px; }
QLabel#subt { color:#4a5578; font-size:11px; padding:0 20px 14px 20px; }
/* Cards */
QFrame#card { background:#fff; border-radius:10px; border:1px solid #e4e8f0; }
QWidget#import_foot { background:#fff; border-top:1px solid #e4e8f0; }
/* Buttons */
QPushButton#btn_primary { background:#3d6fdb; color:#fff; border:none;
    border-radius:6px; padding:7px 18px; font-weight:bold; }
QPushButton#btn_primary:hover { background:#2d5dc8; }
QPushButton#btn_red { background:#ff4d4f; color:#fff; border:none;
    border-radius:6px; padding:7px 14px; }
QPushButton#btn_red:hover { background:#e63b3d; }
QPushButton#btn_green { background:#52c41a; color:#fff; border:none;
    border-radius:6px; padding:7px 14px; }
QPushButton#btn_outline { background:transparent; color:#3d6fdb;
    border:1px solid #3d6fdb; border-radius:6px; padding:7px 14px; }
QPushButton#btn_outline:hover { background:#eef3ff; }
QPushButton#btn_gray { background:#f5f5f5; color:#666; border:1px solid #d9d9d9;
    border-radius:6px; padding:7px 14px; }
QPushButton#btn_gray:hover { background:#e8e8e8; }
/* Inputs */
QLineEdit,QDateEdit,QComboBox,QDoubleSpinBox,QSpinBox,QTextEdit {
    background:#fff; border:1px solid #d9d9d9; border-radius:5px;
    padding:6px 10px; }
QSpinBox,QDoubleSpinBox { qproperty-buttonSymbols: NoButtons; }
QLineEdit:focus,QDateEdit:focus,QDoubleSpinBox:focus { border:1.5px solid #3d6fdb; }
QComboBox::drop-down { border:none; width:22px; }
/* Tables */
QTableWidget { background:#fff; border:none; gridline-color:#f0f2f5;
    selection-background-color:#e6f0ff; selection-color:#1e2130; }
QTableWidget::item { padding:8px 10px; }
QHeaderView::section { background:#fafafa; color:#8b93ae; border:none;
    border-bottom:1px solid #e8ecf2; padding:8px 10px;
    font-size:12px; font-weight:bold; }
/* Tabs */
QTabBar::tab { background:transparent; color:#8b93ae; padding:9px 18px;
    border:none; border-bottom:2px solid transparent; }
QTabBar::tab:selected { color:#3d6fdb; border-bottom:2px solid #3d6fdb; }
QTabWidget::pane { border:none; }
/* Scrollbar */
QScrollBar:vertical { width:5px; background:transparent; }
QScrollBar::handle:vertical { background:#dde1ea; border-radius:2px; min-height:24px; }
QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical { height:0; }
QScrollBar:horizontal { height:5px; background:transparent; }
QScrollBar::handle:horizontal { background:#dde1ea; border-radius:2px; }
QScrollBar::add-line:horizontal,QScrollBar::sub-line:horizontal { width:0; }
"""

def lbl(text, bold=False, color=None, size=None):
    w = QLabel(text)
    st = ""
    if bold: st += "font-weight:bold;"
    if color: st += f"color:{color};"
    if size: st += f"font-size:{size}px;"
    if st: w.setStyleSheet(st)
    return w

def sep():
    f = QFrame(); f.setFrameShape(QFrame.HLine)
    f.setStyleSheet("color:#e8ecf2; margin:4px 0;"); return f

def card(widget=None):
    f = QFrame(); f.setObjectName("card")
    if widget:
        vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        vl.addWidget(widget)
    return f

class NoScrollSpinBox(QSpinBox):
    """QSpinBox that ignores mouse-wheel events to prevent accidental value changes."""
    def wheelEvent(self, event):
        event.ignore()

class NoScrollDoubleSpinBox(QDoubleSpinBox):
    """QDoubleSpinBox that ignores mouse-wheel events to prevent accidental value changes."""
    def wheelEvent(self, event):
        event.ignore()


def fmt_amt(v):
    if v == 0 or v is None: return ""
    return f"{v:,.2f}"

# Contra-asset account root codes (credit-normal)
_CONTRA_ASSET_ROOTS = {
    "1231","1471","1472","1482","1502","1512",
    "1602","1603","1608","1609","1622","1632","1702","1703"
}

def infer_account_type_direction(code, name=""):
    """Infer account type and normal-balance direction from code and name.
    Matches the standard chart of accounts template (4xxx=equity, 5xxx=cost, 6xxx=income/expense).
    """
    code = (code or "").strip()
    name = (name or "").strip()
    if not code:
        return "资产", "借"

    prefix = code[0]
    parts = code.split(".")
    root4 = code[:4] if len(code) >= 4 else code

    if prefix == "1":
        # Check contra-asset (credit normal): traverse ancestors
        for i in range(len(parts), 0, -1):
            ancestor = ".".join(parts[:i])
            if ancestor in _CONTRA_ASSET_ROOTS:
                return "资产", "贷"
        return "资产", "借"

    if prefix == "2":
        # 2702 未确认融资费用 is a contra-liability (debit normal)
        if code.startswith("2702"):
            return "负债", "借"
        return "负债", "贷"

    if prefix == "3":
        # In this template 3xxx are special clearing/hedging accounts (asset-like)
        return "资产", "借"

    if prefix == "4":
        if code.startswith("4201"):   # 库存股 — debit normal
            return "所有者权益", "借"
        return "所有者权益", "贷"

    if prefix == "5":
        if code.startswith("5402"):   # 工程结算 — credit normal
            return "成本", "贷"
        return "成本", "借"

    if prefix == "6":
        try:
            top4 = int(code[:4])
        except ValueError:
            top4 = 9999
        if top4 <= 6301:              # 6001-6301 are income/gain accounts
            return "收入", "贷"
        return "费用", "借"           # 6401+ are expense accounts

    return "资产", "借"

def cn_amount(n):
    units = ["", "拾", "佰", "仟", "万", "拾万", "佰万", "仟万", "亿"]
    digits = "零壹贰叁肆伍陆柒捌玖"
    if n == 0: return "零元整"
    sign = "负" if n < 0 else ""
    n = abs(round(n, 2))
    i_part = int(n); d_part = round((n - i_part) * 100)
    fen = digits[d_part % 10]; jiao = digits[d_part // 10]
    result = ""
    s = str(i_part); idx = 0
    for ch in reversed(s):
        d = int(ch)
        result = (digits[d] + units[idx] if d != 0 else "零") + result
        idx += 1
    result = result.rstrip("零") or "零"
    if d_part == 0: result += "元整"
    elif d_part % 10 == 0: result += f"元{jiao}角"
    else: result += f"元{jiao}角{fen}分"
    return sign + result

# ── Dialogs ────────────────────────────────────────────────────────────────

class ClientDialog(QDialog):
    def __init__(self, parent=None, client=None):
        super().__init__(parent)
        self.client = client
        self.setWindowTitle("新建客户" if not client else "编辑客户")
        self.setMinimumWidth(400)
        self._build()
        if client: self._load()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(24,24,24,24); L.setSpacing(14)
        L.addWidget(lbl("客户信息", bold=True, size=15))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.name = QLineEdit(); self.name.setPlaceholderText("公司全称（必填）")
        self.code = QLineEdit(); self.code.setPlaceholderText("如 ZY")
        self.typ = QComboBox(); self.typ.addItems(["小规模纳税人","一般纳税人","其他"])
        self.taxid = QLineEdit(); self.taxid.setPlaceholderText("统一社会信用代码")
        self.contact = QLineEdit(); self.phone = QLineEdit()
        F.addRow("公司名称 *", self.name); F.addRow("助记码", self.code)
        F.addRow("客户类型", self.typ);   F.addRow("税号", self.taxid)
        F.addRow("联系人", self.contact); F.addRow("电话", self.phone)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        b_cancel = QPushButton("取消"); b_cancel.setObjectName("btn_gray")
        b_save = QPushButton("保 存"); b_save.setObjectName("btn_primary")
        b_cancel.clicked.connect(self.reject); b_save.clicked.connect(self._save)
        row.addWidget(b_cancel); row.addWidget(b_save); L.addLayout(row)

    def _load(self):
        c = self.client
        self.name.setText(c["name"] or ""); self.code.setText(c["short_code"] or "")
        self.taxid.setText(c["tax_id"] or ""); self.contact.setText(c["contact"] or "")
        self.phone.setText(c["phone"] or "")
        idx = self.typ.findText(c["client_type"] or ""); self.typ.setCurrentIndex(max(0,idx))

    def _save(self):
        n = self.name.text().strip()
        if not n: QMessageBox.warning(self,"提示","公司名称不能为空"); return
        conn = get_db(); c = conn.cursor()
        d = (n, self.code.text().strip(), self.typ.currentText(),
             self.taxid.text().strip(), self.contact.text().strip(), self.phone.text().strip())
        if self.client:
            c.execute("UPDATE clients SET name=?,short_code=?,client_type=?,tax_id=?,contact=?,phone=? WHERE id=?",
                      d+(self.client["id"],))
        else:
            c.execute("INSERT INTO clients(name,short_code,client_type,tax_id,contact,phone) VALUES(?,?,?,?,?,?)",d)
            cid = c.lastrowid
            seed_client_accounts(cid, conn)   # reuse same connection — no lock
        conn.commit(); conn.close(); self.accept()


class AccountInitDialog(QDialog):
    """科目期初设置"""
    def __init__(self, parent, client_id, period):
        super().__init__(parent)
        self.client_id = client_id; self.period = period
        self.setWindowTitle("科目期初余额"); self.setMinimumSize(680, 500)
        self._build()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(16,16,16,16); L.setSpacing(10)
        L.addWidget(lbl("科目期初余额设置", bold=True, size=14))
        L.addWidget(lbl(f"期间：{self.period}  （仅显示末级科目）", color="#888"))
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["科目编号","科目名称","期初借方","期初贷方"])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        hh.setMinimumSectionSize(80)
        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(2, 160)
        self.table.setColumnWidth(3, 160)
        self.table.verticalHeader().setVisible(False)
        self._load()
        L.addWidget(self.table)
        row = QHBoxLayout(); row.addStretch()
        bs = QPushButton("保存期初"); bs.setObjectName("btn_primary")
        bc = QPushButton("关闭"); bc.setObjectName("btn_gray")
        bs.clicked.connect(self._save); bc.clicked.connect(self.accept)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT id,code,name,opening_debit,opening_credit FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        rows = c.fetchall(); conn.close()
        self.table.setRowCount(len(rows))
        self._ids = []
        for i,r in enumerate(rows):
            self.table.setRowHeight(i, 40)          # 行高足够容纳输入框
            self._ids.append(r["id"])
            code_it = QTableWidgetItem(r["code"])
            code_it.setForeground(QColor("#3d6fdb"))
            name_it = QTableWidgetItem(r["name"])
            self.table.setItem(i,0,code_it)
            self.table.setItem(i,1,name_it)
            # Spinbox with explicit minimum size so numbers are readable
            def make_spin(val):
                sp = NoScrollDoubleSpinBox()
                sp.setRange(0, 9999999999)
                sp.setDecimals(2)
                sp.setValue(val or 0)
                sp.setMinimumHeight(32)
                sp.setMinimumWidth(140)
                sp.setAlignment(Qt.AlignRight)
                sp.setStyleSheet("QDoubleSpinBox{padding:4px 8px;font-size:13px;}")
                return sp
            d_spin  = make_spin(r["opening_debit"])
            cr_spin = make_spin(r["opening_credit"])
            self.table.setCellWidget(i,2,d_spin)
            self.table.setCellWidget(i,3,cr_spin)

    def _save(self):
        conn = get_db(); c = conn.cursor()
        for i,aid in enumerate(self._ids):
            d = self.table.cellWidget(i,2).value()
            cr = self.table.cellWidget(i,3).value()
            c.execute("UPDATE accounts SET opening_debit=?,opening_credit=? WHERE id=?", (d,cr,aid))
        conn.commit(); conn.close()
        QMessageBox.information(self,"成功","期初余额已保存")


class VoucherDialog(QDialog):
    """新增/编辑凭证 — 智一会计风格"""
    def __init__(self, parent=None, client_id=None, period=None, voucher_id=None):
        super().__init__(parent)
        self.client_id = client_id; self.period = period; self.voucher_id = voucher_id
        self.setWindowTitle("编辑凭证" if voucher_id else "新增凭证")
        self.setMinimumSize(860, 580)
        self._accounts = self._fetch_accounts()
        self._templates = VOUCHER_TEMPLATES
        from PySide6.QtCore import QStringListModel
        self._account_completion_model = QStringListModel(
            [f"{a['code']}  {a['full_name']}" for a in self._accounts], self)
        self._row_completers = []
        self._build()
        if voucher_id: self._load_voucher()
        else: self._add_row(); self._add_row()

    def _fetch_accounts(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name,full_name,direction FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        r = [dict(x) for x in c.fetchall()]
        # Load aux config: account_code -> list of (dim_id, dim_name, items)
        c.execute("""SELECT ac.account_code, ad.id as dim_id, ad.name as dim_name
            FROM account_aux_config ac
            JOIN aux_dimensions ad ON ad.id=ac.dimension_id
            WHERE ac.client_id=? ORDER BY ac.account_code, ad.sort_order""",
                  (self.client_id,))
        self._aux_config = {}   # account_code -> [{dim_id, dim_name}]
        for row in c.fetchall():
            self._aux_config.setdefault(row["account_code"], []).append(
                {"dim_id": row["dim_id"], "dim_name": row["dim_name"]})
        # Load all aux items per dimension
        c.execute("SELECT id, dimension_id, name, code FROM aux_items WHERE client_id=? AND is_active=1 ORDER BY dimension_id,id",
                  (self.client_id,))
        self._aux_items = {}    # dim_id -> [{id, name, code}]
        for row in c.fetchall():
            self._aux_items.setdefault(row["dimension_id"], []).append(
                {"id": row["id"], "name": row["name"], "code": row["code"] or ""})
        conn.close(); return r

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(16,14,16,14); L.setSpacing(10)

        # Header bar
        hdr = QHBoxLayout()
        # Voucher number + date + preparer
        self.lbl_no = lbl("新 建", bold=True, color="#3d6fdb", size=14)
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True); self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.preparer_lbl = lbl("未来", color="#888")
        self.attach_spin = NoScrollSpinBox(); self.attach_spin.setRange(0,999)
        self.attach_spin.setSuffix(" 张"); self.attach_spin.setFixedWidth(70)
        hdr.addWidget(self.lbl_no); hdr.addSpacing(16)
        hdr.addWidget(lbl("日期：")); hdr.addWidget(self.date_edit)
        hdr.addWidget(lbl("  制单：")); hdr.addWidget(self.preparer_lbl)
        hdr.addWidget(lbl("  附单据")); hdr.addWidget(self.attach_spin)
        hdr.addStretch()
        # Template button
        tpl_btn = QPushButton("凭证模板 ▼"); tpl_btn.setObjectName("btn_outline")
        tpl_btn.clicked.connect(self._show_template_menu)
        b_save_new = QPushButton("保存并新增"); b_save_new.setObjectName("btn_primary")
        b_save_new.clicked.connect(lambda: self._save(and_new=True))
        b_save = QPushButton("保 存"); b_save.setObjectName("btn_primary")
        b_save.clicked.connect(lambda: self._save(and_new=False))
        hdr.addWidget(tpl_btn); hdr.addWidget(b_save_new); hdr.addWidget(b_save)
        L.addLayout(hdr)
        L.addWidget(sep())

        # Entry table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["摘  要","科  目","核算对象","借方金额","贷方金额"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(2,160); self.table.setColumnWidth(3,130); self.table.setColumnWidth(4,130)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setMinimumHeight(280)
        L.addWidget(self.table)

        # Add/del row buttons
        row_btns = QHBoxLayout()
        b_add = QPushButton("＋ 增行"); b_add.setObjectName("btn_outline")
        b_add.clicked.connect(lambda: self._add_row())
        b_del = QPushButton("－ 删行"); b_del.setObjectName("btn_gray")
        b_del.clicked.connect(lambda: self._del_row())
        row_btns.addWidget(b_add); row_btns.addWidget(b_del); row_btns.addStretch()
        L.addLayout(row_btns)
        L.addWidget(sep())

        # Total row
        tot = QHBoxLayout()
        tot.addWidget(lbl("合计金额：")); self.lbl_cn = lbl("零元整", bold=True, color="#1e2130")
        tot.addWidget(self.lbl_cn); tot.addStretch()
        self.lbl_debit = lbl("借方合计：0.00", color="#3d6fdb", bold=True)
        self.lbl_credit = lbl("贷方合计：0.00", color="#e05252", bold=True)
        self.lbl_balance = lbl("✓ 借贷平衡", color="#52c41a", bold=True)
        tot.addWidget(self.lbl_debit); tot.addSpacing(20)
        tot.addWidget(self.lbl_credit); tot.addSpacing(20)
        tot.addWidget(self.lbl_balance)
        L.addLayout(tot)

    def _add_row(self, summary="", acct_code="", acct_name="", debit=0, credit=0, aux_data=None):
        """aux_data: list of (dim_id, item_id, item_name) saved previously"""
        i = self.table.rowCount(); self.table.insertRow(i)
        self.table.setRowHeight(i, 44)

        summary_edit = QLineEdit(summary)
        summary_edit.setPlaceholderText("摘要...")

        # ── 科目搜索框（自定义 completer，支持编号/名称任意位置模糊匹配） ──
        acct_edit = QLineEdit()
        acct_edit.setPlaceholderText("输入编号或名称搜索...")
        acct_edit._code = acct_code  # store selected code

        # Build display list and lookup maps
        display_list = [f"{a['code']}  {a['full_name']}" for a in self._accounts]
        code_by_display = {f"{a['code']}  {a['full_name']}": a['code'] for a in self._accounts}
        display_by_code = {a['code']: f"{a['code']}  {a['full_name']}" for a in self._accounts}

        from PySide6.QtWidgets import QCompleter
        completer = QCompleter(self._account_completion_model, acct_edit)
        completer.setFilterMode(Qt.MatchContains)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.popup().setMinimumWidth(380)
        acct_edit.setCompleter(completer)
        acct_edit._completer = completer
        self._row_completers.append(completer)

        # Restore pre-filled value
        if acct_code and acct_code in display_by_code:
            acct_edit.setText(display_by_code[acct_code])

        def on_acct_selected(text):
            code = code_by_display.get(text, "")
            acct_edit._code = code
            rebuild_aux(code)

        def on_acct_edited(text):
            # If user cleared the field, clear code
            if not text:
                acct_edit._code = ""
                rebuild_aux("")
            # If exact match, update code
            if text in code_by_display:
                acct_edit._code = code_by_display[text]
                rebuild_aux(acct_edit._code)
            if text:
                completer.setCompletionPrefix(text)
                completer.complete()

        completer.activated.connect(on_acct_selected)
        acct_edit.textChanged.connect(on_acct_edited)

        # ── 辅助核算 ──
        aux_container = QWidget()
        aux_layout = QHBoxLayout(aux_container)
        aux_layout.setContentsMargins(2,2,2,2); aux_layout.setSpacing(4)
        aux_container._combos = []

        def rebuild_aux(code, saved=None):
            for w in aux_container._combos:
                w[1].setParent(None)
            aux_container._combos.clear()
            dims = self._aux_config.get(code, [])
            for dim in dims:
                cb = QComboBox(); cb.setMinimumWidth(130)
                cb.addItem("— 选核算对象 —", None)
                for it in self._aux_items.get(dim["dim_id"], []):
                    display = f"{it['code']} {it['name']}" if it['code'] else it['name']
                    cb.addItem(display, it["id"])
                cb.setToolTip(dim["dim_name"])
                if saved:
                    for sv_dim, sv_id, sv_name in saved:
                        if sv_dim == dim["dim_id"]:
                            for ki in range(cb.count()):
                                if cb.itemData(ki) == sv_id:
                                    cb.setCurrentIndex(ki); break
                aux_layout.addWidget(cb)
                aux_container._combos.append((dim["dim_id"], cb))
            if not dims:
                aux_layout.addWidget(QLabel("—"))

        rebuild_aux(acct_code, aux_data)

        # ── 借方金额 ──
        d_spin = NoScrollDoubleSpinBox(); d_spin.setRange(0,999999999); d_spin.setDecimals(2)
        d_spin.setSpecialValueText(""); d_spin.setValue(debit)
        d_spin.valueChanged.connect(self._update_totals)

        # ── 贷方金额（支持 = 键自动补平） ──
        cr_spin = NoScrollDoubleSpinBox(); cr_spin.setRange(0,999999999); cr_spin.setDecimals(2)
        cr_spin.setSpecialValueText(""); cr_spin.setValue(credit)
        cr_spin.valueChanged.connect(self._update_totals)

        # Patch keyPressEvent on cr_spin to handle "=" auto-balance
        _row_ref = [i]  # mutable ref so lambda can find the row
        orig_key = cr_spin.keyPressEvent
        def cr_key(event, _cr=cr_spin):
            if event.text() == "=":
                self._auto_balance_credit(_cr)
            else:
                orig_key(event)
        cr_spin.keyPressEvent = cr_key

        self.table.setCellWidget(i,0,summary_edit)
        self.table.setCellWidget(i,1,acct_edit)
        self.table.setCellWidget(i,2,aux_container)
        self.table.setCellWidget(i,3,d_spin)
        self.table.setCellWidget(i,4,cr_spin)
        self._update_totals()

    def _auto_balance_credit(self, target_spin):
        """按 = 键时，将 target_spin 设为令借贷平衡所需的金额。"""
        td = tc_other = 0
        for i in range(self.table.rowCount()):
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            if dw: td += dw.value()
            if cw and cw is not target_spin: tc_other += cw.value()
        needed = max(0, round(td - tc_other, 2))
        target_spin.setValue(needed)
        self._update_totals()

    def _del_row(self):
        row = self.table.currentRow()
        if row < 0:
            row = self.table.rowCount() - 1
        if row >= 0 and self.table.rowCount() > 1:
            self.table.removeRow(row)
            self._update_totals()

    def _update_totals(self):
        td = tc = 0
        for i in range(self.table.rowCount()):
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            if dw: td += dw.value()
            if cw: tc += cw.value()
        self.lbl_debit.setText(f"借方合计：{td:,.2f}")
        self.lbl_credit.setText(f"贷方合计：{tc:,.2f}")
        balanced = abs(td-tc) < 0.005
        self.lbl_balance.setText("✓ 借贷平衡" if balanced else f"✗ 差额 {abs(td-tc):,.2f}")
        self.lbl_balance.setStyleSheet(f"color:{'#52c41a' if balanced else '#ff4d4f'};font-weight:bold;")
        self.lbl_cn.setText(cn_amount(td))

    def _load_templates_from_db(self):
        """Load user-saved templates from database, merged with built-in templates."""
        conn = get_db(); c = conn.cursor()
        try:
            c.execute("SELECT name, entries FROM voucher_templates WHERE client_id=? OR client_id IS NULL ORDER BY id",
                      (self.client_id,))
            db_tmpls = c.fetchall()
        except Exception:
            db_tmpls = []
        conn.close()
        import json
        user_tmpls = []
        for row in db_tmpls:
            try:
                entries = json.loads(row["entries"])
                user_tmpls.append((row["name"], entries))
            except Exception:
                pass
        # Merge: built-in first, then user-saved
        self._templates = list(VOUCHER_TEMPLATES) + user_tmpls

    def _show_template_menu(self):
        self._load_templates_from_db()
        menu = QMenu(self)
        for name, _ in self._templates:
            menu.addAction(name)
        menu.addSeparator()
        b_save_tpl = menu.addAction("📌 存为模板...")
        b_del_tpl  = menu.addAction("🗑 删除模板...")
        act = menu.exec(self.sender().mapToGlobal(self.sender().rect().bottomLeft()))
        if not act: return

        if act is b_save_tpl:
            self._save_as_template()
            return
        if act is b_del_tpl:
            self._delete_template()
            return

        # Apply selected template
        for name, entries in self._templates:
            if act.text() == name:
                self.table.setRowCount(0)
                for entry in entries:
                    # Support both old tuple format and new dict format
                    if isinstance(entry, dict):
                        self._add_row(entry.get("summary",""), entry.get("code",""),
                                      entry.get("name",""), entry.get("debit",0), entry.get("credit",0))
                    else:
                        s, code, _, d, cr = entry
                        self._add_row(s, code, "", d, cr)
                break

    def _save_as_template(self):
        """Save current voucher entries as a reusable template."""
        import json
        # Collect current entries
        entries = []
        for i in range(self.table.rowCount()):
            sw = self.table.cellWidget(i,0); aw = self.table.cellWidget(i,1)
            dw = self.table.cellWidget(i,3); cw = self.table.cellWidget(i,4)
            code = getattr(aw, '_code', "") or "" if aw else ""
            summary = sw.text().strip() if sw else ""
            d = dw.value() if dw else 0
            cr = cw.value() if cw else 0
            if not code and d == 0 and cr == 0: continue
            # Find account name
            aname = ""
            for a in self._accounts:
                if a['code'] == code: aname = a['full_name']; break
            entries.append({"summary": summary, "code": code, "name": aname,
                            "debit": d, "credit": cr})
        if not entries:
            QMessageBox.warning(self, "提示", "请先填写凭证分录再保存为模板"); return

        name, ok = QInputDialog.getText(self, "保存为模板", "模板名称：",
                                         text=entries[0].get("summary","") or "新模板")
        if not ok or not name.strip(): return
        name = name.strip()

        # Check duplicate name
        all_names = [t[0] for t in self._templates]
        if name in all_names:
            if QMessageBox.question(self, "覆盖确认", f"模板【{name}】已存在，是否覆盖？",
                    QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return
            conn = get_db()
            conn.execute("DELETE FROM voucher_templates WHERE name=? AND (client_id=? OR client_id IS NULL)",
                         (name, self.client_id))
            conn.commit(); conn.close()

        conn = get_db()
        conn.execute("INSERT INTO voucher_templates(client_id, name, entries) VALUES(?,?,?)",
                     (self.client_id, name, json.dumps(entries, ensure_ascii=False)))
        conn.commit(); conn.close()
        QMessageBox.information(self, "成功", f"已保存模板：{name}")

    def _delete_template(self):
        """Delete a user-saved template."""
        conn = get_db(); c = conn.cursor()
        try:
            c.execute("SELECT id, name FROM voucher_templates WHERE client_id=? OR client_id IS NULL ORDER BY id",
                      (self.client_id,))
            user_tmpls = c.fetchall()
        except Exception:
            user_tmpls = []
        conn.close()
        if not user_tmpls:
            QMessageBox.information(self, "提示", "没有可删除的自定义模板（内置模板不能删除）")
            return
        names = [r["name"] for r in user_tmpls]
        name, ok = QInputDialog.getItem(self, "删除模板", "选择要删除的模板：", names, editable=False)
        if not ok or not name: return
        conn = get_db()
        conn.execute("DELETE FROM voucher_templates WHERE name=? AND (client_id=? OR client_id IS NULL)",
                     (name, self.client_id))
        conn.commit(); conn.close()
        QMessageBox.information(self, "成功", f"已删除模板：{name}")

    def _load_voucher(self):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM vouchers WHERE id=?", (self.voucher_id,))
        v = c.fetchone()
        self.lbl_no.setText(v["voucher_no"])
        self.date_edit.setDate(QDate.fromString(v["date"],"yyyy-MM-dd"))
        self.attach_spin.setValue(v["attachment_count"] or 0)
        c.execute("SELECT * FROM voucher_entries WHERE voucher_id=? ORDER BY line_no", (self.voucher_id,))
        entries = c.fetchall()
        # Load aux data per entry
        entry_ids = [e["id"] for e in entries]
        aux_by_entry = {}
        if entry_ids:
            placeholders = ",".join("?" * len(entry_ids))
            c.execute(f"SELECT entry_id,dimension_id,aux_item_id,aux_item_name FROM voucher_entry_aux WHERE entry_id IN ({placeholders})",
                      entry_ids)
            for row in c.fetchall():
                aux_by_entry.setdefault(row["entry_id"], []).append(
                    (row["dimension_id"], row["aux_item_id"], row["aux_item_name"]))
        conn.close()
        for e in entries:
            aux_data = aux_by_entry.get(e["id"], None)
            self._add_row(e["summary"] or "", e["account_code"] or "", e["account_name"] or "",
                          e["debit"] or 0, e["credit"] or 0, aux_data=aux_data)

    def _save(self, and_new=False):
        # Validate
        entries = []
        for i in range(self.table.rowCount()):
            sw  = self.table.cellWidget(i,0); aw  = self.table.cellWidget(i,1)
            auxw= self.table.cellWidget(i,2)
            dw  = self.table.cellWidget(i,3); cw  = self.table.cellWidget(i,4)
            if not aw: continue
            code = getattr(aw, '_code', "") or ""
            d = dw.value() if dw else 0; cr = cw.value() if cw else 0
            if not code and d == 0 and cr == 0: continue
            if not code: QMessageBox.warning(self,"提示",f"第{i+1}行科目不能为空"); return
            aname = ""
            for a in self._accounts:
                if a['code'] == code: aname = a['full_name']; break
            # Collect aux selections
            aux_sel = []   # (dim_id, item_id, item_name)
            if auxw:
                for dim_id, cb in auxw._combos:
                    item_id = cb.currentData()
                    item_name = cb.currentText() if item_id else ""
                    if item_id:
                        aux_sel.append((dim_id, item_id, item_name))
            entries.append((i, sw.text().strip() if sw else "", code, aname, d, cr, aux_sel))

        if not entries: QMessageBox.warning(self,"提示","请至少填写一行分录"); return
        td = sum(e[4] for e in entries); tc = sum(e[5] for e in entries)
        if abs(td-tc) > 0.005:
            QMessageBox.warning(self,"借贷不平",f"借方合计 {td:.2f} ≠ 贷方合计 {tc:.2f}\n请检查金额"); return

        conn = get_db(); c = conn.cursor()
        dt = self.date_edit.date().toString("yyyy-MM-dd")
        reverted_to_pending = False

        if self.voucher_id:
            c.execute("SELECT voucher_no, status FROM vouchers WHERE id=?", (self.voucher_id,))
            voucher = c.fetchone()
            if not voucher:
                conn.close()
                QMessageBox.warning(self, "提示", "凭证不存在或已被删除，请刷新后重试")
                return
            vno = voucher["voucher_no"]
            new_status = voucher["status"]
            if voucher["status"] == "已审核":
                new_status = "待审核"
                reverted_to_pending = True
            c.execute("UPDATE vouchers SET date=?,attachment_count=?,status=? WHERE id=?",
                      (dt, self.attach_spin.value(), new_status, self.voucher_id))
            c.execute("DELETE FROM voucher_entries WHERE voucher_id=?", (self.voucher_id,))
            vid = self.voucher_id
        else:
            c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=?",
                      (self.client_id, self.period))
            n = c.fetchone()[0] + 1
            vno = f"记-{n:03d}"
            c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,attachment_count) VALUES(?,?,?,?,?)",
                      (self.client_id, self.period, vno, dt, self.attach_spin.value()))
            vid = c.lastrowid

        for ln, summary, code, aname, d, cr, aux_sel in entries:
            c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                      (vid, ln, summary, code, aname, d, cr))
            entry_id = c.lastrowid
            for dim_id, item_id, item_name in aux_sel:
                c.execute("INSERT INTO voucher_entry_aux(entry_id,dimension_id,aux_item_id,aux_item_name) VALUES(?,?,?,?)",
                          (entry_id, dim_id, item_id, item_name))
        action = "编辑凭证" if self.voucher_id else "新增凭证"
        detail = f"凭证号:{vno} 借方合计:{td:.2f}"
        if reverted_to_pending:
            detail += " 修改后自动回退为待审核"
        log_action(conn, self.client_id,
                   action, "voucher", vid, detail)
        conn.commit(); conn.close()
        self.saved_and_new = and_new
        if reverted_to_pending:
            QMessageBox.information(self, "已回退待审核", "该凭证原状态为“已审核”，修改后已自动回退为“待审核”，请重新审核。")
        self.accept()

# ── Pages ──────────────────────────────────────────────────────────────────

# ── 辅助核算管理页（嵌入 VoucherPage 的 Tab） ─────────────────────────────

class AuxItemDialog(QDialog):
    """新增/编辑 核算对象"""
    def __init__(self, parent, client_id, dimension_id, item=None):
        super().__init__(parent)
        self.client_id = client_id
        self.dimension_id = dimension_id
        self.item = item
        self.setWindowTitle("编辑核算对象" if item else "新增核算对象")
        self.setMinimumWidth(360)
        self._build()
        if item: self._load()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(22,20,22,20); L.setSpacing(12)
        L.addWidget(lbl("核算对象信息", bold=True, size=14))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.f_name    = QLineEdit(); self.f_name.setPlaceholderText("名称（必填）")
        self.f_code    = QLineEdit(); self.f_code.setPlaceholderText("编码（可选）")
        self.f_contact = QLineEdit(); self.f_contact.setPlaceholderText("联系人")
        self.f_phone   = QLineEdit(); self.f_phone.setPlaceholderText("电话")
        F.addRow("名称 *",  self.f_name)
        F.addRow("编码",    self.f_code)
        F.addRow("联系人",  self.f_contact)
        F.addRow("电话",    self.f_phone)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        bc = QPushButton("取消"); bc.setObjectName("btn_gray")
        bs = QPushButton("保存"); bs.setObjectName("btn_primary")
        bc.clicked.connect(self.reject); bs.clicked.connect(self._save)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        self.f_name.setText(self.item["name"] or "")
        self.f_code.setText(self.item["code"] or "")
        self.f_contact.setText(self.item["contact"] or "")
        self.f_phone.setText(self.item["phone"] or "")

    def _save(self):
        name = self.f_name.text().strip()
        if not name: QMessageBox.warning(self, "提示", "名称不能为空"); return
        conn = get_db(); c = conn.cursor()
        d = (name, self.f_code.text().strip(), self.f_contact.text().strip(),
             self.f_phone.text().strip())
        if self.item:
            c.execute("UPDATE aux_items SET name=?,code=?,contact=?,phone=? WHERE id=?",
                      d + (self.item["id"],))
        else:
            c.execute("INSERT INTO aux_items(client_id,dimension_id,name,code,contact,phone)"
                      " VALUES(?,?,?,?,?,?)",
                      (self.client_id, self.dimension_id) + d)
        conn.commit(); conn.close(); self.accept()


class AuxPage(QWidget):
    """辅助核算管理：维度 + 对象 + 科目绑定"""
    def __init__(self):
        super().__init__()
        self.client_id = None
        self._cur_dim_id = None
        L = QHBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # ── 左栏：维度列表 ──
        left = QWidget(); left.setFixedWidth(200)
        left.setStyleSheet("background:#f7f9fc; border-right:1px solid #e8ecf2;")
        ll = QVBoxLayout(left); ll.setContentsMargins(0,0,0,0); ll.setSpacing(0)
        hdr_l = QWidget(); hdr_l.setStyleSheet("background:#fff; border-bottom:1px solid #eee;")
        hl = QHBoxLayout(hdr_l); hl.setContentsMargins(12,10,8,10)
        hl.addWidget(lbl("核算维度", bold=True)); hl.addStretch()
        b_adddim = QPushButton("＋"); b_adddim.setObjectName("btn_primary")
        b_adddim.setFixedSize(28,28); b_adddim.setToolTip("新增维度")
        b_adddim.clicked.connect(self._add_dim)
        hl.addWidget(b_adddim); ll.addWidget(hdr_l)

        self.dim_list = QListWidget()
        self.dim_list.setStyleSheet(
            "QListWidget{border:none;background:#f7f9fc;}"
            "QListWidget::item{padding:10px 14px;border-bottom:1px solid #eef0f4;}"
            "QListWidget::item:selected{background:#e6f0ff;color:#3d6fdb;font-weight:bold;}")
        self.dim_list.currentRowChanged.connect(self._on_dim_changed)
        ll.addWidget(self.dim_list)

        # 维度右键菜单
        self.dim_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.dim_list.customContextMenuRequested.connect(self._dim_context_menu)
        L.addWidget(left)

        # ── 右栏：对象列表 + 科目绑定 ──
        right = QWidget(); rl = QVBoxLayout(right); rl.setContentsMargins(0,0,0,0); rl.setSpacing(0)

        # Right-side tab: 对象管理 | 往来对账
        self.right_tabs = QTabWidget()
        self.right_tabs.setStyleSheet(
            "QTabBar::tab{padding:8px 20px;color:#888;border:none;background:transparent;"
            "border-bottom:2px solid transparent;}"
            "QTabBar::tab:selected{color:#3d6fdb;border-bottom:2px solid #3d6fdb;}"
            "QTabWidget::pane{border:none;}")

        # ── Tab A: 核算对象管理 ──
        tab_mgr = QWidget(); tl = QVBoxLayout(tab_mgr); tl.setContentsMargins(20,14,20,14); tl.setSpacing(10)

        hdr_r = QHBoxLayout()
        self.dim_title = lbl("请选择左侧维度", bold=True, size=15)
        hdr_r.addWidget(self.dim_title); hdr_r.addStretch()
        self.b_additem = QPushButton("＋ 新增对象"); self.b_additem.setObjectName("btn_primary")
        self.b_additem.clicked.connect(self._add_item)
        b_exp = QPushButton("导出Excel"); b_exp.setObjectName("btn_outline")
        b_exp.clicked.connect(self._export_items)
        hdr_r.addWidget(b_exp); hdr_r.addWidget(self.b_additem)
        tl.addLayout(hdr_r)

        f1 = card(); v1 = QVBoxLayout(f1); v1.setContentsMargins(0,0,0,0)
        self.item_tbl = QTableWidget(); self.item_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.item_tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.item_tbl.setShowGrid(False); self.item_tbl.verticalHeader().setVisible(False)
        self.item_tbl.setColumnCount(5)
        self.item_tbl.setHorizontalHeaderLabels(["编码","名称","联系人","电话","操作"])
        hh = self.item_tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        self.item_tbl.setColumnWidth(0,80); self.item_tbl.setColumnWidth(2,90)
        self.item_tbl.setColumnWidth(3,110); self.item_tbl.setColumnWidth(4,160)
        v1.addWidget(self.item_tbl); tl.addWidget(f1)

        # 科目绑定区
        tl.addWidget(lbl("绑定科目（凭证录入时，选中这些科目将显示本维度的对象选择）",
                         color="#666", size=12))
        f2 = card(); v2 = QVBoxLayout(f2); v2.setContentsMargins(12,10,12,10); v2.setSpacing(8)
        bind_hdr = QHBoxLayout()
        bind_hdr.addWidget(lbl("已绑定科目", bold=True)); bind_hdr.addStretch()
        b_bind = QPushButton("＋ 绑定科目"); b_bind.setObjectName("btn_outline")
        b_bind.clicked.connect(self._bind_account)
        bind_hdr.addWidget(b_bind); v2.addLayout(bind_hdr)
        self.bind_list = QListWidget()
        self.bind_list.setStyleSheet(
            "QListWidget{border:1px solid #eee;border-radius:5px;background:#fff;}"
            "QListWidget::item{padding:6px 10px;}"
            "QListWidget::item:selected{background:#e6f0ff;}")
        self.bind_list.setMaximumHeight(120)
        self.bind_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.bind_list.customContextMenuRequested.connect(self._unbind_context_menu)
        v2.addWidget(self.bind_list)
        tl.addWidget(f2)
        self.right_tabs.addTab(tab_mgr, "核算对象管理")

        # ── Tab B: 往来对账报表 ──
        tab_rpt = QWidget(); tr = QVBoxLayout(tab_rpt); tr.setContentsMargins(20,14,20,14); tr.setSpacing(10)

        rpt_hdr = QHBoxLayout()
        rpt_hdr.addWidget(lbl("往来对账表", bold=True, size=15)); rpt_hdr.addStretch()
        b_rpt_exp = QPushButton("导出Excel"); b_rpt_exp.setObjectName("btn_outline")
        b_rpt_exp.clicked.connect(self._export_aux_report)
        rpt_hdr.addWidget(b_rpt_exp); tr.addLayout(rpt_hdr)

        fr = QHBoxLayout(); fr.setSpacing(8)
        fr.addWidget(lbl("维度:"))
        self.rpt_dim_combo = QComboBox(); self.rpt_dim_combo.setMinimumWidth(120)
        fr.addWidget(self.rpt_dim_combo)
        fr.addWidget(lbl("期间:"))
        self.rpt_period_edit = QLineEdit(); self.rpt_period_edit.setFixedWidth(100)
        self.rpt_period_edit.setPlaceholderText("如 2026-03")
        fr.addWidget(self.rpt_period_edit)
        fr.addWidget(lbl("科目:"))
        self.rpt_acct_combo = QComboBox(); self.rpt_acct_combo.setMinimumWidth(200)
        self.rpt_acct_combo.addItem("全部科目", "")
        fr.addWidget(self.rpt_acct_combo)
        b_q = QPushButton("查询"); b_q.setObjectName("btn_primary")
        b_q.clicked.connect(self._load_aux_report)
        fr.addWidget(b_q); fr.addStretch()
        tr.addLayout(fr)

        f3 = card(); v3 = QVBoxLayout(f3); v3.setContentsMargins(0,0,0,0)
        self.aux_rpt_tbl = QTableWidget(); self.aux_rpt_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.aux_rpt_tbl.setShowGrid(True); self.aux_rpt_tbl.verticalHeader().setVisible(False)
        self.aux_rpt_tbl.setColumnCount(6)
        self.aux_rpt_tbl.setHorizontalHeaderLabels(["核算对象","科目","期初余额","本期借方","本期贷方","期末余额"])
        hh3 = self.aux_rpt_tbl.horizontalHeader()
        hh3.setSectionResizeMode(QHeaderView.Interactive)
        hh3.setSectionResizeMode(0, QHeaderView.Stretch)
        hh3.setSectionResizeMode(1, QHeaderView.Stretch)
        for ci in range(2,6): self.aux_rpt_tbl.setColumnWidth(ci, 110)
        v3.addWidget(self.aux_rpt_tbl); tr.addWidget(f3)
        self.right_tabs.addTab(tab_rpt, "往来对账")

        rl.addWidget(self.right_tabs)
        L.addWidget(right)

    def set_client(self, client_id, period=""):
        self.client_id = client_id
        self._period = period
        self._cur_dim_id = None
        self._dims = []
        self._items = []
        self._bindings = []
        self._load_dims()
        self._refresh_rpt_combos()

    def _refresh_rpt_combos(self):
        if not self.client_id: return
        # Period
        from datetime import datetime
        now = datetime.now()
        if self._period:
            self.rpt_period_edit.setText(self._period)
        else:
            self.rpt_period_edit.setText(f"{now.year}-{now.month:02d}")
        # Dimensions
        self.rpt_dim_combo.clear()
        for d in self._dims:
            self.rpt_dim_combo.addItem(d["name"], d["id"])
        # Accounts
        self.rpt_acct_combo.clear(); self.rpt_acct_combo.addItem("全部科目", "")
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code", (self.client_id,))
        for a in c.fetchall():
            self.rpt_acct_combo.addItem(f"{a['code']} {a['name']}", a['code'])
        conn.close()

    def _load_dims(self):
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM aux_dimensions WHERE client_id=? ORDER BY sort_order,id",
                  (self.client_id,))
        dims = c.fetchall(); conn.close()
        self.dim_list.clear()
        self._dims = [dict(d) for d in dims]
        for d in self._dims:
            self.dim_list.addItem(d["name"])
        if self._dims:
            self.dim_list.setCurrentRow(0)

    def _on_dim_changed(self, row):
        if row < 0 or row >= len(self._dims): return
        self._cur_dim_id = self._dims[row]["id"]
        self.dim_title.setText(f"【{self._dims[row]['name']}】核算对象")
        self._load_items()
        self._load_bindings()

    def _add_dim(self):
        if not self.client_id: return
        name, ok = QInputDialog.getText(self, "新增维度", "维度名称（如：客户、员工、项目）：")
        if not ok or not name.strip(): return
        conn = get_db()
        try:
            conn.execute("INSERT INTO aux_dimensions(client_id,name) VALUES(?,?)",
                         (self.client_id, name.strip()))
            conn.commit()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"维度已存在或保存失败：{e}")
        finally:
            conn.close()
        self._load_dims()

    def _dim_context_menu(self, pos):
        row = self.dim_list.currentRow()
        if row < 0: return
        menu = QMenu(self)
        menu.addAction("重命名").triggered.connect(lambda: self._rename_dim(row))
        menu.addAction("删除维度").triggered.connect(lambda: self._del_dim(row))
        menu.exec(self.dim_list.mapToGlobal(pos))

    def _rename_dim(self, row):
        old = self._dims[row]["name"]
        name, ok = QInputDialog.getText(self, "重命名维度", "新名称：", text=old)
        if not ok or not name.strip(): return
        conn = get_db()
        conn.execute("UPDATE aux_dimensions SET name=? WHERE id=?",
                     (name.strip(), self._dims[row]["id"]))
        conn.commit(); conn.close(); self._load_dims()

    def _del_dim(self, row):
        if QMessageBox.question(self, "确认",
                f"删除维度【{self._dims[row]['name']}】及其所有对象和绑定？",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db(); did = self._dims[row]["id"]
        conn.execute("DELETE FROM voucher_entry_aux WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM account_aux_config WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM aux_items WHERE dimension_id=?", (did,))
        conn.execute("DELETE FROM aux_dimensions WHERE id=?", (did,))
        conn.commit(); conn.close(); self._load_dims()

    def _load_items(self):
        if not self._cur_dim_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM aux_items WHERE dimension_id=? AND client_id=? ORDER BY id",
                  (self._cur_dim_id, self.client_id))
        rows = c.fetchall(); conn.close()
        self._items = [dict(r) for r in rows]
        self.item_tbl.setRowCount(len(rows))
        for i, r in enumerate(self._items):
            self.item_tbl.setRowHeight(i, 40)
            for j, v in enumerate([r["code"] or "", r["name"], r["contact"] or "", r["phone"] or ""]):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter if j != 1 else Qt.AlignLeft | Qt.AlignVCenter)
                self.item_tbl.setItem(i, j, it)
            bw = QWidget(); bl = QHBoxLayout(bw); bl.setContentsMargins(4,3,4,3); bl.setSpacing(4)
            b_ed = QPushButton("✏ 编辑"); b_ed.setObjectName("btn_outline"); b_ed.setFixedSize(64,26)
            b_ed.clicked.connect(lambda _, rr=r: self._edit_item(rr))
            b_dl = QPushButton("🗑"); b_dl.setObjectName("btn_red"); b_dl.setFixedSize(30,26)
            b_dl.clicked.connect(lambda _, rid=r["id"]: self._del_item(rid))
            bl.addWidget(b_ed); bl.addWidget(b_dl); bl.addStretch()
            self.item_tbl.setCellWidget(i, 4, bw)

    def _add_item(self):
        if not self._cur_dim_id:
            QMessageBox.information(self, "提示", "请先选择左侧维度"); return
        d = AuxItemDialog(self, self.client_id, self._cur_dim_id)
        if d.exec(): self._load_items()

    def _edit_item(self, r):
        d = AuxItemDialog(self, self.client_id, self._cur_dim_id, item=r)
        if d.exec(): self._load_items()

    def _del_item(self, item_id):
        if QMessageBox.question(self, "确认", "删除该核算对象？",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db()
        conn.execute("DELETE FROM aux_items WHERE id=?", (item_id,))
        conn.commit(); conn.close(); self._load_items()

    def _load_aux_report(self):
        if not self.client_id: return
        dim_id = self.rpt_dim_combo.currentData()
        period = self.rpt_period_edit.text().strip()
        acct_filter = self.rpt_acct_combo.currentData() or ""
        if not dim_id or not period:
            QMessageBox.information(self, "提示", "请选择维度并填写期间"); return
        conn = get_db(); c = conn.cursor()
        # Get account direction info for opening balance computation
        c.execute("SELECT code,direction,opening_debit,opening_credit FROM accounts WHERE client_id=?",
                  (self.client_id,))
        acct_info = {r["code"]: r for r in c.fetchall()}
        sql = """
            SELECT ai.name AS item_name, ai.id AS item_id,
                   e.account_code, e.account_name,
                   SUM(e.debit) AS td, SUM(e.credit) AS tc
            FROM voucher_entry_aux ea
            JOIN aux_items ai ON ai.id=ea.aux_item_id
            JOIN voucher_entries e ON e.id=ea.entry_id
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE ea.dimension_id=? AND v.client_id=? AND v.period=?
        """
        params = [dim_id, self.client_id, period]
        if acct_filter:
            sql += " AND e.account_code=?"; params.append(acct_filter)
        sql += " GROUP BY ai.id, e.account_code ORDER BY ai.name, e.account_code"
        c.execute(sql, params)
        rows = c.fetchall(); conn.close()

        self.aux_rpt_tbl.setRowCount(len(rows) + 1)
        td_tot = tc_tot = 0
        for i, r in enumerate(rows):
            self.aux_rpt_tbl.setRowHeight(i, 36)
            td = r["td"] or 0; tc = r["tc"] or 0
            ai = acct_info.get(r["account_code"])
            od = (ai["opening_debit"] or 0) - (ai["opening_credit"] or 0) if ai else 0
            if ai and ai["direction"] == "贷": od = -od
            ending = od + td - tc
            td_tot += td; tc_tot += tc
            vals = [r["item_name"], f"{r['account_code']} {r['account_name']}",
                    fmt_amt(od), fmt_amt(td), fmt_amt(tc), fmt_amt(ending)]
            for j, v in enumerate(vals):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter if j<=1 else Qt.AlignRight|Qt.AlignVCenter)
                if j == 3 and td: it.setForeground(QColor("#3d6fdb"))
                if j == 4 and tc: it.setForeground(QColor("#e05252"))
                if j == 5:
                    it.setForeground(QColor("#3d6fdb") if ending > 0 else
                                     QColor("#e05252") if ending < 0 else QColor("#888"))
                self.aux_rpt_tbl.setItem(i, j, it)
        # Totals
        n = len(rows); self.aux_rpt_tbl.setRowHeight(n, 38)
        for j, v in enumerate(["合计", "", "", fmt_amt(td_tot), fmt_amt(tc_tot), fmt_amt(td_tot-tc_tot)]):
            it = QTableWidgetItem(v)
            it.setFont(QFont("", weight=QFont.Bold)); it.setBackground(QColor("#f5f7fa"))
            it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter if j >= 2 else Qt.AlignLeft|Qt.AlignVCenter)
            self.aux_rpt_tbl.setItem(n, j, it)

    def _export_aux_report(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        period = self.rpt_period_edit.text().strip() or "report"
        path, _ = QFileDialog.getSaveFileName(self, "保存", f"往来对账_{period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "往来对账"
        hdrs = ["核算对象","科目","期初余额","本期借方","本期贷方","期末余额"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(1, ci, h); cell.font = XFont(bold=True, color="FFFFFF")
            cell.fill = fill; cell.alignment = Alignment(horizontal="center")
        for ri in range(self.aux_rpt_tbl.rowCount()):
            ws.append([self.aux_rpt_tbl.item(ri, ci).text() if self.aux_rpt_tbl.item(ri, ci) else ""
                       for ci in range(6)])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _export_items(self):
        if not self._cur_dim_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        path, _ = QFileDialog.getSaveFileName(self, "保存", "核算对象.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["编码", "名称", "联系人", "电话"])
        for r in self._items:
            ws.append([r["code"], r["name"], r["contact"], r["phone"]])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _load_bindings(self):
        if not self._cur_dim_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT a.code, a.name FROM account_aux_config ac
            JOIN accounts a ON a.client_id=ac.client_id AND a.code=ac.account_code
            WHERE ac.client_id=? AND ac.dimension_id=? ORDER BY a.code""",
                  (self.client_id, self._cur_dim_id))
        rows = c.fetchall(); conn.close()
        self.bind_list.clear()
        self._bindings = [dict(r) for r in rows]
        for r in self._bindings:
            self.bind_list.addItem(f"{r['code']}  {r['name']}")

    def _bind_account(self):
        if not self._cur_dim_id:
            QMessageBox.information(self, "提示", "请先选择维度"); return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code, name FROM accounts WHERE client_id=? ORDER BY code",
                  (self.client_id,))
        accts = c.fetchall(); conn.close()
        # Already bound codes
        bound = {r["code"] for r in self._bindings}
        items = [f"{a['code']}  {a['name']}" for a in accts if a['code'] not in bound]
        if not items:
            QMessageBox.information(self, "提示", "所有科目均已绑定该维度"); return
        item, ok = QInputDialog.getItem(self, "绑定科目", "选择要绑定的科目：", items, editable=True)
        if not ok or not item: return
        code = item.split()[0]
        conn = get_db()
        try:
            conn.execute("INSERT INTO account_aux_config(client_id,account_code,dimension_id) VALUES(?,?,?)",
                         (self.client_id, code, self._cur_dim_id))
            conn.commit()
        except Exception:
            pass
        finally:
            conn.close()
        self._load_bindings()

    def _unbind_context_menu(self, pos):
        row = self.bind_list.currentRow()
        if row < 0: return
        menu = QMenu(self); menu.addAction("解除绑定")
        act = menu.exec(self.bind_list.mapToGlobal(pos))
        if act:
            code = self._bindings[row]["code"]
            conn = get_db()
            conn.execute("DELETE FROM account_aux_config WHERE client_id=? AND account_code=? AND dimension_id=?",
                         (self.client_id, code, self._cur_dim_id))
            conn.commit(); conn.close(); self._load_bindings()


class ImportAccountSetDialog(QDialog):
    """账套导入向导 — 四步骤，支持科目余额表/凭证/银行日记账"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("导入账套")
        self.setMinimumSize(860, 640)
        self.resize(960, 720)
        self._files = {}        # key -> path
        self._client_id = None
        self._build()

    def showEvent(self, event):
        super().showEvent(event)
        self._fit_to_screen()

    def _fit_to_screen(self):
        # macOS 下模态对话框有时会按默认几何打开，导致底部按钮栏超出可视区。
        screen = self.windowHandle().screen() if self.windowHandle() else QApplication.primaryScreen()
        if not screen:
            return
        avail = screen.availableGeometry()
        max_width = max(520, avail.width() - 80)
        max_height = max(420, avail.height() - 80)
        width = min(max(self.width(), 860), max_width)
        height = min(max(self.height(), 640), max_height)
        self.resize(width, height)
        x = avail.x() + (avail.width() - width) // 2
        y = avail.y() + (avail.height() - height) // 2
        self.move(max(avail.x(), x), max(avail.y(), y))

    # ─────────────────────────── UI 骨架 ───────────────────────────
    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # 顶部标题
        title_bar = QWidget()
        title_bar.setStyleSheet("background:#1c2340;")
        tl = QHBoxLayout(title_bar); tl.setContentsMargins(24,16,24,16)
        tl.addWidget(lbl("📥  账套导入向导", bold=True, color="#fff", size=16))
        tl.addSpacing(12)
        tl.addWidget(lbl("从用友 / 金蝶导出文件一键建立账套", color="#8b93ae", size=12))
        tl.addStretch(); L.addWidget(title_bar)

        # 左侧步骤导航 + 右侧内容
        body = QWidget(); bl = QHBoxLayout(body)
        bl.setContentsMargins(0,0,0,0); bl.setSpacing(0)

        nav = QWidget(); nav.setFixedWidth(176)
        nav.setStyleSheet("background:#f7f9fc;border-right:1px solid #e4e8f0;")
        nl = QVBoxLayout(nav); nl.setContentsMargins(0,14,0,14); nl.setSpacing(2)
        self._step_btns = []
        for num, text in [("1","客户信息"),("2","选择文件"),("3","预览确认"),("4","导入完成")]:
            sb = QPushButton(f"  {num}   {text}")
            sb.setStyleSheet("""QPushButton{background:transparent;color:#888;border:none;
                text-align:left;padding:11px 14px;border-left:3px solid transparent;}
                QPushButton[active=true]{background:#e6f0ff;color:#3d6fdb;
                border-left:3px solid #3d6fdb;font-weight:bold;}""")
            sb.setProperty("active","false"); nl.addWidget(sb)
            self._step_btns.append(sb)
        nl.addStretch(); bl.addWidget(nav)

        self.right_scroll = QScrollArea()
        self.right_scroll.setWidgetResizable(True)
        self.right_scroll.setFrameShape(QFrame.NoFrame)
        self.right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.right = QStackedWidget()
        self.right_scroll.setWidget(self.right)
        bl.addWidget(self.right_scroll)
        # stretch=1 让 body 撑满剩余空间，foot 保持固定高度不被挤走
        L.addWidget(body, 1)

        # 底部按钮栏 — 固定高度，确保始终可见
        foot = QWidget()
        foot.setObjectName("import_foot")
        foot.setMinimumHeight(76)
        foot.setMaximumHeight(76)
        fl = QHBoxLayout(foot)
        fl.setContentsMargins(20, 14, 20, 14)
        fl.setSpacing(10)
        fl.addStretch()
        self.btn_back = QPushButton("← 上一步"); self.btn_back.setObjectName("btn_gray")
        self.btn_back.setMinimumHeight(36)
        self.btn_back.setMinimumWidth(120)
        self.btn_back.clicked.connect(self._prev_step)
        self.btn_next = QPushButton("下一步 →"); self.btn_next.setObjectName("btn_primary")
        self.btn_next.setMinimumHeight(36)
        self.btn_next.setMinimumWidth(120)
        self.btn_next.clicked.connect(self._next_step)
        self.btn_close = QPushButton("关 闭"); self.btn_close.setObjectName("btn_gray")
        self.btn_close.setMinimumHeight(36)
        self.btn_close.setMinimumWidth(96)
        self.btn_close.clicked.connect(self.accept); self.btn_close.setVisible(False)
        fl.addWidget(self.btn_back); fl.addWidget(self.btn_next); fl.addWidget(self.btn_close)
        L.addWidget(foot, 0)   # stretch=0，高度由 setFixedHeight 决定

        self._build_step1(); self._build_step2(); self._build_step3(); self._build_step4()
        self._goto_step(0)

    # ─────────────────── Step 1 · 客户信息 ───────────────────
    def _build_step1(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(14)
        L.addWidget(lbl("填写客户（账套）基本信息", bold=True, size=14))
        hint = QLabel("  新建账套，或选择已有客户追加导入。同名账套自动追加，不会重复创建。")
        hint.setStyleSheet("background:#f6f8ff;color:#444;border-radius:6px;"
                           "padding:8px 12px;font-size:12px;")
        hint.setWordWrap(True); L.addWidget(hint)

        mode_row = QHBoxLayout()
        self.mode_new   = QPushButton("✦ 新建账套")
        self.mode_exist = QPushButton("◈ 追加到已有账套")
        for b in [self.mode_new, self.mode_exist]:
            b.setCheckable(True)
            b.setStyleSheet("""QPushButton{background:#f5f7fa;color:#555;
                border:1px solid #d9d9d9;border-radius:6px;padding:8px 18px;}
                QPushButton:checked{background:#3d6fdb;color:#fff;border-color:#3d6fdb;}""")
        self.mode_new.setChecked(True)
        self.mode_new.clicked.connect(lambda: self._toggle_mode(True))
        self.mode_exist.clicked.connect(lambda: self._toggle_mode(False))
        mode_row.addWidget(self.mode_new); mode_row.addWidget(self.mode_exist)
        mode_row.addStretch(); L.addLayout(mode_row)

        # 新建表单
        self.form_new = QWidget(); fn = QFormLayout(self.form_new)
        fn.setSpacing(10); fn.setLabelAlignment(Qt.AlignRight)
        self.f_name  = QLineEdit(); self.f_name.setPlaceholderText("公司全称（必填）")
        self.f_code  = QLineEdit(); self.f_code.setPlaceholderText("助记码，如 ZY")
        self.f_type  = QComboBox()
        self.f_type.addItems(["小规模纳税人","一般纳税人","其他"])
        self.f_taxid = QLineEdit(); self.f_taxid.setPlaceholderText("统一社会信用代码")
        fn.addRow("公司名称 *", self.f_name); fn.addRow("助记码", self.f_code)
        fn.addRow("客户类型", self.f_type);   fn.addRow("税号",    self.f_taxid)
        L.addWidget(self.form_new)

        # 已有账套
        self.form_exist = QWidget(); fe = QFormLayout(self.form_exist)
        fe.setSpacing(10); fe.setLabelAlignment(Qt.AlignRight)
        self.f_exist_combo = QComboBox(); self.f_exist_combo.setMinimumWidth(280)
        fe.addRow("选择已有账套", self.f_exist_combo)
        warn = QLabel("  ⚠ 追加模式：同期间同凭证号自动跳过，不会重复写入。")
        warn.setStyleSheet("color:#ad6800;font-size:12px;")
        fe.addRow("", warn)
        self.form_exist.setVisible(False); L.addWidget(self.form_exist)
        L.addStretch()
        self.right.addWidget(w)
        self._reload_exist_clients()

    def _reload_exist_clients(self):
        self.f_exist_combo.clear()
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT id,name FROM clients ORDER BY name")
        for r in c.fetchall():
            self.f_exist_combo.addItem(r["name"], r["id"])
        conn.close()

    def _toggle_mode(self, is_new):
        self.mode_new.setChecked(is_new); self.mode_exist.setChecked(not is_new)
        self.form_new.setVisible(is_new); self.form_exist.setVisible(not is_new)

    # ─────────────────── Step 2 · 选择文件 ───────────────────
    def _build_step2(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(10)
        L.addWidget(lbl("选择要导入的账务文件", bold=True, size=14))
        hint = QLabel("  支持用友T3/T6/U8、金蝶KIS/K3导出的 XLS/XLSX 文件。"
                       "所有类型均可选，选得越多导入越完整。")
        hint.setStyleSheet("background:#f6f8ff;color:#444;border-radius:6px;"
                           "padding:8px 12px;font-size:12px;")
        hint.setWordWrap(True); L.addWidget(hint)

        file_types = [
            ("balance", "📊 科目余额表",
             "含科目编号、名称、期初余额 — 自动建科目并写期初数（必选）"),
            ("voucher", "📝 记账凭证",
             "全部凭证分录，自动识别用友 / 通用模板格式"),
            ("bank",    "🏦 银行存款日记账",
             "银行流水逐笔明细，写入 bank_statements 表供对账"),
            ("ledger",  "📒 总账 / 明细账",
             "按科目汇总的发生额与余额，用于核对"),
            ("income",  "💹 利润表",
             "收入费用汇总，仅用于人工核对，不写入账套"),
            ("bs",      "🏛 资产负债表",
             "资产负债期末数，仅用于人工核对，不写入账套"),
        ]
        self._file_path_lbls = {}
        for key, label, desc in file_types:
            row_w = QFrame()
            row_w.setStyleSheet("QFrame{background:#fff;border:1px solid #e8ecf2;"
                                "border-radius:8px;}")
            rl = QHBoxLayout(row_w); rl.setContentsMargins(16,10,16,10); rl.setSpacing(10)
            vv = QVBoxLayout()
            vv.addWidget(lbl(label, bold=True))
            vv.addWidget(lbl(desc, color="#888", size=12))
            rl.addLayout(vv); rl.addStretch()
            path_lbl = QLabel("未选择")
            path_lbl.setStyleSheet("color:#bbb;font-size:11px;min-width:160px;max-width:220px;")
            b_sel = QPushButton("选择"); b_sel.setObjectName("btn_outline"); b_sel.setFixedWidth(72)
            b_clr = QPushButton("✕");   b_clr.setObjectName("btn_gray");    b_clr.setFixedSize(26,26)
            b_sel.clicked.connect(lambda _,k=key,pl=path_lbl: self._pick_file(k, pl))
            b_clr.clicked.connect(lambda _,k=key,pl=path_lbl: self._clear_file(k, pl))
            rl.addWidget(path_lbl); rl.addWidget(b_sel); rl.addWidget(b_clr)
            self._file_path_lbls[key] = path_lbl
            L.addWidget(row_w)
        L.addStretch()
        self.right.addWidget(w)

    def _pick_file(self, key, lbl_w):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择文件", "", "Excel 文件 (*.xls *.xlsx)")
        if not path: return
        self._files[key] = path
        import os; lbl_w.setText(os.path.basename(path))
        lbl_w.setStyleSheet("color:#3d6fdb;font-size:11px;")

    def _clear_file(self, key, lbl_w):
        self._files.pop(key, None)
        lbl_w.setText("未选择"); lbl_w.setStyleSheet("color:#bbb;font-size:11px;")

    # ─────────────────── Step 3 · 预览确认 ───────────────────
    def _build_step3(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(10)
        L.addWidget(lbl("预览与确认", bold=True, size=14))
        self.preview_info = QLabel("")
        self.preview_info.setStyleSheet("background:#f6f8ff;color:#333;border-radius:6px;"
                                        "padding:10px 14px;font-size:12px;")
        self.preview_info.setWordWrap(True); L.addWidget(self.preview_info)
        L.addWidget(lbl("数据预览（首个文件前 20 行）:", bold=True, size=12))
        self.preview_tbl = QTableWidget()
        self.preview_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.preview_tbl.setShowGrid(True)
        self.preview_tbl.verticalHeader().setVisible(False)
        L.addWidget(self.preview_tbl)
        self.right.addWidget(w)

    def _refresh_preview(self):
        lines = []
        if self.mode_new.isChecked():
            lines.append(f"新建账套：{self.f_name.text().strip() or '（未填）'}")
        else:
            lines.append(f"追加到：{self.f_exist_combo.currentText()}")
        labels = {"balance":"科目余额表","voucher":"记账凭证","bank":"银行日记账",
                  "ledger":"总账","income":"利润表","bs":"资产负债表"}
        for k, v in labels.items():
            import os
            mark = f"✓  {os.path.basename(self._files[k])}" if k in self._files else "○  跳过"
            lines.append(f"{v}：{mark}")
        self.preview_info.setText("\n".join(lines))

        first = next((k for k in ["balance","voucher","bank","ledger"] if k in self._files), None)
        if not first:
            self.preview_tbl.setRowCount(0); self.preview_tbl.setColumnCount(0); return
        try:
            df = self._read_df(self._files[first])
            df = df.iloc[:20]
            cols = min(df.shape[1], 10)
            self.preview_tbl.setColumnCount(cols); self.preview_tbl.setRowCount(len(df))
            for ri, row in df.iterrows():
                self.preview_tbl.setRowHeight(ri, 26)
                for ci in range(cols):
                    self.preview_tbl.setItem(ri, ci, QTableWidgetItem(str(row.iloc[ci])))
            self.preview_tbl.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeToContents)
        except Exception as e:
            self.preview_tbl.setRowCount(1); self.preview_tbl.setColumnCount(1)
            self.preview_tbl.setItem(0, 0, QTableWidgetItem(f"预览失败：{e}"))

    # ─────────────────── Step 4 · 导入完成 ───────────────────
    def _build_step4(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(32,24,32,24); L.setSpacing(10)
        self.result_icon  = lbl("⏳", size=40, color="#3d6fdb")
        self.result_icon.setAlignment(Qt.AlignCenter)
        self.result_title = lbl("正在导入…", bold=True, size=15)
        self.result_title.setAlignment(Qt.AlignCenter)
        L.addStretch()
        L.addWidget(self.result_icon); L.addWidget(self.result_title); L.addSpacing(8)
        self.log_box = QTextEdit(); self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(
            "font-size:12px;font-family:monospace;background:#fafafa;border-radius:6px;")
        L.addWidget(self.log_box); L.addStretch()
        self.right.addWidget(w)

    # ─────────────────── 步骤控制 ───────────────────
    def _goto_step(self, idx):
        self._cur_step = idx
        self.right.setCurrentIndex(idx)
        for i, b in enumerate(self._step_btns):
            b.setProperty("active", "true" if i == idx else "false")
            b.style().unpolish(b); b.style().polish(b)
        self.btn_back.setVisible(0 < idx < 3)
        self.btn_next.setVisible(idx < 3)
        self.btn_close.setVisible(idx == 3)
        if idx == 2: self._refresh_preview()
        if idx == 3: QTimer.singleShot(100, self._do_import)

    def _prev_step(self):
        if self._cur_step > 0: self._goto_step(self._cur_step - 1)

    def _next_step(self):
        if self._cur_step == 0:
            if self.mode_new.isChecked() and not self.f_name.text().strip():
                QMessageBox.warning(self, "提示", "请填写公司名称"); return
        if self._cur_step == 1 and not self._files:
            QMessageBox.information(self, "提示", "请至少选择一个文件"); return
        self._goto_step(self._cur_step + 1)

    # ─────────────────── 工具方法 ───────────────────
    def _read_df(self, path):
        import pandas as pd
        engine = "xlrd" if path.endswith(".xls") else "openpyxl"
        try:
            return pd.read_excel(path, header=None, dtype=str,
                                 engine=engine).fillna("")
        except Exception:
            import xlrd
            wb = xlrd.open_workbook(path, ignore_workbook_corruption=True)
            ws = wb.sheet_by_index(0)
            data = [[str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                    for r in range(ws.nrows)]
            import pandas as pd
            return pd.DataFrame(data).fillna("")

    def _flt(self, v):
        try: return float(str(v).replace(",", "").strip())
        except: return 0.0

    def _log(self, text):
        self.log_box.append(text)
        QApplication.processEvents()

    # ─────────────────── 核心导入 ───────────────────
    def _do_import(self):
        log = self._log
        log("🚀 开始导入账套…")
        conn = get_db(); c = conn.cursor()
        ok_total = err_total = 0
        try:
            # ── 创建 / 获取客户 ──
            if self.mode_new.isChecked():
                name  = self.f_name.text().strip()
                c.execute("SELECT id FROM clients WHERE name=?", (name,))
                row = c.fetchone()
                if row:
                    self._client_id = row[0]
                    log(f"ℹ 账套已存在，追加数据到：{name}")
                else:
                    c.execute(
                        "INSERT INTO clients(name,short_code,client_type,tax_id)"
                        " VALUES(?,?,?,?)",
                        (name, self.f_code.text().strip(),
                         self.f_type.currentText(),
                         self.f_taxid.text().strip()))
                    self._client_id = c.lastrowid
                    seed_client_accounts(self._client_id, conn)
                    conn.commit()
                    log(f"✓ 新建账套：{name}（ID={self._client_id}）")
            else:
                self._client_id = self.f_exist_combo.currentData()
                log(f"ℹ 追加到已有账套 ID={self._client_id}")

            # ── 1. 科目余额表 ──
            if "balance" in self._files:
                log("\n─── 导入科目余额表 ───")
                ok, err = self._imp_balance(conn, c)
                ok_total += ok; err_total += err

            # ── 2. 记账凭证 ──
            if "voucher" in self._files:
                log("\n─── 导入记账凭证 ───")
                ok, err = self._imp_voucher(conn, c)
                ok_total += ok; err_total += err

            # ── 3. 银行日记账 ──
            if "bank" in self._files:
                log("\n─── 导入银行日记账 ───")
                ok, err = self._imp_bank(conn, c)
                ok_total += ok; err_total += err

            # ── 4. 总账（补充科目期初） ──
            if "ledger" in self._files:
                log("\n─── 导入总账（补充期初） ───")
                ok, err = self._imp_ledger(conn, c)
                ok_total += ok; err_total += err

            # ── 5. 报表类 — 仅提示 ──
            for k, v in [("income","利润表"),("bs","资产负债表")]:
                if k in self._files:
                    log(f"\n○ {v} 已选择 — 仅供人工核对，不写入账套")

            log_action(conn, self._client_id, "账套导入", "import",
                       str(self._client_id),
                       f"文件:{list(self._files.keys())} 成功:{ok_total} 失败:{err_total}")
            conn.commit()
            log(f"\n✅ 导入完成！成功 {ok_total} 项，跳过/失败 {err_total} 项。")
            self.result_icon.setText("✅")
            self.result_title.setText("导入成功！")
        except Exception as e:
            conn.rollback()
            import traceback
            log(f"\n❌ 导入异常：{e}\n{traceback.format_exc()}")
            self.result_icon.setText("❌")
            self.result_title.setText("导入失败，请查看日志")
        finally:
            conn.close()

    # ── 科目余额表 ──
    def _imp_balance(self, conn, c):
        import re
        df = self._read_df(self._files["balance"])
        ok = err = 0
        for ri in range(len(df)):
            row = df.iloc[ri]
            code = str(row.iloc[1]).strip()
            name = str(row.iloc[2]).strip() if df.shape[1] > 2 else ""
            if not code or not name: continue
            if not re.match(r"^\d[\d._]*$", code): continue
            if len(code) < 4 or not code[:4].isdigit(): continue
            od = self._flt(row.iloc[3]) if df.shape[1] > 3 else 0
            oc = self._flt(row.iloc[4]) if df.shape[1] > 4 else 0
            # 推断类型 & 方向
            atype, direction = infer_account_type_direction(code, name)
            normalized = code.replace("_",".")
            level  = normalized.count(".") + 1
            parent = (code.rsplit(".",1)[0] if "." in code
                      else code.rsplit("_",1)[0] if "_" in code else None)
            c.execute("SELECT id FROM accounts WHERE client_id=? AND code=?",
                      (self._client_id, code))
            ex = c.fetchone()
            if ex:
                c.execute("UPDATE accounts SET name=?,full_name=?,type=?,direction=?,"
                          "opening_debit=?,opening_credit=? WHERE id=?",
                          (name, name, atype, direction, od, oc, ex[0]))
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,
                    direction,parent_code,level,opening_debit,opening_credit)
                    VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self._client_id,code,name,name,atype,direction,parent,level,od,oc))
            ok += 1
            if ok <= 5 or ok % 30 == 0:
                self._log(f"  {code}  {name}  借={od:.2f}  贷={oc:.2f}")
        self._log(f"  → 处理科目 {ok} 个")
        return ok, err

    # ── 记账凭证 ──
    def _imp_voucher(self, conn, c):
        import re
        df = self._read_df(self._files["voucher"])
        ok = skip = err = 0

        # 检测格式
        is_yonyou = any(
            "凭证字号" in str(df.iloc[r, 1] if df.shape[1] > 1 else "")
            for r in range(min(10, len(df))))

        title = str(df.iloc[0, 1]) if df.shape[1] > 1 else ""
        pm = re.search(r"(\d{4})年(\d{2})期", title)
        default_period = f"{pm.group(1)}-{pm.group(2)}" if pm else "2026-01"

        if is_yonyou:
            cur_vno = None; cur_date = ""; entries = []

            def flush(vno, date, ents):
                nonlocal ok, skip, err
                if not vno or not ents: return
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self._client_id, default_period, vno))
                if c.fetchone(): skip += 1; return
                td = sum(e[3] for e in ents); tc = sum(e[4] for e in ents)
                if abs(td - tc) > 0.01: err += 1; return
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status)"
                          " VALUES(?,?,?,?,?)",
                          (self._client_id, default_period, vno, date, "已审核"))
                vid = c.lastrowid
                for ln, ent in enumerate(ents, 1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,"
                              "account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid, ln) + ent)
                ok += 1
                if ok <= 5 or ok % 20 == 0:
                    self._log(f"  ✓ {vno}  {len(ents)} 行  借={td:.2f}")

            for ri in range(len(df)):
                row = df.iloc[ri]
                cell1 = str(row.iloc[1] if df.shape[1] > 1 else "").strip()
                if "凭证字号" in cell1:
                    flush(cur_vno, cur_date, entries); entries = []
                    dm = re.search(r"日期:(\S+)", cell1)
                    nm = re.search(r"凭证字号:(\S+)", cell1)
                    cur_date = dm.group(1) if dm else default_period + "-28"
                    cur_vno  = nm.group(1).split()[0] if nm else None
                elif not cell1 or cell1.startswith("合计"): continue
                else:
                    af = str(row.iloc[2] if df.shape[1] > 2 else "").strip()
                    if not af: continue
                    parts = af.split(" ", 1)
                    # Preserve auxiliary-account separators from source files.
                    code = parts[0]
                    aname = parts[1] if len(parts) > 1 else af
                    d  = self._flt(row.iloc[3]) if df.shape[1] > 3 else 0
                    cr = self._flt(row.iloc[4]) if df.shape[1] > 4 else 0
                    if d == 0 and cr == 0: continue
                    entries.append((cell1, code, aname, d, cr))
            flush(cur_vno, cur_date, entries)
        else:
            # 通用模板格式
            from collections import OrderedDict
            vouchers = OrderedDict()
            n = df.shape[1]
            for ri in range(1, len(df)):
                row = df.iloc[ri]
                def gcol(i, default=""):
                    try:
                        v = str(row.iloc[i]).strip() if i < n else default
                        return v if v not in ("nan","") else default
                    except: return default
                def gcolf(i):
                    try:
                        v = row.iloc[i] if i < n else 0
                        return self._flt(v) if v and str(v) not in ("nan","") else 0
                    except: return 0
                period = gcol(0); vno = gcol(1); date = gcol(2)
                summ   = gcol(3); code = gcol(4); aname = gcol(5)
                d = gcolf(6); cr = gcolf(7)
                if not period or not vno or not code: continue
                key = (period, vno)
                if key not in vouchers:
                    vouchers[key] = {"period":period,"vno":vno,"date":date,"entries":[]}
                vouchers[key]["entries"].append((summ, code, aname, d, cr))
            for (period, vno), v in vouchers.items():
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self._client_id, period, vno))
                if c.fetchone(): skip += 1; continue
                ents = v["entries"]
                td = sum(e[3] for e in ents); tc = sum(e[4] for e in ents)
                if abs(td - tc) > 0.01: err += 1; continue
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status)"
                          " VALUES(?,?,?,?,?)",
                          (self._client_id, period, vno, v["date"], "已审核"))
                vid = c.lastrowid
                for ln, ent in enumerate(ents, 1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,"
                              "account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid, ln) + ent)
                ok += 1

        self._log(f"  → 凭证导入 {ok} 张，跳过 {skip}，失败 {err}")
        return ok, skip + err

    # ── 银行日记账 → bank_statements ──
    def _imp_bank(self, conn, c):
        import re
        df = self._read_df(self._files["bank"])
        ok = skip = err = 0

        # 自动检测数据列布局
        # 用友明细账格式：col1=科目, col2=日期, col3=凭证号, col4=摘要,
        #                 col5=借方, col6=贷方, col8=余额
        # 通用银行流水格式：col0=日期, col1=摘要, col2=借/收入, col3=贷/支出, col4=余额

        # 先扫描找到第一个有 YYYY-MM-DD 格式日期的列和行
        date_col = None; data_start = 0
        for ri in range(min(15, len(df))):
            for ci in range(min(df.shape[1], 6)):
                v = str(df.iloc[ri, ci]).strip()
                if re.match(r"\d{4}[-/]\d{2}[-/]\d{2}", v):
                    date_col = ci; data_start = ri; break
            if date_col is not None: break

        if date_col is None:
            self._log("  ✗ 未能识别日期列，请确认文件格式"); return 0, 1

        # 根据 date_col 位置判断布局
        if date_col == 2:
            # 用友明细账：日期在 col2
            acct_col=1; vno_col=3; summ_col=4; d_col=5; cr_col=6; bal_col=8
        elif date_col == 0:
            # 通用银行流水：日期在 col0
            acct_col=None; vno_col=None; summ_col=1; d_col=2; cr_col=3; bal_col=4
        else:
            acct_col=None; vno_col=None; summ_col=date_col+1
            d_col=date_col+2; cr_col=date_col+3; bal_col=date_col+4

        def gcol(row, ci, default=""):
            try:
                v = str(row.iloc[ci]).strip() if ci is not None and ci < len(row) else default
                return v if v not in ("nan","") else default
            except: return default

        for ri in range(data_start, len(df)):
            row = df.iloc[ri]
            raw_date = gcol(row, date_col)
            # 标准化日期
            raw_date = raw_date.replace("/","-")
            if not re.match(r"\d{4}-\d{2}-\d{2}", raw_date): continue
            summary = gcol(row, summ_col)
            if summary in ("期初余额","本月合计","本年累计","合计",""): continue
            d  = self._flt(gcol(row, d_col))
            cr = self._flt(gcol(row, cr_col))
            bal = self._flt(gcol(row, bal_col)) if bal_col and bal_col < df.shape[1] else None
            if d == 0 and cr == 0: continue
            vno  = gcol(row, vno_col) if vno_col else ""
            # 科目信息
            acct_raw  = gcol(row, acct_col) if acct_col else "1002"
            parts = acct_raw.split(" ", 1)
            acct_code = parts[0] if re.match(r"^\d[\d.]*$", parts[0]) else "1002"
            acct_name = parts[1] if len(parts) > 1 else "银行存款"
            # 去重检查（日期+摘要+借+贷）
            c.execute("""SELECT id FROM bank_statements
                WHERE client_id=? AND date=? AND description=?
                AND debit=? AND credit=?""",
                (self._client_id, raw_date, summary, d, cr))
            if c.fetchone(): skip += 1; continue
            c.execute("""INSERT INTO bank_statements
                (client_id,account_code,account_name,date,voucher_no,
                 description,debit,credit,balance,source)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (self._client_id, acct_code, acct_name, raw_date,
                 vno, summary, d, cr, bal, "import"))
            ok += 1
            if ok <= 5 or ok % 50 == 0:
                self._log(f"  ✓ {raw_date}  {summary[:18]}  "
                          f"借={d:.2f}  贷={cr:.2f}")

        self._log(f"  → 银行流水导入 {ok} 条，跳过重复 {skip} 条")
        return ok, skip

    # ── 总账（补充期初） ──
    def _imp_ledger(self, conn, c):
        import re
        df = self._read_df(self._files["ledger"])
        ok = 0
        for ri in range(len(df)):
            row = df.iloc[ri]
            code = str(row.iloc[1]).strip() if df.shape[1] > 1 else ""
            name = str(row.iloc[2]).strip() if df.shape[1] > 2 else ""
            if not code or not name: continue
            if not re.match(r"^\d[\d._]*$", code): continue
            if len(code) < 4 or not code[:4].isdigit(): continue
            od = self._flt(row.iloc[3]) if df.shape[1] > 3 else 0
            oc = self._flt(row.iloc[4]) if df.shape[1] > 4 else 0
            if od == 0 and oc == 0: continue
            c.execute("SELECT id,opening_debit,opening_credit FROM accounts"
                      " WHERE client_id=? AND code=?", (self._client_id, code))
            ex = c.fetchone()
            if ex and ex["opening_debit"] == 0 and ex["opening_credit"] == 0:
                c.execute("UPDATE accounts SET opening_debit=?,opening_credit=? WHERE id=?",
                          (od, oc, ex["id"]))
                ok += 1
        self._log(f"  → 补充科目期初 {ok} 个")
        return ok, 0


class ClientPage(QWidget):
    client_opened = Signal(int, str, str)

    def __init__(self):
        super().__init__()
        L = QVBoxLayout(self); L.setContentsMargins(24,20,24,20); L.setSpacing(14)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("客户列表", bold=True, size=18)); hdr.addStretch()
        b_imp = QPushButton("📥 导入账套"); b_imp.setObjectName("btn_outline")
        b_imp.clicked.connect(self._import_account_set)
        b = QPushButton("＋ 新建客户"); b.setObjectName("btn_primary"); b.clicked.connect(self._add)
        hdr.addWidget(b_imp); hdr.addWidget(b); L.addLayout(hdr)
        self.search = QLineEdit(); self.search.setPlaceholderText("搜索客户名称或助记码...")
        self.search.textChanged.connect(self.load)
        L.addWidget(self.search)
        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.tbl = QTableWidget(); self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows); self.tbl.setShowGrid(False)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setColumnCount(7)
        self.tbl.setHorizontalHeaderLabels(["","客户名称","助记码","客户类型","税号","联系人","操作"])
        hh = self.tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)   # all columns user-draggable
        hh.setSectionResizeMode(1, QHeaderView.Stretch)    # name column stretches
        hh.setMinimumSectionSize(40)
        self.tbl.setColumnWidth(0, 44); self.tbl.setColumnWidth(2, 80)
        self.tbl.setColumnWidth(3, 110); self.tbl.setColumnWidth(4, 140)
        self.tbl.setColumnWidth(5, 90); self.tbl.setColumnWidth(6, 300)
        vl.addWidget(self.tbl); L.addWidget(f)

    def load(self):
        kw = self.search.text().strip()
        conn = get_db(); c = conn.cursor()
        if kw:
            c.execute("SELECT * FROM clients WHERE name LIKE ? OR short_code LIKE ? ORDER BY id",
                      (f"%{kw}%",f"%{kw}%"))
        else:
            c.execute("SELECT * FROM clients ORDER BY id")
        rows = c.fetchall(); conn.close()
        self.tbl.setRowCount(len(rows))
        for i,r in enumerate(rows):
            self.tbl.setRowHeight(i,50)
            # Index badge
            badge = QLabel(f"  {r['id']:02d}  ")
            badge.setStyleSheet("background:#f0f4ff;color:#3d6fdb;border-radius:4px;font-size:11px;")
            badge.setAlignment(Qt.AlignCenter)
            self.tbl.setCellWidget(i,0,badge)
            for j,v in enumerate([r['name'],r['short_code'] or '',r['client_type'] or '',
                                   r['tax_id'] or '',r['contact'] or ''],1):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter)
                it.setData(Qt.UserRole, r['id']); self.tbl.setItem(i,j,it)
            # Buttons
            bw = QWidget()
            bw.setAutoFillBackground(True)
            pal = bw.palette(); pal.setColor(QPalette.Window, QColor("#ffffff")); bw.setPalette(pal)
            bl = QHBoxLayout(bw); bl.setContentsMargins(8,4,8,4); bl.setSpacing(8)
            b1 = QPushButton("📂 进账簿"); b1.setObjectName("btn_primary")
            b1.setFixedSize(94, 30)
            b2 = QPushButton("✏ 编辑"); b2.setObjectName("btn_outline")
            b2.setFixedSize(68, 30)
            b3 = QPushButton("🗑 删除"); b3.setObjectName("btn_red")
            b3.setFixedSize(68, 30)
            b1.clicked.connect(lambda _,rr=r: self.client_opened.emit(rr['id'],rr['name'],rr['short_code'] or ''))
            b2.clicked.connect(lambda _,rr=r: self._edit(rr))
            b3.clicked.connect(lambda _,rr=r: self._del(rr))
            bl.addWidget(b1); bl.addWidget(b2); bl.addWidget(b3); bl.addStretch()
            self.tbl.setCellWidget(i,6,bw)

    def _import_account_set(self):
        d = ImportAccountSetDialog(self)
        d.exec(); self.load()

    def _add(self):
        d = ClientDialog(self)
        if d.exec(): self.load()

    def _edit(self,r):
        d = ClientDialog(self, r)
        if d.exec(): self.load()

    def _del(self,r):
        if QMessageBox.question(self,"确认",f"删除 [{r['name']}]？所有账目数据一并删除。",
                                QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
            conn = get_db()
            try:
                client_id = r['id']
                # Delete dependent rows explicitly because most FKs are NO ACTION.
                conn.execute("DELETE FROM voucher_entries WHERE voucher_id IN (SELECT id FROM vouchers WHERE client_id=?)",
                             (client_id,))
                conn.execute("DELETE FROM voucher_templates WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM bank_statements WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM account_aux_config WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM aux_items WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM aux_dimensions WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM periods WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM audit_log WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM vouchers WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM accounts WHERE client_id=?", (client_id,))
                conn.execute("DELETE FROM clients WHERE id=?", (client_id,))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
            self.load()


class VoucherPage(QWidget):
    """凭证管理 — 新增/查凭证/科目余额表/明细账"""

    def __init__(self):
        super().__init__()
        self.client_id = None; self.client_name = ""; self.period = ""
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)
        # Top toolbar (智一风格横向 tab)
        self.toolbar = QWidget()
        self.toolbar.setStyleSheet("background:#fff; border-bottom:1px solid #e8ecf2;")
        tl = QHBoxLayout(self.toolbar); tl.setContentsMargins(16,0,16,0); tl.setSpacing(0)
        self._tabs = []
        for name in ["新增凭证","查凭证","科目余额表","明细账","科目期初","辅助核算"]:
            b = QPushButton(name); b.setObjectName("nav_tab")
            b.setStyleSheet("""QPushButton{background:transparent;color:#888;border:none;
                padding:12px 16px;border-bottom:2px solid transparent;}
                QPushButton:hover{color:#3d6fdb;}
                QPushButton[active=true]{color:#3d6fdb;border-bottom:2px solid #3d6fdb;}""")
            b.clicked.connect(lambda _,n=name:self._switch_tab(n))
            tl.addWidget(b); self._tabs.append(b)
        tl.addStretch()
        self.client_lbl = lbl("", color="#3d6fdb", bold=True)
        self.period_combo = QComboBox(); self.period_combo.setMinimumWidth(110)
        self.period_combo.currentTextChanged.connect(self._on_period_change)
        tl.addWidget(self.client_lbl); tl.addSpacing(12); tl.addWidget(lbl("期间:"))
        tl.addWidget(self.period_combo)
        L.addWidget(self.toolbar)

        self.stack = QStackedWidget(); L.addWidget(self.stack)
        self._build_voucher_list(); self._build_balance(); self._build_ledger()
        self._aux_page = AuxPage()
        self.stack.addWidget(self._aux_page)   # index 3
        self._switch_tab("查凭证")

    def _switch_tab(self, name):
        mapping = {"新增凭证":None,"查凭证":0,"科目余额表":1,"明细账":2,
                   "科目期初":None,"辅助核算":3}
        for b in self._tabs:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        if name == "新增凭证": self._new_voucher()
        elif name == "科目期初": self._open_period_init()
        elif name == "辅助核算":
            self.stack.setCurrentIndex(3)
            if self.client_id: self._aux_page.set_client(self.client_id, self.period)
        elif mapping.get(name) is not None:
            self.stack.setCurrentIndex(mapping[name])
            if name == "查凭证": self._load_vouchers()
            elif name == "科目余额表": self._load_balance()
            elif name == "明细账": self._load_ledger()

    # ─ Voucher list ─
    def _build_voucher_list(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(10)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("凭证列表", bold=True, size=15))
        self.lock_lbl = QLabel("")  # shows 🔒 已封账 when period is closed
        hdr.addSpacing(12); hdr.addWidget(self.lock_lbl)
        hdr.addStretch()
        b_new = QPushButton("＋ 新增凭证"); b_new.setObjectName("btn_primary")
        b_new.clicked.connect(self._new_voucher)
        b_exp_doc = QPushButton("导出记账凭证"); b_exp_doc.setObjectName("btn_outline")
        b_exp_doc.clicked.connect(self._export_voucher_docs)
        b_exp = QPushButton("导出Excel"); b_exp.setObjectName("btn_outline")
        b_exp.clicked.connect(self._export_vouchers)
        hdr.addWidget(b_exp_doc); hdr.addWidget(b_exp); hdr.addWidget(b_new); L.addLayout(hdr)

        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.v_tbl = QTableWidget(); self.v_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.v_tbl.setSelectionBehavior(QTableWidget.SelectRows); self.v_tbl.setShowGrid(False)
        self.v_tbl.verticalHeader().setVisible(False)
        self.v_tbl.setColumnCount(7)
        self.v_tbl.setHorizontalHeaderLabels(["凭证号","日期","摘要","科目","借方合计","贷方合计","操作"])
        vh = self.v_tbl.horizontalHeader()
        vh.setSectionResizeMode(QHeaderView.Interactive)   # ALL columns user-draggable
        vh.setStretchLastSection(False)                    # don't auto-stretch last col
        vh.setMinimumSectionSize(60)
        self.v_tbl.setColumnWidth(0, 90)   # 凭证号
        self.v_tbl.setColumnWidth(1, 105)  # 日期
        self.v_tbl.setColumnWidth(2, 180)  # 摘要 — now Interactive, fully draggable
        self.v_tbl.setColumnWidth(3, 210)  # 科目
        self.v_tbl.setColumnWidth(4, 100)  # 借方合计
        self.v_tbl.setColumnWidth(5, 100)  # 贷方合计
        self.v_tbl.setColumnWidth(6, 340)  # 操作
        vl.addWidget(self.v_tbl); L.addWidget(f)
        self.stack.addWidget(w)

    def _load_vouchers(self):
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        # Show lock banner if period is closed
        c.execute("SELECT is_closed FROM periods WHERE client_id=? AND period=?",
                  (self.client_id, self.period))
        row = c.fetchone()
        is_closed = bool(row and row["is_closed"])
        c.execute("""SELECT v.id,v.voucher_no,v.date,v.status,
            (SELECT summary FROM voucher_entries WHERE voucher_id=v.id ORDER BY line_no LIMIT 1) as summ,
            (SELECT group_concat(account_name,'/') FROM voucher_entries WHERE voucher_id=v.id) as accts,
            (SELECT SUM(debit) FROM voucher_entries WHERE voucher_id=v.id) as td,
            (SELECT SUM(credit) FROM voucher_entries WHERE voucher_id=v.id) as tc
            FROM vouchers v WHERE v.client_id=? AND v.period=? ORDER BY v.voucher_no""",
                  (self.client_id, self.period))
        rows = c.fetchall(); conn.close()
        # Update period lock indicator in toolbar
        if hasattr(self, 'lock_lbl'):
            if is_closed:
                self.lock_lbl.setText("  🔒 已封账  ")
                self.lock_lbl.setStyleSheet("color:#ff4d4f;font-weight:bold;background:#fff1f0;border-radius:4px;padding:2px 6px;")
            else:
                self.lock_lbl.setText("")
                self.lock_lbl.setStyleSheet("")
        self.v_tbl.setRowCount(len(rows))
        for i,r in enumerate(rows):
            self.v_tbl.setRowHeight(i,46)
            status = r['status']
            status_color = {"待审核":"#fa8c16","已审核":"#52c41a","已拒绝":"#ff4d4f"}.get(status,'#888')
            no_w = QLabel(f"  {r['voucher_no']}  ")
            no_w.setStyleSheet("color:#3d6fdb;font-weight:bold;padding:0 8px;")
            no_w.setAlignment(Qt.AlignCenter)
            self.v_tbl.setCellWidget(i,0,no_w)
            for j,v in enumerate([r['date'],r['summ'] or '',r['accts'] or ''],1):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter)
                it.setData(Qt.UserRole,r['id']); self.v_tbl.setItem(i,j,it)
            for j,v in enumerate([fmt_amt(r['td']),fmt_amt(r['tc'])],4):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                it.setForeground(QColor("#3d6fdb")); self.v_tbl.setItem(i,j,it)
            # Action cell — status pill + audit buttons + edit + delete
            bw = QWidget(); bl = QHBoxLayout(bw); bl.setContentsMargins(6,3,6,3); bl.setSpacing(5)
            s_lbl = QLabel(f" {status} ")
            s_lbl.setStyleSheet(
                f"color:{status_color};font-size:11px;"
                f"border:1px solid {status_color};border-radius:3px;padding:2px 6px;")
            bl.addWidget(s_lbl)
            if status == "待审核":
                b_ok = QPushButton("✓ 审核通过"); b_ok.setObjectName("btn_green"); b_ok.setFixedSize(88,28)
                b_ok.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"已审核"))
                b_no = QPushButton("✗ 拒绝"); b_no.setObjectName("btn_red"); b_no.setFixedSize(60,28)
                b_no.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"已拒绝"))
                bl.addWidget(b_ok); bl.addWidget(b_no)
            elif status == "已拒绝":
                b_re = QPushButton("↩ 重新提交"); b_re.setObjectName("btn_outline"); b_re.setFixedSize(88,28)
                b_re.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"待审核"))
                bl.addWidget(b_re)
            elif status == "已审核":
                b_un = QPushButton("↩ 撤销审核"); b_un.setObjectName("btn_gray"); b_un.setFixedSize(88,28)
                b_un.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"待审核"))
                bl.addWidget(b_un)
            b_edit = QPushButton("✏ 编辑"); b_edit.setObjectName("btn_outline"); b_edit.setFixedSize(68,28)
            b_del = QPushButton("🗑 删除"); b_del.setObjectName("btn_red"); b_del.setFixedSize(60,28)
            b_edit.clicked.connect(lambda _,rid=r['id']:self._edit_voucher(rid))
            b_del.clicked.connect(lambda _,rid=r['id']:self._del_voucher(rid))
            bl.addWidget(b_edit); bl.addWidget(b_del); bl.addStretch()
            self.v_tbl.setCellWidget(i,6,bw)

    def _set_voucher_status(self, vid, new_status):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT voucher_no, status, period FROM vouchers WHERE id=?", (vid,))
        v = c.fetchone()
        if not v: conn.close(); return
        # 已结账期间禁止变更状态
        if self._is_period_closed(v["period"]):
            conn.close()
            QMessageBox.warning(self,"期间已封账","该凭证所在期间已封账，禁止修改审核状态。\n如需修改请先执行反结账。"); return
        # 审核前检查借贷平衡
        if new_status == "已审核":
            c.execute("SELECT ABS(SUM(debit)-SUM(credit)) AS diff FROM voucher_entries WHERE voucher_id=?", (vid,))
            diff = c.fetchone()["diff"] or 0
            if diff > 0.005:
                conn.close()
                QMessageBox.warning(self,"借贷不平",f"该凭证借贷差额 {diff:.2f}，不能审核通过。\n请先编辑修正后再审核。"); return
        conn.execute("UPDATE vouchers SET status=? WHERE id=?", (new_status, vid))
        log_action(conn, self.client_id, f"凭证审核:{new_status}", "voucher", vid, f"状态变更为{new_status}")
        conn.commit(); conn.close()
        self._load_vouchers()

    def _new_voucher(self):
        if not self.client_id:
            QMessageBox.information(self,"提示","请先从客户列表选择一个客户进入账簿"); return
        if self._is_period_closed():
            QMessageBox.warning(self,"期间已封账","该期间已结账封账，禁止新增凭证。\n如需修改请到【期末结账】页面执行反结账。"); return
        d = VoucherDialog(self, self.client_id, self.period)
        if d.exec():
            self._switch_tab("查凭证")
            if getattr(d,'saved_and_new',False): self._new_voucher()

    def _edit_voucher(self, vid):
        if self._is_period_closed():
            QMessageBox.warning(self,"期间已封账","该期间已结账封账，禁止修改凭证。\n如需修改请到【期末结账】页面执行反结账。"); return
        d = VoucherDialog(self, self.client_id, self.period, vid)
        if d.exec(): self._load_vouchers()

    def _del_voucher(self, vid):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT voucher_no, status, period FROM vouchers WHERE id=?", (vid,))
        v = c.fetchone(); conn.close()
        if not v: return
        # 已结账期间禁止删除
        if self._is_period_closed(v["period"]):
            QMessageBox.warning(self,"期间已封账","该凭证所在期间已封账，禁止删除。"); return
        # 已审核凭证需二次确认
        if v["status"] == "已审核":
            reply = QMessageBox.question(self, "⚠ 删除已审核凭证",
                f"凭证【{v['voucher_no']}】状态为【已审核】，删除将影响账务数据。\n\n确认要永久删除该凭证吗？",
                QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes: return
            # 记审计日志
            conn = get_db()
            log_action(conn, self.client_id, "删除已审核凭证", "voucher", vid,
                       f"凭证号:{v['voucher_no']} 期间:{v['period']}")
            conn.execute("DELETE FROM vouchers WHERE id=?", (vid,))
            conn.commit(); conn.close()
        else:
            if QMessageBox.question(self,"确认","删除该凭证？",
                    QMessageBox.Yes|QMessageBox.No) != QMessageBox.Yes: return
            conn = get_db()
            conn.execute("DELETE FROM vouchers WHERE id=?", (vid,))
            conn.commit(); conn.close()
        self._load_vouchers()

    # ─ Balance table (科目余额表) ─
    def _build_balance(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(10)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("科目余额表", bold=True, size=15)); hdr.addStretch()
        b_dl = QPushButton("下载"); b_dl.setObjectName("btn_outline"); b_dl.clicked.connect(self._export_balance)
        hdr.addWidget(b_dl); L.addLayout(hdr)
        # Filter row
        fr = QHBoxLayout(); fr.setSpacing(10)
        fr.addWidget(lbl("期间段:"))
        self.bal_start_period = QComboBox(); self.bal_start_period.setMinimumWidth(100)
        self.bal_end_period = QComboBox(); self.bal_end_period.setMinimumWidth(100)
        fr.addWidget(self.bal_start_period); fr.addWidget(lbl("至")); fr.addWidget(self.bal_end_period)
        b_refresh = QPushButton("刷新"); b_refresh.setObjectName("btn_primary"); b_refresh.clicked.connect(self._load_balance)
        fr.addWidget(b_refresh); fr.addSpacing(20)
        self.bal_aux = QCheckBox("辅助核算科目"); self.bal_detail = QCheckBox("明细科目")
        self.bal_zero = QCheckBox("0值科目")
        fr.addWidget(self.bal_aux); fr.addWidget(self.bal_detail); fr.addWidget(self.bal_zero)
        fr.addStretch(); L.addLayout(fr)
        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.bal_tbl = QTableWidget(); self.bal_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.bal_tbl.setShowGrid(True); self.bal_tbl.verticalHeader().setVisible(False)
        self.bal_tbl.setColumnCount(8)
        self.bal_tbl.setHorizontalHeaderLabels([
            "科目编号","科目名称","期初借方","期初贷方",
            "本期借方","本期贷方","期末借方","期末贷方"])
        self.bal_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        for ci in range(2,8): self.bal_tbl.setColumnWidth(ci,110)
        vl.addWidget(self.bal_tbl); L.addWidget(f)
        self.stack.addWidget(w)

    def _load_balance(self):
        if not self.client_id: return
        start_period = self.bal_start_period.currentData()
        end_period = self.bal_end_period.currentData()
        if not start_period or not end_period: return

        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM accounts WHERE client_id=? ORDER BY code", (self.client_id,))
        accts = {r['code']:dict(r) for r in c.fetchall()}
        # Aggregate voucher entries for period range
        c.execute("""SELECT e.account_code, SUM(e.debit) td, SUM(e.credit) tc
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period >= ? AND v.period <= ? GROUP BY e.account_code""",
                  (self.client_id, start_period, end_period))
        activity = {r['account_code']:(r['td'] or 0, r['tc'] or 0) for r in c.fetchall()}
        conn.close()

        rows = []
        for code, a in sorted(accts.items()):
            od = a['opening_debit'] or 0; oc = a['opening_credit'] or 0
            td, tc = activity.get(code,(0,0))
            # Compute ending balance
            if a['direction'] == '借':
                end_d = od + td - tc; end_c = 0
                if end_d < 0: end_c = -end_d; end_d = 0
            else:
                end_c = oc + tc - td; end_d = 0
                if end_c < 0: end_d = -end_c; end_c = 0
            rows.append((code, a['name'], od, oc, td, tc, end_d, end_c))

        # Totals
        totals = [sum(r[i] for r in rows) for i in range(2,8)]

        self.bal_tbl.setRowCount(len(rows)+1)
        for i,r in enumerate(rows):
            self.bal_tbl.setRowHeight(i,36)
            for j,v in enumerate(r):
                text = v if j < 2 else fmt_amt(v)
                it = QTableWidgetItem(text); it.setTextAlignment(Qt.AlignCenter if j<2 else Qt.AlignRight|Qt.AlignVCenter)
                if j==0: it.setForeground(QColor("#3d6fdb"))
                self.bal_tbl.setItem(i,j,it)
        # Total row
        self.bal_tbl.setRowHeight(len(rows),38)
        it0 = QTableWidgetItem(""); it1 = QTableWidgetItem("合  计")
        it1.setFont(QFont("",weight=QFont.Bold))
        it0.setBackground(QColor("#f5f7fa")); it1.setBackground(QColor("#f5f7fa"))
        it1.setTextAlignment(Qt.AlignCenter)
        self.bal_tbl.setItem(len(rows),0,it0); self.bal_tbl.setItem(len(rows),1,it1)
        for j,v in enumerate(totals,2):
            it = QTableWidgetItem(fmt_amt(v)); it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
            it.setBackground(QColor("#f5f7fa")); it.setFont(QFont("",weight=QFont.Bold))
            self.bal_tbl.setItem(len(rows),j,it)

    # ─ Ledger (明细账) ─
    def _build_ledger(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(10)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("明细账", bold=True, size=15)); hdr.addStretch()
        hdr.addWidget(lbl("期间段:"))
        self.ldg_start_period = QComboBox(); self.ldg_start_period.setMinimumWidth(100)
        self.ldg_end_period = QComboBox(); self.ldg_end_period.setMinimumWidth(100)
        hdr.addWidget(self.ldg_start_period); hdr.addWidget(lbl("至")); hdr.addWidget(self.ldg_end_period)
        hdr.addSpacing(10); hdr.addWidget(lbl("科目:"))
        self.ldg_acct = QComboBox(); self.ldg_acct.setMinimumWidth(220)
        b_query = QPushButton("查询"); b_query.setObjectName("btn_primary"); b_query.clicked.connect(self._load_ledger)
        hdr.addWidget(self.ldg_acct); hdr.addWidget(b_query)
        L.addLayout(hdr)
        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.ldg_tbl = QTableWidget(); self.ldg_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.ldg_tbl.setShowGrid(True); self.ldg_tbl.verticalHeader().setVisible(False)
        self.ldg_tbl.setColumnCount(6)
        self.ldg_tbl.setHorizontalHeaderLabels(["日期","摘要","借方","贷方","方向","余额"])
        self.ldg_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        vl.addWidget(self.ldg_tbl); L.addWidget(f)
        self.stack.addWidget(w)

    def _load_ledger(self):
        if not self.client_id: return
        start_period = self.ldg_start_period.currentData()
        end_period = self.ldg_end_period.currentData()
        if not start_period or not end_period: return

        # Populate account combo
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code",(self.client_id,))
        accts = c.fetchall()
        cur = self.ldg_acct.currentData()
        self.ldg_acct.clear()
        for a in accts:
            self.ldg_acct.addItem(f"{a['code']}  {a['name']}", a['code'])
        if cur:
            for i in range(self.ldg_acct.count()):
                if self.ldg_acct.itemData(i)==cur: self.ldg_acct.setCurrentIndex(i); break

        sel_code = self.ldg_acct.currentData()
        if not sel_code: conn.close(); return
        c.execute("SELECT opening_debit,opening_credit,direction FROM accounts WHERE client_id=? AND code=?",
                  (self.client_id, sel_code))
        acct = c.fetchone()
        c.execute("""SELECT v.date,e.summary,e.debit,e.credit FROM voucher_entries e
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period >= ? AND v.period <= ? AND e.account_code=? ORDER BY v.date,v.voucher_no,e.line_no""",
                  (self.client_id, start_period, end_period, sel_code))
        entries = c.fetchall(); conn.close()

        direction = acct['direction'] if acct else '借'
        balance = (acct['opening_debit'] or 0) - (acct['opening_credit'] or 0) if acct else 0

        rows = [("期初余额","",0,0,direction,balance)]
        for e in entries:
            d = e['debit'] or 0; cr = e['credit'] or 0
            balance += d - cr
            dir_str = "借" if balance >= 0 else "贷"
            rows.append((e['date'],e['summary'] or '',d,cr,dir_str,abs(balance)))
        rows.append(("本期合计","",sum(r[2] for r in rows[1:]),sum(r[3] for r in rows[1:]),"",None))

        self.ldg_tbl.setRowCount(len(rows))
        for i,(dt,summary,d,cr,dirr,bal) in enumerate(rows):
            self.ldg_tbl.setRowHeight(i,36)
            is_header = dt in ("期初余额","本期合计")
            vals = [dt,summary,fmt_amt(d),fmt_amt(cr),dirr,
                    fmt_amt(bal) if bal is not None else ""]
            for j,v in enumerate(vals):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignCenter if j<2 else Qt.AlignRight|Qt.AlignVCenter)
                if is_header:
                    it.setBackground(QColor("#f5f7fa")); it.setFont(QFont("",weight=QFont.Bold))
                if j==2 and d: it.setForeground(QColor("#3d6fdb"))
                if j==3 and cr: it.setForeground(QColor("#e05252"))
                self.ldg_tbl.setItem(i,j,it)

    def _open_period_init(self):
        if not self.client_id:
            QMessageBox.information(self,"提示","请先选择客户"); return
        d = AccountInitDialog(self, self.client_id, self.period)
        d.exec()

    # ─ Aux report (往来对账) ─
    def set_client(self, client_id, client_name, period):
        self.client_id = client_id; self.client_name = client_name; self.period = period
        self.client_lbl.setText(f"【{client_name}】")
        self._refresh_periods()
        # Initialize period ranges for balance and ledger
        self._init_period_ranges()
        # Auto refresh current tab
        idx = self.stack.currentIndex()
        if idx == 0: self._load_vouchers()
        elif idx == 1: self._load_balance()
        elif idx == 2: self._load_ledger()
        elif idx == 3 and self.client_id: self._aux_page.set_client(self.client_id, self.period)

    def _is_period_closed(self, period=None):
        """Return True if the given (or current) period is closed."""
        p = period or self.period
        if not self.client_id or not p: return False
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT is_closed FROM periods WHERE client_id=? AND period=?",
                  (self.client_id, p))
        row = c.fetchone(); conn.close()
        return bool(row and row["is_closed"])

    def _refresh_periods(self):
        self.period_combo.blockSignals(True)
        self.period_combo.clear()
        now = datetime.now()
        for y in range(now.year, now.year-3, -1):
            for m in range(12,0,-1):
                self.period_combo.addItem(f"{y}年{m:02d}期", f"{y}-{m:02d}")
        # Select current
        target = f"{now.year}-{now.month:02d}"
        for i in range(self.period_combo.count()):
            if self.period_combo.itemData(i)==target:
                self.period_combo.setCurrentIndex(i); break
        self.period = target
        self.period_combo.blockSignals(False)

    def _init_period_ranges(self):
        """Initialize period range selectors for balance and ledger pages"""
        # Balance page periods
        self.bal_start_period.clear()
        self.bal_end_period.clear()
        now = datetime.now()
        periods = []
        for y in range(now.year, now.year-3, -1):
            for m in range(12,0,-1):
                period_str = f"{y}-{m:02d}"
                display_str = f"{y}年{m:02d}期"
                periods.append((period_str, display_str))

        for period_str, display_str in periods:
            self.bal_start_period.addItem(display_str, period_str)
            self.bal_end_period.addItem(display_str, period_str)

        # Set default to current period
        current_period = f"{now.year}-{now.month:02d}"
        for i in range(self.bal_start_period.count()):
            if self.bal_start_period.itemData(i) == current_period:
                self.bal_start_period.setCurrentIndex(i)
                self.bal_end_period.setCurrentIndex(i)
                break

        # Ledger page periods
        self.ldg_start_period.clear()
        self.ldg_end_period.clear()
        for period_str, display_str in periods:
            self.ldg_start_period.addItem(display_str, period_str)
            self.ldg_end_period.addItem(display_str, period_str)

        # Set default to current period
        for i in range(self.ldg_start_period.count()):
            if self.ldg_start_period.itemData(i) == current_period:
                self.ldg_start_period.setCurrentIndex(i)
                self.ldg_end_period.setCurrentIndex(i)
                break

    def _on_period_change(self):
        self.period = self.period_combo.currentData() or self.period
        idx = self.stack.currentIndex()
        if idx==0: self._load_vouchers()
        elif idx==1: self._load_balance()
        elif idx==2: self._load_ledger()

    def _export_vouchers(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        path,_ = QFileDialog.getSaveFileName(self,"保存","凭证汇总.xlsx","Excel(*.xlsx)")
        if not path: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT v.voucher_no,v.date,e.summary,e.account_code,e.account_name,e.debit,e.credit
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period=? ORDER BY v.voucher_no,e.line_no""",
                  (self.client_id,self.period))
        rows = c.fetchall(); conn.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="凭证汇总"
        hdrs = ["凭证号","日期","摘要","科目编码","科目名称","借方","贷方"]
        fill = PatternFill("solid",fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell = ws.cell(1,ci,h)
            cell.font=XFont(bold=True,color="FFFFFF"); cell.fill=fill
            cell.alignment=Alignment(horizontal="center")
        for r in rows:
            ws.append([r['voucher_no'],r['date'],r['summary'],r['account_code'],
                       r['account_name'],r['debit'] or 0,r['credit'] or 0])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=16
        wb.save(path); QMessageBox.information(self,"成功",f"已导出:\n{path}")

    def _export_voucher_docs(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        period_text = (self.period or "").replace("-", "年", 1) + "期"
        path,_ = QFileDialog.getSaveFileName(
            self, "保存", f"记账凭证({period_text}).xlsx", "Excel(*.xlsx)")
        if not path: return

        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT id,voucher_no,date,status,attachment_count
            FROM vouchers
            WHERE client_id=? AND period=?
            ORDER BY date,voucher_no,id""",
            (self.client_id, self.period))
        vouchers = c.fetchall()
        if not vouchers:
            conn.close()
            QMessageBox.information(self, "提示", "当前账期没有可导出的凭证")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "记账凭证"

        title_fill = PatternFill("solid", fgColor="1C2340")
        sub_fill = PatternFill("solid", fgColor="EEF3FF")
        head_fill = PatternFill("solid", fgColor="F5F7FA")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = '#,##0.00'

        ws.merge_cells("B1:E1")
        ws["B1"] = f"{period_text} 记账凭证"
        ws["B1"].font = XFont(bold=True, color="FFFFFF", size=14)
        ws["B1"].fill = title_fill
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        row_idx = 3
        for v in vouchers:
            c.execute("""SELECT line_no,summary,account_code,account_name,debit,credit
                FROM voucher_entries
                WHERE voucher_id=?
                ORDER BY line_no""", (v["id"],))
            entries = c.fetchall()
            if not entries:
                continue

            ws.cell(row_idx, 2, f"日期:{v['date']}    凭证字号:{v['voucher_no']}")
            ws.cell(row_idx, 2).font = XFont(bold=True)
            ws.cell(row_idx, 2).fill = sub_fill
            ws.cell(row_idx, 2).alignment = Alignment(horizontal="left", vertical="center")
            for col in range(2, 6):
                ws.cell(row_idx, col).border = border
                if col > 2:
                    ws.cell(row_idx, col).fill = sub_fill
            row_idx += 1

            headers = ["摘要", "科目", "借方金额", "贷方金额"]
            for ci, text in enumerate(headers, start=2):
                cell = ws.cell(row_idx, ci, text)
                cell.font = XFont(bold=True)
                cell.fill = head_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1

            for ent in entries:
                ws.cell(row_idx, 2, ent["summary"] or "")
                ws.cell(row_idx, 3, f"{ent['account_code'] or ''} {ent['account_name'] or ''}".strip())
                ws.cell(row_idx, 4, ent["debit"] or 0)
                ws.cell(row_idx, 5, ent["credit"] or 0)
                ws.cell(row_idx, 4).number_format = money_fmt
                ws.cell(row_idx, 5).number_format = money_fmt
                for col in range(2, 6):
                    ws.cell(row_idx, col).border = border
                    ws.cell(row_idx, col).alignment = Alignment(vertical="center")
                row_idx += 1

            total_debit = sum((ent["debit"] or 0) for ent in entries)
            total_credit = sum((ent["credit"] or 0) for ent in entries)
            ws.cell(row_idx, 2, "合计：")
            ws.cell(row_idx, 4, total_debit)
            ws.cell(row_idx, 5, total_credit)
            ws.cell(row_idx, 4).number_format = money_fmt
            ws.cell(row_idx, 5).number_format = money_fmt
            for col in range(2, 6):
                ws.cell(row_idx, col).border = border
                ws.cell(row_idx, col).fill = sub_fill
                ws.cell(row_idx, col).font = XFont(bold=True)
            row_idx += 1

            ws.cell(row_idx, 2, f"附单据 {v['attachment_count'] or 0} 张")
            ws.cell(row_idx, 2).font = XFont(bold=True)
            ws.cell(row_idx, 2).alignment = Alignment(horizontal="left", vertical="center")
            for col in range(2, 6):
                ws.cell(row_idx, col).border = border
            row_idx += 2

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 36
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14

        conn.close()
        wb.save(path)
        QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _export_balance(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        path,_ = QFileDialog.getSaveFileName(self,"保存",f"科目余额表_{self.period}.xlsx","Excel(*.xlsx)")
        if not path: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM accounts WHERE client_id=? ORDER BY code",(self.client_id,))
        accts = {r['code']:dict(r) for r in c.fetchall()}
        c.execute("""SELECT e.account_code,SUM(e.debit),SUM(e.credit)
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period=? GROUP BY e.account_code""",
                  (self.client_id,self.period))
        activity = {r[0]:(r[1] or 0,r[2] or 0) for r in c.fetchall()}
        conn.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="科目余额表"
        hdrs = ["科目编号","科目名称","期初借方","期初贷方","本期借方","本期贷方","期末借方","期末贷方"]
        fill = PatternFill("solid",fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell = ws.cell(1,ci,h); cell.font=XFont(bold=True,color="FFFFFF")
            cell.fill=fill; cell.alignment=Alignment(horizontal="center")
        for code,a in sorted(accts.items()):
            od=a['opening_debit'] or 0; oc=a['opening_credit'] or 0
            td,tc=activity.get(code,(0,0))
            if a['direction']=='借': end_d=max(0,od+td-tc); end_c=0
            else: end_c=max(0,oc+tc-td); end_d=0
            ws.append([code,a['name'],od,oc,td,tc,end_d,end_c])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=14
        wb.save(path); QMessageBox.information(self,"成功",f"已导出:\n{path}")



class AccountPage(QWidget):
    """科目管理 — 查看/新增/编辑/删除二三级科目"""

    def __init__(self):
        super().__init__()
        self.client_id = None
        L = QVBoxLayout(self); L.setContentsMargins(24,20,24,20); L.setSpacing(14)

        hdr = QHBoxLayout()
        hdr.addWidget(lbl("会计科目管理", bold=True, size=18)); hdr.addStretch()
        b_add = QPushButton("＋ 新增科目"); b_add.setObjectName("btn_primary")
        b_add.clicked.connect(self._add)
        b_imp = QPushButton("从Excel导入历史数据"); b_imp.setObjectName("btn_outline")
        b_imp.clicked.connect(self._import_excel)
        hdr.addWidget(b_imp); hdr.addWidget(b_add); L.addLayout(hdr)

        info = QLabel("  提示：可新增子科目（如 6602.100 管理费用-其他）。系统默认科目不可删除，但可冻结或添加子科目。")
        info.setStyleSheet("background:#fffbe6;color:#ad6800;border-radius:6px;padding:8px 12px;font-size:12px;")
        L.addWidget(info)

        # Filter
        fr = QHBoxLayout(); fr.setSpacing(10)
        self.search_acct = QLineEdit(); self.search_acct.setPlaceholderText("搜索科目编号或名称...")
        self.search_acct.textChanged.connect(self.load)
        self.type_filter = QComboBox()
        self.type_filter.addItems(["全部类型","资产","负债","所有者权益","成本","收入","费用"])
        self.type_filter.currentIndexChanged.connect(self.load)
        fr.addWidget(self.search_acct); fr.addWidget(self.type_filter); fr.addStretch()
        L.addLayout(fr)

        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.tbl = QTableWidget(); self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows); self.tbl.setShowGrid(False)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setColumnCount(5)
        self.tbl.setHorizontalHeaderLabels(["科目编号","科目名称","类型","方向","操作"])
        hh = self.tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setStretchLastSection(True)
        hh.setMinimumSectionSize(55)
        self.tbl.setColumnWidth(0,120); self.tbl.setColumnWidth(1,240)
        self.tbl.setColumnWidth(2,80);  self.tbl.setColumnWidth(3,55)
        self.tbl.setColumnWidth(4,350)
        self.tbl.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        vl.addWidget(self.tbl); L.addWidget(f)

    def set_client(self, client_id):
        self.client_id = client_id
        self.load()

    def load(self):
        if not self.client_id: return
        kw = self.search_acct.text().strip()
        tf = self.type_filter.currentText()
        conn = get_db(); c = conn.cursor()
        sql = "SELECT * FROM accounts WHERE client_id=?"
        params = [self.client_id]
        if kw: sql += " AND (code LIKE ? OR name LIKE ?)"; params += [f"%{kw}%",f"%{kw}%"]
        if tf != "全部类型": sql += " AND type=?"; params.append(tf)
        sql += " ORDER BY code"
        c.execute(sql, params)
        rows = c.fetchall()
        
        # Check which accounts have been used
        used_accounts = set()
        c.execute("SELECT DISTINCT account_code FROM voucher_entries")
        for row in c.fetchall():
            used_accounts.add(row['account_code'])
        conn.close()

        self.tbl.setRowCount(len(rows))
        type_colors = {"资产":"#3d6fdb","负债":"#e05252","所有者权益":"#722ed1",
                       "成本":"#fa8c16","收入":"#52c41a","费用":"#eb5757"}
        for i,r in enumerate(rows):
            self.tbl.setRowHeight(i,68)
            level = r["level"] or 1
            indent = "    " * (level-1)
            code_it = QTableWidgetItem(r["code"])
            code_it.setForeground(QColor("#3d6fdb")); code_it.setTextAlignment(Qt.AlignCenter)
            name_it = QTableWidgetItem(indent + r["name"])
            if level == 1: name_it.setFont(QFont("",weight=QFont.Bold))
            
            # If account is frozen, show gray text
            try:
                is_frozen = r['is_frozen']
            except (KeyError, IndexError):
                is_frozen = 0
            if is_frozen:
                code_it.setForeground(QColor("#ccc"))
                name_it.setForeground(QColor("#ccc"))
            
            type_it = QTableWidgetItem(r["type"] or "")
            type_it.setForeground(QColor(type_colors.get(r["type"],"#888")))
            type_it.setTextAlignment(Qt.AlignCenter)
            dir_it = QTableWidgetItem(r["direction"] or "借"); dir_it.setTextAlignment(Qt.AlignCenter)
            for j,it in enumerate([code_it, name_it, type_it, dir_it]):
                self.tbl.setItem(i,j,it)

            bw = QWidget()
            bw.setAutoFillBackground(True)
            pal = bw.palette()
            pal.setColor(QPalette.Window, QColor("#ffffff"))
            bw.setPalette(pal)
            bl = QHBoxLayout(bw); bl.setContentsMargins(8,10,8,10); bl.setSpacing(12)
            
            # Button style to ensure text is visible on Windows
            outline_style = "color:#3d6fdb; border:1px solid #3d6fdb; background:transparent; border-radius:4px; padding:6px 14px; font-size:12px; font-weight:bold;"
            red_style = "color:#fff; background:#ff4d4f; border:none; border-radius:4px; padding:6px 14px; font-size:12px; font-weight:bold;"
            
            # If frozen, show frozen status
            if is_frozen:
                frozen_lbl = lbl("已冻结", color="#ccc", bold=True)
                bl.addWidget(frozen_lbl)
                bl.addStretch()
                self.tbl.setCellWidget(i,4,bw)
                continue
            
            b_sub = QPushButton("＋ 子科目")
            b_sub.setMinimumWidth(110); b_sub.setMaximumWidth(120); b_sub.setStyleSheet(outline_style)
            b_sub.clicked.connect(lambda _,rr=r: self._add_sub(rr))
            b_ed = QPushButton("✏ 编辑")
            b_ed.setMinimumWidth(90); b_ed.setMaximumWidth(100); b_ed.setStyleSheet(outline_style)
            b_ed.clicked.connect(lambda _,rr=r: self._edit(rr))
            bl.addWidget(b_sub); bl.addWidget(b_ed)
            
            if level > 1:
                is_used = r['code'] in used_accounts
                if is_used:
                    # Account has been used, show freeze button instead of delete
                    b_freeze = QPushButton("❄ 冻结")
                    b_freeze.setMinimumWidth(90); b_freeze.setMaximumWidth(100); b_freeze.setStyleSheet(outline_style)
                    b_freeze.setToolTip("冻结此科目，不再允许使用")
                    b_freeze.clicked.connect(lambda _,rid=r["id"]: self._freeze(rid))
                    bl.addWidget(b_freeze)
                else:
                    # Account not used, show delete button
                    b_del = QPushButton("🗑 删除")
                    b_del.setMinimumWidth(90); b_del.setMaximumWidth(100); b_del.setStyleSheet(red_style)
                    b_del.setToolTip("删除此科目")
                    b_del.clicked.connect(lambda _,rid=r["id"]: self._del(rid))
                    bl.addWidget(b_del)
            bl.addStretch()
            self.tbl.setCellWidget(i,4,bw)

    def _add(self):
        dlg = AccountEditDialog(self, self.client_id)
        if dlg.exec(): self.load()

    def _add_sub(self, parent_acct):
        # Check if parent account has been used (has voucher entries)
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM voucher_entries WHERE account_code=?", (parent_acct['code'],))
        used_count = c.fetchone()[0]
        conn.close()
        
        if used_count > 0:
            QMessageBox.warning(self, "无法添加下级科目", 
                f"上级科目 【{parent_acct['code']} {parent_acct['name']}】 已有凭证使用，不允许添加下级科目。")
            return
        
        dlg = AccountEditDialog(self, self.client_id, parent_acct=parent_acct)
        if dlg.exec(): self.load()

    def _edit(self, r):
        dlg = AccountEditDialog(self, self.client_id, account=r)
        if dlg.exec(): self.load()

    def _freeze(self, aid):
        """Freeze an account to prevent further use"""
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code, name, is_default FROM accounts WHERE id=?", (aid,))
        acct = c.fetchone()
        conn.close()
        
        if acct:
            is_def = acct['is_default'] if 'is_default' in acct.keys() else 0
            extra = "\n（系统默认科目冻结后仍保留，不可删除）" if is_def else ""
            # Second confirmation for freezing
            reply = QMessageBox.question(self, "冻结确认", 
                f"确认冻结科目 【{acct['code']} {acct['name']}】 吗？\n\n冻结后将不再允许使用此科目。{extra}",
                QMessageBox.Yes | QMessageBox.No)
            
            if reply == QMessageBox.Yes:
                conn = get_db()
                conn.execute("UPDATE accounts SET is_frozen=1 WHERE id=?", (aid,))
                conn.commit(); conn.close()
                QMessageBox.information(self, "成功", f"科目 【{acct['code']} {acct['name']}】 已冻结。")
                self.load()

    def _del(self, aid):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code, name, is_default FROM accounts WHERE id=?", (aid,))
        acct = c.fetchone()
        if not acct: conn.close(); return
        
        acct_code, acct_name = acct['code'], acct['name']
        try:
            is_def = acct['is_default']
        except (KeyError, IndexError):
            is_def = 0
        
        if is_def:
            conn.close()
            QMessageBox.warning(self, "系统默认科目", f"科目【{acct_code} {acct_name}】是系统默认科目，不可删除。如不需要使用，可以点击【冻结科目】。")
            return
        
        # Check if account has been used
        c.execute("SELECT COUNT(*) FROM voucher_entries WHERE account_code=?", (acct_code,))
        used_count = c.fetchone()[0]
        conn.close()
        
        if used_count > 0:
            # Account has been used, offer freeze or delete confirmation
            msg_box = QMessageBox(QMessageBox.Warning, "科目已使用", 
                f"科目 【{acct_code} {acct_name}】 已有凭证使用。\n\n选择操作：")
            btn_freeze = msg_box.addButton("冻结科目", QMessageBox.AcceptRole)
            btn_cancel = msg_box.addButton("取消", QMessageBox.RejectRole)
            msg_box.exec()
            
            if msg_box.clickedButton() == btn_freeze:
                # Freeze the account
                conn = get_db()
                conn.execute("UPDATE accounts SET is_frozen=1 WHERE id=?", (aid,))
                conn.commit(); conn.close()
                QMessageBox.information(self, "成功", f"科目 【{acct_code} {acct_name}】 已冻结，不再允许使用。")
                self.load()
            return
        
        # Account not used, show delete confirmation dialog
        reply = QMessageBox.question(self, "二次确认", 
            f"确认删除科目 【{acct_code} {acct_name}】 吗？\n\n此操作不可撤销。",
            QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            conn = get_db()
            conn.execute("DELETE FROM accounts WHERE id=?", (aid,))
            conn.commit(); conn.close()
            QMessageBox.information(self, "成功", f"科目 【{acct_code} {acct_name}】 已删除。")
            self.load()

    def _import_excel(self):
        """Import historical vouchers from Excel."""
        if not self.client_id:
            QMessageBox.information(self,"提示","请先选择客户"); return
        dlg = ImportExcelDialog(self, self.client_id)
        dlg.exec()
        self.load()


class AccountEditDialog(QDialog):
    def __init__(self, parent, client_id, account=None, parent_acct=None):
        super().__init__(parent)
        self.client_id = client_id; self.account = account; self.parent_acct = parent_acct
        self.setWindowTitle("编辑科目" if account else "新增科目")
        self.setMinimumWidth(420); self._build()
        if account: self._load()
        elif parent_acct: self._prefill_from_parent()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(24,20,24,20); L.setSpacing(12)
        L.addWidget(lbl("科目信息", bold=True, size=15))
        F = QFormLayout(); F.setSpacing(10); F.setLabelAlignment(Qt.AlignRight)
        self.code = QLineEdit(); self.code.setPlaceholderText("如 1002.01")
        self.name = QLineEdit(); self.name.setPlaceholderText("科目名称（必填）")
        self.full_name = QLineEdit(); self.full_name.setPlaceholderText("完整名称，如 银行存款-工商银行")
        self.type_cb = QComboBox()
        self.type_cb.addItems(["资产","负债","所有者权益","成本","收入","费用"])
        self.dir_cb = QComboBox(); self.dir_cb.addItems(["借","贷"])
        self.parent_cb = QComboBox(); self.parent_cb.addItem("（无上级）","")
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code",(self.client_id,))
        for r in c.fetchall():
            self.parent_cb.addItem(f"{r['code']}  {r['name']}", r['code'])
        conn.close()
        self.od = NoScrollDoubleSpinBox(); self.od.setRange(0,9999999999); self.od.setDecimals(2); self.od.setPrefix("¥ ")
        self.oc = NoScrollDoubleSpinBox(); self.oc.setRange(0,9999999999); self.oc.setDecimals(2); self.oc.setPrefix("¥ ")
        F.addRow("科目编号 *", self.code); F.addRow("科目名称 *", self.name)
        F.addRow("完整名称", self.full_name); F.addRow("科目类型", self.type_cb)
        F.addRow("余额方向", self.dir_cb); F.addRow("上级科目", self.parent_cb)
        F.addRow("期初借方", self.od); F.addRow("期初贷方", self.oc)
        L.addLayout(F)
        row = QHBoxLayout(); row.addStretch()
        bc = QPushButton("取消"); bc.setObjectName("btn_gray")
        bs = QPushButton("保存"); bs.setObjectName("btn_primary")
        bc.clicked.connect(self.reject); bs.clicked.connect(self._save)
        row.addWidget(bc); row.addWidget(bs); L.addLayout(row)

    def _load(self):
        r = self.account
        self.code.setText(r["code"] or ""); self.name.setText(r["name"] or "")
        self.full_name.setText(r["full_name"] or "")
        idx = self.type_cb.findText(r["type"] or ""); self.type_cb.setCurrentIndex(max(0,idx))
        idx2 = self.dir_cb.findText(r["direction"] or "借"); self.dir_cb.setCurrentIndex(max(0,idx2))
        self.od.setValue(r["opening_debit"] or 0); self.oc.setValue(r["opening_credit"] or 0)
        if r["parent_code"]:
            for i in range(self.parent_cb.count()):
                if self.parent_cb.itemData(i) == r["parent_code"]:
                    self.parent_cb.setCurrentIndex(i); break

    def _prefill_from_parent(self):
        p = self.parent_acct
        self.code.setText(p["code"] + ".01")
        idx = self.type_cb.findText(p["type"] or ""); self.type_cb.setCurrentIndex(max(0,idx))
        idx2 = self.dir_cb.findText(p["direction"] or "借"); self.dir_cb.setCurrentIndex(max(0,idx2))
        for i in range(self.parent_cb.count()):
            if self.parent_cb.itemData(i) == p["code"]:
                self.parent_cb.setCurrentIndex(i); break

    def _save(self):
        code = self.code.text().strip(); name = self.name.text().strip()
        if not code or not name: QMessageBox.warning(self,"提示","编号和名称不能为空"); return
        full = self.full_name.text().strip() or name
        parent = self.parent_cb.currentData() or None
        level = (parent.count(".")+2) if parent else 1
        conn = get_db(); c = conn.cursor()
        try:
            if self.account:
                c.execute("""UPDATE accounts SET code=?,name=?,full_name=?,type=?,direction=?,
                    parent_code=?,level=?,opening_debit=?,opening_credit=? WHERE id=?""",
                    (code,name,full,self.type_cb.currentText(),self.dir_cb.currentText(),
                     parent,level,self.od.value(),self.oc.value(),self.account["id"]))
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,direction,
                    parent_code,level,opening_debit,opening_credit) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self.client_id,code,name,full,self.type_cb.currentText(),
                     self.dir_cb.currentText(),parent,level,self.od.value(),self.oc.value()))
            conn.commit(); conn.close(); self.accept()
        except Exception as e:
            conn.close(); QMessageBox.warning(self,"错误",f"保存失败：{e}")


class ImportExcelDialog(QDialog):
    """三合一导入：凭证 / 科目余额表期初 / 银行日记账"""

    def __init__(self, parent, client_id):
        super().__init__(parent)
        self.client_id = client_id
        self.setWindowTitle("从Excel导入历史数据")
        self.setMinimumSize(780, 560)
        self._build()

    def _build(self):
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)

        # Tab bar
        tab_bar = QWidget(); tab_bar.setStyleSheet("background:#f5f7fa;border-bottom:1px solid #e4e8f0;")
        tl = QHBoxLayout(tab_bar); tl.setContentsMargins(16,0,0,0); tl.setSpacing(0)
        self._itabs = []
        for name in ["记账凭证", "科目期初余额", "银行日记账"]:
            b = QPushButton(name)
            b.setStyleSheet("""QPushButton{background:transparent;color:#888;border:none;
                padding:11px 18px;border-bottom:2px solid transparent;}
                QPushButton:hover{color:#3d6fdb;}
                QPushButton[active=true]{color:#3d6fdb;border-bottom:2px solid #3d6fdb;font-weight:bold;}""")
            b.clicked.connect(lambda _,n=name: self._switch_itab(n))
            tl.addWidget(b); self._itabs.append(b)
        tl.addStretch()
        L.addWidget(tab_bar)

        self.istack = QStackedWidget(); L.addWidget(self.istack)
        self._build_voucher_import()
        self._build_balance_import()
        self._build_bank_import()
        self._switch_itab("记账凭证")

    def _switch_itab(self, name):
        mapping = {"记账凭证":0,"科目期初余额":1,"银行日记账":2}
        for b in self._itabs:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        self.istack.setCurrentIndex(mapping[name])

    # ── helpers ──
    def _make_tab(self, info_text):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,16,20,16); L.setSpacing(10)
        info = QLabel(info_text)
        info.setStyleSheet("background:#f6f8ff;border-radius:6px;padding:10px 14px;font-size:12px;color:#444;")
        info.setWordWrap(True); L.addWidget(info)
        log = QTextEdit(); log.setReadOnly(True)
        log.setStyleSheet("font-size:12px;font-family:monospace;background:#fafafa;")
        log.setPlaceholderText("导入日志将显示在这里…")
        L.addWidget(log)
        close_btn = QPushButton("关闭"); close_btn.setObjectName("btn_gray")
        close_btn.clicked.connect(self.accept)
        row = QHBoxLayout(); row.addStretch(); row.addWidget(close_btn); L.addLayout(row)
        return w, L, log

    def _pick_xls(self):
        path, _ = QFileDialog.getOpenFileName(self,"选择文件","","Excel(*.xls *.xlsx)")
        return path

    # ── Tab 1: 记账凭证 ──
    def _build_voucher_import(self):
        info = ("支持格式：从用友/金蝶等软件导出的记账凭证XLS文件。\n"
                "识别规则：每张凭证以[日期:...凭证字号:记-xxx]开头，下方各行为分录，[合计]行结束。\n"
                "也支持通用模板格式（A=期间 B=凭证号 C=日期 D=摘要 E=科目编号 F=科目名 G=借方 H=贷方）。")
        w, L, self.v_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_tmpl = QPushButton("↓ 下载通用模板"); b_tmpl.setObjectName("btn_outline")
        b_tmpl.clicked.connect(self._dl_voucher_template)
        b_imp = QPushButton("📂 导入凭证文件"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_vouchers)
        btn_row.addWidget(b_tmpl); btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _dl_voucher_template(self):
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        path, _ = QFileDialog.getSaveFileName(self,"保存模板","凭证导入模板.xlsx","Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="凭证数据"
        hdrs = ["期间(YYYY-MM)","凭证号","日期(YYYY-MM-DD)","摘要","科目编号","科目名称","借方","贷方"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell=ws.cell(1,ci,h); cell.font=XFont(bold=True,color="FFFFFF")
            cell.fill=fill; cell.alignment=Alignment(horizontal="center")
        samples = [
            ("2025-01","记-001","2025-01-15","发放工资","2211.001","应付职工薪酬-工资",5000,0),
            ("2025-01","记-001","2025-01-15","发放工资","1002.001","银行存款-光大银行",0,5000),
            ("2025-01","记-002","2025-01-20","收到货款","1002.001","银行存款-光大银行",10000,0),
            ("2025-01","记-002","2025-01-20","收到货款","6001.001","主营业务收入-咨询",0,10000),
        ]
        for r in samples: ws.append(list(r))
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=20
        wb.save(path); self.v_log.append(f"✓ 模板已保存: {path}")

    def _import_vouchers(self):
        path = self._pick_xls()
        if not path: return
        self.v_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.v_log.append(f"✗ 读取失败: {e}"); return

        # Detect format: 用友-style vs generic template
        # 用友 style: row with "日期:…凭证字号:" header rows
        is_yonyou = any("凭证字号" in str(df.iloc[r,1]) for r in range(min(10,len(df))))

        conn = get_db(); c = conn.cursor()
        ok = skip = err = 0

        if is_yonyou:
            # Parse 用友/金蝶 style
            # Extract period from title row (row 0, col 1: "2026年03期 凭证")
            title = str(df.iloc[0,1])
            import re
            pm = re.search(r"(\d{4})年(\d{2})期", title)
            period = f"{pm.group(1)}-{pm.group(2)}" if pm else "2026-01"
            self.v_log.append(f"检测到期间: {period}，开始解析…")

            cur_vno = None; cur_date = ""; entries = []

            def flush(vno, date, ents):
                nonlocal ok, skip, err
                if not vno or not ents: return
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self.client_id, period, vno))
                if c.fetchone():
                    self.v_log.append(f"  跳过 {vno}（已存在）"); skip += 1; return
                td = sum(e[3] for e in ents); tc = sum(e[4] for e in ents)
                if abs(td-tc) > 0.01:
                    self.v_log.append(f"  ✗ {vno} 借贷不平 差{td-tc:.2f}，跳过"); err += 1; return
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status) VALUES(?,?,?,?,?)",
                          (self.client_id,period,vno,date,"已审核"))
                vid = c.lastrowid
                for ln,ent in enumerate(ents,1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln)+ent)
                ok += 1
                self.v_log.append(f"  ✓ {vno}  {len(ents)}行  借={td:.2f}")

            for ri in range(len(df)):
                row = df.iloc[ri]
                cell1 = str(row.iloc[1]).strip()
                if "凭证字号" in cell1:
                    flush(cur_vno, cur_date, entries)
                    entries = []
                    dm = re.search(r"日期:(\S+)", cell1)
                    nm = re.search(r"凭证字号:(\S+)", cell1)
                    cur_date = dm.group(1) if dm else period+"-28"
                    cur_vno  = nm.group(1).split()[0] if nm else None
                elif cell1 in ("合计：","合计:","") or not cell1:
                    continue
                else:
                    # entry row: col1=summary, col2=account, col3=debit, col4=credit
                    acct_full = str(row.iloc[2]).strip()
                    if not acct_full: continue
                    parts = acct_full.split(" ", 1)
                    code = parts[0]; aname = parts[1] if len(parts)>1 else acct_full
                    # Preserve auxiliary-account separators from source files.
                    code_norm = code
                    try: d = float(row.iloc[3]) if row.iloc[3] else 0
                    except: d = 0
                    try: cr = float(row.iloc[4]) if row.iloc[4] else 0
                    except: cr = 0
                    if d == 0 and cr == 0: continue
                    entries.append((cell1, code_norm, aname, d, cr))
            flush(cur_vno, cur_date, entries)

        else:
            # Generic template format (cols: period,vno,date,summary,code,name,debit,credit)
            from collections import OrderedDict
            vouchers = OrderedDict()
            n_cols = df.shape[1]
            for ri in range(1, len(df)):
                row = df.iloc[ri]
                def gcol(i, default=""):
                    try: v = str(row.iloc[i]).strip() if i < n_cols else default; return v if v != "nan" else default
                    except: return default
                def gcol_f(i):
                    try: v = row.iloc[i] if i < n_cols else 0; return float(v) if v and str(v) != "nan" else 0
                    except: return 0
                period  = gcol(0); vno = gcol(1); date = gcol(2)
                summary = gcol(3); code = gcol(4); aname = gcol(5)
                d = gcol_f(6); cr = gcol_f(7)
                if not period or not vno or not code: continue
                key = (period, vno)
                if key not in vouchers:
                    vouchers[key] = {"period":period,"vno":vno,"date":date,"entries":[]}
                vouchers[key]["entries"].append((summary,code,aname,d,cr))
            for (period,vno),v in vouchers.items():
                c.execute("SELECT id FROM vouchers WHERE client_id=? AND period=? AND voucher_no=?",
                          (self.client_id,period,vno))
                if c.fetchone(): skip+=1; continue
                ents = v["entries"]
                td=sum(e[3] for e in ents); tc=sum(e[4] for e in ents)
                if abs(td-tc)>0.01: err+=1; continue
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status) VALUES(?,?,?,?,?)",
                          (self.client_id,period,vno,v["date"],"已审核"))
                vid=c.lastrowid
                for ln,ent in enumerate(ents,1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln)+ent)
                ok+=1

        if ok:
            log_action(conn, self.client_id, "批量导入凭证", "import", "",
                       f"导入{ok}张凭证，跳过{skip}张，失败{err}张")
        conn.commit(); conn.close()
        self.v_log.append(f"\n✅ 完成：导入 {ok} 张，跳过 {skip} 张，失败 {err} 张")

    # ── Tab 2: 科目余额表期初 ──
    def _build_balance_import(self):
        info = ("支持格式：用友/金蝶导出的「科目余额表」XLS。\n"
                "识别列：科目编号（第2列）、科目名称（第3列）、期初借方（第4列）、期初贷方（第5列）。\n"
                "导入后自动创建不存在的科目，并设置期初余额。")
        w, L, self.b_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_imp = QPushButton("📂 导入科目余额表"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_balance)
        btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _import_balance(self):
        path = self._pick_xls()
        if not path: return
        self.b_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.b_log.append(f"✗ 读取失败: {e}"); return

        import re
        # Find data rows: col1=科目编号(numeric-ish), col2=科目名称, col3=期初借方, col4=期初贷方
        # Header rows are rows 0-5 typically; data starts where col1 looks like an account code
        conn = get_db(); c = conn.cursor()
        created = updated = skipped = 0

        for ri in range(len(df)):
            row = df.iloc[ri]
            code = str(row.iloc[1]).strip()
            name = str(row.iloc[2]).strip()
            if not code or not name: continue
            # Account codes: start with digit, may contain dots or underscores (auxiliary dims)
            if not re.match(r"^\d[\d._]*$", code): continue
            # Skip obvious non-account rows (e.g. "2026年03期" matched by looser regex)
            if len(code) < 4 or not code[:4].isdigit(): continue
            try:
                od = float(str(row.iloc[3]).replace(",","")) if row.iloc[3] else 0
            except: od = 0
            try:
                oc = float(str(row.iloc[4]).replace(",","")) if row.iloc[4] else 0
            except: oc = 0

            # Determine account type from code + name
            acct_type, direction = infer_account_type_direction(code, name)

            # Compute level and parent from code (treat _ same as . for hierarchy)
            normalized = code.replace("_", ".")
            level = normalized.count(".") + 1
            parent = code.rsplit(".", 1)[0] if "." in code else (
                     code.rsplit("_", 1)[0] if "_" in code else None)

            c.execute("SELECT id FROM accounts WHERE client_id=? AND code=?",
                      (self.client_id, code))
            existing = c.fetchone()
            if existing:
                c.execute("UPDATE accounts SET opening_debit=?,opening_credit=?,name=?,full_name=?,type=?,direction=? WHERE id=?",
                          (od, oc, name, name, acct_type, direction, existing[0]))
                updated += 1
                self.b_log.append(f"  ↻ {code} {name}  期初借={od:.2f} 贷={oc:.2f}")
            else:
                c.execute("""INSERT INTO accounts(client_id,code,name,full_name,type,direction,
                    parent_code,level,opening_debit,opening_credit) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (self.client_id,code,name,name,acct_type,direction,parent,level,od,oc))
                created += 1
                self.b_log.append(f"  ✓ 新建 {code} {name}  期初借={od:.2f} 贷={oc:.2f}")

        conn.commit(); conn.close()
        self.b_log.append(f"\n✅ 完成：新建科目 {created} 个，更新期初 {updated} 个，跳过 {skipped} 个")

    # ── Tab 3: 银行日记账 ──
    def _build_bank_import(self):
        info = ("支持格式：用友/金蝶导出的「银行存款日记账 / 明细账」XLS。\n"
                "识别列：科目（第2列）、日期（第3列）、凭证号（第4列）、摘要（第5列）、借方（第6列）、贷方（第7列）、余额（第9列）。\n"
                "导入后可在「记账(凭证)→明细账」中查看，并可与实际银行流水对照。")
        w, L, self.k_log = self._make_tab(info)
        btn_row = QHBoxLayout()
        b_imp = QPushButton("📂 导入银行日记账"); b_imp.setObjectName("btn_primary")
        b_imp.clicked.connect(self._import_bank)
        btn_row.addWidget(b_imp); btn_row.addStretch()
        L.insertLayout(1, btn_row)
        self.istack.addWidget(w)

    def _import_bank(self):
        path = self._pick_xls()
        if not path: return
        self.k_log.clear()
        try:
            import pandas as pd
            df = pd.read_excel(path, engine="xlrd" if path.endswith(".xls") else "openpyxl",
                               header=None, dtype=str)
            df = df.fillna("")
        except Exception as e:
            self.k_log.append(f"✗ 读取失败: {e}"); return

        import re
        conn = get_db(); c = conn.cursor()
        ok = skip = 0

        # 自动识别日期列位置
        date_col = None; data_start = 0
        for ri in range(min(15, len(df))):
            for ci in range(min(df.shape[1], 6)):
                v = str(df.iloc[ri, ci]).strip()
                if re.match(r"\d{4}[-/]\d{2}[-/]\d{2}", v):
                    date_col = ci; data_start = ri; break
            if date_col is not None: break

        if date_col is None:
            self.k_log.append("✗ 未能识别日期列，请确认文件格式"); conn.close(); return

        # 根据 date_col 判断布局
        if date_col == 2:   # 用友明细账
            acct_col=1; vno_col=3; summ_col=4; d_col=5; cr_col=6; bal_col=8
        elif date_col == 0: # 通用银行流水
            acct_col=None; vno_col=None; summ_col=1; d_col=2; cr_col=3; bal_col=4
        else:
            acct_col=None; vno_col=None; summ_col=date_col+1
            d_col=date_col+2; cr_col=date_col+3; bal_col=date_col+4

        def gcol(row, ci, default=""):
            try:
                v = str(row.iloc[ci]).strip() if ci is not None and ci < len(row) else default
                return v if v not in ("nan","") else default
            except: return default

        def flt(v):
            try: return float(str(v).replace(",","").strip())
            except: return 0.0

        for ri in range(data_start, len(df)):
            row = df.iloc[ri]
            raw_date = gcol(row, date_col).replace("/","-")
            if not re.match(r"\d{4}-\d{2}-\d{2}", raw_date): continue
            summary = gcol(row, summ_col)
            if summary in ("期初余额","本月合计","本年累计","合计",""): continue
            d  = flt(gcol(row, d_col))
            cr = flt(gcol(row, cr_col))
            bal_v = gcol(row, bal_col) if bal_col and bal_col < df.shape[1] else ""
            bal = flt(bal_v) if bal_v else None
            if d == 0 and cr == 0: continue
            vno = gcol(row, vno_col) if vno_col else ""
            acct_raw = gcol(row, acct_col) if acct_col else "1002"
            parts = acct_raw.split(" ", 1)
            acct_code = parts[0] if re.match(r"^\d[\d.]*$", parts[0]) else "1002"
            acct_name = parts[1] if len(parts) > 1 else "银行存款"
            # 去重
            c.execute("""SELECT id FROM bank_statements
                WHERE client_id=? AND date=? AND description=? AND debit=? AND credit=?""",
                (self.client_id, raw_date, summary, d, cr))
            if c.fetchone(): skip += 1; continue
            c.execute("""INSERT INTO bank_statements
                (client_id,account_code,account_name,date,voucher_no,
                 description,debit,credit,balance,source)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (self.client_id, acct_code, acct_name, raw_date,
                 vno, summary, d, cr, bal, "import"))
            ok += 1
            self.k_log.append(f"  ✓ {raw_date}  {summary[:20]}  借={d:.2f}  贷={cr:.2f}")

        conn.commit(); conn.close()
        self.k_log.append(f"\n✅ 完成：导入 {ok} 条，跳过重复 {skip} 条")


class SettlePage(QWidget):
    """期末结账"""
    carryforward_done = Signal()   # emitted after vouchers created

    def __init__(self):
        super().__init__()
        self.client_id = None; self.client_name = ""; self.period = ""
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0,0,0,0)
        outer.setSpacing(0)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        outer.addWidget(scroll)

        content = QWidget()
        scroll.setWidget(content)
        L = QVBoxLayout(content); L.setContentsMargins(24,20,24,20); L.setSpacing(14)

        # Step indicator
        step_row = QHBoxLayout(); step_row.setSpacing(0)
        s1 = self._step_box("1","期末结转","active"); s2 = self._step_box("2","结账检测","")
        step_row.addWidget(s1); step_row.addWidget(lbl("  ➔  ", color="#bbb"))
        step_row.addWidget(s2); step_row.addStretch()
        L.addLayout(step_row)

        # Period / client row
        pr = QHBoxLayout()
        self.period_combo = QComboBox()
        self.period_combo.setMinimumWidth(130)
        self.period_combo.currentIndexChanged.connect(self._on_period_change)
        self.client_lbl = lbl("请先从客户管理进入账簿", color="#888")
        self.status_lbl = lbl("", bold=True)   # 显示已结账/未结账状态
        self.do_btn = QPushButton("生成结转凭证"); self.do_btn.setObjectName("btn_primary")
        self.do_btn.clicked.connect(self._do_carryforward)
        self.close_btn = QPushButton("🔒 结账封账"); self.close_btn.setObjectName("btn_primary")
        self.close_btn.clicked.connect(self._close_period)
        self.reopen_btn = QPushButton("🔓 反结账"); self.reopen_btn.setObjectName("btn_red")
        self.reopen_btn.clicked.connect(self._reopen_period)
        pr.addWidget(lbl("结账期间：")); pr.addWidget(self.period_combo)
        pr.addSpacing(10); pr.addWidget(self.client_lbl)
        pr.addSpacing(12); pr.addWidget(self.status_lbl)
        pr.addStretch()
        pr.addWidget(self.do_btn); pr.addWidget(self.close_btn); pr.addWidget(self.reopen_btn)
        L.addLayout(pr)
        L.addWidget(sep())

        # Carry cards
        cards_row = QHBoxLayout(); cards_row.setSpacing(16)
        self.card_income  = self._carry_card("结转本期损益(收入)",  "0.00", "#3d6fdb")
        self.card_expense = self._carry_card("结转本期损益(成本费用)","0.00", "#e05252")
        cards_row.addWidget(self.card_income); cards_row.addWidget(self.card_expense)
        cards_row.addStretch(); L.addLayout(cards_row)

        # ── New: period activity summary table ──
        L.addWidget(lbl("本期收入/费用科目发生额（仅显示已审核凭证）", bold=True, size=13))
        hint = QLabel("  结转的前提：凭证中需要有 5001-5899 或 6001-6899 收入/费用科目，且凭证状态为【已审核】。")
        hint.setStyleSheet("color:#ad6800;background:#fffbe6;border-radius:5px;padding:6px 10px;font-size:12px;")
        L.addWidget(hint)
        f = card(); vl2 = QVBoxLayout(f); vl2.setContentsMargins(0,0,0,0)
        self.activity_tbl = QTableWidget()
        self.activity_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.activity_tbl.setShowGrid(False); self.activity_tbl.verticalHeader().setVisible(False)
        self.activity_tbl.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.activity_tbl.setColumnCount(5)
        self.activity_tbl.setHorizontalHeaderLabels(["科目编号","科目名称","类型","本期借方","本期贷方"])
        self.activity_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.activity_tbl.setColumnWidth(0,90); self.activity_tbl.setColumnWidth(2,70)
        self.activity_tbl.setColumnWidth(3,110); self.activity_tbl.setColumnWidth(4,110)
        self.activity_tbl.setMaximumHeight(200)
        vl2.addWidget(self.activity_tbl); L.addWidget(f)
        L.addWidget(sep())

        # Check list
        L.addWidget(lbl("结账检测", bold=True, size=14))
        self.check_list = QTableWidget(); self.check_list.setColumnCount(3)
        self.check_list.setHorizontalHeaderLabels(["序号","检测项目","状态"])
        self.check_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.check_list.verticalHeader().setVisible(False); self.check_list.setShowGrid(False)
        self.check_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        L.addWidget(self.check_list); L.addStretch()

    def _step_box(self, num, text, state):
        w = QFrame()
        color = "#3d6fdb" if state=="active" else "#ddd"
        w.setStyleSheet(f"background:{color};border-radius:8px;padding:4px;")
        vl = QHBoxLayout(w); vl.setContentsMargins(16,10,16,10)
        n = lbl(num, bold=True, color="#fff" if state=="active" else "#999", size=18)
        t = lbl(text, color="#fff" if state=="active" else "#999")
        vl.addWidget(n); vl.addWidget(t); return w

    def _carry_card(self, title, amount, color):
        f = QFrame()
        f.setStyleSheet("background:#fff;border-radius:10px;border:1px solid #e4e8f0;")
        f.setFixedWidth(260)
        vl = QVBoxLayout(f); vl.setContentsMargins(20,16,20,16); vl.setSpacing(8)
        # Checkbox + smart label
        row = QHBoxLayout()
        cb = QCheckBox(); cb.setChecked(True)
        smart = lbl("智能生成", color="#52c41a", size=11)
        row.addWidget(cb); row.addStretch(); row.addWidget(smart)
        vl.addLayout(row)
        icon = lbl("⟳", color=color, size=28); icon.setAlignment(Qt.AlignCenter)
        vl.addWidget(icon)
        t = lbl(title, color="#555"); t.setAlignment(Qt.AlignCenter); vl.addWidget(t)
        a = lbl(f"金额：{amount}", bold=True, color=color, size=14); a.setAlignment(Qt.AlignCenter)
        vl.addWidget(a)
        f._amount_lbl = a; f._cb = cb
        return f

    def set_client(self, client_id, client_name, period):
        self.client_id = client_id; self.client_name = client_name; self.period = period
        self.client_lbl.setText(f"【{client_name}】")
        self._refresh_period_options(period)
        self._refresh_period_view()

    def _refresh_period_options(self, selected_period=None):
        self.period_combo.blockSignals(True)
        self.period_combo.clear()
        if not self.client_id:
            self.period_combo.addItem("请先从客户管理进入账簿", "")
            self.period_combo.setEnabled(False)
            self.period_combo.blockSignals(False)
            return

        periods = set()
        now = datetime.now()
        year, month = now.year, now.month
        for _ in range(36):
            periods.add(f"{year}-{month:02d}")
            month -= 1
            if month == 0:
                year -= 1
                month = 12

        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT period FROM vouchers WHERE client_id=?
                     UNION
                     SELECT period FROM periods WHERE client_id=?""",
                  (self.client_id, self.client_id))
        periods.update(row["period"] for row in c.fetchall() if row["period"])
        conn.close()

        target = selected_period or self.period
        if target:
            periods.add(target)

        for p in sorted(periods, reverse=True):
            self.period_combo.addItem(f"{p[:4]}年{p[5:]}期", p)

        idx = self.period_combo.findData(target)
        if idx < 0 and self.period_combo.count():
            idx = 0
        if idx >= 0:
            self.period_combo.setCurrentIndex(idx)
            self.period = self.period_combo.itemData(idx)
        self.period_combo.setEnabled(self.period_combo.count() > 0)
        self.period_combo.blockSignals(False)

    def _on_period_change(self):
        period = self.period_combo.currentData()
        if not period:
            return
        self.period = period
        self._refresh_period_view()

    def _refresh_period_view(self):
        if not self.client_id or not self.period:
            return
        self._refresh_lock_state()
        self._refresh_carry_amounts()
        self._load_activity()
        self._run_checks()

    def _fit_table_height(self, table, extra=8):
        height = table.horizontalHeader().height()
        for row in range(table.rowCount()):
            height += table.rowHeight(row)
        if table.rowCount() == 0:
            height += table.verticalHeader().defaultSectionSize()
        table.setMinimumHeight(height + extra)
        table.setMaximumHeight(height + extra)

    def _is_period_closed(self):
        """Check if current period is closed."""
        if not self.client_id or not self.period: return False
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT is_closed FROM periods WHERE client_id=? AND period=?",
                  (self.client_id, self.period))
        row = c.fetchone(); conn.close()
        return bool(row and row["is_closed"])

    def _refresh_lock_state(self):
        """Update UI buttons and status label based on period close state."""
        closed = self._is_period_closed()
        if closed:
            self.status_lbl.setText("🔒 已结账封账")
            self.status_lbl.setStyleSheet("color:#ff4d4f;font-weight:bold;")
            self.do_btn.setEnabled(False)
            self.close_btn.setVisible(False)
            self.reopen_btn.setVisible(True)
        else:
            self.status_lbl.setText("○ 未结账")
            self.status_lbl.setStyleSheet("color:#888;")
            self.do_btn.setEnabled(True)
            self.close_btn.setVisible(True)
            self.reopen_btn.setVisible(False)

    def _close_period(self):
        """结账封账：检查无待审核凭证后锁定期间。"""
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        # 检查待审核凭证
        c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=? AND status='待审核'",
                  (self.client_id, self.period))
        pending = c.fetchone()[0]
        if pending > 0:
            conn.close()
            QMessageBox.warning(self, "无法结账",
                f"本期还有 {pending} 张【待审核】凭证，请先全部审核后再结账封账。")
            return
        # 检查借贷不平凭证（额外保障）
        c.execute("""SELECT v.voucher_no,
            ABS(SUM(e.debit)-SUM(e.credit)) AS diff
            FROM vouchers v JOIN voucher_entries e ON e.voucher_id=v.id
            WHERE v.client_id=? AND v.period=?
            GROUP BY v.id HAVING diff > 0.005""", (self.client_id, self.period))
        unbal = c.fetchall()
        if unbal:
            conn.close()
            nos = "、".join(r["voucher_no"] for r in unbal[:5])
            QMessageBox.warning(self, "无法结账",
                f"以下凭证借贷不平衡，请先修正：{nos}")
            return
        if QMessageBox.question(self, "确认结账封账",
                f"确认对期间【{self.period}】进行结账封账？\n\n封账后该期间凭证将无法新增或修改，如需修改请先反结账。",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            conn.close(); return
        # 写入或更新 periods 表
        c.execute("INSERT OR REPLACE INTO periods(client_id,period,is_closed,closed_at) VALUES(?,?,1,datetime('now'))",
                  (self.client_id, self.period))
        log_action(conn, self.client_id, "期间结账封账", "period", self.period,
                   f"期间 {self.period} 封账")
        conn.commit(); conn.close()
        self._refresh_lock_state()
        self._run_checks()
        QMessageBox.information(self, "结账成功", f"期间【{self.period}】已结账封账。")

    def _reopen_period(self):
        """反结账：解除期间锁定。"""
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT period FROM periods
                     WHERE client_id=? AND is_closed=1 AND period>=?
                     ORDER BY period""",
                  (self.client_id, self.period))
        closed_periods = [row["period"] for row in c.fetchall()]
        if not closed_periods:
            conn.close()
            QMessageBox.information(self, "提示", f"期间【{self.period}】当前未封账，无需反结账。")
            self._refresh_lock_state()
            self._run_checks()
            return

        msg = QMessageBox(self)
        msg.setWindowTitle("确认反结账")
        msg.setIcon(QMessageBox.Warning)
        msg.setText(f"请选择期间【{self.period}】的反结账方式。")
        if len(closed_periods) > 1:
            msg.setInformativeText(
                f"当前期间之后还有 {len(closed_periods) - 1} 个已结账期间："
                f"{'、'.join(closed_periods[1:4])}"
                f"{' 等' if len(closed_periods) > 4 else ''}\n\n"
                "你可以只反当前期间，也可以连同后续已结账期间一起反结账。")
        else:
            msg.setInformativeText("当前期间之后没有其他已结账期间。")
        current_btn = msg.addButton("只反当前期间", QMessageBox.AcceptRole)
        cascade_btn = None
        if len(closed_periods) > 1:
            cascade_btn = msg.addButton("反当前及后续期间", QMessageBox.DestructiveRole)
        cancel_btn = msg.addButton("取消", QMessageBox.RejectRole)
        msg.setDefaultButton(current_btn)
        msg.exec()

        clicked = msg.clickedButton()
        if clicked == cancel_btn or clicked is None:
            conn.close()
            return

        if clicked == cascade_btn:
            target_periods = closed_periods
            c.execute("""UPDATE periods
                         SET is_closed=0, closed_at=NULL
                         WHERE client_id=? AND is_closed=1 AND period>=?""",
                      (self.client_id, self.period))
            detail = f"期间 {self.period} 反结账，并级联解除后续 {len(target_periods) - 1} 个期间封账"
            success_msg = (
                f"期间【{self.period}】及后续共 {len(target_periods)} 个已结账期间"
                "已解除封账，可重新修改凭证。")
        else:
            target_periods = [self.period]
            c.execute("""UPDATE periods
                         SET is_closed=0, closed_at=NULL
                         WHERE client_id=? AND period=?""",
                      (self.client_id, self.period))
            detail = f"期间 {self.period} 反结账"
            success_msg = f"期间【{self.period}】已解除封账，可重新修改凭证。"

        log_action(conn, self.client_id, "期间反结账", "period", self.period, detail)
        conn.commit(); conn.close()
        self._refresh_lock_state()
        self._run_checks()
        QMessageBox.information(self, "反结账成功", success_msg)

    def _refresh_carry_amounts(self):
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        # Only count APPROVED vouchers
        c.execute("""SELECT SUM(e.credit)-SUM(e.debit) FROM voucher_entries e
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period=? AND v.status='已审核'
            AND ((e.account_code >= '5001' AND e.account_code < '5400')
              OR (e.account_code >= '6001' AND e.account_code < '6400'))""",
                  (self.client_id, self.period))
        income = c.fetchone()[0] or 0
        c.execute("""SELECT SUM(e.debit)-SUM(e.credit) FROM voucher_entries e
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period=? AND v.status='已审核'
            AND (e.account_code >= '5401' OR (e.account_code >= '6400' AND e.account_code < '7000'))""",
                  (self.client_id, self.period))
        expense = c.fetchone()[0] or 0
        conn.close()
        self.card_income._amount_lbl.setText(f"金额：{income:,.2f}")
        self.card_expense._amount_lbl.setText(f"金额：{expense:,.2f}")
        self._income_amt = income; self._expense_amt = expense

    def _load_activity(self):
        """Show all 5xxx account activity this period so user can verify before carryforward."""
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        # All income+expense accounts with any activity, approved vouchers only
        c.execute("""SELECT e.account_code, e.account_name,
            CASE WHEN (e.account_code >= '5001' AND e.account_code < '5400')
                   OR (e.account_code >= '6001' AND e.account_code < '6400')
                 THEN '收入' ELSE '费用' END as cat,
            SUM(e.debit) td, SUM(e.credit) tc
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period=? AND v.status='已审核'
            AND (e.account_code >= '5001' OR (e.account_code >= '6001' AND e.account_code < '7000'))
            GROUP BY e.account_code ORDER BY e.account_code""",
                  (self.client_id, self.period))
        rows = c.fetchall()
        # Also check unapproved counts
        c.execute("""SELECT COUNT(*) FROM vouchers
            WHERE client_id=? AND period=? AND status='待审核'""",
                  (self.client_id, self.period))
        pending = c.fetchone()[0]
        conn.close()

        self.activity_tbl.setRowCount(len(rows))
        if not rows:
            self.activity_tbl.setRowCount(1)
            msg = f"本期已审核凭证中无收入/费用科目（5001-5899 或 6001-6899）发生额。"
            if pending:
                msg += f"  ⚠ 有 {pending} 张凭证【待审核】，请先审核后再结转。"
            it = QTableWidgetItem(msg)
            it.setForeground(QColor("#ad6800"))
            self.activity_tbl.setItem(0, 0, it)
            self.activity_tbl.setSpan(0, 0, 1, 5)
            self.activity_tbl.setRowHeight(0, 42)
            self._fit_table_height(self.activity_tbl)
            return

        for i, r in enumerate(rows):
            self.activity_tbl.setRowHeight(i, 34)
            cat_color = "#3d6fdb" if r[2] == "收入" else "#e05252"
            for j, val in enumerate([r[0], r[1], r[2], fmt_amt(r[3]), fmt_amt(r[4])]):
                it = QTableWidgetItem(val)
                it.setTextAlignment(Qt.AlignCenter if j != 1 else Qt.AlignLeft | Qt.AlignVCenter)
                if j == 2: it.setForeground(QColor(cat_color))
                if j == 3 and r[3]: it.setForeground(QColor("#3d6fdb"))
                if j == 4 and r[4]: it.setForeground(QColor("#e05252"))
                self.activity_tbl.setItem(i, j, it)

        if pending:
            warn = QTableWidgetItem(f"  ⚠ 另有 {pending} 张凭证【待审核】未计入，请先审核。")
            warn.setForeground(QColor("#ff4d4f"))
            row = len(rows)
            self.activity_tbl.setRowCount(row + 1)
            self.activity_tbl.setRowHeight(row, 38)
            self.activity_tbl.setItem(row, 0, warn)
            self.activity_tbl.setSpan(row, 0, 1, 5)
        self._fit_table_height(self.activity_tbl)

    def _do_carryforward(self):
        if not self.client_id: return
        # 已结账期间不能结转
        if self._is_period_closed():
            QMessageBox.warning(self, "期间已封账", "该期间已结账封账，请先反结账后再操作。"); return
        # 有待审核凭证不能结转
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=? AND status='待审核'",
                  (self.client_id, self.period))
        pending = c.fetchone()[0]
        if pending > 0:
            conn.close()
            QMessageBox.warning(self, "存在待审核凭证",
                f"本期有 {pending} 张凭证尚未审核，请先全部审核通过后再执行结转。"); return

        def next_vno():
            c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=?",
                      (self.client_id, self.period))
            return f"记-{c.fetchone()[0]+1:03d}"

        # Determine profit account once: use 4103 if exists (6xxx体系), else 3103
        c.execute("SELECT code FROM accounts WHERE client_id=? AND code='4103'", (self.client_id,))
        profit_code = "4103" if c.fetchone() else "3103"
        profit_name = "本年利润"

        generated = []
        # Income carry: debit income accounts, credit 本年利润
        if self.card_income._cb.isChecked() and abs(self._income_amt) > 0.005:
            # Collect all income account balances for this period (5xxx and 6xxx)
            c.execute("""SELECT e.account_code,e.account_name,
                SUM(e.credit)-SUM(e.debit) AS net
                FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
                WHERE v.client_id=? AND v.period=?
                AND ((e.account_code >= '5001' AND e.account_code < '5400')
                  OR (e.account_code >= '6001' AND e.account_code < '6400'))
                GROUP BY e.account_code,e.account_name HAVING net>0.005""",
                (self.client_id, self.period))
            income_rows = c.fetchall()
            if income_rows:
                vno = next_vno()
                date = self.period + "-28"
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status,note) VALUES(?,?,?,?,?,?)",
                          (self.client_id,self.period,vno,date,"已审核","结转收入"))
                vid = c.lastrowid
                for ln,(code,name,net) in enumerate(income_rows,1):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln,"结转本期损益",code,name,net,0))
                total_income = sum(r[2] for r in income_rows)
                c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                          (vid,len(income_rows)+1,"结转本期损益",profit_code,profit_name,0,total_income))
                generated.append(f"{vno}（结转收入 {total_income:,.2f}）")

        # Expense carry: credit expense accounts, debit 本年利润
        if self.card_expense._cb.isChecked() and abs(self._expense_amt) > 0.005:
            c.execute("""SELECT e.account_code,e.account_name,
                SUM(e.debit)-SUM(e.credit) AS net
                FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
                WHERE v.client_id=? AND v.period=?
                AND (e.account_code >= '5401'
                  OR (e.account_code >= '6400' AND e.account_code < '7000'))
                GROUP BY e.account_code,e.account_name HAVING net>0.005""",
                (self.client_id, self.period))
            expense_rows = c.fetchall()
            if expense_rows:
                vno = next_vno()
                date = self.period + "-28"
                c.execute("INSERT INTO vouchers(client_id,period,voucher_no,date,status,note) VALUES(?,?,?,?,?,?)",
                          (self.client_id,self.period,vno,date,"已审核","结转费用"))
                vid = c.lastrowid
                total_expense = sum(r[2] for r in expense_rows)
                c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                          (vid,1,"结转本期损益",profit_code,profit_name,total_expense,0))
                for ln,(code,name,net) in enumerate(expense_rows,2):
                    c.execute("INSERT INTO voucher_entries(voucher_id,line_no,summary,account_code,account_name,debit,credit) VALUES(?,?,?,?,?,?,?)",
                              (vid,ln,"结转本期损益",code,name,0,net))
                generated.append(f"{vno}（结转费用 {total_expense:,.2f}）")

        if generated:
            log_action(conn, self.client_id, "期末结转", "settle", self.period,
                       f"生成{len(generated)}张结转凭证: {'; '.join(generated)}")
        conn.commit(); conn.close()

        if generated:
            detail = "\n".join(generated)
            QMessageBox.information(self,"✓ 结转凭证已生成",
                f"已生成 {len(generated)} 张结转凭证（状态：已审核），保存在本期凭证列表中：\n\n{detail}\n\n请到【记账（凭证）→ 查凭证】查看。")
            self.carryforward_done.emit()   # notify VoucherPage to refresh
        else:
            QMessageBox.information(self,"提示","本期无需结转（收入和费用均为零）。")
        self._refresh_carry_amounts(); self._load_activity(); self._run_checks()

    def _run_checks(self):
        if not self.client_id:
            return
        conn = get_db(); c = conn.cursor()

        # 1. 待审核凭证数
        c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=? AND status='待审核'",
                  (self.client_id, self.period))
        pending = c.fetchone()[0]

        # 2. 借贷不平凭证数
        c.execute("""SELECT COUNT(*) FROM (
            SELECT v.id FROM vouchers v JOIN voucher_entries e ON e.voucher_id=v.id
            WHERE v.client_id=? AND v.period=?
            GROUP BY v.id HAVING ABS(SUM(e.debit)-SUM(e.credit)) > 0.005
        )""", (self.client_id, self.period))
        unbalanced = c.fetchone()[0]

        # 3. 是否已结账封账
        c.execute("SELECT is_closed FROM periods WHERE client_id=? AND period=?",
                  (self.client_id, self.period))
        row = c.fetchone()
        is_closed = bool(row and row["is_closed"])

        # 4. 结转凭证是否存在
        c.execute("SELECT COUNT(*) FROM vouchers WHERE client_id=? AND period=? AND note IN ('结转收入','结转费用')",
                  (self.client_id, self.period))
        carried = c.fetchone()[0]
        conn.close()

        if carried > 0:
            carry_status = "已完成"
        elif abs(getattr(self, "_income_amt", 0)) <= 0.005 and abs(getattr(self, "_expense_amt", 0)) <= 0.005:
            carry_status = "无需结转"
        else:
            carry_status = "未结转"

        checks = [
            ("01", "待审核凭证",   "通过" if pending == 0    else f"风险：{pending}张待审核"),
            ("02", "借贷平衡",     "通过" if unbalanced == 0  else f"风险：{unbalanced}张不平"),
            ("03", "期末结转",     carry_status),
            ("04", "期间封账",     "已封账" if is_closed      else "未封账"),
        ]

        self.check_list.setRowCount(len(checks))
        for i, (no, name, status) in enumerate(checks):
            self.check_list.setRowHeight(i, 40)
            for j, v in enumerate([no, name]):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter)
                self.check_list.setItem(i, j, it)
            is_ok = "风险" not in status
            icon = "✓" if is_ok else "✗"
            color = "#52c41a" if is_ok else "#ff4d4f"
            if status in ("未结转", "未封账", "无需结转"):
                color = "#fa8c16"; icon = "○"
            s_w = QLabel(f"  {icon}  {status}  ")
            s_w.setStyleSheet(f"color:{color};font-weight:bold;")
            self.check_list.setCellWidget(i, 2, s_w)
        self._fit_table_height(self.check_list)


class ReportPage(QWidget):
    """财务报表 — 资产负债表 + 利润表"""

    def __init__(self):
        super().__init__()
        self.client_id = None; self.period = ""
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)
        # Top tabs
        tb = QWidget(); tb.setStyleSheet("background:#fff;border-bottom:1px solid #e8ecf2;")
        tl = QHBoxLayout(tb); tl.setContentsMargins(16,0,16,0); tl.setSpacing(0)
        self._rtabs = []
        for n in ["资产负债表","利润表","所有者权益变动表","现金流量表","收支统计表"]:
            b = QPushButton(n); b.setStyleSheet("""QPushButton{background:transparent;color:#888;
                border:none;padding:12px 16px;border-bottom:2px solid transparent;}
                QPushButton:hover{color:#3d6fdb;}
                QPushButton[active=true]{color:#3d6fdb;border-bottom:2px solid #3d6fdb;}""")
            b.clicked.connect(lambda _,nn=n:self._switch(nn)); tl.addWidget(b); self._rtabs.append(b)
        tl.addStretch()
        # Period selector
        pr = QHBoxLayout(); pr.setSpacing(8)
        pr.addWidget(lbl("报告期间:", color="#666"))
        self.rep_start_period = QComboBox(); self.rep_start_period.setMinimumWidth(100)
        self.rep_end_period = QComboBox(); self.rep_end_period.setMinimumWidth(100)
        pr.addWidget(self.rep_start_period); pr.addWidget(lbl("至", color="#666")); pr.addWidget(self.rep_end_period)
        b_refresh = QPushButton("刷新"); b_refresh.setObjectName("btn_primary"); b_refresh.clicked.connect(self._refresh_reports)
        pr.addWidget(b_refresh); pr.addStretch()
        tl.addLayout(pr)
        self.period_lbl = lbl("", color="#888"); tl.addWidget(self.period_lbl)
        b_dl = QPushButton(" ↓ 下载"); b_dl.setObjectName("btn_outline"); b_dl.clicked.connect(self._export)
        tl.addSpacing(12); tl.addWidget(b_dl)
        L.addWidget(tb)
        self.stack = QStackedWidget(); L.addWidget(self.stack)
        self._build_balance(); self._build_income(); self._build_equity(); self._build_cf_stmt(); self._build_cashflow()
        self._switch("资产负债表")

    def _refresh_reports(self):
        """Refresh current report with selected period range"""
        current_tab = None
        for b in self._rtabs:
            if b.property("active") == "true":
                current_tab = b.text()
                break
        if current_tab:
            self._switch(current_tab)

    def _build_placeholder(self, name):
        w = QWidget(); vl = QVBoxLayout(w)
        vl.addStretch(); vl.addWidget(lbl(f"{name}（生成后显示）", color="#bbb", size=16))
        vl.addStretch(); self.stack.addWidget(w)

    def _switch(self, name):
        mapping = {"资产负债表":0,"利润表":1,"所有者权益变动表":2,"现金流量表":3,"收支统计表":4}
        for b in self._rtabs:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        if name in mapping:
            self.stack.setCurrentIndex(mapping[name])
            if name=="资产负债表": self._load_balance()
            elif name=="利润表": self._load_income()
            elif name=="所有者权益变动表": self._load_equity()
            elif name=="现金流量表": self._load_cf_stmt()
            elif name=="收支统计表": self._load_cashflow()

    def _make_report_table(self, cols, col_widths=None):
        t = QTableWidget(); t.setColumnCount(len(cols))
        t.setHorizontalHeaderLabels(cols)
        t.verticalHeader().setVisible(False); t.setShowGrid(True)
        t.setEditTriggers(QTableWidget.NoEditTriggers)
        if col_widths:
            for i,w in enumerate(col_widths):
                if w == -1: t.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
                else: t.setColumnWidth(i,w)
        return t

    def _build_balance(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14)
        self.bs_tbl = self._make_report_table(
            ["资产项目","行次","期末金额","年初金额","负债和所有者权益","行次","期末金额","年初金额"],
            [-1,40,110,110,-1,40,110,110])
        L.addWidget(self.bs_tbl); self.stack.addWidget(w)

    def _load_balance(self):
        if not self.client_id: return
        end_period = self.rep_end_period.currentData()
        if not end_period: return
        year = end_period[:4]
        year_start = f"{year}-01"   # 本年第一期
        conn = get_db(); c = conn.cursor()

        # 期末：截止所选期间的累计发生额
        c.execute("""SELECT e.account_code, SUM(e.debit)-SUM(e.credit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period<=? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, end_period))
        mv = {r[0]: r[1] or 0 for r in c.fetchall()}

        # 年初：上年年末 = 期初余额 + 本年第一期之前的凭证发生额
        # 若所选期间就是01期，则年初 = 纯期初余额（凭证发生额=0）
        c.execute("""SELECT e.account_code, SUM(e.debit)-SUM(e.credit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period<? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, year_start))
        mv_ys = {r[0]: r[1] or 0 for r in c.fetchall()}

        c.execute("SELECT code,opening_debit,opening_credit,direction FROM accounts WHERE client_id=?",
                  (self.client_id,))
        accts = {r['code']: r for r in c.fetchall()}
        conn.close()

        # 预计算末级科目：没有任何子科目的科目
        all_codes = set(accts.keys())
        leaf_codes = {
            code for code in all_codes
            if not any(
                other != code and (other.startswith(code+".") or other.startswith(code+"_"))
                for other in all_codes
            )
        }

        def _bal_with_mv(code_prefix_list, movements):
            """通用余额计算：末级科目取期初+发生额，父科目只取发生额"""
            total = 0
            for code, a in accts.items():
                if not any(code == p or code.startswith(p+".") or code.startswith(p+"_")
                           for p in code_prefix_list):
                    continue
                net_mv = movements.get(code, 0)
                if code in leaf_codes:
                    od = a['opening_debit'] or 0; oc = a['opening_credit'] or 0
                    if a['direction'] == '借':
                        total += (od - oc) + net_mv
                    else:
                        total += (oc - od) - net_mv
                else:
                    if a['direction'] == '借':
                        total += net_mv
                    else:
                        total -= net_mv
            return total

        def bal(code_prefix_list):
            """期末余额"""
            return _bal_with_mv(code_prefix_list, mv)

        def bal_ys(code_prefix_list):
            """年初余额（上年年末 = 期初 + 本年首期前发生额）"""
            return _bal_with_mv(code_prefix_list, mv_ys)

        # ── 资产方 ──
        cash      = bal(["1001","1002","1012"])
        notes_rec = bal(["1121"])
        acct_rec  = bal(["1122"])
        _prepay_raw = bal(["1123"])
        _advrec_raw = bal(["2203"])
        # 预付账款贷方余额 → 重分类为预收账款（确保资产负债表两边同步）
        prepay  = max(0.0, _prepay_raw)  + max(0.0, -_advrec_raw)
        int_rec   = bal(["1132"])
        div_rec   = bal(["1131"])
        oth_rec   = bal(["1221"])

        # ── 年初余额（同结构，使用 bal_ys） ──
        notes_rec_y = bal_ys(["1121"])
        acct_rec_y  = bal_ys(["1122"])
        _prepay_y   = bal_ys(["1123"])
        _advrec_y   = bal_ys(["2203"])
        prepay_y    = max(0.0, _prepay_y) + max(0.0, -_advrec_y)
        int_rec_y   = bal_ys(["1132"])
        div_rec_y   = bal_ys(["1131"])
        oth_rec_y   = bal_ys(["1221"])
        cash_y      = bal_ys(["1001","1002","1012"])
        inventory_y = bal_ys(["1401","1402","1403","1404","1405","1406","1408","1411"])
        prepd_exp_y = bal_ys(["1461"])
        fa_y        = bal_ys(["1601"]) - abs(bal_ys(["1602"])) - abs(bal_ys(["1603"]))
        wip_y       = bal_ys(["1604"])
        intangible_y= bal_ys(["1701"]) - abs(bal_ys(["1702"]))
        lt_prepaid_y= bal_ys(["1901"])
        deferred_a_y= bal_ys(["1911"])
        lt_equity_y = bal_ys(["1801","1811","1521","1511"])
        cur_asset_y  = cash_y+notes_rec_y+acct_rec_y+prepay_y+int_rec_y+div_rec_y+oth_rec_y+inventory_y+prepd_exp_y
        noncur_asset_y = fa_y+wip_y+intangible_y+lt_prepaid_y+lt_equity_y+deferred_a_y
        total_asset_y  = cur_asset_y + noncur_asset_y

        st_loan_y   = bal_ys(["2001"]); notes_pay_y = bal_ys(["2201"])
        acct_pay_y  = bal_ys(["2202"]); adv_rec_y   = max(0.0, _advrec_y) + max(0.0, -_prepay_y)
        emp_pay_y   = bal_ys(["2211"]); tax_pay_y   = bal_ys(["2221"])
        int_pay_y   = bal_ys(["2231"]); div_pay_y   = bal_ys(["2232"])
        oth_pay_y   = bal_ys(["2241"])
        cur_liab_y  = st_loan_y+notes_pay_y+acct_pay_y+adv_rec_y+emp_pay_y+tax_pay_y+int_pay_y+div_pay_y+oth_pay_y
        lt_loan_y   = bal_ys(["2501"]); bonds_pay_y = bal_ys(["2502"])
        lt_payable_y= bal_ys(["2701"]); est_liab_y  = bal_ys(["2801"])
        deferred_l_y= bal_ys(["2901"])
        noncur_liab_y = lt_loan_y+bonds_pay_y+lt_payable_y+est_liab_y+deferred_l_y
        total_liab_y  = cur_liab_y + noncur_liab_y
        cap_y     = bal_ys(["3001","4001"]); cap_res_y = bal_ys(["3002","4002"])
        surp_res_y= bal_ys(["3101","4101"])
        profit_y  = bal_ys(["3103","4103"]) + bal_ys(["3104","4104"])
        tsy_y     = bal_ys(["3201","4201"])
        total_equity_y = cap_y + cap_res_y + surp_res_y + profit_y - tsy_y
        total_le_y     = total_liab_y + total_equity_y
        inventory = bal(["1401","1402","1403","1404","1405","1406","1408","1411"])
        prepd_exp = bal(["1461"])
        fa        = bal(["1601"]) - abs(bal(["1602"])) - abs(bal(["1603"]))
        wip       = bal(["1604"])
        intangible= bal(["1701"]) - abs(bal(["1702"]))
        lt_prepaid= bal(["1901"])
        deferred_a= bal(["1911"])
        lt_equity = bal(["1801","1811","1521","1511"])
        cur_asset = cash+notes_rec+acct_rec+prepay+int_rec+div_rec+oth_rec+inventory+prepd_exp
        noncur_asset = fa+wip+intangible+lt_prepaid+lt_equity+deferred_a
        total_asset = cur_asset + noncur_asset

        # ── 负债方 ──
        st_loan   = bal(["2001"])
        notes_pay = bal(["2201"])
        acct_pay  = bal(["2202"])
        adv_rec   = max(0.0, _advrec_raw) + max(0.0, -_prepay_raw)
        emp_pay   = bal(["2211"])
        tax_pay   = bal(["2221"])
        int_pay   = bal(["2231"])
        div_pay   = bal(["2232"])
        oth_pay   = bal(["2241"])
        cur_liab  = st_loan+notes_pay+acct_pay+adv_rec+emp_pay+tax_pay+int_pay+div_pay+oth_pay
        lt_loan   = bal(["2501"])
        bonds_pay = bal(["2502"])
        lt_payable= bal(["2601"])
        est_liab  = bal(["2701"])
        deferred_l= bal(["2901"])
        noncur_liab = lt_loan+bonds_pay+lt_payable+est_liab+deferred_l
        total_liab = cur_liab + noncur_liab

        # ── 所有者权益 ──
        # 兼容3xxx（标准）和4xxx（部分软件）两套所有者权益科目体系
        cap       = bal(["3001","4001"])
        cap_res   = bal(["3002","4002"])
        surp_res  = bal(["3101","4101"])
        profit    = bal(["3103","4103"]) + bal(["3104","4104"])
        tsy_stock = bal(["3201","4201"])
        total_equity = cap + cap_res + surp_res + profit - tsy_stock
        total_le     = total_liab + total_equity

        def R(label, rowno, left_val, right_label="", right_rowno="", right_val=None,
              is_header=False, is_total=False,
              left_ys=None, right_ys=None):
            return (label, rowno, left_val, right_label, right_rowno, right_val,
                    is_header, is_total, left_ys, right_ys)

        rows = [
            R("流动资产：","","",  "流动负债：","","",          True),
            R("货币资金","1",cash,            "短期借款","34",st_loan,         left_ys=cash_y,        right_ys=st_loan_y),
            R("以公允价值计量且其变动\n计入当期损益的金融资产","2",0, "以公允价值计量且其变动\n计入当期损益的金融负债","35",0),
            R("衍生金融资产","3",0,            "衍生金融负债","36",0),
            R("应收票据","4",notes_rec,        "应付票据","37",notes_pay,       left_ys=notes_rec_y,   right_ys=notes_pay_y),
            R("应收账款","5",acct_rec,         "应付账款","38",acct_pay,        left_ys=acct_rec_y,    right_ys=acct_pay_y),
            R("预付款项","6",prepay,           "预收款项","39",adv_rec,         left_ys=prepay_y,      right_ys=adv_rec_y),
            R("应收利息","7",int_rec,          "应付职工薪酬","40",emp_pay,     left_ys=int_rec_y,     right_ys=emp_pay_y),
            R("应收股利","8",div_rec,          "应交税费","41",tax_pay,         left_ys=div_rec_y,     right_ys=tax_pay_y),
            R("其他应收款","9",oth_rec,         "应付利息","42",int_pay,        left_ys=oth_rec_y,     right_ys=int_pay_y),
            R("存货","10",inventory,           "应付股利","43",div_pay,         left_ys=inventory_y,   right_ys=div_pay_y),
            R("持有待售资产","11",0,            "其他应付款","44",oth_pay,                              right_ys=oth_pay_y),
            R("一年内到期的非流动资产","12",0,  "持有待售负债","45",0),
            R("其他流动资产","13",prepd_exp,    "一年内到期的非流动负债","46",0, left_ys=prepd_exp_y),
            R("流动资产合计","14",cur_asset,   "其他流动负债","47",0,   False,True, left_ys=cur_asset_y, right_ys=cur_liab_y),
            R("非流动资产：","","",            "流动负债合计","48",cur_liab,True,True,                  right_ys=cur_liab_y),
            R("可供出售金融资产","15",lt_equity,"非流动负债：","","",   False,False, left_ys=lt_equity_y),
            R("持有至到期投资","16",0,          "长期借款","49",lt_loan,                                right_ys=lt_loan_y),
            R("长期应收款","17",0,             "应付债券","50",bonds_pay),
            R("长期股权投资","18",0,            "其中：优先股","51",0),
            R("投资性房地产","19",0,            "永续债","52",0),
            R("固定资产","20",fa,              "长期应付款","53",lt_payable,    left_ys=fa_y,           right_ys=lt_payable_y),
            R("在建工程","21",wip,             "专项应付款","54",0,             left_ys=wip_y),
            R("工程物资","22",0,               "预计负债","55",est_liab,                               right_ys=est_liab_y),
            R("固定资产清理","23",0,            "递延收益","56",0),
            R("生产性生物资产","24",0,          "递延所得税负债","57",deferred_l,                       right_ys=deferred_l_y),
            R("油气资产","25",0,               "其他非流动负债","58",0),
            R("无形资产","26",intangible,       "非流动负债合计","59",noncur_liab, False,True, left_ys=intangible_y, right_ys=noncur_liab_y),
            R("开发支出","27",0,               "负债合计","60",total_liab,    False,True,               right_ys=total_liab_y),
            R("商誉","28",0,                   "所有者权益（或股东权益）：","","",True),
            R("长期待摊费用","29",lt_prepaid,   "实收资本（或股本）","61",cap,  left_ys=lt_prepaid_y,   right_ys=cap_y),
            R("递延所得税资产","30",deferred_a, "其他权益工具","62",0,          left_ys=deferred_a_y),
            R("其他非流动资产","31",0,          "其中：优先股","63",0),
            R("非流动资产合计","32",noncur_asset,"永续债","64",0,          False,True, left_ys=noncur_asset_y),
            R("","","",                        "资本公积","65",cap_res,                                 right_ys=cap_res_y),
            R("","","",                        "减：库存股","66",tsy_stock,                             right_ys=tsy_y),
            R("","","",                        "其他综合收益","67",0),
            R("","","",                        "盈余公积","68",surp_res,                                right_ys=surp_res_y),
            R("","","",                        "未分配利润","69",profit,                                right_ys=profit_y),
            R("","","",                        "所有者权益合计","70",total_equity, False,True,          right_ys=total_equity_y),
            R("资产总计","33",total_asset,     "负债和所有者权益总计","71",total_le,False,True, left_ys=total_asset_y, right_ys=total_le_y),
        ]

        self.bs_tbl.setRowCount(len(rows))
        for i,(l_name,l_row,l_val,r_name,r_row,r_val,is_hdr,is_tot,l_ys,r_ys) in enumerate(rows):
            self.bs_tbl.setRowHeight(i, 32)
            # Left
            for j,(text,align) in enumerate([
                (l_name, Qt.AlignLeft|Qt.AlignVCenter),
                (str(l_row), Qt.AlignCenter),
                (fmt_amt(l_val) if isinstance(l_val,(int,float)) else "", Qt.AlignRight|Qt.AlignVCenter),
                (fmt_amt(l_ys) if isinstance(l_ys,(int,float)) else "", Qt.AlignRight|Qt.AlignVCenter),
            ]):
                it = QTableWidgetItem(text); it.setTextAlignment(align)
                if is_hdr or is_tot:
                    it.setBackground(QColor("#f0f4ff" if is_hdr else "#f5f7fa"))
                    if is_tot: it.setFont(QFont("",weight=QFont.Bold))
                if j==0 and is_hdr: it.setForeground(QColor("#3d6fdb"))
                if j==2 and isinstance(l_val,(int,float)) and l_val<0:
                    it.setForeground(QColor("#e05252"))
                if j==3 and isinstance(l_ys,(int,float)) and l_ys<0:
                    it.setForeground(QColor("#e05252"))
                self.bs_tbl.setItem(i,j,it)
            # Right
            for j,(text,align) in enumerate([
                (r_name, Qt.AlignLeft|Qt.AlignVCenter),
                (str(r_row), Qt.AlignCenter),
                (fmt_amt(r_val) if isinstance(r_val,(int,float)) else "", Qt.AlignRight|Qt.AlignVCenter),
                (fmt_amt(r_ys) if isinstance(r_ys,(int,float)) else "", Qt.AlignRight|Qt.AlignVCenter),
            ],4):
                it = QTableWidgetItem(text); it.setTextAlignment(align)
                if is_hdr or is_tot:
                    it.setBackground(QColor("#f0f4ff" if is_hdr else "#f5f7fa"))
                    if is_tot: it.setFont(QFont("",weight=QFont.Bold))
                if j==4 and is_hdr: it.setForeground(QColor("#3d6fdb"))
                if j==6 and isinstance(r_val,(int,float)) and r_val<0:
                    it.setForeground(QColor("#e05252"))
                if j==7 and isinstance(r_ys,(int,float)) and r_ys<0:
                    it.setForeground(QColor("#e05252"))
                self.bs_tbl.setItem(i,j,it)

    def _build_income(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14)
        self.inc_tbl = self._make_report_table(
            ["项目","行次","本期金额","本年累计"],[-1,40,160,160])
        L.addWidget(self.inc_tbl); self.stack.addWidget(w)

    def _load_income(self):
        if not self.client_id: return
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        conn = get_db(); c = conn.cursor()

        def fetch_period(period_filter):
            c.execute("""SELECT e.account_code, SUM(e.credit)-SUM(e.debit) net
                FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
                WHERE v.client_id=? AND v.period"""+period_filter+""" AND v.status='已审核'
                GROUP BY e.account_code""", (self.client_id,))
            return {r[0]: r[1] or 0 for r in c.fetchall()}

        # Current period: from start to end
        cur = fetch_period(f">='{start_period}' AND v.period<='{end_period}'")
        year = end_period[:4]
        # Year-to-date: from year start to end
        ytd = fetch_period(f" LIKE '{year}%' AND v.period<='{end_period}'")
        conn.close()

        def g(codes, d=None):
            """Sum credit-minus-debit net for all accounts matching any prefix in codes list."""
            if d is None: d = cur
            if isinstance(codes, str): codes = [codes]
            total = 0
            for acct_code, val in d.items():
                for code in codes:
                    if acct_code == code or acct_code.startswith(code+".") or acct_code.startswith(code+"_"):
                        total += val
                        break
            return total
        def gy(codes): return g(codes, ytd)

        # Try 6xxx first, fall back to 5xxx
        use_6xxx = bool(g(["6001","6002","6401","6601","6602"]))

        if use_6xxx:
            # 6xxx科目体系（用友/金蝶新版）
            rev      = g(["6001","6002","6051"])         # 主营+其他业务收入（贷方余额为正）
            cost_n   = -g(["6401","6402"])               # 主营+其他业务成本（借方为正，取负得正数成本）
            tax      = -g(["6403"])                      # 税金及附加
            sell     = -g(["6601"])                      # 销售费用
            mgmt     = -g(["6602"])                      # 管理费用
            rnd      = -g(["6604"])                      # 研发费用
            fin_net  = g(["6603"])                       # 财务费用净额（正=净收益，负=净支出）
            inv_g    = g(["6111"])                       # 投资收益
            fv_g     = g(["6121"])                       # 公允价值变动
            asset_d  = g(["6301"])                       # 营业外收入（此处作资产处置收益）
            op_profit = rev - cost_n - tax - sell - mgmt - rnd + fin_net + inv_g + fv_g
            nop_inc   = g(["6301"])                      # 营业外收入
            nop_exp   = -g(["6711"])                     # 营业外支出
            tax_exp   = -g(["6801"])                     # 所得税费用
            # YTD
            rev_y    = gy(["6001","6002","6051"])
            cost_y   = -gy(["6401","6402"])
            sell_y   = -gy(["6601"]); mgmt_y = -gy(["6602"])
            fin_y    = gy(["6603"]); inv_y = gy(["6111"])
            nop_y    = gy(["6301"]); nopx_y = -gy(["6711"])
            tax_y    = -gy(["6801"])
            op_y     = rev_y - cost_y - gy(["6403"]) - sell_y - mgmt_y + fin_y + inv_y
            net_y    = op_y + nop_y + nopx_y + tax_y
        else:
            # 5xxx科目体系（旧版/标准）
            rev      = g(["5001","5051"])
            cost_n   = -g(["5401","5402"])
            tax      = -g(["5403"])
            sell     = -g(["5501"])
            mgmt     = -g(["5502"])
            rnd      = 0
            fin_net  = g(["5503"])
            inv_g    = g(["5111"])
            fv_g     = g(["5121"])
            asset_d  = 0
            op_profit = rev + cost_n + tax + sell + mgmt + fin_net + inv_g + fv_g
            nop_inc   = g(["5301"])
            nop_exp   = -g(["5601"])
            tax_exp   = -g(["5701"])
            rev_y    = gy(["5001","5051"])
            cost_y   = -gy(["5401"])
            op_y     = rev_y + cost_y - abs(gy(["5501"])) - abs(gy(["5502"])) + gy(["5111"])
            net_y    = op_y + gy(["5301"]) - abs(gy(["5601"])) - abs(gy(["5701"]))

        total_profit = op_profit + nop_inc + nop_exp
        net_profit   = total_profit + tax_exp

        rows_data = [
            ("一、营业收入",           "1",  rev,           rev_y,    True),
            ("  减：营业成本",          "2",  cost_n,        cost_y,   False),
            ("      税金及附加",        "3",  tax,           0,        False),
            ("      销售费用",          "4",  sell,          sell_y if use_6xxx else 0, False),
            ("      管理费用",          "5",  mgmt,          mgmt_y if use_6xxx else 0, False),
            ("      研发费用",          "6",  rnd,           0,        False),
            ("  加：财务费用（收益以-号填列）","7", fin_net, fin_y if use_6xxx else 0, False),
            ("      投资收益",          "8",  inv_g,         inv_y if use_6xxx else 0, False),
            ("      公允价值变动收益",   "9",  fv_g,          0,        False),
            ("      资产处置收益",       "9a", asset_d,       0,        False),
            ("二、营业利润（亏损）",     "10", op_profit,     op_y,     True),
            ("  加：营业外收入",         "11", nop_inc,       nop_y if use_6xxx else 0, False),
            ("  减：营业外支出",         "12", nop_exp,       nopx_y if use_6xxx else 0, False),
            ("三、利润总额（亏损总额）", "13", total_profit,  0,        True),
            ("  减：所得税费用",         "14", tax_exp,       tax_y if use_6xxx else 0, False),
            ("四、净利润（净亏损）",     "15", net_profit,    net_y,    True),
            ("  其中：归属于母公司股东的净利润","16", net_profit, 0, False),
            ("        少数股东损益",    "17", 0,             0,        False),
            ("五、其他综合收益的税后净额","18", 0,            0,        True),
            ("六、综合收益总额",         "19", net_profit,   0,        True),
            ("  其中：归属于母公司股东的综合收益","20", net_profit, 0, False),
            ("        归属于少数股东的综合收益","21", 0,      0,       False),
            ("七、每股收益","","","",True),
            ("  基本每股收益",           "22", 0,             0,        False),
            ("  稀释每股收益",           "23", 0,             0,        False),
        ]

        self.inc_tbl.setRowCount(len(rows_data))
        for i,row_item in enumerate(rows_data):
            name = row_item[0]; rowno = row_item[1]
            cur_v = row_item[2]; ytd_v = row_item[3]
            is_key = row_item[4] if len(row_item)>4 else False
            self.inc_tbl.setRowHeight(i, 34)
            bg = QColor("#f0f4ff") if is_key else None
            for j,v in enumerate([name, str(rowno) if rowno else "",
                                   fmt_amt(cur_v) if isinstance(cur_v,(int,float)) else "",
                                   fmt_amt(ytd_v) if isinstance(ytd_v,(int,float)) else ""]):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter if j==0 else Qt.AlignCenter if j==1 else Qt.AlignRight|Qt.AlignVCenter)
                if is_key:
                    it.setFont(QFont("",weight=QFont.Bold))
                    if bg: it.setBackground(bg)
                if j>=2 and isinstance(cur_v,(int,float)):
                    val = cur_v if j==2 else ytd_v
                    if val and val < 0: it.setForeground(QColor("#ff4d4f"))
                self.inc_tbl.setItem(i,j,it)


    def _build_equity(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(8)
        title_row = QHBoxLayout()
        title_row.addWidget(lbl("所有者权益变动表", bold=True, size=15))
        title_row.addStretch()
        title_row.addWidget(lbl("（企业会计准则格式）", color="#888", size=12))
        L.addLayout(title_row)
        L.addWidget(lbl("单位：元", color="#aaa", size=11))

        self.eq_tbl = self._make_report_table(
            ["项目",
             "实收资本(股本)",
             "资本公积",
             "其他综合收益",
             "盈余公积",
             "未分配利润",
             "合计"],
            [-1, 110, 110, 100, 100, 110, 110]
        )
        self.eq_tbl.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.eq_tbl.setWordWrap(True)
        self.eq_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.eq_tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        L.addWidget(self.eq_tbl)
        self.stack.addWidget(w)

    def _load_equity(self):
        if not self.client_id: return
        end_period = self.rep_end_period.currentData()
        if not end_period: return
        conn = get_db(); c = conn.cursor()
        year = end_period[:4]

        # Fetch year-to-date balances from voucher entries (approved only)
        c.execute("""SELECT e.account_code, SUM(e.debit)-SUM(e.credit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period LIKE ? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, f"{year}%"))
        ytd = {r[0]: -(r[1] or 0) for r in c.fetchall()}  # credit-normal for equity

        # Opening balances from accounts table
        c.execute("SELECT code, opening_credit FROM accounts WHERE client_id=?", (self.client_id,))
        opening = {r[0]: r[1] or 0 for r in c.fetchall()}
        conn.close()

        def op(code):
            # Support both 3xxx and 4xxx equity accounts
            alt = {"3001":"4001","3002":"4002","3101":"4101","3103":"4103","3104":"4104","3201":"4201"}
            return opening.get(code, 0) or opening.get(alt.get(code,""), 0)
        def mv(code):
            alt = {"3001":"4001","3002":"4002","3101":"4101","3103":"4103","3104":"4104","3201":"4201"}
            return ytd.get(code, 0) or ytd.get(alt.get(code,""), 0)

        cap_op  = op("3001"); cap_mv  = mv("3001")
        cprs_op = op("3002"); cprs_mv = mv("3002")
        oci_op  = 0;          oci_mv  = 0          # 其他综合收益（暂无专用科目）
        surp_op = op("3101"); surp_mv = mv("3101")
        re_op   = op("3103") + op("3104")
        re_mv   = mv("3103") + mv("3104")          # 本年利润 + 利润分配

        def row_data(label, c1, c2, c3, c4, c5, bold=False, bg=None):
            total = c1+c2+c3+c4+c5
            return (label, c1, c2, c3, c4, c5, total, bold, bg)

        rows = [
            row_data("一、上年年末余额",    cap_op,  cprs_op, oci_op,  surp_op, re_op,  bold=True,  bg="#f0f4ff"),
            row_data("  加：会计政策变更",  0, 0, 0, 0, 0),
            row_data("     前期差错更正",   0, 0, 0, 0, 0),
            row_data("二、本年年初余额",    cap_op,  cprs_op, oci_op,  surp_op, re_op,  bold=True,  bg="#f0f4ff"),
            row_data("三、本年增减变动",    cap_mv,  cprs_mv, oci_mv,  surp_mv, re_mv,  bold=True,  bg="#fafafa"),
            row_data("  (一)综合收益总额",  0,       0,       oci_mv,  0,       re_mv),
            row_data("  (二)所有者投入",    cap_mv,  cprs_mv, 0,       0,       0),
            row_data("  (三)利润分配",      0,       0,       0,       surp_mv, re_mv - re_mv),
            row_data("四、本年年末余额",
                     cap_op+cap_mv, cprs_op+cprs_mv, oci_op+oci_mv,
                     surp_op+surp_mv, re_op+re_mv,    bold=True, bg="#e6f0ff"),
        ]

        self.eq_tbl.setRowCount(len(rows))
        for i, (label,c1,c2,c3,c4,c5,total,bold,bg) in enumerate(rows):
            self.eq_tbl.setRowHeight(i, 38)
            vals = [label, c1, c2, c3, c4, c5, total]
            for j, v in enumerate(vals):
                text = v if j == 0 else (fmt_amt(v) if v else "")
                it = QTableWidgetItem(text)
                it.setTextAlignment(Qt.AlignLeft|Qt.AlignVCenter if j==0 else Qt.AlignRight|Qt.AlignVCenter)
                if bold: it.setFont(QFont("", weight=QFont.Bold))
                if bg:   it.setBackground(QColor(bg))
                if j > 0 and isinstance(v, float) and v < 0:
                    it.setForeground(QColor("#e05252"))
                self.eq_tbl.setItem(i, j, it)

    def _build_cf_stmt(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(8)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("现金流量表", bold=True, size=15)); hdr.addStretch()
        b_dl = QPushButton("导出Excel"); b_dl.setObjectName("btn_outline")
        b_dl.clicked.connect(self._export_cf_stmt)
        hdr.addWidget(b_dl); L.addLayout(hdr)
        L.addWidget(lbl("（采用直接法，现金及现金等价物 = 库存现金+银行存款+其他货币资金）",
                         color="#888", size=12))
        self.cf_stmt_tbl = self._make_report_table(
            ["项目", "行次", "本期金额", "本年累计金额"],
            [-1, 40, 140, 140])
        L.addWidget(self.cf_stmt_tbl)
        self.stack.addWidget(w)

    def _get_cash_balance(self, c, client_id, period_end):
        """期末现金余额 = 期初 + 本年至今净发生额"""
        # Opening balance from accounts
        c.execute("""SELECT SUM(opening_debit - opening_credit) FROM accounts
            WHERE client_id=? AND (code='1001' OR code LIKE '1001.%' OR code LIKE '1001_%'
              OR code='1002' OR code LIKE '1002.%' OR code LIKE '1002_%'
              OR code='1012' OR code LIKE '1012.%' OR code LIKE '1012_%')""",
            (client_id,))
        opening = c.fetchone()[0] or 0
        if not period_end:
            return opening
        year = period_end[:4]
        c.execute("""SELECT SUM(e.debit) - SUM(e.credit) FROM voucher_entries e
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period<=? AND v.period LIKE ? AND v.status='已审核'
            AND (e.account_code='1001' OR e.account_code LIKE '1001.%' OR e.account_code LIKE '1001_%'
              OR e.account_code='1002' OR e.account_code LIKE '1002.%' OR e.account_code LIKE '1002_%'
              OR e.account_code='1012' OR e.account_code LIKE '1012.%' OR e.account_code LIKE '1012_%')""",
            (client_id, period_end, f"{year}%"))
        ytd_net = c.fetchone()[0] or 0
        return opening + ytd_net

    def _compute_cf(self, c, client_id, start_period, end_period):
        """
        Compute cash flow by analyzing cash account counterparts in vouchers.
        Returns dict: row_key -> amount (positive = inflow, negative = outflow shown as positive)
        Two dicts returned: current_period and ytd.
        """
        year = end_period[:4]

        def _analyze(p_start, p_end):
            """Analyze cash flows for a period range."""
            # Get all voucher IDs with cash account entries in range
            c.execute("""SELECT DISTINCT v.id FROM vouchers v
                JOIN voucher_entries e ON e.voucher_id=v.id
                WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
                AND (e.account_code='1001' OR e.account_code LIKE '1001.%' OR e.account_code LIKE '1001_%'
                  OR e.account_code='1002' OR e.account_code LIKE '1002.%' OR e.account_code LIKE '1002_%'
                  OR e.account_code='1012' OR e.account_code LIKE '1012.%' OR e.account_code LIKE '1012_%')""",
                (client_id, p_start, p_end))
            vids = [r[0] for r in c.fetchall()]

            rows = {}  # row_number -> amount

            def add(key, amt):
                rows[key] = rows.get(key, 0) + amt

            for vid in vids:
                c.execute("SELECT account_code, debit, credit FROM voucher_entries WHERE voucher_id=?", (vid,))
                entries = c.fetchall()

                cash_in = 0; cash_out = 0
                non_cash = []
                for e in entries:
                    code = e[0] or ""
                    d = e[1] or 0; cr = e[2] or 0
                    if (code == '1001' or code.startswith('1001.') or code.startswith('1001_') or
                        code == '1002' or code.startswith('1002.') or code.startswith('1002_') or
                        code == '1012' or code.startswith('1012.') or code.startswith('1012_')):
                        cash_in += d; cash_out += cr
                    else:
                        non_cash.append((code, d, cr))

                # Classify inflows (cash debited)
                if cash_in > 0:
                    for code, d, cr in non_cash:
                        amt = cr  # credit side = source of cash
                        if amt <= 0: continue
                        p = code[:4]
                        # Revenue accounts → 销售商品收到现金
                        if (code.startswith('6001') or code.startswith('6002') or
                            code.startswith('6051') or code.startswith('5001') or
                            code.startswith('5051') or code.startswith('1122') or
                            code.startswith('2203')):
                            add('r1', amt)
                        elif code.startswith('2221') or code.startswith('1321'):
                            add('r2', amt)  # 税费返还
                        elif code.startswith('6301') or code.startswith('5301'):
                            add('r3', amt)  # 其他经营收入
                        elif (code.startswith('6111') or code.startswith('5111') or
                              code.startswith('1511') or code.startswith('1521') or
                              code.startswith('1131') or code.startswith('1132')):
                            add('r12', amt)  # 取得投资收益
                        elif code.startswith('1601') or code.startswith('1604'):
                            add('r13', amt)  # 处置固定资产
                        elif code.startswith('2001') or code.startswith('2501'):
                            add('r24', amt)  # 取得借款
                        elif code.startswith('3001') or code.startswith('4001'):
                            add('r23', amt)  # 吸收投资
                        else:
                            add('r3', amt)   # 其他经营收入

                # Classify outflows (cash credited)
                if cash_out > 0:
                    for code, d, cr in non_cash:
                        amt = d  # debit side = destination of cash
                        if amt <= 0: continue
                        if (code.startswith('1403') or code.startswith('1401') or
                            code.startswith('1405') or code.startswith('6401') or
                            code.startswith('6402') or code.startswith('5401') or
                            code.startswith('5402') or code.startswith('2202') or
                            code.startswith('1221')):
                            add('r5', amt)   # 购买商品
                        elif code.startswith('2211'):
                            add('r6', amt)   # 支付员工
                        elif code.startswith('2221') or code.startswith('2231'):
                            add('r7', amt)   # 支付税费
                        elif (code.startswith('6601') or code.startswith('6602') or
                              code.startswith('6603') or code.startswith('5501') or
                              code.startswith('5502') or code.startswith('5503') or
                              code.startswith('2241') or code.startswith('1461')):
                            add('r8', amt)   # 其他经营支出
                        elif (code.startswith('1601') or code.startswith('1604') or
                              code.startswith('1605') or code.startswith('1701')):
                            add('r17', amt)  # 购建固定资产
                        elif (code.startswith('1801') or code.startswith('1511') or
                              code.startswith('1521') or code.startswith('1531')):
                            add('r18', amt)  # 投资支出
                        elif code.startswith('2001') or code.startswith('2501'):
                            add('r27', amt)  # 偿还借款
                        elif (code.startswith('3104') or code.startswith('4104') or
                              code.startswith('2232')):
                            add('r28', amt)  # 分配股利
                        else:
                            add('r8', amt)   # 其他经营支出

            return rows

        cur = _analyze(start_period, end_period)
        ytd = _analyze(f"{year}-01", end_period)
        return cur, ytd

    def _load_cf_stmt(self):
        if not self.client_id: return
        start_period = self.rep_start_period.currentData()
        end_period   = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        year = end_period[:4]

        conn = get_db(); c = conn.cursor()

        # Cash balances
        cash_end  = self._get_cash_balance(c, self.client_id, end_period)
        cash_beg  = self._get_cash_balance(c, self.client_id,
                        f"{year}-01" if start_period[:4] == year else start_period)
        cash_open = self._get_cash_balance(c, self.client_id, None)  # opening from accounts

        # Compute cash flow amounts
        cur, ytd = self._compute_cf(c, self.client_id, start_period, end_period)

        # Subtotals
        def g(d, *keys): return sum(d.get(k, 0) for k in keys)

        # Current period
        ci  = g(cur,'r1','r2','r3')       # 经营流入
        co  = g(cur,'r5','r6','r7','r8')  # 经营流出
        cn  = ci - co                      # 经营净额
        ii  = g(cur,'r11','r12','r13','r14','r15')
        io_ = g(cur,'r17','r18','r19','r20')
        inv_n = ii - io_
        fi  = g(cur,'r23','r24','r25')
        fo  = g(cur,'r27','r28','r29')
        fin_n = fi - fo
        net_cur = cn + inv_n + fin_n

        # YTD
        ci_y  = g(ytd,'r1','r2','r3')
        co_y  = g(ytd,'r5','r6','r7','r8')
        cn_y  = ci_y - co_y
        ii_y  = g(ytd,'r11','r12','r13','r14','r15')
        io_y  = g(ytd,'r17','r18','r19','r20')
        inv_ny = ii_y - io_y
        fi_y  = g(ytd,'r23','r24','r25')
        fo_y  = g(ytd,'r27','r28','r29')
        fin_ny = fi_y - fo_y
        net_ytd = cn_y + inv_ny + fin_ny

        # Net profit for supplementary
        c.execute("""SELECT e.account_code, SUM(e.credit)-SUM(e.debit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, f"{year}-01", end_period))
        mv_ytd = {r[0]: r[1] or 0 for r in c.fetchall()}
        c.execute("""SELECT e.account_code, SUM(e.credit)-SUM(e.debit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, start_period, end_period))
        mv_cur = {r[0]: r[1] or 0 for r in c.fetchall()}
        conn.close()

        def net_profit(mv):
            use6 = any(k.startswith('6') for k in mv)
            if use6:
                rev  = sum(v for k,v in mv.items() if k[:4]<'6400' and k[0]=='6')
                cost = -sum(v for k,v in mv.items() if k[:4]>='6400' and k[0]=='6')
                return rev + cost
            else:
                rev  = sum(v for k,v in mv.items() if k[0]=='5' and k[:4]<'5400')
                cost = -sum(v for k,v in mv.items() if k[0]=='5' and k[:4]>='5400')
                return rev + cost

        np_cur = net_profit(mv_cur)
        np_ytd = net_profit(mv_ytd)

        # AR/AP changes for supplementary (ytd)
        def bal_chg(mv, codes):
            total = 0
            for k, v in mv.items():
                for code in codes:
                    if k == code or k.startswith(code+'.') or k.startswith(code+'_'):
                        total += v; break
            return total
        ar_chg  = -bal_chg(mv_ytd, ['1122','1123','1131','1132','1221'])
        ap_chg  =  bal_chg(mv_ytd, ['2202','2203','2211','2241'])

        # ── Build table rows ──
        BOLD_BG = "#f0f4ff"; HDR_BG = "#e6ecf8"

        def R(label, rowno, cur_val, ytd_val, style="normal"):
            return (label, str(rowno) if rowno else "", cur_val, ytd_val, style)

        rows = [
            # ── 经营活动 ──
            R("一、经营活动产生的现金流量：",  "", None, None, "header"),
            R("  销售商品、提供劳务收到的现金","1", cur.get('r1',0), ytd.get('r1',0)),
            R("  收到的税费返还",              "2", cur.get('r2',0), ytd.get('r2',0)),
            R("  收到的其他与经营活动有关的现金","3",cur.get('r3',0),ytd.get('r3',0)),
            R("  经营活动现金流入小计",         "4", ci,   ci_y,  "subtotal"),
            R("  购买商品、接受劳务支付的现金", "5", cur.get('r5',0), ytd.get('r5',0)),
            R("  支付给职工以及为职工支付的现金","6",cur.get('r6',0),ytd.get('r6',0)),
            R("  支付的各项税费",               "7", cur.get('r7',0), ytd.get('r7',0)),
            R("  支付的其他与经营活动有关的现金","8",cur.get('r8',0),ytd.get('r8',0)),
            R("  经营活动现金流出小计",          "9", co,   co_y,  "subtotal"),
            R("  经营活动产生的现金流量净额",   "10", cn,   cn_y,  "total"),
            # ── 投资活动 ──
            R("二、投资活动产生的现金流量：",   "", None, None,   "header"),
            R("  收回投资收到的现金",           "11", cur.get('r11',0), ytd.get('r11',0)),
            R("  取得投资收益收到的现金",        "12", cur.get('r12',0), ytd.get('r12',0)),
            R("  处置固定资产收回的现金净额",   "13", cur.get('r13',0), ytd.get('r13',0)),
            R("  处置子公司收到的现金净额",     "14", cur.get('r14',0), ytd.get('r14',0)),
            R("  收到的其他与投资活动有关的现金","15",cur.get('r15',0),ytd.get('r15',0)),
            R("  投资活动现金流入小计",         "16", ii,   ii_y,  "subtotal"),
            R("  购建固定资产支付的现金",        "17", cur.get('r17',0), ytd.get('r17',0)),
            R("  投资支付的现金",               "18", cur.get('r18',0), ytd.get('r18',0)),
            R("  取得子公司支付的现金净额",     "19", cur.get('r19',0), ytd.get('r19',0)),
            R("  支付的其他与投资活动有关的现金","20",cur.get('r20',0),ytd.get('r20',0)),
            R("  投资活动现金流出小计",         "21", io_,  io_y,  "subtotal"),
            R("  投资活动产生的现金流量净额",   "22", inv_n,inv_ny,"total"),
            # ── 筹资活动 ──
            R("三、筹资活动产生的现金流量：",   "", None, None,   "header"),
            R("  吸收投资收到的现金",           "23", cur.get('r23',0), ytd.get('r23',0)),
            R("  取得借款收到的现金",           "24", cur.get('r24',0), ytd.get('r24',0)),
            R("  收到的其他与筹资活动有关的现金","25",cur.get('r25',0),ytd.get('r25',0)),
            R("  筹资活动现金流入小计",         "26", fi,   fi_y,  "subtotal"),
            R("  偿还债务支付的现金",           "27", cur.get('r27',0), ytd.get('r27',0)),
            R("  分配股利或偿付利息支付的现金", "28", cur.get('r28',0), ytd.get('r28',0)),
            R("  支付的其他与筹资活动有关的现金","29",cur.get('r29',0),ytd.get('r29',0)),
            R("  筹资活动现金流出小计",         "30", fo,   fo_y,  "subtotal"),
            R("  筹资活动产生的现金流量净额",   "31", fin_n,fin_ny,"total"),
            R("四、汇率变动对现金及现金等价物的影响","32",0,0),
            R("五、现金及现金等价物净增加额",   "33", net_cur, net_ytd, "total"),
            R("  加：期初现金及现金等价物余额", "34", cash_open, cash_open),
            R("六、期末现金及现金等价物余额",   "35", cash_end, cash_end, "total"),
            # ── 补充资料分隔 ──
            R("━━━━ 补充资料 ━━━━",            "",  None, None,  "header"),
            R("一、将净利润调节为经营活动现金流量：","", None, None, "header"),
            R("  净利润",                       "1",  np_cur, np_ytd),
            R("  加：资产减值准备",             "2",  0, 0),
            R("  固定资产折旧",                 "3",  0, 0),
            R("  无形资产摊销",                 "4",  0, 0),
            R("  长期待摊费用摊销",             "5",  0, 0),
            R("  处置固定资产损失（收益-）",    "6",  0, 0),
            R("  公允价值变动损失（收益-）",    "8",  0, 0),
            R("  财务费用（收益-）",            "9",  0, 0),
            R("  投资损失（收益-）",            "10", 0, 0),
            R("  经营性应收项目的减少（增加-）","14", 0, ar_chg),
            R("  经营性应付项目的增加（减少-）","15", 0, ap_chg),
            R("  其他",                         "16", 0, 0),
            R("  经营活动产生的现金流量净额",   "17", cn, cn_y, "total"),
            R("三、现金及现金等价物净变动情况：","", None, None, "header"),
            R("  现金的期末余额",               "21", cash_end, cash_end),
            R("  减：现金的期初余额",           "22", cash_open, cash_open),
            R("  现金及现金等价物净增加额",     "25", cash_end - cash_open, cash_end - cash_open, "total"),
        ]

        self.cf_stmt_tbl.setRowCount(len(rows))
        for i, (label, rowno, cur_v, ytd_v, style) in enumerate(rows):
            self.cf_stmt_tbl.setRowHeight(i, 32)
            is_hdr    = (style == "header")
            is_sub    = (style == "subtotal")
            is_tot    = (style == "total")
            bg = QColor(HDR_BG) if is_hdr else QColor(BOLD_BG) if is_sub or is_tot else None

            for j, (text, align) in enumerate([
                (label,  Qt.AlignLeft|Qt.AlignVCenter),
                (rowno,  Qt.AlignCenter),
                (fmt_amt(cur_v) if isinstance(cur_v, (int,float)) else "",
                         Qt.AlignRight|Qt.AlignVCenter),
                (fmt_amt(ytd_v) if isinstance(ytd_v, (int,float)) else "",
                         Qt.AlignRight|Qt.AlignVCenter),
            ]):
                it = QTableWidgetItem(text); it.setTextAlignment(align)
                if is_hdr:
                    it.setBackground(QColor(HDR_BG))
                    if j == 0: it.setForeground(QColor("#3d6fdb"))
                    it.setFont(QFont("", weight=QFont.Bold))
                elif is_sub or is_tot:
                    it.setBackground(QColor(BOLD_BG))
                    it.setFont(QFont("", weight=QFont.Bold))
                if j >= 2 and isinstance(cur_v if j==2 else ytd_v, (int,float)):
                    val = cur_v if j == 2 else ytd_v
                    if val and val < 0:
                        it.setForeground(QColor("#ff4d4f"))
                self.cf_stmt_tbl.setItem(i, j, it)

    def _export_cf_stmt(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        end_period = self.rep_end_period.currentData() or self.period
        path, _ = QFileDialog.getSaveFileName(self, "保存",
            f"现金流量表_{end_period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "现金流量表"
        hdrs = ["项目","行次","本期金额","本年累计金额"]
        fill_hdr = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(1, ci, h)
            cell.font = XFont(bold=True, color="FFFFFF"); cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center")
        for ri in range(self.cf_stmt_tbl.rowCount()):
            row_vals = []
            for ci in range(4):
                it = self.cf_stmt_tbl.item(ri, ci)
                row_vals.append(it.text() if it else "")
            ws.append(row_vals)
        ws.column_dimensions['A'].width = 45
        for col in ['B','C','D']: ws.column_dimensions[col].width = 16
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _build_cashflow(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(8)
        L.addWidget(lbl("收支统计表（本期科目发生额汇总）", bold=True, size=15))
        L.addWidget(lbl("按资产/负债/收入/费用分类展示本期所有科目的借贷发生额", color="#888", size=12))
        self.cf_tbl = self._make_report_table(
            ["科目编号","科目名称","类型","本期借方","本期贷方","净额"],
            [90,-1,70,110,110,110])
        self.cf_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.cf_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        L.addWidget(self.cf_tbl); self.stack.addWidget(w)

    def _load_cashflow(self):
        if not self.client_id: return
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT e.account_code, e.account_name,
            SUM(e.debit) td, SUM(e.credit) tc
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
            GROUP BY e.account_code ORDER BY e.account_code""",
            (self.client_id, start_period, end_period))
        entries = c.fetchall()
        # Get account types
        c.execute("SELECT code,type FROM accounts WHERE client_id=?",(self.client_id,))
        acct_types = {r[0]:r[1] for r in c.fetchall()}
        conn.close()

        type_colors = {"资产":"#3d6fdb","负债":"#e05252","所有者权益":"#722ed1",
                       "成本":"#fa8c16","收入":"#52c41a","费用":"#eb5757"}

        self.cf_tbl.setRowCount(len(entries))
        td_total = tc_total = 0
        for i,r in enumerate(entries):
            self.cf_tbl.setRowHeight(i,34)
            d=r['td'] or 0; cr=r['tc'] or 0; net=d-cr
            td_total+=d; tc_total+=cr
            atype = acct_types.get(r['account_code'],"")
            tcolor = type_colors.get(atype,"#555")
            vals = [r['account_code'],r['account_name'] or "",atype,fmt_amt(d),fmt_amt(cr),fmt_amt(net)]
            for j,v in enumerate(vals):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignCenter if j!=1 else Qt.AlignLeft|Qt.AlignVCenter)
                if j==2: it.setForeground(QColor(tcolor))
                if j==5: it.setForeground(QColor("#3d6fdb") if net>0 else QColor("#ff4d4f") if net<0 else QColor("#888"))
                self.cf_tbl.setItem(i,j,it)
        # Add totals row
        n = len(entries)
        self.cf_tbl.setRowCount(n+1)
        self.cf_tbl.setRowHeight(n,38)
        for j,v in enumerate(["","合计","",fmt_amt(td_total),fmt_amt(tc_total),fmt_amt(td_total-tc_total)]):
            it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter if j!=1 else Qt.AlignLeft|Qt.AlignVCenter)
            it.setFont(QFont("",weight=QFont.Bold)); it.setBackground(QColor("#f5f7fa"))
            self.cf_tbl.setItem(n,j,it)

    def set_client(self, client_id, client_name, period):
        self.client_id = client_id; self.period = period
        self.period_lbl.setText(f"【{client_name}】{period}")
        # Initialize period ranges
        self.rep_start_period.clear()
        self.rep_end_period.clear()
        now = datetime.now()
        periods = []
        for y in range(now.year, now.year-3, -1):
            for m in range(12,0,-1):
                period_str = f"{y}-{m:02d}"
                display_str = f"{y}年{m:02d}期"
                periods.append((period_str, display_str))

        for period_str, display_str in periods:
            self.rep_start_period.addItem(display_str, period_str)
            self.rep_end_period.addItem(display_str, period_str)

        # Set default to current period
        current_period = f"{now.year}-{now.month:02d}"
        for i in range(self.rep_start_period.count()):
            if self.rep_start_period.itemData(i) == current_period:
                self.rep_start_period.setCurrentIndex(i)
                self.rep_end_period.setCurrentIndex(i)
                break

        idx = self.stack.currentIndex()
        if idx==0: self._load_balance()
        elif idx==1: self._load_income()
        elif idx==2: self._load_equity()
        elif idx==3: self._load_cf_stmt()
        elif idx==4: self._load_cashflow()

    def _export(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        end_period = self.rep_end_period.currentData()
        if not end_period: return
        path,_=QFileDialog.getSaveFileName(self,"保存",f"财务报表_{end_period}.xlsx","Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook()
        # Income sheet
        ws = wb.active; ws.title="利润表"
        ws.append(["项目","行次","本期金额","本年累计"])
        conn = get_db(); c = conn.cursor()
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period:
            QMessageBox.warning(self, "错误", "请选择报告期间")
            return
        c.execute("""SELECT e.account_code,SUM(e.credit)-SUM(e.debit) FROM voucher_entries e
            JOIN vouchers v ON v.id=e.voucher_id WHERE v.client_id=? AND v.period>=? AND v.period<=? GROUP BY e.account_code""",
                  (self.client_id, start_period, end_period))
        cur = {r[0]:r[1] or 0 for r in c.fetchall()}
        def g(code):
            total = 0
            for k, v in cur.items():
                if k == code or k.startswith(code+".") or k.startswith(code+"_"):
                    total += (v or 0)
            return total
        use_6 = bool(g("6001") or g("6401"))
        if use_6:
            income = g("6001") + g("6051")
            cost   = -(g("6401") + g("6402"))
            ops    = income + cost - abs(g("6403")) - abs(g("6601")) - abs(g("6602")) + g("6603") - abs(g("6604"))
            net    = ops + g("6301") - abs(g("6711")) - abs(g("6801"))
        else:
            income = g("5001") + g("5051"); cost = -g("5401") - g("5402")
            ops    = income + cost - abs(g("5501")) - abs(g("5502")) + g("5503")
            net    = ops + g("5301") - abs(g("5601")) - abs(g("5701"))
        ws.append(["营业收入","1",income,""]); ws.append(["营业成本","2",cost,""])
        ws.append(["营业利润","10",ops,""])
        ws.append(["净利润","17",net,""])
        conn.close()
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=20
        wb.save(path); QMessageBox.information(self,"成功",f"报表已导出:\n{path}")


# ── Main Window ─────────────────────────────────────────────────────────────

class AuditPage(QWidget):
    """操作审计日志 — 记录所有关键操作，可导出为审计报告"""

    def __init__(self):
        super().__init__()
        self.client_id = None
        L = QVBoxLayout(self); L.setContentsMargins(24,20,24,20); L.setSpacing(14)

        hdr = QHBoxLayout()
        hdr.addWidget(lbl("操作审计日志", bold=True, size=18)); hdr.addStretch()
        b_exp = QPushButton("导出审计报告(Excel)"); b_exp.setObjectName("btn_outline")
        b_exp.clicked.connect(self._export)
        b_clr = QPushButton("清空日志"); b_clr.setObjectName("btn_red")
        b_clr.clicked.connect(self._clear)
        hdr.addWidget(b_exp); hdr.addWidget(b_clr); L.addLayout(hdr)

        info = QLabel("  记录所有凭证新增/审核/删除、期末结账、数据导入等操作，可作为内部审计依据。")
        info.setStyleSheet("background:#f6f8ff;color:#444;border-radius:6px;padding:8px 12px;font-size:12px;")
        L.addWidget(info)

        # Filter row
        fr = QHBoxLayout(); fr.setSpacing(10)
        self.action_filter = QComboBox()
        self.action_filter.addItems(["全部操作","新增凭证","编辑凭证","凭证审核:已审核","凭证审核:已拒绝",
                                      "凭证审核:待审核","批量导入凭证","期末结转","期间结账封账",
                                      "期间反结账","删除凭证"])
        self.action_filter.currentIndexChanged.connect(self.load)
        self.date_from = QDateEdit(); self.date_from.setDisplayFormat("yyyy-MM-dd")
        self.date_from.setDate(QDate.currentDate().addMonths(-3))
        self.date_from.setCalendarPopup(True)
        self.date_to = QDateEdit(); self.date_to.setDisplayFormat("yyyy-MM-dd")
        self.date_to.setDate(QDate.currentDate()); self.date_to.setCalendarPopup(True)
        self.date_from.dateChanged.connect(self.load)
        self.date_to.dateChanged.connect(self.load)
        fr.addWidget(lbl("操作类型:")); fr.addWidget(self.action_filter)
        fr.addWidget(lbl("从:")); fr.addWidget(self.date_from)
        fr.addWidget(lbl("到:")); fr.addWidget(self.date_to)
        b_q = QPushButton("查询"); b_q.setObjectName("btn_primary"); b_q.clicked.connect(self.load)
        fr.addWidget(b_q); fr.addStretch()
        L.addLayout(fr)

        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.tbl = QTableWidget(); self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows); self.tbl.setShowGrid(False)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setColumnCount(6)
        self.tbl.setHorizontalHeaderLabels(["时间","操作人","操作类型","对象类型","对象ID","详情"])
        hh = self.tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(5, QHeaderView.Stretch)
        self.tbl.setColumnWidth(0,155); self.tbl.setColumnWidth(1,80)
        self.tbl.setColumnWidth(2,130); self.tbl.setColumnWidth(3,80); self.tbl.setColumnWidth(4,80)
        vl.addWidget(self.tbl)

        # Summary bar
        self.summary_bar = QLabel("  共 0 条记录")
        self.summary_bar.setStyleSheet("color:#888;font-size:12px;padding:6px 12px;background:#fafafa;border-top:1px solid #f0f0f0;")
        vl.addWidget(self.summary_bar)
        L.addWidget(f)

    def set_client(self, client_id):
        self.client_id = client_id
        self.load()

    def load(self):
        conn = get_db(); c = conn.cursor()
        action_f = self.action_filter.currentText()
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to   = self.date_to.date().toString("yyyy-MM-dd") + " 23:59:59"

        where = ["created_at >= ? AND created_at <= ?"]
        params = [date_from, date_to]
        if self.client_id:
            where.append("(client_id=? OR client_id IS NULL)")
            params.append(self.client_id)
        if action_f != "全部操作":
            where.append("action=?"); params.append(action_f)

        c.execute(f"""SELECT created_at,operator,action,target_type,target_id,detail
            FROM audit_log WHERE {" AND ".join(where)} ORDER BY created_at DESC LIMIT 500""",
            params)
        rows = c.fetchall(); conn.close()

        action_colors = {
            "新增凭证":"#3d6fdb","编辑凭证":"#fa8c16",
            "凭证审核:已审核":"#52c41a","凭证审核:已拒绝":"#ff4d4f",
            "凭证审核:待审核":"#888","批量导入凭证":"#722ed1",
            "期末结转":"#eb2f96","期间结账封账":"#fa8c16",
            "期间反结账":"#13c2c2","删除凭证":"#ff4d4f",
        }

        self.tbl.setRowCount(len(rows))
        for i,r in enumerate(rows):
            self.tbl.setRowHeight(i,36)
            action = r[2]
            color  = action_colors.get(action,"#555")
            for j,(val,align) in enumerate([
                (r[0][:19],    Qt.AlignCenter),
                (r[1] or "",   Qt.AlignCenter),
                (action,       Qt.AlignCenter),
                (r[3] or "",   Qt.AlignCenter),
                (str(r[4] or ""), Qt.AlignCenter),
                (r[5] or "",   Qt.AlignLeft|Qt.AlignVCenter),
            ]):
                it = QTableWidgetItem(val); it.setTextAlignment(align)
                if j==2: it.setForeground(QColor(color))
                self.tbl.setItem(i,j,it)

        self.summary_bar.setText(f"  共 {len(rows)} 条记录")

    def _export(self):
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        path,_ = QFileDialog.getSaveFileName(self,"保存审计报告",
            f"审计报告_{date.today()}.xlsx","Excel(*.xlsx)")
        if not path: return
        conn = get_db(); c = conn.cursor()
        where = "1=1"
        params = []
        if self.client_id:
            where = "client_id=?"; params = [self.client_id]
        c.execute(f"""SELECT created_at,operator,action,target_type,target_id,detail
            FROM audit_log WHERE {where} ORDER BY created_at DESC""", params)
        rows = c.fetchall(); conn.close()

        wb = openpyxl.Workbook(); ws = wb.active; ws.title="审计日志"
        hdrs = ["操作时间","操作人","操作类型","对象类型","对象ID","详情"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell = ws.cell(1,ci,h)
            cell.font=XFont(bold=True,color="FFFFFF"); cell.fill=fill
            cell.alignment=Alignment(horizontal="center")
        for r in rows:
            ws.append([r[0],r[1],r[2],r[3],r[4],r[5]])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22
        wb.save(path)
        QMessageBox.information(self,"成功",f"审计报告已导出:\n{path}")

    def _clear(self):
        if QMessageBox.question(self,"确认","清空全部审计日志？此操作不可恢复。",
                QMessageBox.Yes|QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db()
        where = "client_id=?" if self.client_id else "1=1"
        params = [self.client_id] if self.client_id else []
        conn.execute(f"DELETE FROM audit_log WHERE {where}", params)
        conn.commit(); conn.close()
        self.load()


class SystemPage(QWidget):
    """系统管理：数据备份/恢复、关于"""

    def __init__(self):
        super().__init__()
        L = QVBoxLayout(self); L.setContentsMargins(40,32,40,32); L.setSpacing(24)
        L.addWidget(lbl("系统管理", bold=True, size=18))

        # ── 数据备份 ──
        grp1 = QFrame(); grp1.setObjectName("card")
        g1 = QVBoxLayout(grp1); g1.setContentsMargins(24,20,24,20); g1.setSpacing(12)
        g1.addWidget(lbl("数据备份", bold=True, size=14))
        g1.addWidget(lbl(
            f"数据库位置：{self._db_path()}\n"
            "备份会将当前数据库完整复制到你指定的位置，建议定期备份到云盘或移动硬盘。",
            color="#666"))
        row1 = QHBoxLayout(); row1.setSpacing(12)
        b_backup = QPushButton("📦 立即备份"); b_backup.setObjectName("btn_primary")
        b_backup.setFixedWidth(140); b_backup.clicked.connect(self._backup)
        row1.addWidget(b_backup); row1.addStretch()
        g1.addLayout(row1)
        self.backup_log = QLabel(""); self.backup_log.setStyleSheet("color:#52c41a;font-size:12px;")
        g1.addWidget(self.backup_log)
        L.addWidget(grp1)

        # ── 数据恢复 ──
        grp2 = QFrame(); grp2.setObjectName("card")
        g2 = QVBoxLayout(grp2); g2.setContentsMargins(24,20,24,20); g2.setSpacing(12)
        g2.addWidget(lbl("数据恢复", bold=True, size=14))
        warn = QLabel("⚠ 恢复将覆盖当前全部数据，操作不可撤销。请确保已备份当前数据！")
        warn.setStyleSheet("background:#fff7e6;color:#d46b08;border-radius:6px;padding:8px 12px;font-size:12px;")
        warn.setWordWrap(True); g2.addWidget(warn)
        row2 = QHBoxLayout(); row2.setSpacing(12)
        b_restore = QPushButton("📂 从备份文件恢复"); b_restore.setObjectName("btn_red")
        b_restore.setFixedWidth(180); b_restore.clicked.connect(self._restore)
        row2.addWidget(b_restore); row2.addStretch()
        g2.addLayout(row2)
        self.restore_log = QLabel(""); self.restore_log.setStyleSheet("color:#666;font-size:12px;")
        g2.addWidget(self.restore_log)
        L.addWidget(grp2)

        # ── 关于 ──
        grp3 = QFrame(); grp3.setObjectName("card")
        g3 = QVBoxLayout(grp3); g3.setContentsMargins(24,20,24,20); g3.setSpacing(6)
        g3.addWidget(lbl("关于 智一会计", bold=True, size=14))
        g3.addWidget(lbl("版本：1.0.0    企业会计准则（2006）    支持 macOS / Windows", color="#666"))
        g3.addWidget(lbl(f"数据目录：{self._db_path()}", color="#aaa", size=11))
        L.addWidget(grp3)
        L.addStretch()

    def _db_path(self):
        from db import DB_PATH
        return DB_PATH

    def _backup(self):
        import shutil, datetime
        default = f"智一会计备份_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        path, _ = QFileDialog.getSaveFileName(self, "选择备份位置", default, "数据库文件(*.db)")
        if not path: return
        try:
            # Use SQLite backup API for a safe online backup
            import sqlite3
            from db import DB_PATH
            src = sqlite3.connect(DB_PATH)
            dst = sqlite3.connect(path)
            src.backup(dst)
            src.close(); dst.close()
            self.backup_log.setText(f"✓ 备份成功：{path}")
            self.backup_log.setStyleSheet("color:#52c41a;font-size:12px;")
        except Exception as e:
            self.backup_log.setText(f"✗ 备份失败：{e}")
            self.backup_log.setStyleSheet("color:#ff4d4f;font-size:12px;")

    def _restore(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择备份文件", "", "数据库文件(*.db)")
        if not path: return
        if QMessageBox.question(self, "确认恢复",
                "恢复将覆盖当前所有数据，此操作不可撤销！\n\n确定要从所选备份文件恢复吗？",
                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        try:
            import sqlite3
            from db import DB_PATH
            # Validate: check it's a valid SQLite db
            test = sqlite3.connect(path)
            test.execute("SELECT name FROM sqlite_master LIMIT 1")
            test.close()
            # Backup current first
            import shutil, datetime
            auto_bak = DB_PATH + f".autobak_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
            shutil.copy2(DB_PATH, auto_bak)
            # Restore
            src = sqlite3.connect(path)
            dst = sqlite3.connect(DB_PATH)
            src.backup(dst)
            src.close(); dst.close()
            self.restore_log.setText(f"✓ 恢复成功！原数据已自动备份至：{auto_bak}\n请重启应用使数据生效。")
            self.restore_log.setStyleSheet("color:#52c41a;font-size:12px;")
            QMessageBox.information(self, "恢复成功",
                "数据恢复成功！\n请关闭并重新启动应用以加载恢复的数据。")
        except Exception as e:
            self.restore_log.setText(f"✗ 恢复失败：{e}")
            self.restore_log.setStyleSheet("color:#ff4d4f;font-size:12px;")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("智一会计 · 本地版")
        self.setMinimumSize(1150, 720)
        self._cur_client = None; self._cur_name = ""; self._cur_period = ""
        self._build()

    def _build(self):
        root = QWidget(); root.setObjectName("root")
        self.setCentralWidget(root)
        row = QHBoxLayout(root); row.setSpacing(0); row.setContentsMargins(0,0,0,0)

        # Sidebar
        sb = QWidget(); sb.setObjectName("sidebar")
        sb.setFixedWidth(196)
        sl = QVBoxLayout(sb); sl.setContentsMargins(0,0,0,0); sl.setSpacing(0)
        logo = QLabel("智一会计"); logo.setObjectName("logo")
        logo.setStyleSheet("color:#fff;font-size:18px;font-weight:bold;padding:22px 20px 4px 20px;")
        sub = QLabel("本地专业版"); sub.setObjectName("subt")
        sub.setStyleSheet("color:#445;font-size:11px;padding:0 20px 14px 20px;")
        sl.addWidget(logo); sl.addWidget(sub)
        div = QFrame(); div.setFrameShape(QFrame.HLine)
        div.setStyleSheet("background:#2a3255;max-height:1px;margin:0 16px 8px 16px;")
        sl.addWidget(div)
        self._nav_btns = []
        for name in ["客户管理","记账（凭证）","科目管理","期末结账","财务报表","审计日志","系统管理"]:
            b = QPushButton(name); b.setObjectName("nav"); b.setProperty("active","false")
            b.clicked.connect(lambda _,n=name: self._nav(n))
            sl.addWidget(b); self._nav_btns.append(b)
        sl.addStretch()
        self._client_info = QLabel(""); self._client_info.setWordWrap(True)
        self._client_info.setStyleSheet("color:#556;font-size:11px;padding:10px 16px;")
        sl.addWidget(self._client_info)
        row.addWidget(sb)

        # Content
        self.stack = QStackedWidget(); row.addWidget(self.stack)
        self.pg_clients = ClientPage()
        self.pg_vouchers = VoucherPage()
        self.pg_accounts = AccountPage()
        self.pg_settle = SettlePage()
        self.pg_reports = ReportPage()
        self.pg_audit = AuditPage()
        self.pg_system = SystemPage()
        for pg in [self.pg_clients, self.pg_vouchers, self.pg_accounts,
                   self.pg_settle, self.pg_reports, self.pg_audit, self.pg_system]:
            self.stack.addWidget(pg)
        self.pg_clients.client_opened.connect(self._open_client)
        self.pg_settle.carryforward_done.connect(self._on_carryforward_done)
        self._nav("客户管理")

    def _on_carryforward_done(self):
        """After carryforward, switch to voucher page and refresh so user can see new vouchers."""
        self.pg_vouchers._switch_tab("查凭证")
        self._nav("记账（凭证）")

    def _nav(self, name):
        mapping = {"客户管理":0,"记账（凭证）":1,"科目管理":2,"期末结账":3,
                   "财务报表":4,"审计日志":5,"系统管理":6}
        self.stack.setCurrentIndex(mapping[name])
        for b in self._nav_btns:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        if name=="客户管理": self.pg_clients.load()

    def _open_client(self, client_id, name, code):
        self._cur_client = client_id; self._cur_name = name
        now = datetime.now()
        self._cur_period = f"{now.year}-{now.month:02d}"
        self._client_info.setText(f"当前客户:\n{name}\n({code})")
        self.pg_vouchers.set_client(client_id, name, self._cur_period)
        self.pg_accounts.set_client(client_id)
        self.pg_settle.set_client(client_id, name, self._cur_period)
        self.pg_reports.set_client(client_id, name, self._cur_period)
        self.pg_audit.set_client(client_id)
        # Log client open
        conn = get_db()
        log_action(conn, client_id, "打开账套", "client", client_id, f"客户: {name}")
        conn.commit(); conn.close()
        self._nav("记账（凭证）")


def main():
    import traceback, atexit
    # Write log file next to the script for silent-crash debugging
    _log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "startup.log")
    def _wlog(msg):
        try:
            with open(_log_path, "a", encoding="utf-8") as _f:
                _f.write(msg + "\n")
            print(msg, file=sys.stderr, flush=True)
        except Exception:
            pass
    from datetime import datetime as _dt
    _wlog(f"\n=== 启动 {_dt.now()} ===")
    try:
        _wlog("step 1: init_db")
        init_db()
        _wlog("step 2: QApplication")
        app = QApplication(sys.argv)
        _wlog("step 3: setStyleSheet")
        app.setStyleSheet(SS)
        _wlog("step 4: MainWindow()")
        w = MainWindow()
        _wlog("step 5: w.show()")
        w.show()
        _wlog("step 6: entering event loop")
        sys.exit(app.exec())
    except Exception:
        tb = traceback.format_exc()
        _wlog("EXCEPTION: " + tb)
        try:
            _app = QApplication.instance() or QApplication(sys.argv)
            QMessageBox.critical(None, "启动错误", tb[:2000])
        except Exception as e2:
            _wlog("Dialog also failed: " + str(e2))
        sys.exit(1)

if __name__ == "__main__":
    main()