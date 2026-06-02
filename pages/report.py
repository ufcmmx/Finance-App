"""pages/report.py — ReportPage — 财务报表"""
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, Signal, QTimer
from PySide6.QtGui import QColor, QFont, QBrush, QPalette

from db import get_db, log_action
from utils import lbl, sep, card, fmt_amt, make_export_button, NoScrollSpinBox, NoScrollDoubleSpinBox

# openpyxl imported lazily inside each export function


class _PrintDialog(QDialog):
    """
    一体化打印窗口：左侧打印设置 + 右侧实时预览。
    布局参考 Chrome / Office 打印面板风格。
    """

    def __init__(self, report_title: str, html: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"打印 — {report_title}")
        self.resize(1060, 700)
        self._html = html
        self._report_title = report_title

        from PySide6.QtPrintSupport import QPrinter, QPrintPreviewWidget, QPrinterInfo
        from PySide6.QtGui import QPageLayout
        from PySide6.QtCore import QMarginsF

        # ── 初始化打印机 ──
        self._printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        self._printer.setPageMargins(QMarginsF(15, 15, 15, 15), QPageLayout.Unit.Millimeter)
        wide = report_title in ("资产负债表", "所有者权益变动表")
        self._printer.setPageOrientation(
            QPageLayout.Orientation.Landscape if wide else QPageLayout.Orientation.Portrait)

        # ═══════════════════════════════════════════════════════
        # 布局：左侧设置面板（固定宽） | 右侧预览
        # ═══════════════════════════════════════════════════════
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── 左侧面板 ──────────────────────────────────────────
        panel = QWidget()
        panel.setFixedWidth(260)
        panel.setStyleSheet("background:#f8f9fc;border-right:1px solid #e0e0e0;")
        pl = QVBoxLayout(panel)
        pl.setContentsMargins(16, 16, 16, 16)
        pl.setSpacing(14)

        def _sec(text):
            l = QLabel(text)
            l.setStyleSheet("font-weight:bold;color:#333;font-size:12px;margin-top:4px;")
            return l

        # 打印机选择
        pl.addWidget(_sec("打印机"))
        self._printer_combo = QComboBox()
        self._printer_combo.setStyleSheet("padding:4px;")
        default_name = QPrinterInfo.defaultPrinter().printerName()
        for info in QPrinterInfo.availablePrinters():
            self._printer_combo.addItem(info.printerName())
        if default_name:
            idx = self._printer_combo.findText(default_name)
            if idx >= 0:
                self._printer_combo.setCurrentIndex(idx)
        pl.addWidget(self._printer_combo)

        # 份数
        pl.addWidget(_sec("份数"))
        self._copies_spin = QSpinBox()
        self._copies_spin.setRange(1, 999)
        self._copies_spin.setValue(1)
        self._copies_spin.setStyleSheet("padding:4px;")
        pl.addWidget(self._copies_spin)

        # 方向
        pl.addWidget(_sec("方向"))
        self._orient_port = QRadioButton("纵向")
        self._orient_land = QRadioButton("横向")
        self._orient_port.setChecked(not wide)
        self._orient_land.setChecked(wide)
        orient_grp = QButtonGroup(self)
        orient_grp.addButton(self._orient_port)
        orient_grp.addButton(self._orient_land)
        pl.addWidget(self._orient_port)
        pl.addWidget(self._orient_land)

        # 纸张
        pl.addWidget(_sec("纸张大小"))
        self._paper_combo = QComboBox()
        self._paper_combo.setStyleSheet("padding:4px;")
        from PySide6.QtGui import QPageSize
        paper_list = [
            ("A4",     QPageSize.PageSizeId.A4),
            ("A3",     QPageSize.PageSizeId.A3),
            ("Letter", QPageSize.PageSizeId.Letter),
        ]
        for name, pid in paper_list:
            self._paper_combo.addItem(name, pid)
        pl.addWidget(self._paper_combo)

        # 边距
        pl.addWidget(_sec("边距（毫米）"))
        margin_row = QHBoxLayout()
        self._margin_spin = QSpinBox()
        self._margin_spin.setRange(5, 50)
        self._margin_spin.setValue(15)
        self._margin_spin.setStyleSheet("padding:4px;")
        margin_row.addWidget(QLabel("四边:"))
        margin_row.addWidget(self._margin_spin)
        pl.addLayout(margin_row)

        pl.addStretch()

        # 打印 / 取消 按钮
        btn_row = QHBoxLayout()
        btn_cancel = QPushButton("取消")
        btn_cancel.setStyleSheet(
            "QPushButton{padding:8px 20px;border:1px solid #ccc;border-radius:5px;background:#fff;}"
            "QPushButton:hover{background:#f0f0f0;}")
        btn_cancel.clicked.connect(self.reject)
        btn_print = QPushButton("打印")
        btn_print.setStyleSheet(
            "QPushButton{padding:8px 20px;border:none;border-radius:5px;"
            "background:#3d6fdb;color:#fff;font-weight:bold;}"
            "QPushButton:hover{background:#2d5dc8;}")
        btn_print.clicked.connect(self._do_print)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_print)
        pl.addLayout(btn_row)

        root.addWidget(panel)

        # ── 右侧预览 ──────────────────────────────────────────
        self._preview = QPrintPreviewWidget(self._printer, self)
        self._preview.setStyleSheet("background:#e8e8e8;")
        self._preview.paintRequested.connect(self._render)
        root.addWidget(self._preview, 1)

        # ── 信号连接（控件全部建好后再连，避免初始化期间误触发）──
        self._orient_port.toggled.connect(self._update_preview)
        self._orient_land.toggled.connect(self._update_preview)
        self._paper_combo.currentIndexChanged.connect(self._update_preview)
        self._margin_spin.valueChanged.connect(self._update_preview)

    # ── 渲染函数（paintRequested 信号回调，预览 & 实际打印共用）─────────────
    def _render(self, printer):
        from PySide6.QtGui import QTextDocument, QPageSize
        from PySide6.QtPrintSupport import QPrinter
        doc = QTextDocument()
        doc.setHtml(self._html)
        # pageRect(Unit.Point) 在 Qt6 中仍可用，但需通过类名访问
        rect = printer.pageRect(QPrinter.Unit.Point)
        doc.setPageSize(rect.size())
        doc.print_(printer)

    # ── 设置变更 → 更新打印机参数 → 刷新预览 ────────────────────────────────
    def _apply_settings(self):
        """将左侧面板的设置写入 self._printer，不触发预览刷新。"""
        from PySide6.QtGui import QPageLayout, QPageSize
        from PySide6.QtCore import QMarginsF
        orient = (QPageLayout.Orientation.Landscape if self._orient_land.isChecked()
                  else QPageLayout.Orientation.Portrait)
        self._printer.setPageOrientation(orient)
        pid = self._paper_combo.currentData()
        if pid is not None:
            self._printer.setPageSize(QPageSize(pid))
        m = self._margin_spin.value()
        self._printer.setPageMargins(QMarginsF(m, m, m, m), QPageLayout.Unit.Millimeter)

    def _update_preview(self):
        """设置变更时：先写入打印机，再刷新预览（仅刷新一次）。"""
        self._apply_settings()
        self._preview.updatePreview()

    # ── 实际打印 ──────────────────────────────────────────────────────────────
    def _do_print(self):
        from PySide6.QtPrintSupport import QPrinterInfo

        # 应用选定打印机
        selected = self._printer_combo.currentText()
        for info in QPrinterInfo.availablePrinters():
            if info.printerName() == selected:
                self._printer.setPrinterName(selected)
                break

        self._printer.setCopyCount(self._copies_spin.value())
        self._apply_settings()   # 确保最终设置写入，不重复刷新预览

        # 渲染并送打印机（直接调用，不经过 paintRequested 信号）
        self._render(self._printer)
        self.accept()


class ReportPage(QWidget):
    """财务报表 — 资产负债表 + 利润表"""

    def _log(self, msg):
        import sys, os
        try:
            _log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "startup.log")
            with open(_log_path, "a", encoding="utf-8") as f:
                from datetime import datetime as _dt
                f.write(f"[{_dt.now()}] {msg}\n")
            print(msg, file=sys.stderr, flush=True)
        except:
            pass

    def __init__(self, parent=None):
        super().__init__(parent)
        self.client_id = None; self.period = ""; self._acct_std = "企业会计准则"
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)
        # Top tabs
        # ── 第一行：Tab 导航 ──
        tb = QWidget(); tb.setStyleSheet("background:#fff;border-bottom:1px solid #e8ecf2;")
        tl = QHBoxLayout(tb); tl.setContentsMargins(16,0,16,0); tl.setSpacing(0)
        self._rtabs = []
        for n in ["资产负债表","利润表","所有者权益变动表","现金流量表","收支统计表"]:
            b = QPushButton(n); b.setStyleSheet("""QPushButton{background:transparent;color:#888;
                border:none;padding:12px 16px;border-bottom:2px solid transparent;}
                QPushButton:hover{color:#3d6fdb;}
                QPushButton[active=true]{color:#3d6fdb;border-bottom:2px solid #3d6fdb;}""")
            b.clicked.connect(lambda _,nn=n:self._switch(nn)); tl.addWidget(b); self._rtabs.append(b)
        tl.addStretch()
        L.addWidget(tb)

        # ── 第二行：期间选择 + 刷新 + 下载 ──
        tb2 = QWidget()
        tb2.setStyleSheet("background:#f8f9fc;border-bottom:1px solid #e8ecf2;")
        tl2 = QHBoxLayout(tb2); tl2.setContentsMargins(16,6,16,6); tl2.setSpacing(10)
        self.period_lbl = lbl("", color="#3d6fdb", bold=True); tl2.addWidget(self.period_lbl)
        tl2.addSpacing(16)
        tl2.addWidget(lbl("报告期间:", color="#666"))
        self.rep_start_period = QComboBox(); self.rep_start_period.setMinimumWidth(140)
        self.rep_end_period   = QComboBox(); self.rep_end_period.setMinimumWidth(140)
        self.rep_start_period.currentIndexChanged.connect(self._refresh_reports)
        self.rep_end_period.currentIndexChanged.connect(self._refresh_reports)
        tl2.addWidget(self.rep_start_period)
        tl2.addWidget(lbl("至", color="#666"))
        tl2.addWidget(self.rep_end_period)
        b_refresh = QPushButton("刷新"); b_refresh.setObjectName("btn_primary")
        b_refresh.setStyleSheet("QPushButton{background:#3d6fdb;color:#fff;border:none;border-radius:6px;padding:7px 18px;font-weight:bold;}QPushButton:hover{background:#2d5dc8;}")
        b_refresh.clicked.connect(self._refresh_reports)
        tl2.addWidget(b_refresh)
        tl2.addStretch()
        b_dl = make_export_button([
            ("Excel (.xlsx)", self._export),
            ("PDF",           self._export_pdf),
        ], label="⬇ 下载")
        tl2.addWidget(b_dl)
        b_print = QPushButton("🖨 打印"); b_print.setObjectName("btn_outline")
        b_print.clicked.connect(self._print_report)
        tl2.addWidget(b_print)
        L.addWidget(tb2)
        self.stack = QStackedWidget(); L.addWidget(self.stack)

        # ── 浮动取数公式卡片（Qt.Popup：点击外部自动关闭）──
        self._tip = QFrame(None)
        self._tip.setWindowFlags(Qt.Popup | Qt.FramelessWindowHint)
        self._tip.setStyleSheet("""
            QFrame { background:#1c2340; border-radius:8px; border:1.5px solid #3d6fdb; }
            QLabel { background:transparent; color:#e8eeff; }
        """)
        _tl = QVBoxLayout(self._tip); _tl.setContentsMargins(14,10,14,12); _tl.setSpacing(6)
        self._tip_name = QLabel(""); self._tip_name.setFont(QFont("", 10, QFont.Bold))
        self._tip_name.setWordWrap(True)
        _sep = QFrame(); _sep.setFrameShape(QFrame.HLine)
        _sep.setStyleSheet("background:#3d6fdb; max-height:1px; border:none;")
        self._tip_fml = QLabel(""); self._tip_fml.setWordWrap(True)
        self._tip_fml.setStyleSheet("color:#a8c4f0; font-size:11px;")
        _tl.addWidget(self._tip_name); _tl.addWidget(_sep); _tl.addWidget(self._tip_fml)
        self._tip.setFixedWidth(320)

        self._build_balance(); self._build_income(); self._build_equity(); self._build_cf_stmt(); self._build_cashflow()
        self._switch("资产负债表")

    def _refresh_reports(self):
        """Refresh current report with selected period range"""
        current_tab = None
        for b in self._rtabs:
            if b.property("active") == "true":
                current_tab = b.text()
                break
        if current_tab:
            self._switch(current_tab)

    def _build_placeholder(self, name):
        w = QWidget(); vl = QVBoxLayout(w)
        vl.addStretch(); vl.addWidget(lbl(f"{name}（生成后显示）", color="#bbb", size=16))
        vl.addStretch(); self.stack.addWidget(w)

    def _switch(self, name):
        self._log(f"_switch: tab={name}, client_id={self.client_id}")
        mapping = {"资产负债表":0,"利润表":1,"所有者权益变动表":2,"现金流量表":3,"收支统计表":4}
        for b in self._rtabs:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        if name in mapping:
            self.stack.setCurrentIndex(mapping[name])
            if name=="资产负债表": self._load_balance()
            elif name=="利润表": self._load_income()
            elif name=="所有者权益变动表": self._load_equity()
            elif name=="现金流量表": self._load_cf_stmt()
            elif name=="收支统计表": self._load_cashflow()

    # ── 公式 Tooltip ──────────────────────────────────────────────
    def _show_formula_tip(self, name, formula, table, item):
        """点击项目名称列 → 弹出取数公式卡片（点击其他位置自动关闭）。"""
        self._tip_name.setText(name)
        self._tip_fml.setText(formula if formula else "（暂无取数规则）")
        self._tip.setFixedWidth(320)
        self._tip.adjustSize()
        rect = table.visualItemRect(item)
        gpos = table.viewport().mapToGlobal(rect.bottomLeft())
        from PySide6.QtGui import QGuiApplication
        screen = QGuiApplication.primaryScreen().availableGeometry()
        tx = gpos.x()
        ty = gpos.y() + 4
        tip_h = self._tip.sizeHint().height() or 80
        if tx + 320 > screen.right() - 4:
            tx = screen.right() - 324
        if ty + tip_h > screen.bottom() - 4:
            ty = gpos.y() - tip_h - rect.height() - 4
        self._tip.move(tx, ty)
        self._tip.show()

    def _on_bs_cell_clicked(self, row, col):
        """点击资产负债表项目名称列 → 显示取数公式。"""
        if col not in (0, 4):
            return
        it = self.bs_tbl.item(row, col)
        if not it:
            return
        data = it.data(Qt.ItemDataRole.UserRole)
        if not data:
            return
        self._show_formula_tip(data[0], data[1], self.bs_tbl, it)

    def _on_inc_cell_clicked(self, row, col):
        """点击利润表项目名称列 → 显示取数公式。"""
        if col != 0:
            return
        it = self.inc_tbl.item(row, col)
        if not it:
            return
        data = it.data(Qt.ItemDataRole.UserRole)
        if not data:
            return
        self._show_formula_tip(data[0], data[1], self.inc_tbl, it)

    def _make_report_table(self, cols, col_widths=None):
        t = QTableWidget(); t.setColumnCount(len(cols))
        t.setHorizontalHeaderLabels(cols)
        t.verticalHeader().setVisible(False); t.setShowGrid(True)
        t.setEditTriggers(QTableWidget.NoEditTriggers)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)  # 所有列均可拖拽
        t.horizontalHeader().setStretchLastSection(False)
        t.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        if col_widths:
            for i, w in enumerate(col_widths):
                # -1 原为 Stretch（不可调），改为 Interactive 并给合适默认宽度
                t.setColumnWidth(i, 220 if w == -1 else w)
        return t

    def _build_balance(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14)
        self.bs_tbl = self._make_report_table(
            ["资产项目","行次","期末金额","年初金额","负债和所有者权益","行次","期末金额","年初金额"],
            [-1,40,110,110,-1,40,110,110])
        self.bs_tbl.cellClicked.connect(self._on_bs_cell_clicked)
        L.addWidget(self.bs_tbl); self.stack.addWidget(w)

    def _load_balance(self):
        if not self.client_id: return
        end_period = self.rep_end_period.currentData()
        if not end_period: return
        is_small = getattr(self, '_acct_std', '企业会计准则') == '小企业会计制度'
        year = end_period[:4]
        year_start = f"{year}-01"   # 本年第一期
        conn = get_db(); c = conn.cursor()

        # 期末：截止所选期间的累计发生额
        c.execute("""SELECT e.account_code, SUM(e.debit)-SUM(e.credit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period<=? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, end_period))
        mv = {r[0]: r[1] or 0 for r in c.fetchall()}

        # 年初：上年年末 = 期初余额 + 本年第一期之前的凭证发生额
        # 若所选期间就是01期，则年初 = 纯期初余额（凭证发生额=0）
        c.execute("""SELECT e.account_code, SUM(e.debit)-SUM(e.credit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period<? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, year_start))
        mv_ys = {r[0]: r[1] or 0 for r in c.fetchall()}

        c.execute("SELECT code, name, opening_debit, opening_credit, direction FROM accounts WHERE client_id=?",
                  (self.client_id,))
        accts = {r['code']: r for r in c.fetchall()}
        conn.close()

        # 预计算末级科目：没有任何子科目的科目
        all_codes = set(accts.keys())
        leaf_codes = {
            code for code in all_codes
            if not any(
                other != code and (other.startswith(code+".") or other.startswith(code+"_"))
                for other in all_codes
            )
        }

        def _vat_split(movements):
            """
            拆分 2221 应交税费，返回 (tax_pay_net, to_oth_cur_asset, to_oth_cur_liab)
            - 借方余额明细（留底税/待抵扣/待认证）→ 其他流动资产
            - 含"待转"的明细科目贷方余额         → 其他流动负债
            - 其余贷方余额                       → 应交税费
            """
            to_asset = 0.0; to_liab = 0.0; to_tax = 0.0
            leaf_2221 = [code for code in accts
                         if code in leaf_codes and
                         (code == "2221" or code.startswith("2221.") or code.startswith("2221_"))]
            if not leaf_2221:
                # 无子科目，按整体余额处理
                b = _bal_with_mv(["2221"], movements)
                if b < 0: to_asset += -b
                else: to_tax += b
                return to_tax, to_asset, to_liab
            for code in leaf_2221:
                aname = (accts[code]['name'] or "").strip()
                b = _bal_with_mv([code], movements)
                if "待转" in aname:
                    to_liab += max(0.0, b)   # 待转销项税 → 其他流动负债
                elif b < 0:
                    to_asset += -b           # 借方余额（留底税等）→ 其他流动资产
                else:
                    to_tax += b              # 正常贷方余额 → 应交税费
            return to_tax, to_asset, to_liab

        def _bal_with_mv(code_prefix_list, movements):
            """通用余额计算：末级科目取期初+发生额，父科目只取发生额"""
            total = 0
            for code, a in accts.items():
                if not any(code == p or code.startswith(p+".") or code.startswith(p+"_")
                           for p in code_prefix_list):
                    continue
                net_mv = movements.get(code, 0)
                if code in leaf_codes:
                    od = a['opening_debit'] or 0; oc = a['opening_credit'] or 0
                    if a['direction'] == '借':
                        total += (od - oc) + net_mv
                    else:
                        total += (oc - od) - net_mv
                else:
                    if a['direction'] == '借':
                        total += net_mv
                    else:
                        total -= net_mv
            return total

        def bal(code_prefix_list):
            """期末余额"""
            return _bal_with_mv(code_prefix_list, mv)

        def bal_ys(code_prefix_list):
            """年初余额（上年年末 = 期初 + 本年首期前发生额）"""
            return _bal_with_mv(code_prefix_list, mv_ys)

        # ── 应交税费重分类（须在 _bal_with_mv 定义后调用）──
        tax_pay_net,   vat_to_asset,   vat_to_liab   = _vat_split(mv)
        tax_pay_net_y, vat_to_asset_y, vat_to_liab_y = _vat_split(mv_ys)

        # ── 资产方 ──
        cash      = bal(["1001","1002","1012"])
        notes_rec = bal(["1121"])
        # ── 重分类（应收↔预收、预付↔应付、其他应收↔其他应付）──
        _ar_raw   = bal(["1122"]); _ad_raw   = bal(["2203"])
        _pr_raw   = bal(["1123"]); _ap_raw   = bal(["2202"])
        _or_raw   = bal(["1221"]); _op_raw   = bal(["2241"])
        acct_rec = max(0.0, _ar_raw) + max(0.0, -_ad_raw)
        prepay   = max(0.0, _pr_raw) + max(0.0, -_ap_raw)
        oth_rec  = max(0.0, _or_raw) + max(0.0, -_op_raw)
        int_rec   = bal(["1132"])
        div_rec   = bal(["1131"])

        # ── 年初余额（同结构，使用 bal_ys） ──
        notes_rec_y = bal_ys(["1121"])
        # ── 重分类（应收↔预收、预付↔应付、其他应收↔其他应付）──
        _ar_y    = bal_ys(["1122"]); _ad_y    = bal_ys(["2203"])
        _pr_y    = bal_ys(["1123"]); _ap_y    = bal_ys(["2202"])
        _or_y    = bal_ys(["1221"]); _op_y    = bal_ys(["2241"])
        acct_rec_y = max(0.0, _ar_y) + max(0.0, -_ad_y)
        prepay_y   = max(0.0, _pr_y) + max(0.0, -_ap_y)
        oth_rec_y  = max(0.0, _or_y) + max(0.0, -_op_y)
        int_rec_y   = bal_ys(["1132"])
        div_rec_y   = bal_ys(["1131"])
        cash_y      = bal_ys(["1001","1002","1012"])
        inventory_y = (bal_ys(["1401","1402","1403","1404","1405","1406","1407","1408","1409","1411","1415","1421"])
                      - abs(bal_ys(["1471","1472"])))
        prepd_exp_y = bal_ys(["1901"])
        fa_y        = bal_ys(["1601"]) - abs(bal_ys(["1602"])) - abs(bal_ys(["1603"]))
        wip_y       = bal_ys(["1604"])
        intangible_y= bal_ys(["1701"]) - abs(bal_ys(["1702"])) - abs(bal_ys(["1703"]))
        lt_prepaid_y= bal_ys(["1801"])
        deferred_a_y= bal_ys(["1811"])
        avail_sale_y   = bal_ys(["1503"])
        held_to_mat_y  = bal_ys(["1501"]) - abs(bal_ys(["1502"]))
        lt_eq_invest_y = bal_ys(["1511"])
        invest_prop_y  = bal_ys(["1521"])
        lt_equity_y    = avail_sale_y + held_to_mat_y + lt_eq_invest_y + invest_prop_y
        noncur_asset_y = fa_y+wip_y+intangible_y+lt_prepaid_y+lt_equity_y+deferred_a_y

        st_loan_y   = bal_ys(["2001"]); notes_pay_y = bal_ys(["2201"])
        acct_pay_y  = max(0.0, _ap_y) + max(0.0, -_pr_y)
        adv_rec_y   = max(0.0, _ad_y) + max(0.0, -_ar_y)
        emp_pay_y   = bal_ys(["2211"]); tax_pay_y   = tax_pay_net_y
        int_pay_y   = bal_ys(["2231"]); div_pay_y   = bal_ys(["2232"])
        oth_pay_y   = max(0.0, _op_y) + max(0.0, -_or_y)
        # 其他流动资产（年初）= 待处理财产损溢 + 待摊费用 + 应交税费借方余额重分类
        oth_cur_asset_y = prepd_exp_y + bal_ys(["1461"]) + vat_to_asset_y
        # 其他流动负债（年初）= 待转销项税额
        oth_cur_liab_y  = vat_to_liab_y
        cur_asset_y  = (cash_y+notes_rec_y+acct_rec_y+prepay_y+int_rec_y+div_rec_y
                        +oth_rec_y+inventory_y+oth_cur_asset_y)
        cur_liab_y  = (st_loan_y+notes_pay_y+acct_pay_y+adv_rec_y+emp_pay_y+tax_pay_y
                       +int_pay_y+div_pay_y+oth_pay_y+oth_cur_liab_y)
        total_asset_y  = cur_asset_y + noncur_asset_y
        lt_loan_y   = bal_ys(["2501"]); bonds_pay_y = bal_ys(["2502"])
        lt_payable_y= bal_ys(["2701"]); est_liab_y  = bal_ys(["2801"])
        deferred_l_y= bal_ys(["2901"])
        noncur_liab_y = lt_loan_y+bonds_pay_y+lt_payable_y+est_liab_y+deferred_l_y
        total_liab_y  = cur_liab_y + noncur_liab_y
        if is_small:
            cap_y     = bal_ys(["3001"]); cap_res_y = bal_ys(["3002"])
            surp_res_y= bal_ys(["3101"])
            profit_y  = bal_ys(["3103"]) + bal_ys(["3104"])
            tsy_y     = 0.0
        else:
            cap_y     = bal_ys(["4001"]); cap_res_y = bal_ys(["4002"])
            surp_res_y= bal_ys(["4101"])
            profit_y  = bal_ys(["4103"]) + bal_ys(["4104"])
            tsy_y     = bal_ys(["4201"])
        total_equity_y = cap_y + cap_res_y + surp_res_y + profit_y - tsy_y
        total_le_y     = total_liab_y + total_equity_y
        # 存货 = 各存货科目合计 - 存货跌价准备 - 消耗性生物资产跌价准备
        inventory = (bal(["1401","1402","1403","1404","1405","1406","1407","1408","1409","1411","1415","1421"])
                     - abs(bal(["1471","1472"])))
        prepd_exp = bal(["1901"])   # 待处理财产损溢
        fa        = bal(["1601"]) - abs(bal(["1602"])) - abs(bal(["1603"]))
        wip       = bal(["1604"])
        intangible= bal(["1701"]) - abs(bal(["1702"])) - abs(bal(["1703"]))
        lt_prepaid= bal(["1801"])   # 长期待摊费用
        deferred_a= bal(["1811"])   # 递延所得税资产
        avail_sale  = bal(["1503"])                    # 可供出售金融资产
        held_to_mat = bal(["1501"]) - abs(bal(["1502"]))  # 持有至到期投资净额
        lt_eq_invest= bal(["1511"])                    # 长期股权投资
        invest_prop = bal(["1521"])                    # 投资性房地产
        lt_equity   = avail_sale + held_to_mat + lt_eq_invest + invest_prop
        noncur_asset = fa+wip+intangible+lt_prepaid+lt_equity+deferred_a

        # ── 负债方 ──
        st_loan   = bal(["2001"])
        notes_pay = bal(["2201"])
        acct_pay  = max(0.0, _ap_raw) + max(0.0, -_pr_raw)
        adv_rec   = max(0.0, _ad_raw) + max(0.0, -_ar_raw)
        emp_pay   = bal(["2211"])
        tax_pay   = tax_pay_net   # 应交税费贷方净额（借方余额已重分类到其他流动资产）
        int_pay   = bal(["2231"])
        div_pay   = bal(["2232"])
        oth_pay   = max(0.0, _op_raw) + max(0.0, -_or_raw)
        # 其他流动资产 = 待处理财产损溢(1901) + 待摊费用(1461) + 应交税费借方余额重分类
        oth_cur_asset = prepd_exp + bal(["1461"]) + vat_to_asset
        # 其他流动负债 = 待转销项税额（应交税费下贷方余额重分类）
        oth_cur_liab  = vat_to_liab
        cur_asset = (cash+notes_rec+acct_rec+prepay+int_rec+div_rec
                     +oth_rec+inventory+oth_cur_asset)
        cur_liab  = (st_loan+notes_pay+acct_pay+adv_rec+emp_pay+tax_pay
                     +int_pay+div_pay+oth_pay+oth_cur_liab)
        total_asset = cur_asset + noncur_asset
        lt_loan   = bal(["2501"])
        bonds_pay = bal(["2502"])
        lt_payable= bal(["2701"])   # 长期应付款
        est_liab  = bal(["2801"])   # 预计负债
        deferred_l= bal(["2901"])   # 递延所得税负债
        noncur_liab = lt_loan+bonds_pay+lt_payable+est_liab+deferred_l
        total_liab = cur_liab + noncur_liab

        # ── 所有者权益 ──
        if is_small:
            cap       = bal(["3001"])
            cap_res   = bal(["3002"])
            surp_res  = bal(["3101"])
            profit    = bal(["3103"]) + bal(["3104"])
            tsy_stock = 0.0
        else:
            cap       = bal(["4001"])
            cap_res   = bal(["4002"])
            surp_res  = bal(["4101"])
            profit    = bal(["4103"]) + bal(["4104"])
            tsy_stock = bal(["4201"])
        total_equity = cap + cap_res + surp_res + profit - tsy_stock
        total_le     = total_liab + total_equity

        def R(label, rowno, left_val, right_label="", right_rowno="", right_val=None,
              is_header=False, is_total=False,
              left_ys=None, right_ys=None,
              left_formula="", right_formula=""):
            return (label, rowno, left_val, right_label, right_rowno, right_val,
                    is_header, is_total, left_ys, right_ys, left_formula, right_formula)

        # ── 制度相关公式字符串 ──
        _eq = "小企业会计制度" if is_small else "企业会计准则"
        cap_fml    = ("贷方余额 | 3001 实收资本" if is_small
                      else "贷方余额 | 4001 实收资本（或股本）")
        cap_res_fml= ("贷方余额 | 3002 资本公积" if is_small
                      else "贷方余额 | 4002 资本公积")
        surp_fml   = ("贷方余额 | 3101 盈余公积" if is_small
                      else "贷方余额 | 4101 盈余公积")
        profit_fml = ("贷方余额 | 3103 本年利润 + 3104 利润分配" if is_small
                      else "贷方余额 | 4103 本年利润 + 4104 利润分配")
        tsy_fml    = "借方余额（取负）| 4201 库存股"

        rows = [
            R("流动资产：","","",  "流动负债：","","",          True),
            R("货币资金","1",cash,            "短期借款","34",st_loan,
              left_ys=cash_y,        right_ys=st_loan_y,
              left_formula ="借方余额合计 | 1001 库存现金 + 1002 银行存款 + 1012 其他货币资金",
              right_formula="贷方余额 | 2001 短期借款"),
            R("以公允价值计量且其变动\n计入当期损益的金融资产","2",0,
              "以公允价值计量且其变动\n计入当期损益的金融负债","35",0,
              left_formula ="借方余额 | 1101 交易性金融资产",
              right_formula="贷方余额 | 2101 交易性金融负债"),
            R("衍生金融资产","3",0,            "衍生金融负债","36",0,
              left_formula ="借方余额 | 1103 衍生工具（资产方向）",
              right_formula="贷方余额 | 2103 衍生工具（负债方向）"),
            R("应收票据","4",notes_rec,        "应付票据","37",notes_pay,
              left_ys=notes_rec_y,   right_ys=notes_pay_y,
              left_formula ="借方余额 | 1121 应收票据",
              right_formula="贷方余额 | 2201 应付票据"),
            R("应收账款","5",acct_rec,         "应付账款","38",acct_pay,
              left_ys=acct_rec_y,    right_ys=acct_pay_y,
              left_formula ="1122 借方余额 + 2203 贷方余额重分类（预收款贷方计入预收款项）",
              right_formula="2202 贷方余额 + 1123 借方余额重分类（预付款借方计入预付款项）"),
            R("预付款项","6",prepay,           "预收款项","39",adv_rec,
              left_ys=prepay_y,      right_ys=adv_rec_y,
              left_formula ="1123 借方余额 + 2202 贷方余额重分类（应付账款贷方计入应付账款）",
              right_formula="2203 贷方余额 + 1122 借方余额重分类（应收账款借方计入应收账款）"),
            R("应收利息","7",int_rec,          "应付职工薪酬","40",emp_pay,
              left_ys=int_rec_y,     right_ys=emp_pay_y,
              left_formula ="借方余额 | 1132 应收利息",
              right_formula="贷方余额 | 2211 应付职工薪酬"),
            R("应收股利","8",div_rec,          "应交税费","41",tax_pay,
              left_ys=div_rec_y,     right_ys=tax_pay_y,
              left_formula ="借方余额 | 1131 应收股利",
              right_formula="2221 贷方净余额（借方留底税额重分类至其他流动资产）"),
            R("其他应收款","9",oth_rec,         "应付利息","42",int_pay,
              left_ys=oth_rec_y,     right_ys=int_pay_y,
              left_formula ="1221 借方余额 + 2241 贷方余额重分类",
              right_formula="贷方余额 | 2231 应付利息"),
            R("存货","10",inventory,           "应付股利","43",div_pay,
              left_ys=inventory_y,   right_ys=div_pay_y,
              left_formula ="1401~1421 合计 − |1471 存货跌价准备| − |1472 消耗性生物资产跌价准备|",
              right_formula="贷方余额 | 2232 应付股利"),
            R("持有待售资产","11",0,            "其他应付款","44",oth_pay,
              right_ys=oth_pay_y,
              left_formula ="重分类项目：将于一年内处置的非流动资产转入",
              right_formula="2241 贷方余额 + 1221 借方余额重分类"),
            R("一年内到期的非流动资产","12",0,  "持有待售负债","45",0,
              left_formula ="重分类项目：将于一年内到期的长期资产转入",
              right_formula="重分类项目：与持有待售资产相关的负债"),
            R("其他流动资产","13",oth_cur_asset, "一年内到期的非流动负债","46",0,
              left_ys=oth_cur_asset_y,
              left_formula ="1901 待处理财产损溢 + 1461 待摊费用\n+ 2221 借方余额重分类（留底税/待抵扣/待认证）",
              right_formula="重分类项目：将于一年内到期的长期负债转入"),
            R("流动资产合计","14",cur_asset,   "其他流动负债","47",oth_cur_liab,
              False,True, left_ys=cur_asset_y, right_ys=cur_liab_y,
              left_formula ="以上各流动资产项目合计",
              right_formula="2221「待转销项税额」明细贷方余额"),
            R("非流动资产：","","",            "流动负债合计","48",cur_liab,
              True,True, right_ys=cur_liab_y,
              right_formula="以上各流动负债项目合计"),
            R("可供出售金融资产","15",avail_sale,   "非流动负债：","","",
              False,False, left_ys=avail_sale_y,
              left_formula="借方余额 | 1503 可供出售金融资产"),
            R("持有至到期投资","16",held_to_mat,    "长期借款","49",lt_loan,
              left_ys=held_to_mat_y,  right_ys=lt_loan_y,
              left_formula ="1501 持有至到期投资 − |1502 持有至到期投资减值准备|",
              right_formula="贷方余额 | 2501 长期借款"),
            R("长期应收款","17",0,                  "应付债券","50",bonds_pay,
              right_formula="贷方余额 | 2502 应付债券",
              left_formula ="1231 长期应收款 − |1232 未实现融资收益|"),
            R("长期股权投资","18",lt_eq_invest,     "其中：优先股","51",0,
              left_ys=lt_eq_invest_y,
              left_formula="借方余额 | 1511 长期股权投资"),
            R("投资性房地产","19",invest_prop,       "永续债","52",0,
              left_ys=invest_prop_y,
              left_formula="借方余额 | 1521 投资性房地产"),
            R("固定资产","20",fa,              "长期应付款","53",lt_payable,
              left_ys=fa_y,           right_ys=lt_payable_y,
              left_formula ="1601 固定资产原值 − |1602 累计折旧| − |1603 固定资产减值准备|",
              right_formula="贷方余额 | 2701 长期应付款"),
            R("在建工程","21",wip,             "专项应付款","54",0,
              left_ys=wip_y,
              left_formula="借方余额 | 1604 在建工程"),
            R("工程物资","22",0,               "预计负债","55",est_liab,
              right_ys=est_liab_y,
              left_formula ="借方余额 | 1605 工程物资",
              right_formula="贷方余额 | 2801 预计负债"),
            R("固定资产清理","23",0,            "递延收益","56",0,
              left_formula ="1606 固定资产清理净余额（借方为待处理损失，贷方为待转收益）",
              right_formula="贷方余额 | 2401 递延收益"),
            R("生产性生物资产","24",0,          "递延所得税负债","57",deferred_l,
              right_ys=deferred_l_y,
              left_formula ="1611 生产性生物资产 − |1612 生产性生物资产累计折旧|",
              right_formula="贷方余额 | 2901 递延所得税负债"),
            R("油气资产","25",0,               "其他非流动负债","58",0,
              left_formula ="1621 油气资产 − |1622 油气资产折耗| − |1623 油气资产减值准备|",
              right_formula="其他长期负债科目贷方余额"),
            R("无形资产","26",intangible,       "非流动负债合计","59",noncur_liab,
              False,True, left_ys=intangible_y, right_ys=noncur_liab_y,
              left_formula ="1701 无形资产原值 − |1702 累计摊销| − |1703 无形资产减值准备|",
              right_formula="以上各非流动负债项目合计"),
            R("开发支出","27",0,               "负债合计","60",total_liab,
              False,True, right_ys=total_liab_y,
              left_formula ="借方余额 | 1712 研发支出（资本化部分）",
              right_formula="流动负债合计 + 非流动负债合计"),
            R("商誉","28",0,                   "所有者权益（或股东权益）：","","",True,
              left_formula="借方余额 | 1711 商誉 − |商誉减值准备|"),
            R("长期待摊费用","29",lt_prepaid,   "实收资本（或股本）","61",cap,
              left_ys=lt_prepaid_y,   right_ys=cap_y,
              left_formula ="借方余额 | 1801 长期待摊费用",
              right_formula=cap_fml),
            R("递延所得税资产","30",deferred_a, "其他权益工具","62",0,
              left_ys=deferred_a_y,
              left_formula ="借方余额 | 1811 递延所得税资产",
              right_formula="贷方余额 | 4004 其他权益工具"),
            R("其他非流动资产","31",0,          "其中：优先股","63",0,
              left_formula="其他未单独列示的非流动资产科目借方余额"),
            R("非流动资产合计","32",noncur_asset,"永续债","64",0,
              False,True, left_ys=noncur_asset_y,
              left_formula="以上各非流动资产项目合计"),
            R("","","",                        "资本公积","65",cap_res,
              right_ys=cap_res_y, right_formula=cap_res_fml),
            R("","","",                        "" if is_small else "减：库存股",
              "" if is_small else "66",
              None if is_small else tsy_stock,
              right_ys=None if is_small else tsy_y,
              right_formula="" if is_small else tsy_fml),
            R("","","",                        "其他综合收益","67",0,
              right_formula="贷方余额 | 4003 其他综合收益（小企业：3003）"),
            R("","","",                        "盈余公积","68",surp_res,
              right_ys=surp_res_y, right_formula=surp_fml),
            R("","","",                        "未分配利润","69",profit,
              right_ys=profit_y, right_formula=profit_fml),
            R("","","",                        "所有者权益合计","70",total_equity,
              False,True, right_ys=total_equity_y,
              right_formula="实收资本 + 资本公积 + 盈余公积 + 未分配利润 − 库存股"),
            R("资产总计","33",total_asset,     "负债和所有者权益总计","71",total_le,
              False,True, left_ys=total_asset_y, right_ys=total_le_y,
              left_formula ="流动资产合计 + 非流动资产合计",
              right_formula="负债合计 + 所有者权益合计"),
        ]

        self.bs_tbl.setRowCount(len(rows))
        for i,(l_name,l_row,l_val,r_name,r_row,r_val,is_hdr,is_tot,l_ys,r_ys,l_fml,r_fml) in enumerate(rows):
            self.bs_tbl.setRowHeight(i, 32)
            # Left
            for j,(text,align) in enumerate([
                (l_name, Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter),
                (str(l_row), Qt.AlignmentFlag.AlignCenter),
                (fmt_amt(l_val) if isinstance(l_val,(int,float)) else "", Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter),
                (fmt_amt(l_ys) if isinstance(l_ys,(int,float)) else "", Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter),
            ]):
                it = QTableWidgetItem(text); it.setTextAlignment(align)
                if is_hdr or is_tot:
                    it.setBackground(QColor("#f0f4ff" if is_hdr else "#f5f7fa"))
                    if is_tot: it.setFont(QFont("",weight=QFont.Bold))
                if j==0 and is_hdr: it.setForeground(QColor("#3d6fdb"))
                if j==2 and isinstance(l_val,(int,float)) and l_val<0:
                    it.setForeground(QColor("#e05252"))
                if j==3 and isinstance(l_ys,(int,float)) and l_ys<0:
                    it.setForeground(QColor("#e05252"))
                # 名称列存公式，鼠标指针改为问号提示可点击
                if j == 0 and l_fml and not is_hdr:
                    it.setData(Qt.ItemDataRole.UserRole, (l_name.replace("\n"," "), l_fml))
                    it.setToolTip("点击查看取数公式")
                self.bs_tbl.setItem(i,j,it)
            # Right
            for j,(text,align) in enumerate([
                (r_name, Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter),
                (str(r_row), Qt.AlignmentFlag.AlignCenter),
                (fmt_amt(r_val) if isinstance(r_val,(int,float)) else "", Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter),
                (fmt_amt(r_ys) if isinstance(r_ys,(int,float)) else "", Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter),
            ],4):
                it = QTableWidgetItem(text); it.setTextAlignment(align)
                if is_hdr or is_tot:
                    it.setBackground(QColor("#f0f4ff" if is_hdr else "#f5f7fa"))
                    if is_tot: it.setFont(QFont("",weight=QFont.Bold))
                if j==4 and is_hdr: it.setForeground(QColor("#3d6fdb"))
                if j==6 and isinstance(r_val,(int,float)) and r_val<0:
                    it.setForeground(QColor("#e05252"))
                if j==7 and isinstance(r_ys,(int,float)) and r_ys<0:
                    it.setForeground(QColor("#e05252"))
                # 名称列存公式，鼠标指针改为问号提示可点击
                if j == 4 and r_fml and not is_hdr:
                    it.setData(Qt.ItemDataRole.UserRole, (r_name.replace("\n"," "), r_fml))
                    it.setToolTip("点击查看取数公式")
                self.bs_tbl.setItem(i,j,it)

    def _build_income(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14)
        self.inc_tbl = self._make_report_table(
            ["项目","行次","本期金额","本年累计金额"],[-1,40,160,160])
        self.inc_tbl.cellClicked.connect(self._on_inc_cell_clicked)
        L.addWidget(self.inc_tbl); self.stack.addWidget(w)

    def _load_income(self):
        if not self.client_id:
            self._log("_load_income: no client_id, returning")
            return
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period:
            self._log(f"_load_income: no periods (start={start_period}, end={end_period}), returning")
            return
        self._log(f"_load_income: client={self.client_id}, start={start_period}, end={end_period}")
        conn = get_db(); c = conn.cursor()

        # ── 检测科目体系：根据客户选择的会计准则，而非活动数据 ──
        c.execute("SELECT accounting_std FROM clients WHERE id=?",
                  (self.client_id,))
        row = c.fetchone()
        use_6xxx = (row["accounting_std"] == "企业会计准则" if row else True)
        self._log(f"_load_income: use_6xxx={use_6xxx} (std={row['accounting_std'] if row else 'NOT FOUND'})")

        def fetch_period(p_start, p_end):
            c.execute("""SELECT e.account_code, SUM(e.credit)-SUM(e.debit) net
                FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
                WHERE v.client_id=? AND v.period>=? AND v.period<=?
                AND v.status='已审核'
                AND v.id NOT IN (
                    SELECT ve.voucher_id FROM voucher_entries ve
                    WHERE ve.account_code IN ('4103','3103')
                    AND EXISTS (
                        SELECT 1 FROM voucher_entries ve2
                        WHERE ve2.voucher_id = ve.voucher_id
                        AND ve2.account_code >= '6001' AND ve2.account_code < '7000'
                    )
                )
                GROUP BY e.account_code""", (self.client_id, p_start, p_end))
            return {r[0]: r[1] or 0 for r in c.fetchall()}

        # 本期：start_period ~ end_period
        cur = fetch_period(start_period, end_period)
        year = end_period[:4]
        year_start = f"{year}-01"
        # 本年累计：本年1月 ~ end_period（与start_period无关，始终从年初算起）
        ytd = fetch_period(year_start, end_period)
        conn.close()
        self._log(f"_load_income: cur entries={len(cur)}, ytd entries={len(ytd)}")

        def g(codes, d=None):
            """Sum credit-minus-debit net for all accounts matching any prefix in codes list."""
            if d is None: d = cur
            if isinstance(codes, str): codes = [codes]
            total = 0
            for acct_code, val in d.items():
                for code in codes:
                    if acct_code == code or acct_code.startswith(code+".") or acct_code.startswith(code+"_"):
                        total += val
                        break
            return total
        def gy(codes): return g(codes, ytd)

        if use_6xxx:
            # ── 企业会计准则 ──
            rev      = g(["6001","6051"])          # 营业收入
            cost_n   = -g(["6401","6402"])         # 营业成本
            tax      = -g(["6403"])                # 税金及附加
            sell     = -g(["6601"])                # 销售费用
            mgmt     = -g(["6602"])                # 管理费用
            rnd      = -g(["6604"])                # 研发费用
            fin_exp  = -g(["6603"])                # 财务费用（费用化列示）
            impair   = -g(["6701"])                # 资产减值损失
            oth_inc  = g(["6117"])                 # 其他收益
            inv_g    = g(["6111"])                 # 投资收益
            fv_g     = g(["6101"])                 # 公允价值变动收益
            asset_d  = g(["6115"])                 # 资产处置收益
            op_profit = (rev - cost_n - tax - sell - mgmt - rnd - fin_exp - impair
                         + oth_inc + inv_g + fv_g + asset_d)
            nop_inc   = g(["6301"])
            nop_exp   = -g(["6711"])
            tax_exp   = -g(["6801"])
            total_profit = op_profit + nop_inc + nop_exp
            net_profit   = total_profit - tax_exp
            # ── 本年累计 ──
            rev_y     = gy(["6001","6051"])
            cost_y    = -gy(["6401","6402"])
            tax_y_al  = -gy(["6403"])
            sell_y    = -gy(["6601"])
            mgmt_y    = -gy(["6602"])
            rnd_y     = -gy(["6604"])
            fin_exp_y = -gy(["6603"])
            impair_y  = -gy(["6701"])
            oth_inc_y = gy(["6117"])
            inv_y     = gy(["6111"])
            fv_y      = gy(["6101"])
            asset_d_y = gy(["6115"])
            op_y      = (rev_y - cost_y - tax_y_al - sell_y - mgmt_y - rnd_y
                         - fin_exp_y - impair_y + oth_inc_y + inv_y + fv_y + asset_d_y)
            nop_y     = gy(["6301"])
            nopx_y    = -gy(["6711"])
            tax_y     = -gy(["6801"])
            total_y   = op_y + nop_y + nopx_y
            net_y     = total_y - tax_y
        else:
            # ── 小企业会计制度 ──
            rev      = g(["5001","5051"])
            cost_n   = -g(["5401","5402"])
            tax      = -g(["5403"])
            sell     = -g(["5501"])
            mgmt     = -g(["5502"])
            rnd      = 0
            fin_exp  = -g(["5503"])
            impair   = 0
            oth_inc  = 0
            inv_g    = g(["5111"])
            fv_g     = 0
            asset_d  = 0
            op_profit = rev - cost_n - tax - sell - mgmt - fin_exp + inv_g
            nop_inc   = g(["5301"])
            nop_exp   = -g(["5601"])
            tax_exp   = -g(["5701"])
            total_profit = op_profit + nop_inc + nop_exp
            net_profit   = total_profit - tax_exp
            # ── 本年累计 ──
            rev_y     = gy(["5001","5051"])
            cost_y    = -gy(["5401","5402"])
            tax_y_al  = -gy(["5403"])
            sell_y    = -gy(["5501"])
            mgmt_y    = -gy(["5502"])
            rnd_y     = 0
            fin_exp_y = -gy(["5503"])
            impair_y  = 0
            oth_inc_y = 0
            inv_y     = gy(["5111"])
            fv_y      = 0
            asset_d_y = 0
            op_y      = rev_y - cost_y - tax_y_al - sell_y - mgmt_y - fin_exp_y + inv_y
            nop_y     = gy(["5301"])
            nopx_y    = -gy(["5601"])
            tax_y     = -gy(["5701"])
            total_y   = op_y + nop_y + nopx_y
            net_y     = total_y - tax_y

        # ── 取数公式字符串（按制度区分）──
        if use_6xxx:
            _f_rev    = "贷方净额 | 6001 主营业务收入 + 6051 其他业务收入"
            _f_cost   = "借方净额（取负）| 6401 主营业务成本 + 6402 其他业务成本"
            _f_tax    = "借方净额（取负）| 6403 税金及附加"
            _f_sell   = "借方净额（取负）| 6601 销售费用"
            _f_mgmt   = "借方净额（取负）| 6602 管理费用"
            _f_rnd    = "借方净额（取负）| 6604 研发费用"
            _f_fin    = "借方净额（取负）| 6603 财务费用"
            _f_imp    = "借方净额（取负）| 6701 资产减值损失"
            _f_oinc   = "贷方净额 | 6117 其他收益"
            _f_inv    = "贷方净额 | 6111 投资收益"
            _f_fv     = "贷方净额 | 6101 公允价值变动损益"
            _f_adisp  = "贷方净额 | 6115 资产处置损益"
            _f_noi    = "贷方净额 | 6301 营业外收入"
            _f_noe    = "借方净额（取负）| 6711 营业外支出"
            _f_te     = "借方净额（取负）| 6801 所得税费用"
        else:
            _f_rev    = "贷方净额 | 5001 主营业务收入 + 5051 其他业务收入"
            _f_cost   = "借方净额（取负）| 5401 主营业务成本 + 5402 其他业务成本"
            _f_tax    = "借方净额（取负）| 5403 税金及附加"
            _f_sell   = "借方净额（取负）| 5501 销售费用"
            _f_mgmt   = "借方净额（取负）| 5502 管理费用"
            _f_rnd    = ""
            _f_fin    = "借方净额（取负）| 5503 财务费用"
            _f_imp    = ""
            _f_oinc   = ""
            _f_inv    = "贷方净额 | 5111 投资收益"
            _f_fv     = ""
            _f_adisp  = ""
            _f_noi    = "贷方净额 | 5301 营业外收入"
            _f_noe    = "借方净额（取负）| 5601 营业外支出"
            _f_te     = "借方净额（取负）| 5701 所得税费用"
        _f_op  = "营业收入 + 其他收益 + 投资收益 ± 公允价值变动 ± 资产处置\n− 营业成本 − 税金 − 三费 − 资产减值"
        _f_tp  = "营业利润 + 营业外收入 − 营业外支出"
        _f_np  = "利润总额 − 所得税费用"

        rows_data = [
            ("一、营业收入",                                   "1",  rev,          rev_y,      True,  _f_rev),
            ("减：营业成本",                                   "2",  cost_n,       cost_y,     False, _f_cost),
            ("    税金及附加",                                 "3",  tax,          tax_y_al,   False, _f_tax),
            ("    销售费用",                                   "4",  sell,         sell_y,     False, _f_sell),
            ("    管理费用",                                   "5",  mgmt,         mgmt_y,     False, _f_mgmt),
            ("    研发费用",                                   "6",  rnd,          rnd_y,      False, _f_rnd),
            ("    财务费用",                                   "7",  fin_exp,      fin_exp_y,  False, _f_fin),
            ("    其中：利息费用",                             "8",  0,            0,          False, ""),
            ("          利息收入",                             "9",  0,            0,          False, ""),
            ("    资产减值损失",                               "10", impair,       impair_y,   False, _f_imp),
            ("加：其他收益",                                   "11", oth_inc,      oth_inc_y,  False, _f_oinc),
            ("    投资收益（损失以\"-\"号填列）",              "12", inv_g,        inv_y,      False, _f_inv),
            ("    其中：对联营企业和合营企业的投资收益",        "13", 0,            0,          False, ""),
            ("    公允价值变动收益（损失以\"-\"号填列）",      "14", fv_g,         fv_y,       False, _f_fv),
            ("    资产处置收益（损失以\"-\"号填列）",          "15", asset_d,      asset_d_y,  False, _f_adisp),
            ("二、营业利润（亏损以\"-\"号填列）",              "16", op_profit,    op_y,       True,  _f_op),
            ("加：营业外收入",                                 "17", nop_inc,      nop_y,      False, _f_noi),
            ("减：营业外支出",                                 "18", nop_exp,      nopx_y,     False, _f_noe),
            ("三、利润总额（亏损总额以\"-\"号填列）",          "19", total_profit, total_y,    True,  _f_tp),
            ("减：所得税费用",                                 "20", tax_exp,      tax_y,      False, _f_te),
            ("四、净利润（净亏损以\"-\"号填列）",              "21", net_profit,   net_y,      True,  _f_np),
            ("（一）持续经营净利润（净亏损以\"-\"号填列）",    "22", net_profit,   net_y,      False, _f_np),
            ("（二）终止经营净利润（净亏损以\"-\"号填列）",    "23", 0,            0,          False, ""),
            ("五、其他综合收益的税后净额",                     "24", 0,            0,          True,  ""),
            ("（一）以后不能重分类进损益的其他综合收益",       "25", 0,            0,          False, ""),
            ("    1.重新计量设定受益计划净负债或净资产的变动", "26", 0,            0,          False, ""),
            ("    2.权益法下在被投资单位不能重分类进损益的\n其他综合收益中享有的份额", "27", 0, 0, False, ""),
            ("（二）以后将重分类进损益的其他综合收益",         "28", 0,            0,          False, ""),
            ("    1.权益法下在被投资单位以后将重分类进损益的\n其他综合收益中享有的份额", "29", 0, 0, False, ""),
            ("    2.可供出售金融资产公允价值变动损益",         "30", 0,            0,          False, ""),
            ("    3.持有至到期投资重分类为可供出售金融资产损益","31", 0,            0,          False, ""),
            ("    4.现金流量套期损益的有效部分",               "32", 0,            0,          False, ""),
            ("    5.外币财务报表折算差额",                     "33", 0,            0,          False, ""),
            ("六、综合收益总额",                               "34", net_profit,   net_y,      True,  _f_np),
            ("七、每股收益",                                   "35", "",           "",         True,  ""),
            ("（一）基本每股收益",                             "36", 0,            0,          False, ""),
            ("（二）稀释每股收益",                             "37", 0,            0,          False, ""),
        ]

        self.inc_tbl.setRowCount(len(rows_data))
        self._log(f"_load_income: rendering {len(rows_data)} rows, rev={rev:,.2f}, cost={cost_n:,.2f}, op_profit={op_profit:,.2f}, net={net_profit:,.2f}")
        for i,row_item in enumerate(rows_data):
            name = row_item[0]; rowno = row_item[1]
            cur_v = row_item[2]; ytd_v = row_item[3]
            is_key = row_item[4] if len(row_item)>4 else False
            formula = row_item[5] if len(row_item)>5 else ""
            self.inc_tbl.setRowHeight(i, 34)
            bg = QColor("#f0f4ff") if is_key else None
            for j,v in enumerate([name, str(rowno) if rowno else "",
                                   fmt_amt(cur_v) if isinstance(cur_v,(int,float)) else "",
                                   fmt_amt(ytd_v) if isinstance(ytd_v,(int,float)) else ""]):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter if j==0 else Qt.AlignmentFlag.AlignCenter if j==1 else Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter)
                if is_key:
                    it.setFont(QFont("",weight=QFont.Bold))
                    if bg: it.setBackground(bg)
                if j>=2 and isinstance(cur_v,(int,float)):
                    val = cur_v if j==2 else ytd_v
                    if val and val < 0: it.setForeground(QColor("#ff4d4f"))
                # 名称列存公式
                if j == 0 and formula:
                    it.setData(Qt.ItemDataRole.UserRole, (name.strip(), formula))
                    it.setToolTip("点击查看取数公式")
                self.inc_tbl.setItem(i,j,it)


    def _build_equity(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(8)
        title_row = QHBoxLayout()
        title_row.addWidget(lbl("所有者权益变动表", bold=True, size=15))
        title_row.addStretch()
        title_row.addWidget(lbl("（企业会计准则格式）", color="#888", size=12))
        L.addLayout(title_row)
        L.addWidget(lbl("单位：元", color="#aaa", size=11))

        self.eq_tbl = self._make_report_table(
            ["项目",
             "实收资本(股本)",
             "资本公积",
             "其他综合收益",
             "盈余公积",
             "未分配利润",
             "合计"],
            [-1, 110, 110, 100, 100, 110, 110]
        )
        self.eq_tbl.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.eq_tbl.setWordWrap(True)
        # col 0 保持 Interactive（已在 _make_report_table 中统一设置，不再覆盖为 Stretch）
        L.addWidget(self.eq_tbl)
        self.stack.addWidget(w)

    def _load_equity(self):
        if not self.client_id: return
        end_period = self.rep_end_period.currentData()
        if not end_period: return
        conn = get_db(); c = conn.cursor()
        year = end_period[:4]

        # Fetch year-to-date balances from voucher entries (approved only)
        c.execute("""SELECT e.account_code, SUM(e.debit)-SUM(e.credit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period LIKE ? AND v.status='已审核'
            GROUP BY e.account_code""", (self.client_id, f"{year}%"))
        ytd = {r[0]: -(r[1] or 0) for r in c.fetchall()}  # credit-normal for equity

        # Opening balances from accounts table
        c.execute("SELECT code, opening_credit FROM accounts WHERE client_id=?", (self.client_id,))
        opening = {r[0]: r[1] or 0 for r in c.fetchall()}
        conn.close()

        def op(code):
            return opening.get(code, 0)
        def mv(code):
            return ytd.get(code, 0)

        use_4xxx = getattr(self, '_acct_std', '企业会计准则') == '企业会计准则'
        if use_4xxx:
            cap_op  = op("4001"); cap_mv  = mv("4001")
            cprs_op = op("4002"); cprs_mv = mv("4002")
            surp_op = op("4101"); surp_mv = mv("4101")
            re_op   = op("4103") + op("4104")
            re_mv   = mv("4103") + mv("4104")
        else:
            cap_op  = op("3001"); cap_mv  = mv("3001")
            cprs_op = op("3002"); cprs_mv = mv("3002")
            surp_op = op("3101"); surp_mv = mv("3101")
            re_op   = op("3103") + op("3104")
            re_mv   = mv("3103") + mv("3104")
        oci_op = 0; oci_mv = 0   # 其他综合收益（暂无专用科目）

        def row_data(label, c1, c2, c3, c4, c5, bold=False, bg=None):
            total = c1+c2+c3+c4+c5
            return (label, c1, c2, c3, c4, c5, total, bold, bg)

        rows = [
            row_data("一、上年年末余额",    cap_op,  cprs_op, oci_op,  surp_op, re_op,  bold=True,  bg="#f0f4ff"),
            row_data("  加：会计政策变更",  0, 0, 0, 0, 0),
            row_data("     前期差错更正",   0, 0, 0, 0, 0),
            row_data("二、本年年初余额",    cap_op,  cprs_op, oci_op,  surp_op, re_op,  bold=True,  bg="#f0f4ff"),
            row_data("三、本年增减变动",    cap_mv,  cprs_mv, oci_mv,  surp_mv, re_mv,  bold=True,  bg="#fafafa"),
            row_data("  (一)综合收益总额",  0,       0,       oci_mv,  0,       re_mv),
            row_data("  (二)所有者投入",    cap_mv,  cprs_mv, 0,       0,       0),
            row_data("  (三)利润分配",      0,       0,       0,       surp_mv, re_mv - re_mv),
            row_data("四、本年年末余额",
                     cap_op+cap_mv, cprs_op+cprs_mv, oci_op+oci_mv,
                     surp_op+surp_mv, re_op+re_mv,    bold=True, bg="#e6f0ff"),
        ]

        self.eq_tbl.setRowCount(len(rows))
        for i, (label,c1,c2,c3,c4,c5,total,bold,bg) in enumerate(rows):
            self.eq_tbl.setRowHeight(i, 38)
            vals = [label, c1, c2, c3, c4, c5, total]
            for j, v in enumerate(vals):
                text = v if j == 0 else (fmt_amt(v) if v else "")
                it = QTableWidgetItem(text)
                it.setTextAlignment(Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter if j==0 else Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter)
                if bold: it.setFont(QFont("", weight=QFont.Bold))
                if bg:   it.setBackground(QColor(bg))
                if j > 0 and isinstance(v, float) and v < 0:
                    it.setForeground(QColor("#e05252"))
                self.eq_tbl.setItem(i, j, it)

    def _build_cf_stmt(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(8)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("现金流量表", bold=True, size=15)); hdr.addStretch()
        b_dl = QPushButton("导出Excel"); b_dl.setObjectName("btn_outline")
        b_dl.clicked.connect(self._export_cf_stmt)
        hdr.addWidget(b_dl); L.addLayout(hdr)
        L.addWidget(lbl("（采用直接法，现金及现金等价物 = 库存现金+银行存款+其他货币资金）",
                         color="#888", size=12))
        self.cf_stmt_tbl = self._make_report_table(
            ["项目", "行次", "本期金额", "本年累计金额"],
            [-1, 40, 140, 140])
        L.addWidget(self.cf_stmt_tbl)
        self.stack.addWidget(w)

    def _get_cash_balance(self, c, client_id, period_end):
        """期末现金余额 = 期初 + 本年至今净发生额"""
        # Opening balance from accounts
        c.execute("""SELECT SUM(opening_debit - opening_credit) FROM accounts
            WHERE client_id=? AND (code='1001' OR code LIKE '1001.%' OR code LIKE '1001_%'
              OR code='1002' OR code LIKE '1002.%' OR code LIKE '1002_%'
              OR code='1012' OR code LIKE '1012.%' OR code LIKE '1012_%')""",
            (client_id,))
        opening = c.fetchone()[0] or 0
        if not period_end:
            return opening
        year = period_end[:4]
        c.execute("""SELECT SUM(e.debit) - SUM(e.credit) FROM voucher_entries e
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period<=? AND v.period LIKE ? AND v.status='已审核'
            AND (e.account_code='1001' OR e.account_code LIKE '1001.%' OR e.account_code LIKE '1001_%'
              OR e.account_code='1002' OR e.account_code LIKE '1002.%' OR e.account_code LIKE '1002_%'
              OR e.account_code='1012' OR e.account_code LIKE '1012.%' OR e.account_code LIKE '1012_%')""",
            (client_id, period_end, f"{year}%"))
        ytd_net = c.fetchone()[0] or 0
        return opening + ytd_net

    def _compute_cf(self, c, client_id, start_period, end_period):
        """
        Compute cash flow by analyzing cash account counterparts in vouchers.
        Returns dict: row_key -> amount (positive = inflow, negative = outflow shown as positive)
        Two dicts returned: current_period and ytd.
        """
        year = end_period[:4]

        def _analyze(p_start, p_end):
            """Analyze cash flows for a period range."""
            # Get all voucher IDs with cash account entries in range
            c.execute("""SELECT DISTINCT v.id FROM vouchers v
                JOIN voucher_entries e ON e.voucher_id=v.id
                WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
                AND (e.account_code='1001' OR e.account_code LIKE '1001.%' OR e.account_code LIKE '1001_%'
                  OR e.account_code='1002' OR e.account_code LIKE '1002.%' OR e.account_code LIKE '1002_%'
                  OR e.account_code='1012' OR e.account_code LIKE '1012.%' OR e.account_code LIKE '1012_%')""",
                (client_id, p_start, p_end))
            vids = [r[0] for r in c.fetchall()]

            rows = {}  # row_number -> amount

            def add(key, amt):
                rows[key] = rows.get(key, 0) + amt

            for vid in vids:
                c.execute("SELECT account_code, debit, credit FROM voucher_entries WHERE voucher_id=?", (vid,))
                entries = c.fetchall()

                cash_in = 0; cash_out = 0
                non_cash = []
                for e in entries:
                    code = e[0] or ""
                    d = e[1] or 0; cr = e[2] or 0
                    if (code == '1001' or code.startswith('1001.') or code.startswith('1001_') or
                        code == '1002' or code.startswith('1002.') or code.startswith('1002_') or
                        code == '1012' or code.startswith('1012.') or code.startswith('1012_')):
                        cash_in += d; cash_out += cr
                    else:
                        non_cash.append((code, d, cr))

                # Classify inflows (cash debited)
                if cash_in > 0:
                    for code, d, cr in non_cash:
                        amt = cr  # credit side = source of cash
                        if amt <= 0: continue
                        p = code[:4]
                        # Revenue accounts → 销售商品收到现金
                        if (code.startswith('6001') or code.startswith('6002') or
                            code.startswith('6051') or code.startswith('5001') or
                            code.startswith('5051') or code.startswith('1122') or
                            code.startswith('2203')):
                            add('r1', amt)
                        elif code.startswith('2221') or code.startswith('1321'):
                            add('r2', amt)  # 税费返还
                        elif code.startswith('6301') or code.startswith('5301'):
                            add('r3', amt)  # 其他经营收入
                        elif (code.startswith('6111') or code.startswith('5111') or
                              code.startswith('1511') or code.startswith('1521') or
                              code.startswith('1131') or code.startswith('1132')):
                            add('r12', amt)  # 取得投资收益
                        elif code.startswith('1601') or code.startswith('1604'):
                            add('r13', amt)  # 处置固定资产
                        elif code.startswith('2001') or code.startswith('2501'):
                            add('r24', amt)  # 取得借款
                        elif code.startswith('3001') or code.startswith('4001'):
                            add('r23', amt)  # 吸收投资
                        else:
                            add('r3', amt)   # 其他经营收入

                # Classify outflows (cash credited)
                if cash_out > 0:
                    for code, d, cr in non_cash:
                        amt = d  # debit side = destination of cash
                        if amt <= 0: continue
                        if (code.startswith('1403') or code.startswith('1401') or
                            code.startswith('1405') or code.startswith('6401') or
                            code.startswith('6402') or code.startswith('5401') or
                            code.startswith('5402') or code.startswith('2202') or
                            code.startswith('1221')):
                            add('r5', amt)   # 购买商品
                        elif code.startswith('2211'):
                            add('r6', amt)   # 支付员工
                        elif code.startswith('2221') or code.startswith('2231'):
                            add('r7', amt)   # 支付税费
                        elif (code.startswith('6601') or code.startswith('6602') or
                              code.startswith('6603') or code.startswith('5501') or
                              code.startswith('5502') or code.startswith('5503') or
                              code.startswith('2241') or code.startswith('1461')):
                            add('r8', amt)   # 其他经营支出
                        elif (code.startswith('1601') or code.startswith('1604') or
                              code.startswith('1605') or code.startswith('1701')):
                            add('r17', amt)  # 购建固定资产
                        elif (code.startswith('1801') or code.startswith('1511') or
                              code.startswith('1521') or code.startswith('1531')):
                            add('r18', amt)  # 投资支出
                        elif code.startswith('2001') or code.startswith('2501'):
                            add('r27', amt)  # 偿还借款
                        elif (code.startswith('3104') or code.startswith('4104') or
                              code.startswith('2232')):
                            add('r28', amt)  # 分配股利
                        else:
                            add('r8', amt)   # 其他经营支出

            return rows

        cur = _analyze(start_period, end_period)
        ytd = _analyze(f"{year}-01", end_period)
        return cur, ytd

    def _load_cf_stmt(self):
        if not self.client_id: return
        start_period = self.rep_start_period.currentData()
        end_period   = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        year = end_period[:4]

        conn = get_db(); c = conn.cursor()

        # Cash balances
        cash_end  = self._get_cash_balance(c, self.client_id, end_period)
        cash_beg  = self._get_cash_balance(c, self.client_id,
                        f"{year}-01" if start_period[:4] == year else start_period)
        cash_open = self._get_cash_balance(c, self.client_id, None)  # opening from accounts

        # Compute cash flow amounts
        cur, ytd = self._compute_cf(c, self.client_id, start_period, end_period)

        # Subtotals
        def g(d, *keys): return sum(d.get(k, 0) for k in keys)

        # Current period
        ci  = g(cur,'r1','r2','r3')       # 经营流入
        co  = g(cur,'r5','r6','r7','r8')  # 经营流出
        cn  = ci - co                      # 经营净额
        ii  = g(cur,'r11','r12','r13','r14','r15')
        io_ = g(cur,'r17','r18','r19','r20')
        inv_n = ii - io_
        fi  = g(cur,'r23','r24','r25')
        fo  = g(cur,'r27','r28','r29')
        fin_n = fi - fo
        net_cur = cn + inv_n + fin_n

        # YTD
        ci_y  = g(ytd,'r1','r2','r3')
        co_y  = g(ytd,'r5','r6','r7','r8')
        cn_y  = ci_y - co_y
        ii_y  = g(ytd,'r11','r12','r13','r14','r15')
        io_y  = g(ytd,'r17','r18','r19','r20')
        inv_ny = ii_y - io_y
        fi_y  = g(ytd,'r23','r24','r25')
        fo_y  = g(ytd,'r27','r28','r29')
        fin_ny = fi_y - fo_y
        net_ytd = cn_y + inv_ny + fin_ny

        # Net profit for supplementary - exclude carryforward vouchers
        c.execute("""SELECT e.account_code, SUM(e.credit)-SUM(e.debit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
            AND (v.note IS NULL OR v.note NOT IN ('结转收入','结转费用'))
            GROUP BY e.account_code""", (self.client_id, f"{year}-01", end_period))
        mv_ytd = {r[0]: r[1] or 0 for r in c.fetchall()}
        c.execute("""SELECT e.account_code, SUM(e.credit)-SUM(e.debit) net
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
            AND (v.note IS NULL OR v.note NOT IN ('结转收入','结转费用'))
            GROUP BY e.account_code""", (self.client_id, start_period, end_period))
        mv_cur = {r[0]: r[1] or 0 for r in c.fetchall()}
        conn.close()

        def net_profit(mv):
            use6 = any(k.startswith('6') for k in mv)
            if use6:
                rev  = sum(v for k,v in mv.items() if k[:4]<'6400' and k[0]=='6')
                cost = -sum(v for k,v in mv.items() if k[:4]>='6400' and k[0]=='6')
                return rev + cost
            else:
                rev  = sum(v for k,v in mv.items() if k[0]=='5' and k[:4]<'5400')
                cost = -sum(v for k,v in mv.items() if k[0]=='5' and k[:4]>='5400')
                return rev + cost

        np_cur = net_profit(mv_cur)
        np_ytd = net_profit(mv_ytd)

        # AR/AP changes for supplementary (ytd)
        def bal_chg(mv, codes):
            total = 0
            for k, v in mv.items():
                for code in codes:
                    if k == code or k.startswith(code+'.') or k.startswith(code+'_'):
                        total += v; break
            return total
        ar_chg  = -bal_chg(mv_ytd, ['1122','1123','1131','1132','1221'])
        ap_chg  =  bal_chg(mv_ytd, ['2202','2203','2211','2241'])

        # ── Build table rows ──
        BOLD_BG = "#f0f4ff"; HDR_BG = "#e6ecf8"

        def R(label, rowno, cur_val, ytd_val, style="normal"):
            return (label, str(rowno) if rowno else "", cur_val, ytd_val, style)

        rows = [
            # ── 经营活动 ──
            R("一、经营活动产生的现金流量：",  "", None, None, "header"),
            R("  销售商品、提供劳务收到的现金","1", cur.get('r1',0), ytd.get('r1',0)),
            R("  收到的税费返还",              "2", cur.get('r2',0), ytd.get('r2',0)),
            R("  收到的其他与经营活动有关的现金","3",cur.get('r3',0),ytd.get('r3',0)),
            R("  经营活动现金流入小计",         "4", ci,   ci_y,  "subtotal"),
            R("  购买商品、接受劳务支付的现金", "5", cur.get('r5',0), ytd.get('r5',0)),
            R("  支付给职工以及为职工支付的现金","6",cur.get('r6',0),ytd.get('r6',0)),
            R("  支付的各项税费",               "7", cur.get('r7',0), ytd.get('r7',0)),
            R("  支付的其他与经营活动有关的现金","8",cur.get('r8',0),ytd.get('r8',0)),
            R("  经营活动现金流出小计",          "9", co,   co_y,  "subtotal"),
            R("  经营活动产生的现金流量净额",   "10", cn,   cn_y,  "total"),
            # ── 投资活动 ──
            R("二、投资活动产生的现金流量：",   "", None, None,   "header"),
            R("  收回投资收到的现金",           "11", cur.get('r11',0), ytd.get('r11',0)),
            R("  取得投资收益收到的现金",        "12", cur.get('r12',0), ytd.get('r12',0)),
            R("  处置固定资产收回的现金净额",   "13", cur.get('r13',0), ytd.get('r13',0)),
            R("  处置子公司收到的现金净额",     "14", cur.get('r14',0), ytd.get('r14',0)),
            R("  收到的其他与投资活动有关的现金","15",cur.get('r15',0),ytd.get('r15',0)),
            R("  投资活动现金流入小计",         "16", ii,   ii_y,  "subtotal"),
            R("  购建固定资产支付的现金",        "17", cur.get('r17',0), ytd.get('r17',0)),
            R("  投资支付的现金",               "18", cur.get('r18',0), ytd.get('r18',0)),
            R("  取得子公司支付的现金净额",     "19", cur.get('r19',0), ytd.get('r19',0)),
            R("  支付的其他与投资活动有关的现金","20",cur.get('r20',0),ytd.get('r20',0)),
            R("  投资活动现金流出小计",         "21", io_,  io_y,  "subtotal"),
            R("  投资活动产生的现金流量净额",   "22", inv_n,inv_ny,"total"),
            # ── 筹资活动 ──
            R("三、筹资活动产生的现金流量：",   "", None, None,   "header"),
            R("  吸收投资收到的现金",           "23", cur.get('r23',0), ytd.get('r23',0)),
            R("  取得借款收到的现金",           "24", cur.get('r24',0), ytd.get('r24',0)),
            R("  收到的其他与筹资活动有关的现金","25",cur.get('r25',0),ytd.get('r25',0)),
            R("  筹资活动现金流入小计",         "26", fi,   fi_y,  "subtotal"),
            R("  偿还债务支付的现金",           "27", cur.get('r27',0), ytd.get('r27',0)),
            R("  分配股利或偿付利息支付的现金", "28", cur.get('r28',0), ytd.get('r28',0)),
            R("  支付的其他与筹资活动有关的现金","29",cur.get('r29',0),ytd.get('r29',0)),
            R("  筹资活动现金流出小计",         "30", fo,   fo_y,  "subtotal"),
            R("  筹资活动产生的现金流量净额",   "31", fin_n,fin_ny,"total"),
            R("四、汇率变动对现金及现金等价物的影响","32",0,0),
            R("五、现金及现金等价物净增加额",   "33", net_cur, net_ytd, "total"),
            R("  加：期初现金及现金等价物余额", "34", cash_open, cash_open),
            R("六、期末现金及现金等价物余额",   "35", cash_end, cash_end, "total"),
            # ── 补充资料分隔 ──
            R("━━━━ 补充资料 ━━━━",            "",  None, None,  "header"),
            R("一、将净利润调节为经营活动现金流量：","", None, None, "header"),
            R("  净利润",                       "1",  np_cur, np_ytd),
            R("  加：资产减值准备",             "2",  0, 0),
            R("  固定资产折旧",                 "3",  0, 0),
            R("  无形资产摊销",                 "4",  0, 0),
            R("  长期待摊费用摊销",             "5",  0, 0),
            R("  处置固定资产损失（收益-）",    "6",  0, 0),
            R("  公允价值变动损失（收益-）",    "8",  0, 0),
            R("  财务费用（收益-）",            "9",  0, 0),
            R("  投资损失（收益-）",            "10", 0, 0),
            R("  经营性应收项目的减少（增加-）","14", 0, ar_chg),
            R("  经营性应付项目的增加（减少-）","15", 0, ap_chg),
            R("  其他",                         "16", 0, 0),
            R("  经营活动产生的现金流量净额",   "17", cn, cn_y, "total"),
            R("三、现金及现金等价物净变动情况：","", None, None, "header"),
            R("  现金的期末余额",               "21", cash_end, cash_end),
            R("  减：现金的期初余额",           "22", cash_open, cash_open),
            R("  现金及现金等价物净增加额",     "25", cash_end - cash_open, cash_end - cash_open, "total"),
        ]

        self.cf_stmt_tbl.setRowCount(len(rows))
        for i, (label, rowno, cur_v, ytd_v, style) in enumerate(rows):
            self.cf_stmt_tbl.setRowHeight(i, 32)
            is_hdr    = (style == "header")
            is_sub    = (style == "subtotal")
            is_tot    = (style == "total")
            bg = QColor(HDR_BG) if is_hdr else QColor(BOLD_BG) if is_sub or is_tot else None

            for j, (text, align) in enumerate([
                (label,  Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter),
                (rowno,  Qt.AlignmentFlag.AlignCenter),
                (fmt_amt(cur_v) if isinstance(cur_v, (int,float)) else "",
                         Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter),
                (fmt_amt(ytd_v) if isinstance(ytd_v, (int,float)) else "",
                         Qt.AlignmentFlag.AlignRight|Qt.AlignmentFlag.AlignVCenter),
            ]):
                it = QTableWidgetItem(text); it.setTextAlignment(align)
                if is_hdr:
                    it.setBackground(QColor(HDR_BG))
                    if j == 0: it.setForeground(QColor("#3d6fdb"))
                    it.setFont(QFont("", weight=QFont.Bold))
                elif is_sub or is_tot:
                    it.setBackground(QColor(BOLD_BG))
                    it.setFont(QFont("", weight=QFont.Bold))
                if j >= 2 and isinstance(cur_v if j==2 else ytd_v, (int,float)):
                    val = cur_v if j == 2 else ytd_v
                    if val and val < 0:
                        it.setForeground(QColor("#ff4d4f"))
                self.cf_stmt_tbl.setItem(i, j, it)

    def _export_cf_stmt(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        path, _ = QFileDialog.getSaveFileName(self, "保存",
            f"现金流量表_{end_period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "现金流量表"
        
        # 表头
        ws['A1'] = self.client_name
        ws['A1'].font = XFont(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:D1')
        
        ws['A2'] = "现金流量表"
        ws['A2'].font = XFont(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:D2')
        
        period_text = f"期间：{start_period} 至 {end_period}"
        ws['A3'] = period_text
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:D3')
        
        # 空行
        ws['A4'] = ""
        
        # 数据表头
        hdrs = ["项目","行次","本期金额","本年累计金额"]
        fill_hdr = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(5, ci, h)
            cell.font = XFont(bold=True, color="FFFFFF"); cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        for ri in range(self.cf_stmt_tbl.rowCount()):
            row_vals = []
            for ci in range(4):
                it = self.cf_stmt_tbl.item(ri, ci)
                row_vals.append(it.text() if it else "")
            ws.append(row_vals)
        
        ws.column_dimensions['A'].width = 45
        for col in ['B','C','D']: ws.column_dimensions[col].width = 16
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _build_cashflow(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(8)
        L.addWidget(lbl("收支统计表（本期科目发生额汇总）", bold=True, size=15))
        L.addWidget(lbl("按资产/负债/收入/费用分类展示本期所有科目的借贷发生额", color="#888", size=12))
        self.cf_tbl = self._make_report_table(
            ["科目编号","科目名称","类型","本期借方","本期贷方","净额"],
            [90,-1,70,110,110,110])
        self.cf_tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.cf_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        L.addWidget(self.cf_tbl); self.stack.addWidget(w)

    def _load_cashflow(self):
        if not self.client_id: return
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT e.account_code, e.account_name,
            SUM(e.debit) td, SUM(e.credit) tc
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period>=? AND v.period<=? AND v.status='已审核'
            AND (v.note IS NULL OR v.note NOT IN ('结转收入','结转费用'))
            GROUP BY e.account_code ORDER BY e.account_code""",
            (self.client_id, start_period, end_period))
        entries = c.fetchall()
        # Get account types
        c.execute("SELECT code,type FROM accounts WHERE client_id=?",(self.client_id,))
        acct_types = {r[0]:r[1] for r in c.fetchall()}
        conn.close()

        type_colors = {"资产":"#3d6fdb","负债":"#e05252","所有者权益":"#722ed1",
                       "成本":"#fa8c16","收入":"#52c41a","费用":"#eb5757"}

        self.cf_tbl.setRowCount(len(entries))
        td_total = tc_total = 0
        for i,r in enumerate(entries):
            self.cf_tbl.setRowHeight(i,34)
            d=r['td'] or 0; cr=r['tc'] or 0; net=d-cr
            td_total+=d; tc_total+=cr
            atype = acct_types.get(r['account_code'],"")
            tcolor = type_colors.get(atype,"#555")
            vals = [r['account_code'],r['account_name'] or "",atype,fmt_amt(d),fmt_amt(cr),fmt_amt(net)]
            for j,v in enumerate(vals):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignmentFlag.AlignCenter if j!=1 else Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter)
                if j==2: it.setForeground(QColor(tcolor))
                if j==5: it.setForeground(QColor("#3d6fdb") if net>0 else QColor("#ff4d4f") if net<0 else QColor("#888"))
                self.cf_tbl.setItem(i,j,it)
        # Add totals row
        n = len(entries)
        self.cf_tbl.setRowCount(n+1)
        self.cf_tbl.setRowHeight(n,38)
        for j,v in enumerate(["","合计","",fmt_amt(td_total),fmt_amt(tc_total),fmt_amt(td_total-tc_total)]):
            it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignmentFlag.AlignCenter if j!=1 else Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignVCenter)
            it.setFont(QFont("",weight=QFont.Bold)); it.setBackground(QColor("#f5f7fa"))
            self.cf_tbl.setItem(n,j,it)

    def set_client(self, client_id, client_name, period):
        self.client_id = client_id; self.client_name = client_name; self.period = period
        self.period_lbl.setText(f"【{client_name}】{period}")

        # ── 查询会计制度，控制 Tab 显示 ──
        conn0 = get_db(); c0 = conn0.cursor()
        c0.execute("SELECT accounting_std FROM clients WHERE id=?", (client_id,))
        std_row = c0.fetchone(); conn0.close()
        self._acct_std = (std_row["accounting_std"] if std_row else None) or "企业会计准则"
        is_small = (self._acct_std == "小企业会计制度")
        # 小企业制度不编制"所有者权益变动表"和"现金流量表"
        self._rtabs[2].setVisible(not is_small)   # 所有者权益变动表
        self._rtabs[3].setVisible(not is_small)   # 现金流量表
        # 若当前正显示这两个被隐藏的 Tab，跳回资产负债表
        if is_small and self.stack.currentIndex() in (2, 3):
            self._switch("资产负债表")
            return

        # ── 查询该客户最近有已审核凭证的期间 ──
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT period FROM vouchers
                     WHERE client_id=? AND status='已审核'
                     ORDER BY period DESC LIMIT 1""", (client_id,))
        row = c.fetchone()
        conn.close()
        latest_period = row[0] if row else period

        # ── 填充期间下拉框（blockSignals 防止 addItem 时触发查询） ──
        self.rep_start_period.blockSignals(True)
        self.rep_end_period.blockSignals(True)
        self.rep_start_period.clear()
        self.rep_end_period.clear()
        now = datetime.now()
        periods = []
        for y in range(now.year, 2018-1, -1):
            for m in range(12,0,-1):
                period_str = f"{y}-{m:02d}"
                display_str = f"{y}年{m:02d}期"
                periods.append((period_str, display_str))

        for period_str, display_str in periods:
            self.rep_start_period.addItem(display_str, period_str)
            self.rep_end_period.addItem(display_str, period_str)

        # 默认选最近有数据的期间
        for i in range(self.rep_start_period.count()):
            if self.rep_start_period.itemData(i) == latest_period:
                self.rep_start_period.setCurrentIndex(i)
                self.rep_end_period.setCurrentIndex(i)
                break

        self.rep_start_period.blockSignals(False)
        self.rep_end_period.blockSignals(False)

        idx = self.stack.currentIndex()
        if idx==0: self._load_balance()
        elif idx==1: self._load_income()
        elif idx==2: self._load_equity()
        elif idx==3: self._load_cf_stmt()
        elif idx==4: self._load_cashflow()

    # ── 辅助：获取当前激活 Tab 名称 ────────────────────────────────────────────
    def _get_current_tab(self) -> str | None:
        for b in self._rtabs:
            if b.property("active") == "true":
                return b.text()
        return None

    # ── 辅助：将当前报表渲染为 PDF（不弹对话框，供打印/导出共用）───────────────
    def _render_pdf(self, tab_name: str, path: str) -> None:
        """按 tab_name 将对应 QTableWidget 用 reportlab 渲染到 path。"""
        from print_utils import export_report_pdf
        start, end = self._get_periods()
        if not start:
            return
        _cfg: dict = {
            "资产负债表": dict(
                table_widget      = self.bs_tbl,
                col_headers       = ["资产项目","行次","期末金额","年初金额",
                                     "负债和所有者权益","行次","期末金额","年初金额"],
                col_ratios        = [0.260, 0.047, 0.140, 0.140, 0.230, 0.047, 0.068, 0.068],
                report_title      = "资产负债表",
                is_landscape      = True,
                amount_col_indices= [2, 3, 6, 7],
            ),
            "利润表": dict(
                table_widget      = self.inc_tbl,
                col_headers       = ["项目", "行次", "本期金额", "本年累计金额"],
                col_ratios        = [0.478, 0.067, 0.228, 0.227],
                report_title      = "利润表",
                is_landscape      = False,
                amount_col_indices= [2, 3],
            ),
            "所有者权益变动表": dict(
                table_widget      = self.eq_tbl,
                col_headers       = ["项目","实收资本(股本)","资本公积",
                                     "其他综合收益","盈余公积","未分配利润","合计"],
                col_ratios        = [0.247, 0.126, 0.126, 0.126, 0.126, 0.126, 0.123],
                report_title      = "所有者权益变动表",
                is_landscape      = True,
                amount_col_indices= [1, 2, 3, 4, 5, 6],
            ),
            "现金流量表": dict(
                table_widget      = self.cf_stmt_tbl,
                col_headers       = ["项目", "行次", "本期金额", "本年累计金额"],
                col_ratios        = [0.478, 0.067, 0.228, 0.227],
                report_title      = "现金流量表",
                is_landscape      = False,
                amount_col_indices= [2, 3],
            ),
            "收支统计表": dict(
                table_widget      = self.cf_tbl,
                col_headers       = ["科目编号","科目名称","类型","本期借方","本期贷方","净额"],
                col_ratios        = [0.124, 0.373, 0.101, 0.134, 0.134, 0.134],
                report_title      = "收支统计表",
                is_landscape      = False,
                amount_col_indices= [3, 4, 5],
            ),
        }
        cfg = _cfg.get(tab_name)
        if not cfg:
            return
        export_report_pdf(
            path         = path,
            company_name = self.client_name,
            period_text  = f"{start} 至 {end}",
            **cfg,
        )

    # ── 打印辅助：将 QTableWidget 转为打印用 HTML ─────────────────────────────
    def _build_print_html(self, tab: str) -> str:
        tab_map = {
            "资产负债表":      (self.bs_tbl,      ["资产项目","行次","期末金额","年初金额",
                                                   "负债和所有者权益","行次","期末金额","年初金额"]),
            "利润表":          (self.inc_tbl,     ["项目","行次","本期金额","本年累计金额"]),
            "所有者权益变动表": (self.eq_tbl,      ["项目","实收资本(股本)","资本公积",
                                                   "其他综合收益","盈余公积","未分配利润","合计"]),
            "现金流量表":      (self.cf_stmt_tbl, ["项目","行次","本期金额","本年累计金额"]),
            "收支统计表":      (self.cf_tbl,      ["科目编号","科目名称","类型",
                                                   "本期借方","本期贷方","净额"]),
        }
        tbl, headers = tab_map.get(tab, (None, []))
        if tbl is None:
            return ""
        start, end = self._get_periods()
        period_text = f"{start} 至 {end}" if start else ""
        amount_cols = {
            "资产负债表":      {2, 3, 6, 7},
            "利润表":          {2, 3},
            "所有者权益变动表": {1, 2, 3, 4, 5, 6},
            "现金流量表":      {2, 3},
            "收支统计表":      {3, 4, 5},
        }.get(tab, set())

        rows_html = []
        for ri in range(tbl.rowCount()):
            row_bg = ""; is_sec = False; bold = False
            item0 = tbl.item(ri, 0)
            if item0:
                brush = item0.background()
                if brush.style() != Qt.BrushStyle.NoBrush:
                    c = brush.color()
                    r, g, b = c.red(), c.green(), c.blue()
                    if b > 240 and b > r + 5:          # 蓝调 → section header
                        row_bg = "#f0f4ff"; is_sec = True
                    elif r > 240 and g > 240 and b > 240:  # 浅灰 → 合计行
                        row_bg = "#f5f7fa"
                f = item0.font()
                bold = f.bold() or f.weight() >= 63

            cells = []
            for ci in range(tbl.columnCount()):
                item = tbl.item(ri, ci)
                text = (item.text() if item else "").replace("&", "&amp;").replace("<", "&lt;")
                align = "right" if ci in amount_cols else ("center" if ci == 1 else "left")
                s = f"padding:3px 5px;border:1px solid #ddd;text-align:{align};"
                if bold:
                    s += "font-weight:bold;"
                if is_sec and ci == 0:
                    s += "color:#3d6fdb;"
                if item and ci in amount_cols:
                    fg = item.foreground().color()
                    if fg.isValid() and fg.red() > 180 and fg.green() < 100:
                        s += "color:#e05252;"
                cells.append(f'<td style="{s}">{text}</td>')

            tr_bg = f'style="background:{row_bg};"' if row_bg else ""
            rows_html.append(f"<tr {tr_bg}>{''.join(cells)}</tr>")

        hcells = "".join(
            f'<th style="padding:5px;background:#1c2340;color:#fff;'
            f'text-align:center;border:1px solid #1c2340;">{h}</th>'
            for h in headers
        )
        return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
  body{{font-family:"Microsoft YaHei","PingFang SC",sans-serif;font-size:10px;margin:8px;}}
  h2{{text-align:center;margin:3px 0;font-size:14px;}}
  h3{{text-align:center;margin:2px 0;font-size:12px;color:#333;}}
  p{{text-align:center;color:#777;font-size:9px;margin:2px 0 6px;}}
  table{{border-collapse:collapse;width:100%;}}
</style></head><body>
<h2>{self.client_name}</h2><h3>{tab}</h3>
<p>期间：{period_text}</p>
<table><thead><tr>{hcells}</tr></thead>
<tbody>{''.join(rows_html)}</tbody></table>
</body></html>"""

    # ── 打印：一体化打印窗口（左侧设置 + 右侧实时预览）────────────────────────
    def _print_report(self):
        if not self.client_id:
            return
        tab = self._get_current_tab()
        if not tab:
            return
        html = self._build_print_html(tab)
        dlg = _PrintDialog(tab, html, self)
        dlg.exec()

    def _export(self):
        if not self.client_id: return
        current_tab = self._get_current_tab()
        if not current_tab: return

        # Call corresponding export method
        if current_tab == "资产负债表":
            self._export_balance()
        elif current_tab == "利润表":
            self._export_income()
        elif current_tab == "所有者权益变动表":
            self._export_equity()
        elif current_tab == "现金流量表":
            self._export_cf_stmt()
        elif current_tab == "收支统计表":
            self._export_cashflow()

    def _export_balance(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        path, _ = QFileDialog.getSaveFileName(self, "保存",
            f"资产负债表_{end_period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "资产负债表"
        
        # 表头
        ws['A1'] = self.client_name
        ws['A1'].font = XFont(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = "资产负债表"
        ws['A2'].font = XFont(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:H2')
        
        period_text = f"期间：{start_period} 至 {end_period}"
        ws['A3'] = period_text
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:H3')
        
        # 空行
        ws['A4'] = ""
        
        # 数据表头
        hdrs = ["资产项目","行次","期末金额","年初金额","负债和所有者权益","行次","期末金额","年初金额"]
        fill_hdr = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(5, ci, h)
            cell.font = XFont(bold=True, color="FFFFFF"); cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        for ri in range(self.bs_tbl.rowCount()):
            row_vals = []
            for ci in range(8):
                it = self.bs_tbl.item(ri, ci)
                row_vals.append(it.text() if it else "")
            ws.append(row_vals)
        
        ws.column_dimensions['A'].width = 30; ws.column_dimensions['E'].width = 30
        for col in ['B','C','D','F','G','H']: ws.column_dimensions[col].width = 12
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _export_income(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        path, _ = QFileDialog.getSaveFileName(self, "保存",
            f"利润表_{end_period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "利润表"
        
        # 表头
        ws['A1'] = self.client_name
        ws['A1'].font = XFont(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:D1')
        
        ws['A2'] = "利润表"
        ws['A2'].font = XFont(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:D2')
        
        period_text = f"期间：{start_period} 至 {end_period}"
        ws['A3'] = period_text
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:D3')
        
        # 空行
        ws['A4'] = ""
        
        # 数据表头
        hdrs = ["项目","行次","本期金额","本年累计金额"]
        fill_hdr = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(5, ci, h)
            cell.font = XFont(bold=True, color="FFFFFF"); cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        for ri in range(self.inc_tbl.rowCount()):
            row_vals = []
            for ci in range(4):
                it = self.inc_tbl.item(ri, ci)
                row_vals.append(it.text() if it else "")
            ws.append(row_vals)
        
        ws.column_dimensions['A'].width = 30
        for col in ['B','C','D']: ws.column_dimensions[col].width = 16
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _export_equity(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        path, _ = QFileDialog.getSaveFileName(self, "保存",
            f"所有者权益变动表_{end_period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "所有者权益变动表"
        
        # 表头
        ws['A1'] = self.client_name
        ws['A1'].font = XFont(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:G1')
        
        ws['A2'] = "所有者权益变动表"
        ws['A2'].font = XFont(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:G2')
        
        period_text = f"期间：{start_period} 至 {end_period}"
        ws['A3'] = period_text
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:G3')
        
        # 空行
        ws['A4'] = ""
        
        # 数据表头
        hdrs = ["项目","实收资本(股本)","资本公积","其他综合收益","盈余公积","未分配利润","合计"]
        fill_hdr = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(5, ci, h)
            cell.font = XFont(bold=True, color="FFFFFF"); cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        for ri in range(self.eq_tbl.rowCount()):
            row_vals = []
            for ci in range(7):
                it = self.eq_tbl.item(ri, ci)
                row_vals.append(it.text() if it else "")
            ws.append(row_vals)
        
        ws.column_dimensions['A'].width = 20
        for col in ['B','C','D','E','F','G']: ws.column_dimensions[col].width = 14
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    def _export_cashflow(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        start_period = self.rep_start_period.currentData()
        end_period = self.rep_end_period.currentData()
        if not start_period or not end_period: return
        path, _ = QFileDialog.getSaveFileName(self, "保存",
            f"收支统计表_{end_period}.xlsx", "Excel(*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "收支统计表"
        
        # 表头
        ws['A1'] = self.client_name
        ws['A1'].font = XFont(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:F1')
        
        ws['A2'] = "收支统计表"
        ws['A2'].font = XFont(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:F2')
        
        period_text = f"期间：{start_period} 至 {end_period}"
        ws['A3'] = period_text
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:F3')
        
        # 空行
        ws['A4'] = ""
        
        # 数据表头
        hdrs = ["科目编号","科目名称","类型","本期借方","本期贷方","净额"]
        fill_hdr = PatternFill("solid", fgColor="1C2340")
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(5, ci, h)
            cell.font = XFont(bold=True, color="FFFFFF"); cell.fill = fill_hdr
            cell.alignment = Alignment(horizontal="center")
        
        # 数据行
        for ri in range(self.cf_tbl.rowCount()):
            row_vals = []
            for ci in range(6):
                it = self.cf_tbl.item(ri, ci)
                row_vals.append(it.text() if it else "")
            ws.append(row_vals)
        
        ws.column_dimensions['A'].width = 12; ws.column_dimensions['B'].width = 25
        for col in ['C','D','E','F']: ws.column_dimensions[col].width = 14
        wb.save(path); QMessageBox.information(self, "成功", f"已导出:\n{path}")

    # ══════════════════════════════════════════════════════════════════════════
    # PDF 导出（reportlab）
    # ══════════════════════════════════════════════════════════════════════════

    def _export_pdf(self):
        """根据当前 Tab 弹出保存对话框后导出 PDF。"""
        if not self.client_id:
            return
        if not self._check_pdf_deps():
            return
        tab = self._get_current_tab()
        if not tab:
            return
        _, end = self._get_periods()
        if not end:
            return
        safe = tab.replace("/", "_")
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 PDF", f"{safe}_{end}.pdf", "PDF 文件 (*.pdf)")
        if not path:
            return
        try:
            self._render_pdf(tab, path)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出 PDF 失败：\n{e}")
            return
        QMessageBox.information(self, "成功", f"已导出 PDF：\n{path}")

    def _get_periods(self):
        """返回 (start_period, end_period)；任一为空则返回 (None, None)。"""
        s = self.rep_start_period.currentData()
        e = self.rep_end_period.currentData()
        return (s, e) if (s and e) else (None, None)

    def _check_pdf_deps(self) -> bool:
        """检查 reportlab 是否可用；不可用时弹提示返回 False。"""
        try:
            import reportlab  # noqa: F401
            return True
        except ImportError:
            QMessageBox.warning(
                self, "缺少依赖",
                "导出 PDF 需要安装 reportlab 库。\n\n"
                "请在终端运行：\n  pip install reportlab\n\n安装后重启程序。",
            )
            return False

