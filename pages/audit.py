"""pages/audit.py — AuditPage — 操作审计日志"""
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, Signal, QTimer
from PySide6.QtGui import QColor, QFont, QBrush, QPalette

from db import get_db, log_action
from utils import lbl, sep, card, fmt_amt, NoScrollSpinBox, NoScrollDoubleSpinBox
from datetime import date
# openpyxl imported lazily inside each export function

class AuditPage(QWidget):
    """操作审计日志 — 记录所有关键操作，可导出为审计报告"""

    def __init__(self):
        super().__init__()
        self.client_id = None
        L = QVBoxLayout(self); L.setContentsMargins(24,20,24,20); L.setSpacing(14)

        hdr = QHBoxLayout()
        hdr.addWidget(lbl("操作审计日志", bold=True, size=18)); hdr.addStretch()
        b_exp = QPushButton("导出审计报告(Excel)"); b_exp.setObjectName("btn_outline")
        b_exp.clicked.connect(self._export)
        b_clr = QPushButton("清空日志"); b_clr.setObjectName("btn_red")
        b_clr.clicked.connect(self._clear)
        hdr.addWidget(b_exp); hdr.addWidget(b_clr); L.addLayout(hdr)

        info = QLabel("  记录所有凭证新增/审核/删除、期末结账、数据导入等操作，可作为内部审计依据。")
        info.setStyleSheet("background:#f6f8ff;color:#444;border-radius:6px;padding:8px 12px;font-size:12px;")
        L.addWidget(info)

        # Filter row
        fr = QHBoxLayout(); fr.setSpacing(10)
        self.action_filter = QComboBox()
        self.action_filter.addItems(["全部操作","新增凭证","编辑凭证","凭证审核:已审核","凭证审核:已拒绝",
                                      "凭证审核:待审核","批量导入凭证","期末结转","期间结账封账",
                                      "期间反结账","删除凭证"])
        self.action_filter.currentIndexChanged.connect(self.load)
        self.date_from = QDateEdit(); self.date_from.setDisplayFormat("yyyy-MM-dd")
        self.date_from.setDate(QDate.currentDate().addMonths(-3))
        self.date_from.setCalendarPopup(True)
        self.date_to = QDateEdit(); self.date_to.setDisplayFormat("yyyy-MM-dd")
        self.date_to.setDate(QDate.currentDate()); self.date_to.setCalendarPopup(True)
        self.date_from.dateChanged.connect(self.load)
        self.date_to.dateChanged.connect(self.load)
        fr.addWidget(lbl("操作类型:")); fr.addWidget(self.action_filter)
        fr.addWidget(lbl("从:")); fr.addWidget(self.date_from)
        fr.addWidget(lbl("到:")); fr.addWidget(self.date_to)
        b_q = QPushButton("查询"); b_q.setObjectName("btn_primary"); b_q.clicked.connect(self.load)
        fr.addWidget(b_q); fr.addStretch()
        L.addLayout(fr)

        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.tbl = QTableWidget(); self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows); self.tbl.setShowGrid(False)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setColumnCount(6)
        self.tbl.setHorizontalHeaderLabels(["时间","操作人","操作类型","对象类型","对象ID","详情"])
        hh = self.tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(5, QHeaderView.Stretch)
        self.tbl.setColumnWidth(0,155); self.tbl.setColumnWidth(1,80)
        self.tbl.setColumnWidth(2,130); self.tbl.setColumnWidth(3,80); self.tbl.setColumnWidth(4,80)
        vl.addWidget(self.tbl)

        # Summary bar
        self.summary_bar = QLabel("  共 0 条记录")
        self.summary_bar.setStyleSheet("color:#888;font-size:12px;padding:6px 12px;background:#fafafa;border-top:1px solid #f0f0f0;")
        vl.addWidget(self.summary_bar)
        L.addWidget(f)

    def set_client(self, client_id):
        self.client_id = client_id
        self.load()

    def load(self):
        conn = get_db(); c = conn.cursor()
        action_f = self.action_filter.currentText()
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to   = self.date_to.date().toString("yyyy-MM-dd") + " 23:59:59"

        where = ["created_at >= ? AND created_at <= ?"]
        params = [date_from, date_to]
        if self.client_id:
            where.append("(client_id=? OR client_id IS NULL)")
            params.append(self.client_id)
        if action_f != "全部操作":
            where.append("action=?"); params.append(action_f)

        c.execute(f"""SELECT created_at,operator,action,target_type,target_id,detail
            FROM audit_log WHERE {" AND ".join(where)} ORDER BY created_at DESC LIMIT 500""",
            params)
        rows = c.fetchall(); conn.close()

        action_colors = {
            "新增凭证":"#3d6fdb","编辑凭证":"#fa8c16",
            "凭证审核:已审核":"#52c41a","凭证审核:已拒绝":"#ff4d4f",
            "凭证审核:待审核":"#888","批量导入凭证":"#722ed1",
            "期末结转":"#eb2f96","期间结账封账":"#fa8c16",
            "期间反结账":"#13c2c2","删除凭证":"#ff4d4f",
        }

        self.tbl.setRowCount(len(rows))
        for i,r in enumerate(rows):
            self.tbl.setRowHeight(i,36)
            action = r[2]
            color  = action_colors.get(action,"#555")
            for j,(val,align) in enumerate([
                (r[0][:19],    Qt.AlignCenter),
                (r[1] or "",   Qt.AlignCenter),
                (action,       Qt.AlignCenter),
                (r[3] or "",   Qt.AlignCenter),
                (str(r[4] or ""), Qt.AlignCenter),
                (r[5] or "",   Qt.AlignLeft|Qt.AlignVCenter),
            ]):
                it = QTableWidgetItem(val); it.setTextAlignment(align)
                if j==2: it.setForeground(QColor(color))
                self.tbl.setItem(i,j,it)

        self.summary_bar.setText(f"  共 {len(rows)} 条记录")

    def _export(self):
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
        path,_ = QFileDialog.getSaveFileName(self,"保存审计报告",
            f"审计报告_{date.today()}.xlsx","Excel(*.xlsx)")
        if not path: return
        conn = get_db(); c = conn.cursor()
        where = "1=1"
        params = []
        if self.client_id:
            where = "client_id=?"; params = [self.client_id]
        c.execute(f"""SELECT created_at,operator,action,target_type,target_id,detail
            FROM audit_log WHERE {where} ORDER BY created_at DESC""", params)
        rows = c.fetchall(); conn.close()

        wb = openpyxl.Workbook(); ws = wb.active; ws.title="审计日志"
        hdrs = ["操作时间","操作人","操作类型","对象类型","对象ID","详情"]
        fill = PatternFill("solid", fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell = ws.cell(1,ci,h)
            cell.font=XFont(bold=True,color="FFFFFF"); cell.fill=fill
            cell.alignment=Alignment(horizontal="center")
        for r in rows:
            ws.append([r[0],r[1],r[2],r[3],r[4],r[5]])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22
        wb.save(path)
        QMessageBox.information(self,"成功",f"审计报告已导出:\n{path}")

    def _clear(self):
        if QMessageBox.question(self,"确认","清空全部审计日志？此操作不可恢复。",
                QMessageBox.Yes|QMessageBox.No) != QMessageBox.Yes: return
        conn = get_db()
        where = "client_id=?" if self.client_id else "1=1"
        params = [self.client_id] if self.client_id else []
        conn.execute(f"DELETE FROM audit_log WHERE {where}", params)
        conn.commit(); conn.close()
        self.load()


