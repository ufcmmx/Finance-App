"""pages/voucher.py — VoucherPage — 凭证录入与查询"""
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import Qt, QDate, Signal, QTimer
from PySide6.QtGui import QColor, QFont, QBrush, QPalette

from db import get_db, log_action
from utils import lbl, sep, card, fmt_amt, cn_amount, NoScrollSpinBox, NoScrollDoubleSpinBox
from dialogs import VoucherDialog, AccountInitDialog, AuxPage
# openpyxl imported lazily inside each export function

class VoucherPage(QWidget):
    """凭证管理 — 新增/查凭证/科目余额表/明细账"""

    def __init__(self):
        super().__init__()
        self.client_id = None; self.client_name = ""; self.period = ""
        L = QVBoxLayout(self); L.setContentsMargins(0,0,0,0); L.setSpacing(0)
        # Top toolbar (智一风格横向 tab)
        self.toolbar = QWidget()
        self.toolbar.setStyleSheet("background:#fff; border-bottom:1px solid #e8ecf2;")
        tl = QHBoxLayout(self.toolbar); tl.setContentsMargins(16,0,16,0); tl.setSpacing(0)
        self._tabs = []
        for name in ["新增凭证","查凭证","科目余额表","明细账","科目期初","辅助核算"]:
            b = QPushButton(name); b.setObjectName("nav_tab")
            b.setStyleSheet("""QPushButton{background:transparent;color:#888;border:none;
                padding:12px 16px;border-bottom:2px solid transparent;}
                QPushButton:hover{color:#3d6fdb;}
                QPushButton[active=true]{color:#3d6fdb;border-bottom:2px solid #3d6fdb;}""")
            b.clicked.connect(lambda _,n=name:self._switch_tab(n))
            tl.addWidget(b); self._tabs.append(b)
        tl.addStretch()
        self.client_lbl = lbl("", color="#3d6fdb", bold=True)
        self.period_combo = QComboBox(); self.period_combo.setMinimumWidth(110)
        self.period_combo.currentTextChanged.connect(self._on_period_change)
        tl.addWidget(self.client_lbl); tl.addSpacing(12); tl.addWidget(lbl("期间:"))
        tl.addWidget(self.period_combo)
        L.addWidget(self.toolbar)

        self.stack = QStackedWidget(); L.addWidget(self.stack)
        self._build_voucher_list(); self._build_balance(); self._build_ledger()
        self._aux_page = AuxPage()
        self.stack.addWidget(self._aux_page)   # index 3
        self._switch_tab("查凭证")

    def _switch_tab(self, name):
        mapping = {"新增凭证":None,"查凭证":0,"科目余额表":1,"明细账":2,
                   "科目期初":None,"辅助核算":3}
        for b in self._tabs:
            b.setProperty("active","true" if b.text()==name else "false")
            b.style().unpolish(b); b.style().polish(b)
        if name == "新增凭证": self._new_voucher()
        elif name == "科目期初": self._open_period_init()
        elif name == "辅助核算":
            self.stack.setCurrentIndex(3)
            if self.client_id: self._aux_page.set_client(self.client_id, self.period)
        elif mapping.get(name) is not None:
            self.stack.setCurrentIndex(mapping[name])
            if name == "查凭证": self._load_vouchers()
            elif name == "科目余额表": self._load_balance()
            elif name == "明细账": self._load_ledger()

    # ─ Voucher list ─
    def _build_voucher_list(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(10)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("凭证列表", bold=True, size=15))
        self.lock_lbl = QLabel("")  # shows 🔒 已封账 when period is closed
        hdr.addSpacing(12); hdr.addWidget(self.lock_lbl)
        hdr.addStretch()
        b_new = QPushButton("＋ 新增凭证"); b_new.setObjectName("btn_primary")
        b_new.clicked.connect(self._new_voucher)
        b_exp_doc = QPushButton("导出记账凭证(PDF)"); b_exp_doc.setObjectName("btn_outline")
        b_exp_doc.clicked.connect(self._export_voucher_pdf)
        b_exp = QPushButton("导出Excel"); b_exp.setObjectName("btn_outline")
        b_exp.clicked.connect(self._export_vouchers)
        hdr.addWidget(b_exp_doc); hdr.addWidget(b_exp); hdr.addWidget(b_new); L.addLayout(hdr)

        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.v_tbl = QTableWidget(); self.v_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.v_tbl.setSelectionBehavior(QTableWidget.SelectRows); self.v_tbl.setShowGrid(False)
        self.v_tbl.verticalHeader().setVisible(False)
        self.v_tbl.setColumnCount(7)
        self.v_tbl.setHorizontalHeaderLabels(["凭证号","日期","摘要","科目","借方合计","贷方合计","操作"])
        vh = self.v_tbl.horizontalHeader()
        vh.setSectionResizeMode(QHeaderView.Interactive)   # ALL columns user-draggable
        vh.setStretchLastSection(False)                    # don't auto-stretch last col
        vh.setMinimumSectionSize(60)
        self.v_tbl.setColumnWidth(0, 90)   # 凭证号
        self.v_tbl.setColumnWidth(1, 105)  # 日期
        self.v_tbl.setColumnWidth(2, 180)  # 摘要 — now Interactive, fully draggable
        self.v_tbl.setColumnWidth(3, 210)  # 科目
        self.v_tbl.setColumnWidth(4, 100)  # 借方合计
        self.v_tbl.setColumnWidth(5, 100)  # 贷方合计
        self.v_tbl.setColumnWidth(6, 340)  # 操作
        vl.addWidget(self.v_tbl); L.addWidget(f)
        self.stack.addWidget(w)

    def _load_vouchers(self):
        if not self.client_id: return
        conn = get_db(); c = conn.cursor()
        # Show lock banner if period is closed
        c.execute("SELECT is_closed FROM periods WHERE client_id=? AND period=?",
                  (self.client_id, self.period))
        row = c.fetchone()
        is_closed = bool(row and row["is_closed"])
        c.execute("""SELECT v.id,v.voucher_no,v.date,v.status,
            (SELECT summary FROM voucher_entries WHERE voucher_id=v.id ORDER BY line_no LIMIT 1) as summ,
            (SELECT group_concat(account_code || ' ' || account_name,'/') FROM voucher_entries WHERE voucher_id=v.id) as accts,
            (SELECT SUM(debit) FROM voucher_entries WHERE voucher_id=v.id) as td,
            (SELECT SUM(credit) FROM voucher_entries WHERE voucher_id=v.id) as tc
            FROM vouchers v WHERE v.client_id=? AND v.period=? ORDER BY v.voucher_no""",
                  (self.client_id, self.period))
        rows = c.fetchall(); conn.close()
        # Update period lock indicator in toolbar
        if hasattr(self, 'lock_lbl'):
            if is_closed:
                self.lock_lbl.setText("  🔒 已封账  ")
                self.lock_lbl.setStyleSheet("color:#ff4d4f;font-weight:bold;background:#fff1f0;border-radius:4px;padding:2px 6px;")
            else:
                self.lock_lbl.setText("")
                self.lock_lbl.setStyleSheet("")
        self.v_tbl.setRowCount(len(rows))
        for i,r in enumerate(rows):
            self.v_tbl.setRowHeight(i,46)
            status = r['status']
            status_color = {"待审核":"#fa8c16","已审核":"#52c41a","已拒绝":"#ff4d4f"}.get(status,'#888')
            no_w = QLabel(f"  {r['voucher_no']}  ")
            no_w.setStyleSheet("color:#3d6fdb;font-weight:bold;padding:0 8px;")
            no_w.setAlignment(Qt.AlignCenter)
            self.v_tbl.setCellWidget(i,0,no_w)
            for j,v in enumerate([r['date'],r['summ'] or '',r['accts'] or ''],1):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter)
                it.setData(Qt.UserRole,r['id']); self.v_tbl.setItem(i,j,it)
            for j,v in enumerate([fmt_amt(r['td']),fmt_amt(r['tc'])],4):
                it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                it.setForeground(QColor("#3d6fdb")); self.v_tbl.setItem(i,j,it)
            # Action cell — status pill + audit buttons + edit + delete
            bw = QWidget(); bl = QHBoxLayout(bw); bl.setContentsMargins(6,3,6,3); bl.setSpacing(5)
            s_lbl = QLabel(f" {status} ")
            s_lbl.setStyleSheet(
                f"color:{status_color};font-size:11px;"
                f"border:1px solid {status_color};border-radius:3px;padding:2px 6px;")
            bl.addWidget(s_lbl)
            if status == "待审核":
                b_ok = QPushButton("✓ 审核通过"); b_ok.setObjectName("btn_green"); b_ok.setFixedSize(88,28)
                b_ok.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"已审核"))
                b_no = QPushButton("✗ 拒绝"); b_no.setObjectName("btn_red"); b_no.setFixedSize(60,28)
                b_no.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"已拒绝"))
                bl.addWidget(b_ok); bl.addWidget(b_no)
            elif status == "已拒绝":
                b_re = QPushButton("↩ 重新提交"); b_re.setObjectName("btn_outline"); b_re.setFixedSize(88,28)
                b_re.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"待审核"))
                bl.addWidget(b_re)
            elif status == "已审核":
                b_un = QPushButton("↩ 撤销审核"); b_un.setObjectName("btn_gray"); b_un.setFixedSize(88,28)
                b_un.clicked.connect(lambda _,rid=r['id']:self._set_voucher_status(rid,"待审核"))
                bl.addWidget(b_un)
            b_edit = QPushButton("编辑"); b_edit.setObjectName("btn_outline"); b_edit.setFixedSize(68,28)
            b_del = QPushButton("删除"); b_del.setObjectName("btn_red"); b_del.setFixedSize(60,28)
            b_edit.clicked.connect(lambda _,rid=r['id']:self._edit_voucher(rid))
            b_del.clicked.connect(lambda _,rid=r['id']:self._del_voucher(rid))
            bl.addWidget(b_edit); bl.addWidget(b_del); bl.addStretch()
            self.v_tbl.setCellWidget(i,6,bw)

    def _set_voucher_status(self, vid, new_status):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT voucher_no, status, period FROM vouchers WHERE id=?", (vid,))
        v = c.fetchone()
        if not v: conn.close(); return
        # 已结账期间禁止变更状态
        if self._is_period_closed(v["period"]):
            conn.close()
            QMessageBox.warning(self,"期间已封账","该凭证所在期间已封账，禁止修改审核状态。\n如需修改请先执行反结账。"); return
        # 审核前检查借贷平衡
        if new_status == "已审核":
            c.execute("SELECT ABS(SUM(debit)-SUM(credit)) AS diff FROM voucher_entries WHERE voucher_id=?", (vid,))
            diff = c.fetchone()["diff"] or 0
            if diff > 0.005:
                conn.close()
                QMessageBox.warning(self,"借贷不平",f"该凭证借贷差额 {diff:.2f}，不能审核通过。\n请先编辑修正后再审核。"); return
        conn.execute("UPDATE vouchers SET status=? WHERE id=?", (new_status, vid))
        log_action(conn, self.client_id, f"凭证审核:{new_status}", "voucher", vid, f"状态变更为{new_status}")
        conn.commit(); conn.close()
        self._load_vouchers()

    def _new_voucher(self):
        if not self.client_id:
            QMessageBox.information(self,"提示","请先从客户列表选择一个客户进入账簿"); return
        if self._is_period_closed():
            QMessageBox.warning(self,"期间已封账","该期间已结账封账，禁止新增凭证。\n如需修改请到【期末结账】页面执行反结账。"); return
        d = VoucherDialog(self, self.client_id, self.period)
        if d.exec():
            self._switch_tab("查凭证")
            if getattr(d,'saved_and_new',False): self._new_voucher()

    def _edit_voucher(self, vid):
        if self._is_period_closed():
            QMessageBox.warning(self,"期间已封账","该期间已结账封账，禁止修改凭证。\n如需修改请到【期末结账】页面执行反结账。"); return
        d = VoucherDialog(self, self.client_id, self.period, vid)
        if d.exec(): self._load_vouchers()

    def _del_voucher(self, vid):
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT voucher_no, status, period FROM vouchers WHERE id=?", (vid,))
        v = c.fetchone(); conn.close()
        if not v: return
        # 已结账期间禁止删除
        if self._is_period_closed(v["period"]):
            QMessageBox.warning(self,"期间已封账","该凭证所在期间已封账，禁止删除。"); return
        # 已审核凭证需二次确认
        if v["status"] == "已审核":
            reply = QMessageBox.question(self, "⚠ 删除已审核凭证",
                f"凭证【{v['voucher_no']}】状态为【已审核】，删除将影响账务数据。\n\n确认要永久删除该凭证吗？",
                QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes: return
            # 记审计日志
            conn = get_db()
            log_action(conn, self.client_id, "删除已审核凭证", "voucher", vid,
                       f"凭证号:{v['voucher_no']} 期间:{v['period']}")
            conn.execute("DELETE FROM vouchers WHERE id=?", (vid,))
            conn.commit(); conn.close()
        else:
            if QMessageBox.question(self,"确认","删除该凭证？",
                    QMessageBox.Yes|QMessageBox.No) != QMessageBox.Yes: return
            conn = get_db()
            conn.execute("DELETE FROM vouchers WHERE id=?", (vid,))
            conn.commit(); conn.close()
        self._load_vouchers()

    # ─ Balance table (科目余额表) ─
    def _build_balance(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(10)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("科目余额表", bold=True, size=15)); hdr.addStretch()
        b_dl = QPushButton("下载"); b_dl.setObjectName("btn_outline"); b_dl.clicked.connect(self._export_balance)
        hdr.addWidget(b_dl); L.addLayout(hdr)
        # Filter row
        fr = QHBoxLayout(); fr.setSpacing(10)
        fr.addWidget(lbl("期间段:"))
        self.bal_start_period = QComboBox(); self.bal_start_period.setMinimumWidth(100)
        self.bal_end_period = QComboBox(); self.bal_end_period.setMinimumWidth(100)
        fr.addWidget(self.bal_start_period); fr.addWidget(lbl("至")); fr.addWidget(self.bal_end_period)
        b_refresh = QPushButton("刷新"); b_refresh.setObjectName("btn_primary"); b_refresh.clicked.connect(self._load_balance)
        fr.addWidget(b_refresh); fr.addSpacing(20)
        self.bal_aux = QCheckBox("辅助核算科目"); self.bal_detail = QCheckBox("明细科目")
        self.bal_zero = QCheckBox("0值科目")
        fr.addWidget(self.bal_aux); fr.addWidget(self.bal_detail); fr.addWidget(self.bal_zero)
        fr.addStretch(); L.addLayout(fr)
        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.bal_tbl = QTableWidget(); self.bal_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.bal_tbl.setShowGrid(True); self.bal_tbl.verticalHeader().setVisible(False)
        self.bal_tbl.setColumnCount(8)
        self.bal_tbl.setHorizontalHeaderLabels([
            "科目编号","科目名称","期初借方","期初贷方",
            "本期借方","本期贷方","期末借方","期末贷方"])
        self.bal_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        for ci in range(2,8): self.bal_tbl.setColumnWidth(ci,110)
        vl.addWidget(self.bal_tbl); L.addWidget(f)
        self.stack.addWidget(w)

    def _load_balance(self):
        if not self.client_id: return
        start_period = self.bal_start_period.currentData()
        end_period = self.bal_end_period.currentData()
        if not start_period or not end_period: return

        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM accounts WHERE client_id=? ORDER BY code", (self.client_id,))
        accts = {r['code']: dict(r) for r in c.fetchall()}
        # Aggregate voucher entries for period range
        c.execute("""SELECT e.account_code, SUM(e.debit) td, SUM(e.credit) tc
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period >= ? AND v.period <= ? GROUP BY e.account_code""",
                  (self.client_id, start_period, end_period))
        leaf_activity = {r['account_code']: (r['td'] or 0, r['tc'] or 0)
                         for r in c.fetchall()}
        conn.close()

        all_codes = set(accts.keys())
        # Identify leaf accounts (no children)
        leaf_codes = {c for c in all_codes
                      if not any(o != c and o.startswith(c + '.') for o in all_codes)}

        # ── Roll up opening balances from leaves to parents ──
        open_d = {code: accts[code]['opening_debit'] or 0 for code in all_codes}
        open_c = {code: accts[code]['opening_credit'] or 0 for code in all_codes}
        act_d  = {code: leaf_activity.get(code, (0, 0))[0] for code in all_codes}
        act_c  = {code: leaf_activity.get(code, (0, 0))[1] for code in all_codes}

        # Zero parent accounts, then bubble up from leaves
        for code in all_codes - leaf_codes:
            open_d[code] = open_c[code] = 0.0
            act_d[code]  = act_c[code]  = 0.0
        for code in sorted(leaf_codes, reverse=True):   # deepest first
            parts = code.split('.')
            for depth in range(1, len(parts)):
                parent = '.'.join(parts[:depth])
                if parent in all_codes:
                    open_d[parent] += open_d[code]
                    open_c[parent] += open_c[code]
                    act_d[parent]  += act_d[code]
                    act_c[parent]  += act_c[code]

        # ── Build display rows ──
        rows = []
        for code, a in sorted(accts.items()):
            is_leaf = code in leaf_codes
            od, oc = open_d[code], open_c[code]
            td, tc = act_d[code],  act_c[code]
            direction = a['direction']
            if direction == '借':
                end_d = od + td - tc; end_c = 0
                if end_d < 0: end_c = -end_d; end_d = 0
            else:
                end_c = oc + tc - td; end_d = 0
                if end_c < 0: end_d = -end_c; end_c = 0
            rows.append((code, a['name'], od, oc, td, tc, end_d, end_c, is_leaf))

        # Only sum leaf rows for the grand total (avoid double-counting)
        totals = [sum(r[i] for r in rows if r[8]) for i in range(2, 8)]

        self.bal_tbl.setRowCount(len(rows) + 1)
        for i, r in enumerate(rows):
            self.bal_tbl.setRowHeight(i, 36)
            is_leaf = r[8]
            for j, v in enumerate(r[:8]):
                text = v if j < 2 else fmt_amt(v)
                it = QTableWidgetItem(text)
                it.setTextAlignment(
                    Qt.AlignCenter if j < 2 else Qt.AlignRight | Qt.AlignVCenter)
                if j == 0:
                    it.setForeground(QColor("#3d6fdb" if is_leaf else "#888"))
                if not is_leaf:
                    # Parent rows: bold, light background
                    it.setBackground(QColor("#f9fafc"))
                    it.setFont(QFont("", weight=QFont.Bold))
                self.bal_tbl.setItem(i, j, it)
        # Grand total row (leaf sums only)
        self.bal_tbl.setRowHeight(len(rows), 38)
        it0 = QTableWidgetItem(""); it1 = QTableWidgetItem("合  计（末级合计）")
        it1.setFont(QFont("", weight=QFont.Bold))
        it0.setBackground(QColor("#f5f7fa")); it1.setBackground(QColor("#f5f7fa"))
        it1.setTextAlignment(Qt.AlignCenter)
        self.bal_tbl.setItem(len(rows), 0, it0)
        self.bal_tbl.setItem(len(rows), 1, it1)
        for j, v in enumerate(totals, 2):
            it = QTableWidgetItem(fmt_amt(v))
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it.setBackground(QColor("#f5f7fa"))
            it.setFont(QFont("", weight=QFont.Bold))
            self.bal_tbl.setItem(len(rows), j, it)

    # ─ Ledger (明细账) ─
    def _build_ledger(self):
        w = QWidget(); L = QVBoxLayout(w); L.setContentsMargins(20,14,20,14); L.setSpacing(10)
        hdr = QHBoxLayout()
        hdr.addWidget(lbl("明细账", bold=True, size=15)); hdr.addStretch()
        hdr.addWidget(lbl("期间段:"))
        self.ldg_start_period = QComboBox(); self.ldg_start_period.setMinimumWidth(100)
        self.ldg_end_period = QComboBox(); self.ldg_end_period.setMinimumWidth(100)
        hdr.addWidget(self.ldg_start_period); hdr.addWidget(lbl("至")); hdr.addWidget(self.ldg_end_period)
        hdr.addSpacing(10); hdr.addWidget(lbl("科目:"))
        self.ldg_acct = QComboBox(); self.ldg_acct.setMinimumWidth(220)
        b_query = QPushButton("查询"); b_query.setObjectName("btn_primary"); b_query.clicked.connect(self._load_ledger)
        hdr.addWidget(self.ldg_acct); hdr.addWidget(b_query)
        L.addLayout(hdr)
        f = card(); vl = QVBoxLayout(f); vl.setContentsMargins(0,0,0,0)
        self.ldg_tbl = QTableWidget(); self.ldg_tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.ldg_tbl.setShowGrid(True); self.ldg_tbl.verticalHeader().setVisible(False)
        self.ldg_tbl.setColumnCount(6)
        self.ldg_tbl.setHorizontalHeaderLabels(["日期","摘要","借方","贷方","方向","余额"])
        self.ldg_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        vl.addWidget(self.ldg_tbl); L.addWidget(f)
        self.stack.addWidget(w)

    def _load_ledger(self):
        if not self.client_id: return
        start_period = self.ldg_start_period.currentData()
        end_period = self.ldg_end_period.currentData()
        if not start_period or not end_period: return

        # Populate account combo
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT code,name FROM accounts WHERE client_id=? ORDER BY code",(self.client_id,))
        accts = c.fetchall()
        cur = self.ldg_acct.currentData()
        self.ldg_acct.clear()
        for a in accts:
            self.ldg_acct.addItem(f"{a['code']}  {a['name']}", a['code'])
        if cur:
            for i in range(self.ldg_acct.count()):
                if self.ldg_acct.itemData(i)==cur: self.ldg_acct.setCurrentIndex(i); break

        sel_code = self.ldg_acct.currentData()
        if not sel_code: conn.close(); return
        c.execute("SELECT opening_debit,opening_credit,direction FROM accounts WHERE client_id=? AND code=?",
                  (self.client_id, sel_code))
        acct = c.fetchone()
        c.execute("""SELECT v.date,e.summary,e.debit,e.credit FROM voucher_entries e
            JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period >= ? AND v.period <= ? AND e.account_code=? ORDER BY v.date,v.voucher_no,e.line_no""",
                  (self.client_id, start_period, end_period, sel_code))
        entries = c.fetchall(); conn.close()

        direction = acct['direction'] if acct else '借'
        balance = (acct['opening_debit'] or 0) - (acct['opening_credit'] or 0) if acct else 0

        rows = [("期初余额","",0,0,direction,balance)]
        for e in entries:
            d = e['debit'] or 0; cr = e['credit'] or 0
            balance += d - cr
            dir_str = "借" if balance >= 0 else "贷"
            rows.append((e['date'],e['summary'] or '',d,cr,dir_str,abs(balance)))
        rows.append(("本期合计","",sum(r[2] for r in rows[1:]),sum(r[3] for r in rows[1:]),"",None))

        self.ldg_tbl.setRowCount(len(rows))
        for i,(dt,summary,d,cr,dirr,bal) in enumerate(rows):
            self.ldg_tbl.setRowHeight(i,36)
            is_header = dt in ("期初余额","本期合计")
            vals = [dt,summary,fmt_amt(d),fmt_amt(cr),dirr,
                    fmt_amt(bal) if bal is not None else ""]
            for j,v in enumerate(vals):
                it = QTableWidgetItem(v)
                it.setTextAlignment(Qt.AlignCenter if j<2 else Qt.AlignRight|Qt.AlignVCenter)
                if is_header:
                    it.setBackground(QColor("#f5f7fa")); it.setFont(QFont("",weight=QFont.Bold))
                if j==2 and d: it.setForeground(QColor("#3d6fdb"))
                if j==3 and cr: it.setForeground(QColor("#e05252"))
                self.ldg_tbl.setItem(i,j,it)

    def _open_period_init(self):
        if not self.client_id:
            QMessageBox.information(self,"提示","请先选择客户"); return
        d = AccountInitDialog(self, self.client_id, self.period)
        d.exec()

    # ─ Aux report (往来对账) ─
    def set_client(self, client_id, client_name, period):
        self.client_id = client_id; self.client_name = client_name; self.period = period
        self.client_lbl.setText(f"【{client_name}】")
        self._refresh_periods()
        # Initialize period ranges for balance and ledger
        self._init_period_ranges()
        # Auto refresh current tab
        idx = self.stack.currentIndex()
        if idx == 0: self._load_vouchers()
        elif idx == 1: self._load_balance()
        elif idx == 2: self._load_ledger()
        elif idx == 3 and self.client_id: self._aux_page.set_client(self.client_id, self.period)

    def _is_period_closed(self, period=None):
        """Return True if the given (or current) period is closed."""
        p = period or self.period
        if not self.client_id or not p: return False
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT is_closed FROM periods WHERE client_id=? AND period=?",
                  (self.client_id, p))
        row = c.fetchone(); conn.close()
        return bool(row and row["is_closed"])

    def _refresh_periods(self):
        self.period_combo.blockSignals(True)
        self.period_combo.clear()
        now = datetime.now()
        for y in range(now.year, 2018-1, -1):
            for m in range(12,0,-1):
                self.period_combo.addItem(f"{y}年{m:02d}期", f"{y}-{m:02d}")
        # Select current
        target = f"{now.year}-{now.month:02d}"
        for i in range(self.period_combo.count()):
            if self.period_combo.itemData(i)==target:
                self.period_combo.setCurrentIndex(i); break
        self.period = target
        self.period_combo.blockSignals(False)

    def _init_period_ranges(self):
        """Initialize period range selectors for balance and ledger pages"""
        # Balance page periods
        self.bal_start_period.clear()
        self.bal_end_period.clear()
        now = datetime.now()
        periods = []
        for y in range(now.year, 2018-1, -1):
            for m in range(12,0,-1):
                period_str = f"{y}-{m:02d}"
                display_str = f"{y}年{m:02d}期"
                periods.append((period_str, display_str))

        for period_str, display_str in periods:
            self.bal_start_period.addItem(display_str, period_str)
            self.bal_end_period.addItem(display_str, period_str)

        # Set default to current period
        current_period = f"{now.year}-{now.month:02d}"
        for i in range(self.bal_start_period.count()):
            if self.bal_start_period.itemData(i) == current_period:
                self.bal_start_period.setCurrentIndex(i)
                self.bal_end_period.setCurrentIndex(i)
                break

        # Ledger page periods
        self.ldg_start_period.clear()
        self.ldg_end_period.clear()
        for period_str, display_str in periods:
            self.ldg_start_period.addItem(display_str, period_str)
            self.ldg_end_period.addItem(display_str, period_str)

        # Set default to current period
        for i in range(self.ldg_start_period.count()):
            if self.ldg_start_period.itemData(i) == current_period:
                self.ldg_start_period.setCurrentIndex(i)
                self.ldg_end_period.setCurrentIndex(i)
                break

    def _on_period_change(self):
        self.period = self.period_combo.currentData() or self.period
        idx = self.stack.currentIndex()
        if idx==0: self._load_vouchers()
        elif idx==1: self._load_balance()
        elif idx==2: self._load_ledger()

    def _export_vouchers(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        path,_ = QFileDialog.getSaveFileName(self,"保存","凭证汇总.xlsx","Excel(*.xlsx)")
        if not path: return
        conn = get_db(); c = conn.cursor()
        c.execute("""SELECT v.voucher_no,v.date,e.summary,e.account_code,e.account_name,e.debit,e.credit
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period=? ORDER BY v.voucher_no,e.line_no""",
                  (self.client_id,self.period))
        rows = c.fetchall(); conn.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="凭证汇总"
        hdrs = ["凭证号","日期","摘要","科目编码","科目名称","借方","贷方"]
        fill = PatternFill("solid",fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell = ws.cell(1,ci,h)
            cell.font=XFont(bold=True,color="FFFFFF"); cell.fill=fill
            cell.alignment=Alignment(horizontal="center")
        for r in rows:
            ws.append([r['voucher_no'],r['date'],r['summary'],r['account_code'],
                       r['account_name'],r['debit'] or 0,r['credit'] or 0])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=16
        wb.save(path); QMessageBox.information(self,"成功",f"已导出:\n{path}")

    def _export_voucher_pdf(self):
        """导出记账凭证 PDF（A4 横版单凭证格式）"""
        if not self.client_id: return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib import colors as rl_colors
        except ImportError:
            QMessageBox.warning(self, "缺少依赖",
                "导出PDF需要安装 reportlab 库。\n\n请在终端运行：\npip install reportlab\n\n安装后重启程序。")
            return

        # 优先使用系统 TTF 字体（间距正常），CID 字体作为兜底
        FONT = None
        ttf_candidates = [
            ('SongtiSC', '/System/Library/Fonts/Supplemental/Songti.ttc', 0),
            ('PingFangSC', '/System/Library/Fonts/PingFang.ttc', 0),
            ('STHeitiSC', '/System/Library/Fonts/STHeiti Light.ttc', 0),
        ]
        for font_name, font_path, sub_idx in ttf_candidates:
            try:
                import os
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont(font_name, font_path, subfontIndex=sub_idx))
                    FONT = font_name
                    break
            except Exception:
                continue
        if not FONT:
            try:
                pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
                FONT = 'STSong-Light'
            except Exception as e:
                QMessageBox.warning(self, "字体错误", f"无法加载中文字体：{e}")
                return

        period_text = (self.period or "").replace("-", "年", 1) + "期"
        path, _ = QFileDialog.getSaveFileName(
            self, "保存记账凭证", f"记账凭证_A4横版_{period_text}.pdf", "PDF文件 (*.pdf)")
        if not path: return

        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT name FROM clients WHERE id=?", (self.client_id,))
        row = cur.fetchone()
        company_name = row['name'] if row else ''

        cur.execute("""SELECT id, voucher_no, date, status, attachment_count, preparer
            FROM vouchers WHERE client_id=? AND period=?
            ORDER BY date, voucher_no, id""", (self.client_id, self.period))
        vouchers = cur.fetchall()

        if not vouchers:
            conn.close()
            QMessageBox.information(self, "提示", "当前账期没有可导出的凭证")
            return

        MAX_ROWS = 5

        # ── 构建分页后的凭证页列表 ──
        pages = []
        for v in vouchers:
            cur.execute("""SELECT line_no, summary, account_code, account_name, debit, credit
                FROM voucher_entries WHERE voucher_id=? ORDER BY line_no""", (v['id'],))
            entries = list(cur.fetchall())
            total_debit  = sum(e['debit']  or 0 for e in entries)
            total_credit = sum(e['credit'] or 0 for e in entries)
            chunks = [entries[i:i+MAX_ROWS] for i in range(0, max(len(entries), 1), MAX_ROWS)]
            if not chunks: chunks = [[]]
            for idx, chunk in enumerate(chunks):
                pages.append(dict(
                    voucher=dict(v),
                    entries=chunk,
                    page_no=idx + 1,
                    total_pages=len(chunks),
                    total_debit=total_debit,
                    total_credit=total_credit,
                    company=company_name,
                ))
        conn.close()

        # ── PDF 参数 ──
        page_w, page_h = landscape(A4)
        cv = rl_canvas.Canvas(path, pagesize=(page_w, page_h))
        margin_left = 42
        margin_right = 42
        margin_top = 34
        margin_bottom = 32
        content_w = page_w - margin_left - margin_right
        table_top = page_h - 170
        table_bottom = margin_bottom + 42
        table_h = table_top - table_bottom

        def wrap_text(text, font_name, font_size, max_width, max_lines=2):
            text = str(text or "").strip()
            if not text:
                return [""]
            lines = []
            current = ""
            for ch in text:
                probe = current + ch
                if pdfmetrics.stringWidth(probe, font_name, font_size) <= max_width:
                    current = probe
                else:
                    if current:
                        lines.append(current)
                    current = ch
                    if len(lines) >= max_lines - 1:
                        break
            if len(lines) < max_lines and current:
                lines.append(current)
            leftover = text[len("".join(lines)):]
            if leftover and lines:
                tail = lines[-1]
                while tail and pdfmetrics.stringWidth(tail + "…", font_name, font_size) > max_width:
                    tail = tail[:-1]
                lines[-1] = (tail or "") + "…"
            return lines[:max_lines]

        def draw_cell_text(x, y, w, h, text="", fs=10, align='left', bold=False, max_lines=2):
            if not text:
                return
            font_name = FONT
            cv.setFont(font_name, fs)
            inner_w = max(w - 10, 20)
            lines = wrap_text(text, font_name, fs, inner_w, max_lines=max_lines)
            line_gap = fs + 2
            total_h = line_gap * len(lines)
            base_y = y + (h - total_h) / 2 + (len(lines) - 1) * line_gap
            for idx, line in enumerate(lines):
                line_y = base_y - idx * line_gap
                if align == 'center':
                    cv.drawCentredString(x + w / 2, line_y, line)
                elif align == 'right':
                    cv.drawRightString(x + w - 14, line_y, line)
                else:
                    cv.drawString(x + 8, line_y, line)

        def draw_page(sd):
            cv.setStrokeColor(rl_colors.black)
            cv.setFillColor(rl_colors.black)
            cv.setLineWidth(1)


            # 标题区
            cv.setFont(FONT, 20)
            cv.drawCentredString(page_w / 2, page_h - 88, "记账凭证")

            attach = sd['voucher'].get('attachment_count') or 0
            cv.setFont(FONT, 11)
            cv.drawRightString(page_w - margin_right, page_h - 86, f"附单据数：{attach}")

            company_y = page_h - 136
            date_y = company_y
            no_y = company_y
            cv.setFont(FONT, 11)
            cv.drawString(margin_left, company_y, f"核算单位：{sd['company']}")
            cv.drawCentredString(page_w / 2, date_y, f"日期：{sd['voucher']['date']}")
            vno = sd['voucher']['voucher_no']
            page_no = sd['page_no']
            total_pages = sd['total_pages']
            vno_text = f"{vno}（{page_no}/{total_pages}）" if total_pages > 1 else vno
            cv.drawRightString(page_w - margin_right, no_y, f"凭证号： {vno_text}")

            # 主表格
            col_widths = [content_w * ratio for ratio in (0.28, 0.36, 0.18, 0.18)]
            col_x = [margin_left]
            for width in col_widths[:-1]:
                col_x.append(col_x[-1] + width)
            header_h = 52
            total_h_row = 48
            body_h = table_h - header_h - total_h_row
            row_h = body_h / MAX_ROWS
            y_header_bottom = table_top - header_h

            # ── 先画背景填充（在网格线之下） ──
            cv.setFillColor(rl_colors.Color(0.92, 0.93, 0.95))
            cv.rect(margin_left, y_header_bottom, content_w, header_h, stroke=0, fill=1)   # 表头背景
            cv.rect(margin_left, table_bottom, content_w, total_h_row, stroke=0, fill=1)    # 合计行背景
            cv.setFillColor(rl_colors.black)

            # ── 再画网格线（在背景之上） ──
            cv.rect(margin_left, table_bottom, content_w, table_h, stroke=1, fill=0)
            x_cursor = margin_left
            for col_idx, width in enumerate(col_widths[:-1]):
                x_cursor += width
                if col_idx == 0:
                    # 第一根竖线：合计行不画（让合计文字跨两列）
                    cv.line(x_cursor, table_bottom + total_h_row, x_cursor, table_top)
                else:
                    cv.line(x_cursor, table_bottom, x_cursor, table_top)

            cv.line(margin_left, y_header_bottom, margin_left + content_w, y_header_bottom)
            for row_idx in range(MAX_ROWS - 1):
                y_line = y_header_bottom - (row_idx + 1) * row_h
                cv.line(margin_left, y_line, margin_left + content_w, y_line)
            cv.line(margin_left, table_bottom + total_h_row, margin_left + content_w, table_bottom + total_h_row)

            # ── 表头文字 ──
            headers = ["摘要", "科目", "借方", "贷方"]
            for idx, title in enumerate(headers):
                draw_cell_text(col_x[idx], y_header_bottom, col_widths[idx], header_h,
                               title, fs=16, align='center', bold=True, max_lines=1)

            entries = sd['entries']
            for idx in range(MAX_ROWS):
                row_bottom = table_bottom + total_h_row + (MAX_ROWS - 1 - idx) * row_h
                if idx >= len(entries):
                    continue
                entry = entries[idx]
                account_text = " ".join(
                    part for part in [entry['account_code'] or '', entry['account_name'] or ''] if part
                )
                draw_cell_text(col_x[0], row_bottom, col_widths[0], row_h,
                               entry['summary'] or '', fs=12, align='left', max_lines=2)
                draw_cell_text(col_x[1], row_bottom, col_widths[1], row_h,
                               account_text, fs=12, align='left', max_lines=2)
                draw_cell_text(col_x[2], row_bottom, col_widths[2], row_h,
                               fmt_amt(entry['debit']) if entry['debit'] else '',
                               fs=12, align='right', max_lines=1)
                draw_cell_text(col_x[3], row_bottom, col_widths[3], row_h,
                               fmt_amt(entry['credit']) if entry['credit'] else '',
                               fs=12, align='right', max_lines=1)

            td = sd['total_debit']
            tc = sd['total_credit']
            amount_cn = cn_amount(td) if td else "零元整"
            # ── 合计行文字 ──
            draw_cell_text(margin_left, table_bottom, col_widths[0] + col_widths[1], total_h_row,
                           f"合计： {amount_cn}", fs=14, align='left', bold=True, max_lines=1)
            draw_cell_text(col_x[2], table_bottom, col_widths[2], total_h_row,
                           fmt_amt(td) if td else "0.00", fs=14, align='right', bold=True, max_lines=1)
            draw_cell_text(col_x[3], table_bottom, col_widths[3], total_h_row,
                           fmt_amt(tc) if tc else "0.00", fs=14, align='right', bold=True, max_lines=1)

            # 底部签字区与裁切线
            preparer = sd['voucher'].get('preparer') or ""
            footer_y = margin_bottom + 10
            cv.setFont(FONT, 12)
            cv.drawString(margin_left, footer_y, f"记账：{preparer}")
            cv.drawString(page_w * 0.36, footer_y, "审核：")
            cv.drawString(page_w * 0.54, footer_y, f"制单：{preparer}")


        # ── 分页输出 ──
        for idx, page in enumerate(pages):
            if idx > 0:
                cv.showPage()
            draw_page(page)

        cv.save()
        QMessageBox.information(self, "成功", f"已导出记账凭证 PDF：\n{path}")


    def _export_balance(self):
        if not self.client_id: return
        import openpyxl
        from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side

        path,_ = QFileDialog.getSaveFileName(self,"保存",f"科目余额表_{self.period}.xlsx","Excel(*.xlsx)")
        if not path: return
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM accounts WHERE client_id=? ORDER BY code",(self.client_id,))
        accts = {r['code']:dict(r) for r in c.fetchall()}
        c.execute("""SELECT e.account_code,SUM(e.debit),SUM(e.credit)
            FROM voucher_entries e JOIN vouchers v ON v.id=e.voucher_id
            WHERE v.client_id=? AND v.period=? GROUP BY e.account_code""",
                  (self.client_id,self.period))
        activity = {r[0]:(r[1] or 0,r[2] or 0) for r in c.fetchall()}
        conn.close()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="科目余额表"
        hdrs = ["科目编号","科目名称","期初借方","期初贷方","本期借方","本期贷方","期末借方","期末贷方"]
        fill = PatternFill("solid",fgColor="1C2340")
        for ci,h in enumerate(hdrs,1):
            cell = ws.cell(1,ci,h); cell.font=XFont(bold=True,color="FFFFFF")
            cell.fill=fill; cell.alignment=Alignment(horizontal="center")
        for code,a in sorted(accts.items()):
            od=a['opening_debit'] or 0; oc=a['opening_credit'] or 0
            td,tc=activity.get(code,(0,0))
            if a['direction']=='借': end_d=max(0,od+td-tc); end_c=0
            else: end_c=max(0,oc+tc-td); end_d=0
            ws.append([code,a['name'],od,oc,td,tc,end_d,end_c])
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=14
        wb.save(path); QMessageBox.information(self,"成功",f"已导出:\n{path}")